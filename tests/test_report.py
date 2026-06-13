"""
tests/test_report.py — 다중시트 Report 오케스트레이터 (B).

실행: py -m unittest tests.test_report

검증: 단일시트 sheet-builder 들을 묶어 다중시트 제본 + 크로스시트 tie(메모리값) +
폴리시(표지/목차/탭색/하이퍼링크) + n=1 하위호환. 합성 재무수치 없음(구조 더미).
"""
from __future__ import annotations

import os
import tempfile
import unittest

import fpna._bootstrap  # noqa: F401

from openpyxl import load_workbook

from fpna import house_style as hs
from fpna.conserve import ConserveSpec
from fpna.report import ReportSpec, SheetSpec, build_report, qc_report


# --- 네이티브 sheet-builder (구조 더미) --------------------------------------
_LINES = [("A", 300.0), ("B", 250.0), ("C", 450.0)]   # 합 1000 (의미없는 더미)


def _detail_builder(ws, ctx):
    r = hs.report_frame(ws, "상세 (Detail)", unit="₩", last_col=2)
    total = 0.0
    for name, val in _LINES:
        hs.set_cell(ws, r, 1, name, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, val, role="calc", number_format=hs.FMT_INT)
        total += val
        r += 1
    hs.set_cell(ws, r, 1, "합계", role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 2, total, role="total", number_format=hs.FMT_INT)
    return {"total": total, "n": len(_LINES)}


def _summary_builder(ws, ctx, reported=1000.0):
    r = hs.report_frame(ws, "요약 (Summary)", unit="₩", last_col=2)
    hs.set_cell(ws, r, 1, "총 고정비", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, reported, role="calc", number_format=hs.FMT_INT)
    return {"total": reported}


def _make_spec(summary_reported=1000.0):
    return ReportSpec(
        title="고정비 보드팩 (데모)",
        subtitle="다중시트 워크페이퍼",
        as_of="2026-05-31",
        sheets=[
            SheetSpec("요약", lambda ws, ctx: _summary_builder(ws, ctx, summary_reported),
                      section="summary", title="요약"),
            SheetSpec("상세", _detail_builder, section="detail", title="상세 명세"),
        ],
        # 크로스시트 tie = ConserveSpec 재사용. source = 평탄 facts.
        cross_specs=[ConserveSpec("요약 == 상세 합", raw_sum_fn=lambda f: f["상세.total"],
                                  reported_key="요약.total")],
    )


class ReportBuildTest(unittest.TestCase):
    def test_sheets_and_policy(self):
        wb = build_report(_make_spec())
        titles = [ws.title for ws in wb.worksheets]
        # 표지·목차·요약·상세·검증 모두 존재
        for t in ["표지", "목차", "요약", "상세", "검증"]:
            self.assertTrue(any(x.startswith(t) for x in titles), "%s 시트 없음: %s" % (t, titles))
        # 표지 최상단, 검증 마지막
        self.assertTrue(titles[0].startswith("표지"))
        self.assertTrue(titles[-1].startswith("검증"))

    def test_tab_colors(self):
        wb = build_report(_make_spec())
        cover = next(ws for ws in wb.worksheets if ws.title.startswith("표지"))
        self.assertEqual(str(cover.sheet_properties.tabColor.rgb)[-6:], "2E5A87")

    def test_toc_hyperlinks(self):
        wb = build_report(_make_spec())
        toc = next(ws for ws in wb.worksheets if ws.title.startswith("목차"))
        links = [c.hyperlink.target if c.hyperlink else None
                 for row in toc.iter_rows() for c in row if c.hyperlink]
        self.assertTrue(any(l and l.startswith("#'") for l in links), "내부 하이퍼링크 없음")

    def test_cross_tie_passes_clean(self):
        rep = qc_report(build_report(_make_spec(1000.0)), _make_spec(1000.0))
        self.assertTrue(rep.passed, rep.summary())

    def test_cross_tie_blocks_mismatch(self):
        """요약 보고값을 상세 합과 다르게 → 크로스시트 tie FAIL."""
        spec = _make_spec(summary_reported=1234.0)   # 상세 합 1000 ≠ 1234
        rep = qc_report(build_report(spec), spec)
        self.assertFalse(rep.passed)

    def test_check_sheet_memory_bool(self):
        """검증 시트가 수식이 아니라 메모리값 bool(OK/XX)을 기록."""
        wb = build_report(_make_spec(1000.0))
        chk = next(ws for ws in wb.worksheets if ws.title.startswith("검증"))
        vals = [c.value for row in chk.iter_rows() for c in row if c.value in ("OK", "XX")]
        self.assertIn("OK", vals)

    def test_roundtrip_save(self):
        wb = build_report(_make_spec())
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "pack.xlsx")
            wb.save(p)
            wb2 = load_workbook(p)
            self.assertTrue(any(ws.title.startswith("검증") for ws in wb2.worksheets))


class ReportN1Test(unittest.TestCase):
    def test_single_sheet_report(self):
        """n=1 = 단일 데이터시트(기존 단일시트 경로 특수케이스)."""
        spec = ReportSpec(title="단일", sheets=[SheetSpec("상세", _detail_builder, "detail")])
        wb = build_report(spec)
        # 표지/목차/검증 + 상세 1 = 데이터시트 1개
        data_sheets = [ws.title for ws in wb.worksheets
                       if not ws.title.startswith(("표지", "목차", "검증"))]
        self.assertEqual(len(data_sheets), 1)


# --- R3: B 실행경로(레지스트리 경유 + main.py cmd) ----------------------------
class BoardpackRegistryTest(unittest.TestCase):
    def test_registry_build_cross_tie_passes(self):
        """레지스트리 get_report → make_spec → build_report → qc_report PASS + 검증시트."""
        from fpna.reports import get_report, available
        self.assertIn("fc_boardpack", available())
        spec = get_report("fc_boardpack").make_spec()
        wb = build_report(spec)
        rep = qc_report(wb, spec)
        self.assertTrue(rep.passed, rep.summary())
        # 4시트(표지/요약/상세/검증) — 검증 시트 OK 기록
        chk = next(ws for ws in wb.worksheets if ws.title.startswith("검증"))
        vals = [c.value for row in chk.iter_rows() for c in row if c.value in ("OK", "XX")]
        self.assertIn("OK", vals)
        self.assertNotIn("XX", vals)
        for t in ["표지", "목차", "요약", "상세", "검증"]:
            self.assertTrue(any(x.startswith(t) for x in [w.title for w in wb.worksheets]))

    def test_unknown_report_raises(self):
        from fpna.reports import get_report
        with self.assertRaises(KeyError):
            get_report("nope")

    def test_main_cmd_report_smoke(self):
        """main.cmd_report → rc=0 + 파일 생성(스파인 동등 저장)."""
        from main import cmd_report
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "_p.xlsx")
            rc = cmd_report(["fc_boardpack", p])
            self.assertEqual(rc, 0)
            self.assertTrue(os.path.isfile(p))
            wb = load_workbook(p)
            self.assertTrue(any(w.title.startswith("검증") for w in wb.worksheets))

    def test_main_cmd_report_bad_name(self):
        from main import cmd_report
        self.assertEqual(cmd_report(["no_such_report", "out/x.xlsx"]), 2)


# --- R4: DimContext 주입 + grain 정합 검증 ------------------------------------
class DimContextTest(unittest.TestCase):
    def test_ctx_injected_as_dimcontext(self):
        """build_report 가 ctx=None 일 때 builder 에 DimContext 인스턴스 주입(패스스루 X)."""
        from fpna.report import DimContext
        seen = {}

        def _probe(ws, ctx):
            seen["ctx"] = ctx
            hs.report_frame(ws, "probe", unit="₩", last_col=2)
            return {"total": 0.0}

        spec = ReportSpec(title="ctx", sheets=[
            SheetSpec("상세", _probe, "detail",
                      grain={"accounts": ("6010", "6020")})])
        build_report(spec)
        self.assertIsInstance(seen["ctx"], DimContext)
        # grain 합집합이 축에 흡수됨
        self.assertEqual(seen["ctx"].accounts, ("6010", "6020"))

    def test_from_spec_union(self):
        from fpna.report import DimContext
        spec = ReportSpec(title="u", sheets=[
            SheetSpec("a", _detail_builder, "detail", grain={"accounts": ("6010",)}),
            SheetSpec("b", _detail_builder, "detail", grain={"accounts": ("6020",)}),
        ])
        ctx = DimContext.from_spec(spec)
        self.assertEqual(ctx.accounts, ("6010", "6020"))
        self.assertEqual(ctx.periods, ())

    def test_grain_conflict_fails_qc(self):
        """공통 축에 키 교집합 0인 시트 쌍 → qc FAIL(시트간 grain 불일치)."""
        spec = ReportSpec(title="conflict", sheets=[
            SheetSpec("요약", lambda ws, ctx: _summary_builder(ws, ctx), "summary",
                      grain={"periods": ("2024-01",)}),
            SheetSpec("상세", _detail_builder, "detail",
                      grain={"periods": ("2025-01",)}),   # 같은 축, 키 공유 0
        ])
        rep = qc_report(build_report(spec), spec)
        self.assertFalse(rep.passed)
        self.assertTrue(any("grain" in r[0] and not r[1] for r in rep.checks)
                        if hasattr(rep, "checks") else True, rep.summary())

    def test_grain_compatible_passes(self):
        """공통 축 키 교집합 존재 → grain 정합 PASS."""
        spec = _make_spec(1000.0)
        spec.sheets[0].grain = {"periods": ("2024-01",)}
        spec.sheets[1].grain = {"periods": ("2024-01", "2024-02")}
        rep = qc_report(build_report(spec), spec)
        self.assertTrue(rep.passed, rep.summary())


# --- dispatcher: 보드팩 → stage="report" -------------------------------------
class StageReportRouteTest(unittest.TestCase):
    def test_classify_stage_boardpack(self):
        from fpna.dispatcher import classify_stage
        for txt in ["보드팩 만들어", "제본 보고서", "다중시트 워크페이퍼"]:
            stage, cmd = classify_stage(txt)
            self.assertEqual(stage, "report", txt)
            self.assertIn("report fc_boardpack", cmd)

    def test_route_report(self):
        from fpna.dispatcher import route
        r = route("보드팩 만들어줘")
        self.assertEqual(r["stage"], "report")
        self.assertIsNone(r["template"])


if __name__ == "__main__":
    unittest.main()
