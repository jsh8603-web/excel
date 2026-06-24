#!/usr/bin/env python3
"""tests/test_design_zones.py — 한 시트 내 정형블록+freehand 혼합 영역 게이트 회귀.

구조 픽스처(의미없는 더미 숫자 — 합성 재무수치 아님): 한 시트에 라인아이템 행을
좌우로 ACTUAL/FCST 두 블록이 공유하는 2-D 레이아웃. 마커 2트랙(숨김 값) + 좌표 free 계약.
"""
import unittest

import fpna._bootstrap  # noqa: F401  vendor 주입
import openpyxl
from openpyxl.styles import Font

from fpna import house_style as hs
from fpna import design_zones as dz
from fpna import design_audit as da
from fpna.templates.base import QCReport


def _fixture(first_band_row=6):
    """anchor (4,1). row-band 'li' 시작=first_band_row. col-band actual(2-3)/fcst(4-5).
    데이터 B{br}:E{br+2} 를 set_cell(role=calc)로 채움."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"
    br = first_band_row
    # 데이터(좌우 병렬, 행 공유) — 더미 숫자
    for i, r in enumerate(range(br, br + 3)):
        for c in range(2, 6):
            hs.set_cell(ws, r, c, (i + 1) * c, role="calc", number_format=hs.FMT_INT)
    dz.stamp_zone(
        ws,
        origin=(br, 2),
        row_bands=[(br, "li")],                 # 라인아이템 밴드(공유)
        col_bands=[(2, "actual"), (4, "fcst")], # 좌우 파티션
        marker_col=1, marker_row=4,
    )
    return wb, ws, br


class ZoneMarkerTest(unittest.TestCase):
    def test_marker_roundtrip_and_resolve(self):
        wb, ws, br = _fixture()
        # 저장→재로딩(값 마커 왕복 보존)
        import io
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        wb2 = openpyxl.load_workbook(buf); ws2 = wb2["P&L"]
        row_map, col_map, anchor = dz.read_band_maps(ws2)
        self.assertIsNotNone(anchor)
        self.assertEqual(col_map[2], "actual")
        self.assertEqual(col_map[3], "actual")     # RLE 연속
        self.assertEqual(col_map[4], "fcst")
        self.assertEqual(col_map[5], "fcst")
        self.assertEqual(row_map[br], "li")
        cells, rm, cm = dz.resolve_blocks(ws2, {"blocks": {}, "band_map": {}})
        self.assertEqual(cells[(br, 2)], ("li", "actual", "li::actual"))
        self.assertEqual(cells[(br, 4)], ("li", "fcst", "li::fcst"))

    def test_clean_block_no_drift(self):
        wb, ws, br = _fixture()
        contract = dz.capture_contract(ws, {"actual": "calc", "fcst": "calc"})
        z = da.zone_findings(wb, contract)
        self.assertEqual(z["resolved_drift"], [], "갓 생성한 블록은 drift 0")
        self.assertEqual(z["unsealed"], [])
        rep = QCReport("test")
        self.assertTrue(da.assert_design_standard(rep, wb, contract=contract))

    def test_resolved_drift_detected(self):
        wb, ws, br = _fixture()
        contract = dz.capture_contract(ws, {"actual": "calc", "fcst": "calc"})
        ws.cell(br, 2).font = Font(name="Comic Sans MS", size=20)   # strict 셀 변조
        z = da.zone_findings(wb, contract)
        self.assertTrue(any(coord == ws.cell(br, 2).coordinate for (_t, coord, _b, _c) in z["resolved_drift"]),
                        "변조 셀이 resolved_drift 로 잡혀야")
        rep = QCReport("test")
        self.assertFalse(da.assert_design_standard(rep, wb, contract=contract), "drift 시 strict fail")

    def test_restyle_zone_repairs_value_preserved(self):
        wb, ws, br = _fixture()
        contract = dz.capture_contract(ws, {"actual": "calc", "fcst": "calc"})
        v0 = ws.cell(br, 2).value
        ws.cell(br, 2).font = Font(name="Comic Sans MS", size=20)
        actions = da.restyle_zone(wb, contract)
        self.assertTrue(any(a[0] == "retag" for a in actions))
        self.assertEqual(ws.cell(br, 2).value, v0, "수선은 값 불변")
        z = da.zone_findings(wb, contract)
        self.assertEqual(z["resolved_drift"], [], "수선 후 drift 0")

    def test_unsealed_above_band_start(self):
        # row-band 시작=6, 데이터 B5(마커행4 아래·밴드시작 위) → row 5 미커버 = unsealed
        wb, ws, br = _fixture(first_band_row=6)
        contract = dz.capture_contract(ws, {"actual": "calc", "fcst": "calc"})
        hs.set_cell(ws, 5, 2, 999, role="calc")     # bbox 위쪽 미경유 데이터
        z = da.zone_findings(wb, contract)
        self.assertTrue(any(coord == "B5" for (_t, coord) in z["unsealed"]),
                        "밴드시작 위 데이터셀 = unsealed loud")

    def test_contract_rejects_coordinates(self):
        with self.assertRaises(ValueError):
            dz.load_contract({"blocks": {"x": {"range": "A1:B2"}}})

    def test_legacy_audit_unaffected(self):
        # contract 미지정 = 기존 전체순회 동작 그대로(회귀 0)
        wb, ws, br = _fixture()
        rep = QCReport("test")
        ok = da.assert_design_standard(rep, wb)        # contract 없음
        self.assertTrue(ok)


if __name__ == "__main__":
    unittest.main()
