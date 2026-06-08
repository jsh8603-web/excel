"""
tests/make_fixtures.py — 정형화 테스트용 '누더기' 워크북 생성.

⚠ 실데이터 아님. ingest 파이프라인의 구조 처리(다중헤더/병합/단위/소계/센티넬)를
검증하기 위한 합성 *구조* 픽스처일 뿐, 재무 수치는 의미 없는 더미.
"""
from __future__ import annotations

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.styles import Alignment


def build_messy(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "messy"

    # 제목 행 + 단위 안내(비데이터 행)
    ws["A1"] = "월별 손익 요약"
    ws["A2"] = "(단위: 천원)"

    # 병합 2단 열헤더: 2024 / 2025 각각 매출·비용
    ws["B4"] = "2024"
    ws.merge_cells("B4:C4")
    ws["D4"] = "2025"
    ws.merge_cells("D4:E4")
    ws["B5"] = "매출"
    ws["C5"] = "비용"
    ws["D5"] = "매출"
    ws["E5"] = "비용"
    ws["A5"] = "계정"

    # 행헤더(들여쓰기 계층) + 데이터 + 센티넬 + 괄호음수 + 소계
    rows = [
        ("제품A", 1000, 600, 1200, 700, 0),
        ("제품B", 800, "(50)", 900, "-", 1),     # 괄호음수, 대시 센티넬
        ("소계", 1800, 550, 2100, 700, 0),         # 소계 행
        ("기타", "△30", 10, 40, 20, 0),            # 세모 음수
    ]
    r = 6
    for name, *vals in rows:
        indent = vals[-1]
        c = ws.cell(row=r, column=1, value=name)
        c.alignment = Alignment(indent=indent)
        for j, v in enumerate(vals[:-1], start=2):
            ws.cell(row=r, column=j, value=v)
        r += 1

    # 각주(비데이터 행)
    ws.cell(row=r + 1, column=1, value="주1) 더미 데이터임")

    # 두 번째 표(빈 영역으로 분리) — 단순 KPI
    ws["G4"] = "지표"
    ws["H4"] = "값"
    ws["G5"] = "가동률"
    ws["H5"] = "85%"
    ws["G6"] = "직원수"
    ws["H6"] = 42

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures", "messy_sample.xlsx")
    build_messy(out)
    print("wrote", out)
