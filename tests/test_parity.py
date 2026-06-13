"""
tests/test_parity.py — 스파인 전환 패리티 게이트 (T0).

실행: py -m unittest tests.test_parity

목적(자문 R1·R3 C7): main.py/render() 를 run_report 스파인으로 전환할 때
**골든 회귀 0** 을 증명한다. 신경로(run_report)와 구 본문(_legacy_render)이
같은 골든 입력에 대해 저장된 워크북의 셀(값+수식)이 동일한지 단언한다.

녹색 = 동작보존 리팩터 확정(셀 산출 불변). 적색 = 스파인이 골든에 없던
행동을 추가 → 머지 금지(화해 먼저).
"""
from __future__ import annotations

import os
import tempfile
import unittest

import fpna._bootstrap  # noqa: F401

from openpyxl import load_workbook

from fpna.pipeline import run_report, _hash_workbook, _mint, _MINT
from fpna.render import _legacy_render
from fpna.templates import available, get_template


def _cells(path: str) -> dict:
    """저장된 워크북의 (시트, 좌표) → 값 맵(값+수식 문자열 포함)."""
    wb = load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    out[(ws.title, c.coordinate)] = c.value
    return out


class ParityTest(unittest.TestCase):
    def test_golden_cell_parity(self):
        """전 골든: run_report 산출 셀 == 구 render 본문 산출 셀."""
        types = available()
        self.assertTrue(types, "구현된 템플릿이 없습니다")
        for t in types:
            mod = get_template(t)
            data = mod.golden_sample()
            with tempfile.TemporaryDirectory() as d:
                pa = os.path.join(d, "spine.xlsx")
                pb = os.path.join(d, "legacy.xlsx")
                ra = run_report(mod, data, out_path=pa)
                rb = _legacy_render(t, data, pb)
                self.assertEqual(ra.saved, rb.saved, "%s: saved 불일치" % t)
                if ra.saved and rb.saved:
                    self.assertEqual(_cells(pa), _cells(pb),
                                     "%s: 셀 산출이 스파인↔구경로 불일치" % t)


class ReceiptGateTest(unittest.TestCase):
    def test_forged_receipt_rejected(self):
        """스파인 밖에서 만든 GatePass(토큰 없음)는 저장 거부."""
        from fpna.pipeline import GatePass, _render_with_receipt
        mod = get_template(available()[0])
        data = mod.golden_sample()
        wb = mod.build(data)
        forged = GatePass(mod.TYPE, True, True, _hash_workbook(wb), token=None)
        with tempfile.TemporaryDirectory() as d:
            with self.assertRaises(RuntimeError):
                _render_with_receipt(wb, os.path.join(d, "x.xlsx"), forged, force=False)

    def test_stale_hash_rejected(self):
        """qc_hash 가 다른 wb 의 것이면(stale) 저장 거부."""
        from fpna.pipeline import _render_with_receipt
        mod = get_template(available()[0])
        data = mod.golden_sample()
        wb = mod.build(data)
        stale = _mint(mod.TYPE, "deadbeef" * 8)   # 엉뚱한 해시
        with tempfile.TemporaryDirectory() as d:
            with self.assertRaises(RuntimeError):
                _render_with_receipt(wb, os.path.join(d, "x.xlsx"), stale, force=False)

    def test_no_receipt_rejected(self):
        """receipt None + force False = 우회 차단."""
        from fpna.pipeline import _render_with_receipt
        mod = get_template(available()[0])
        wb = mod.build(mod.golden_sample())
        with tempfile.TemporaryDirectory() as d:
            with self.assertRaises(RuntimeError):
                _render_with_receipt(wb, os.path.join(d, "x.xlsx"), None, force=False)


if __name__ == "__main__":
    unittest.main()
