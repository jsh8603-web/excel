"""
tests/test_freehand_gate.py — freehand-excel-integrity 스킬의 라우터/게이트 회귀.

신규 더티 케이스 2종:
  · 정적(static) — ties.expected(소스 독립총계) 불일치 → xlsx_doctor [6] FAIL → 게이트 FAIL.
  · 수식(formula) — fill-down 파손(한 셀만 다른 행 참조) → formula_lint FAIL → 게이트 FAIL.
라우팅 신호 결정함수 + 환경 다운그레이드 투명성도 검증.

스킬 scripts/ 는 stdlib+openpyxl 만 쓰므로(이식 가능) 이 repo 테스트에서 직접 호출한다.
재계산(recalc)은 실 Excel 을 띄우므로 게이트 통합테스트에선 --no-recalc 로 분리한다.
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest

_REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SKILL = os.path.join(_REPO, "skills", "freehand-excel-integrity", "scripts")
sys.path.insert(0, os.path.join(_REPO, "vendor"))
sys.path.insert(0, _SKILL)

from openpyxl import Workbook  # noqa: E402


def _run_gate(*args: str) -> tuple[int, str]:
    env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
    p = subprocess.run([sys.executable, os.path.join(_SKILL, "verify_workbook.py"), *args],
                       capture_output=True, text=True, encoding="utf-8", errors="replace", env=env)
    return p.returncode, (p.stdout or "") + (p.stderr or "")


class RouteTest(unittest.TestCase):
    def test_signal_routing(self):
        import verify_workbook as vw
        self.assertEqual(vw.route(editing=False, file_open=False, bulk=True, chart=True, excel_feature=False), "xlsxwriter")
        self.assertEqual(vw.route(editing=False, file_open=False, bulk=False, chart=False, excel_feature=False), "openpyxl")
        self.assertEqual(vw.route(editing=True, file_open=True, bulk=False, chart=False, excel_feature=False), "xlwings")
        self.assertEqual(vw.route(editing=True, file_open=False, bulk=False, chart=False, excel_feature=False), "openpyxl")
        self.assertEqual(vw.route(editing=False, file_open=False, bulk=False, chart=False, excel_feature=True), "xlwings")

    def test_downgrade_transparency(self):
        """xlwings 모듈을 가짜로 부재 처리 → pivot 요청은 DOWNGRADE 로 드러나야(침묵 금지)."""
        import verify_workbook as vw
        orig = vw._module_available
        try:
            vw._module_available = lambda name: False if name in ("xlwings", "win32com.client") else orig(name)
            eff, downgrades = vw.resolve("xlwings", {"pivot"})
            self.assertEqual(eff, "openpyxl")
            self.assertTrue(any("DOWNGRADE" in d and "pivot" in d for d in downgrades),
                            "pivot 다운그레이드가 명시되지 않음: %s" % downgrades)
        finally:
            vw._module_available = orig


class StaticDirtyTest(unittest.TestCase):
    def test_static_expected_mismatch_fails(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "static.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "S"
            ws["A1"] = "a"; ws["B1"] = 10
            ws["A2"] = "b"; ws["B2"] = 20
            ws["A3"] = "c"; ws["B3"] = 30
            ws["A4"] = "TOTAL"; ws["B4"] = 60   # 정적 값(수식 아님)
            wb.save(p)
            # 독립 총계 expected 를 일부러 틀리게(999) → [6] tie FAIL 유도
            # xlsx_doctor 계약 자동탐색 = 확장자 치환(static.contract.json), append 아님.
            contract = {"sheet": "S",
                        "ties": [{"name": "T", "total": "S!B4", "expected": 999, "parts": "S!B1:B3"}]}
            with open(os.path.splitext(p)[0] + ".contract.json", "w", encoding="utf-8") as fh:
                json.dump(contract, fh)
            rc, out = _run_gate(p, "--no-recalc")
            self.assertEqual(rc, 1, "정적 expected 불일치인데 게이트 통과:\n" + out)
            self.assertIn("xlsx_doctor", out)


class FormulaDirtyTest(unittest.TestCase):
    def test_filldown_break_fails(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "formula.xlsx")
            wb = Workbook(); ws = wb.active; ws.title = "F"
            for r in (1, 2, 3, 4):
                ws.cell(r, 1, 10 * r)   # A
                ws.cell(r, 2, r)        # B
            ws["C1"] = "=A1-B1"
            ws["C2"] = "=A2-B2"
            ws["C3"] = "=A3-B3"
            ws["C4"] = "=A4-B2"   # 파손: B 가 4행이 아니라 2행 고정
            wb.save(p)
            rc, out = _run_gate(p, "--no-recalc", "--allow-no-contract")
            self.assertEqual(rc, 1, "fill-down 파손인데 게이트 통과:\n" + out)
            self.assertIn("formula_lint", out)

    def test_filldown_consistent_passes_lint(self):
        import formula_lint
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "ok.xlsx")
            wb = Workbook(); ws = wb.active
            for r in (1, 2, 3, 4):
                ws.cell(r, 1, 10 * r); ws.cell(r, 2, r)
                ws.cell(r, 3, "=A%d-B%d" % (r, r))
            wb.save(p)
            self.assertEqual(formula_lint.lint(p), [])


if __name__ == "__main__":
    unittest.main()
