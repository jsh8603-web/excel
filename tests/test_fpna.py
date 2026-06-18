"""
tests/test_fpna.py — stdlib unittest 회귀(pytest 불필요).

실행: py -m unittest tests.test_fpna   또는   py -m unittest discover tests
회사 PC에서도 설치 없이 동작.
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest
from dataclasses import replace

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fpna._bootstrap  # noqa: F401

from fpna import finance
from fpna.dispatcher import dispatch
from fpna.templates import available, get_template
from fpna.render import render

FIXTURE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures", "messy_sample.xlsx")


class TestFinance(unittest.TestCase):
    def test_npv_irr_consistency(self):
        cfs = [-1000, 300, 350, 400, 450, 300]
        r = finance.irr(cfs)
        self.assertIsNotNone(r)
        # IRR 에서 NPV ≈ 0
        self.assertAlmostEqual(finance.npv(r, cfs), 0.0, places=4)

    def test_irr_no_sign_change(self):
        self.assertIsNone(finance.irr([100, 200, 300]))

    def test_discounted_payback(self):
        self.assertIsNotNone(finance.discounted_payback(0.1, [-1000, 300, 350, 400, 450]))
        self.assertIsNone(finance.discounted_payback(0.1, [-1000, 10, 10]))

    def test_safe_div(self):
        self.assertIsNone(finance.safe_div(1, 0))

    def test_mirr(self):
        cfs = [-1000, 300, 350, 400, 450, 300]
        m = finance.mirr(cfs, 0.08, 0.09)
        self.assertIsNotNone(m)
        # MIRR 정의 직접 검증: (FV_pos/-PV_neg)^(1/n)-1
        n = len(cfs) - 1
        fv_pos = sum(cf * (1.09) ** (n - t) for t, cf in enumerate(cfs) if cf > 0)
        pv_neg = sum(cf / (1.08) ** t for t, cf in enumerate(cfs) if cf < 0)
        self.assertAlmostEqual(m, (fv_pos / -pv_neg) ** (1 / n) - 1, places=8)
        # 부호 한쪽만 → 미정
        self.assertIsNone(finance.mirr([100, 200, 300], 0.08, 0.09))

    def test_wacc(self):
        # E600 D400 Re12% Rd5% t22% = .6·.12 + .4·.05·.78 = .0876
        self.assertAlmostEqual(finance.wacc(600, 400, 0.12, 0.05, 0.22), 0.0876, places=6)
        self.assertIsNone(finance.wacc(0, 0, 0.1, 0.05, 0.2))  # V≤0


class TestIngest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not os.path.isfile(FIXTURE):
            from tests.make_fixtures import build_messy
            build_messy(FIXTURE)

    def test_ingest_structure(self):
        from fpna.ingest import ingest_workbook
        res = ingest_workbook(FIXTURE)
        self.assertEqual(res.n_blocks, 2)
        self.assertEqual(res.report.n_rejected, 0)
        rows = {(r.entity, r.period, r.metric): r.value for r in res.tidy_rows}
        # 괄호음수 / 세모음수 / 센티넬 / 소계
        self.assertEqual(rows[("제품B", "2024", "비용")], -50)
        self.assertEqual(rows[("기타", "2024", "매출")], -30)
        self.assertIsNone(rows[("제품B", "2025", "비용")])  # "-" 센티넬
        subtotals = [r for r in res.tidy_rows if r.row_role == "subtotal"]
        self.assertTrue(len(subtotals) >= 1)

    def test_percent_text(self):
        from fpna.ingest import ingest_workbook
        res = ingest_workbook(FIXTURE)
        rows = {r.entity: r.value for r in res.tidy_rows}
        self.assertAlmostEqual(rows["가동률"], 0.85, places=4)

    # ------------------------------------------------------------------
    # 무음 손상 3종 방어 골든(G1 단위전파 / G2 혼재 / G4 ditto / G6 오류값).
    # ⚠ 합성 *구조* 더미 — 재무 수치는 의미 없음(스케일/플래그 로직 검증용).
    # ------------------------------------------------------------------
    @staticmethod
    def _ingest_wb(builder):
        """openpyxl.Workbook 빌더 콜백 → 임시파일 저장 → ingest 결과."""
        import openpyxl
        from fpna.ingest import ingest_workbook
        wb = openpyxl.Workbook()
        builder(wb.active)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
            path = fh.name
        try:
            wb.save(path)
            return ingest_workbook(path)
        finally:
            try:
                os.remove(path)
            except OSError:
                pass

    def test_g1_unit_scale_propagation(self):
        """G1: 블록 내부 '(단위: 백만원)' → 블록 전체 값 base(원) 환산 + raw 보존."""
        def build(ws):
            ws["A1"] = "(단위: 백만원)"
            ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "계정", "2023", "2024", "2025"
            ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "매출", 1000, 1200, 1500
            ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "비용", 600, 800, 900
            ws["A5"], ws["B5"], ws["C5"], ws["D5"] = "이익", 400, 400, 600
        res = self._ingest_wb(build)
        rows = {(r.entity, r.period): r for r in res.tidy_rows}
        tr = rows[("매출", "2024")]
        self.assertEqual(tr.value, 1200 * 1_000_000)   # 백만원 → 원
        self.assertEqual(tr.scale_applied, 1_000_000)
        self.assertEqual(tr.scale_source, "block")
        self.assertEqual(tr.raw_value, 1200)           # 환산 전 원본 보존
        # 단위행이 데이터로 오인되지 않음(period/entity 정상 복원)
        self.assertEqual(rows[("이익", "2025")].value, 600 * 1_000_000)

    def test_g2_scale_heterogeneous(self):
        """G2: 블록단위(백만) 위에 셀 접미(억)가 우선 + 한 열 스케일 2종 → smell."""
        def build(ws):
            ws["A1"] = "(단위: 백만원)"
            ws["A2"], ws["B2"], ws["C2"] = "항목", "2024", "2025"
            ws["A3"], ws["B3"], ws["C3"] = "매출", 1000, "3,400억"
            ws["A4"], ws["B4"], ws["C4"] = "비용", 500, 600
            ws["A5"], ws["B5"], ws["C5"] = "이익", 500, 700
        res = self._ingest_wb(build)
        rows = {(r.entity, r.period): r for r in res.tidy_rows}
        # 셀 접미(억=1e8)가 블록(백만=1e6)을 덮어씀
        cell_tr = rows[("매출", "2025")]
        self.assertEqual(cell_tr.value, 3400 * 100_000_000)
        self.assertEqual(cell_tr.scale_source, "cell")
        # 같은 열 다른 셀은 블록 스케일
        self.assertEqual(rows[("비용", "2025")].scale_source, "block")
        self.assertEqual(rows[("비용", "2025")].value, 600 * 1_000_000)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("SCALE_HETEROGENEOUS", kinds)

    def test_g4_ditto_fill_down(self):
        """G4: 카테고리 열 상동 빈칸을 위→아래 단방향 충전 + DITTO_FILLED 플래그."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"], ws["D1"] = "부문", "제품", "매출", "수량"
            ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "국내", "제품A", 100, 10
            ws["A3"], ws["B3"], ws["C3"], ws["D3"] = None, "제품B", 200, 20
            ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "해외", "제품C", 300, 30
            ws["A5"], ws["B5"], ws["C5"], ws["D5"] = None, "제품D", 400, 40
        res = self._ingest_wb(build)
        # 빈칸이 채워져 국내/해외가 모든 행에 전파
        ents = {r.entity for r in res.tidy_rows}
        self.assertEqual(ents, {"국내", "해외"})
        filled = [r for r in res.tidy_rows if "DITTO_FILLED" in r.flags]
        self.assertTrue(filled)                       # 일부 행은 ditto 충전 표시
        # 충전된 행은 직전 부문 값 상속(제품B → 국내, 제품D → 해외)
        b_rows = [r for r in res.tidy_rows if r.metric and "제품B" in r.metric]
        self.assertTrue(all(r.entity == "국내" for r in b_rows))
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("DITTO_FILLED", kinds)

    def test_g6_error_cell_preserved(self):
        """G6: #DIV/0!/#N/A/#REF! → value=null + raw 보존 + ERROR_CELL smell.

        ⛔ 0/NaN coerce 금지(reject 도 아님 — 행은 유지, 값만 null)."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "계정", "2024", "2025"
            ws["A2"], ws["B2"], ws["C2"] = "매출", 1000, 1200
            ws["A3"], ws["B3"], ws["C3"] = "증가율", "#DIV/0!", "#N/A"
            ws["A4"], ws["B4"], ws["C4"] = "참조", "#REF!", 50
        res = self._ingest_wb(build)
        rows = {(r.entity, r.period): r for r in res.tidy_rows}
        err = rows[("증가율", "2024")]
        self.assertIsNone(err.value)                  # 0/NaN 아님 — null
        self.assertEqual(err.raw_value, "#DIV/0!")    # 원본 보존
        self.assertIn("ERROR_CELL", err.flags)
        # 정상 값은 그대로
        self.assertEqual(rows[("매출", "2024")].value, 1000)
        self.assertEqual(rows[("참조", "2025")].value, 50)
        # reject 되지 않고 smell 로만 표면화
        self.assertEqual(res.report.n_rejected, 0)
        err_smells = [s for s in res.smells if s["kind"] == "ERROR_CELL"]
        self.assertEqual(len(err_smells), 3)

    # ------------------------------------------------------------------
    # 무음손상 외 패턴 골든(G3 들여쓰기계층 / G5 색·볼드소계+색음수 /
    #                       G7 각주마커 / G8 반복헤더).
    # ⚠ 합성 *구조* 더미 — 재무 수치 의미 없음(계층/소계/마커 로직 검증용).
    # ------------------------------------------------------------------
    def test_g3_indent_hierarchy(self):
        """G3: alignment.indent + 선행공백 → level 산출 + 부모==Σ자식 정합 플래그."""
        from openpyxl.styles import Alignment

        def build(ws):
            ws["A1"], ws["B1"] = "계정", "2024"
            ws["A2"], ws["B2"] = "영업비용", 300     # 부모(level0)
            ws["A3"], ws["B3"] = "인건비", 200       # 자식(indent 1)
            ws["A4"], ws["B4"] = "임차료", 100       # 자식(indent 1)
            ws["A2"].alignment = Alignment(indent=0)
            ws["A3"].alignment = Alignment(indent=1)
            ws["A4"].alignment = Alignment(indent=1)
        res = self._ingest_wb(build)
        rows = {r.entity: r for r in res.tidy_rows}
        # 들여쓰기가 level 로 반영
        self.assertEqual(rows["영업비용"].level, 0)
        self.assertEqual(rows["인건비"].level, 1)
        self.assertEqual(rows["임차료"].level, 1)
        # 부모(300) == 자식합(200+100) → PARENT_EQ_CHILDREN_SUM
        self.assertIn("PARENT_EQ_CHILDREN_SUM", rows["영업비용"].flags)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("PARENT_EQ_CHILDREN_SUM", kinds)

    def test_g3_leading_space_level(self):
        """선행공백 2칸=1레벨 환산(indent 속성 없이도 계층 복원)."""
        from fpna.ingest.normalize import leading_space_level
        self.assertEqual(leading_space_level("매출"), 0)
        self.assertEqual(leading_space_level("  매출"), 1)
        self.assertEqual(leading_space_level("    매출"), 2)

    def test_g5_color_bold_subtotal_and_red_negative(self):
        """G5: label+bold+arith 교집합(score≥2) → subtotal + 빨강폰트 음수 보정."""
        from openpyxl.styles import Font

        def build(ws):
            ws["A1"], ws["B1"] = "항목", "2024"
            ws["A2"], ws["B2"] = "제품A", 100
            ws["A3"], ws["B3"] = "제품B", 200
            ws["A4"], ws["B4"] = "합계", 300          # label+arith(=100+200) → subtotal
            ws["A4"].font = Font(bold=True)
            ws["B4"].font = Font(bold=True)
            ws["A5"], ws["B5"] = "손실", 50           # 빨강폰트 → 음수 보정
            ws["B5"].font = Font(color="FFFF0000")
        res = self._ingest_wb(build)
        rows = {r.entity: r for r in res.tidy_rows}
        # 합계 = 소계로 분류(label+bold+arith 3신호)
        self.assertEqual(rows["합계"].row_role, "subtotal")
        # 빨강폰트 양수 → 음수 보정 + 플래그
        self.assertEqual(rows["손실"].value, -50)
        self.assertIn("SIGN_FROM_COLOR", rows["손실"].flags)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("SIGN_FROM_COLOR", kinds)
        self.assertIn("SUBTOTAL_DETECTED", kinds)

    def test_g7_footnote_marker_unify(self):
        """G7: 헤더 각주마커(¹*주N) 제거 → 동일 논리 metric 키 통일."""
        from fpna.ingest.normalize import strip_footnote_marker
        self.assertEqual(strip_footnote_marker("매출¹"), ("매출", True))
        self.assertEqual(strip_footnote_marker("매출*"), ("매출", True))
        self.assertEqual(strip_footnote_marker("매출(주1)"), ("매출", True))
        self.assertEqual(strip_footnote_marker("매출"), ("매출", False))

        def build(ws):
            # 두 표기('영업이익'/'영업이익¹')가 마커 제거 후 같은 metric 으로 통일
            ws["A1"], ws["B1"], ws["C1"] = "계정", "영업이익¹", "당기순이익"
            ws["A2"], ws["B2"], ws["C2"] = "회사A", 100, 80
            ws["A3"], ws["B3"], ws["C3"] = "회사B", 110, 90
            ws["A4"], ws["B4"], ws["C4"] = "회사C", 120, 95
        res = self._ingest_wb(build)
        metrics = {r.metric for r in res.tidy_rows}
        self.assertIn("영업이익", metrics)            # 마커 제거된 키
        self.assertNotIn("영업이익¹", metrics)

    def test_g7_dup_key_conflict(self):
        """G7: 마커 제거가 두 헤더를 같은 키로 붕괴시키면 DUP_KEY smell."""
        def build(ws):
            # '매출¹' 과 '매출²' → 둘 다 '매출' 로 붕괴 → 충돌 표면화
            ws["A1"], ws["B1"], ws["C1"] = "계정", "매출¹", "매출²"
            ws["A2"], ws["B2"], ws["C2"] = "회사A", 100, 200
            ws["A3"], ws["B3"], ws["C3"] = "회사B", 110, 210
            ws["A4"], ws["B4"], ws["C4"] = "회사C", 120, 220
        res = self._ingest_wb(build)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("DUP_KEY", kinds)

    def test_g8_repeated_header_dropped(self):
        """G8: 표 중간 재삽입된 헤더행(첫 헤더와 동일) → 데이터 제외 + 플래그."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "계정", "2024", "2025"
            ws["A2"], ws["B2"], ws["C2"] = "매출", 100, 120
            ws["A3"], ws["B3"], ws["C3"] = "비용", 60, 70
            # 페이지브레이크로 재삽입된 동일 헤더행(텍스트만, 숫자 없음)
            ws["A4"], ws["B4"], ws["C4"] = "계정", "2024", "2025"
            ws["A5"], ws["B5"], ws["C5"] = "이익", 40, 50
        res = self._ingest_wb(build)
        # 반복헤더 행(src_row=4)은 데이터로 안 남음
        src_rows = {r.src_row for r in res.tidy_rows}
        self.assertNotIn(4, src_rows)
        # 진짜 데이터(매출/비용/이익)는 보존
        ents = {r.entity for r in res.tidy_rows}
        self.assertEqual(ents, {"매출", "비용", "이익"})
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("REPEATED_HEADER_DROPPED", kinds)

    def test_fullwidth_and_suffix_normalization(self):
        """전각숫자/NBSP/접미 정규화 — normalize_value_ex 단위 검증."""
        from fpna.ingest.normalize import normalize_value_ex, split_cell_scale
        self.assertEqual(split_cell_scale("1,234천원"), ("1,234", 1000))
        self.assertEqual(split_cell_scale("3,400억"), ("3,400", 100_000_000))
        # 전각숫자 + 접미
        val, sent, neg, scale = normalize_value_ex("１，２３４천원")
        self.assertEqual(val, 1234)
        self.assertEqual(scale, 1000)
        # 퍼센트는 스케일 환산 대상 아님
        val, sent, neg, scale = normalize_value_ex("85%")
        self.assertAlmostEqual(val, 0.85, places=4)
        self.assertEqual(scale, 1)


    def test_g9_text_as_num_and_mixed_format(self):
        """G9(MetaCollector 흡수): @서식 숫자 → TEXT_AS_NUM_SUSPECT, 열 서식 2종 → MIXED."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "계정", "2024", "2025"
            ws["A2"], ws["B2"], ws["C2"] = "매출", 1000, "1200"
            ws["A3"], ws["B3"], ws["C3"] = "비용", 600, "700"
            ws["C2"].number_format = "@"; ws["C3"].number_format = "@"     # C열: 텍스트서식+숫자
            ws["B2"].number_format = "#,##0"; ws["B3"].number_format = "0.00"  # B열: 서식 2종
        res = self._ingest_wb(build)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("TEXT_AS_NUM_SUSPECT", kinds)
        self.assertIn("MIXED_NUMBER_FORMAT", kinds)

    def test_g10_no_header_suspect(self):
        """G10(MetaCollector 흡수): 텍스트 우세 첫 행이 헤더로 먹히나 실은 데이터 → 경고."""
        def build(ws):
            # 첫 행도 데이터(지역/구/값)인데 텍스트 우세라 헤더밴드로 잡힘 → 첫 행 손실 위험.
            ws["A1"], ws["B1"], ws["C1"] = "서울", "강남", 100
            ws["A2"], ws["B2"], ws["C2"] = "부산", "해운대", 200
            ws["A3"], ws["B3"], ws["C3"] = "대구", "수성", 300
        res = self._ingest_wb(build)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("NO_HEADER_SUSPECT", kinds)

    def test_g11_integer_year_header(self):
        """G11: 정수형 연도 헤더(2024)도 문자열처럼 헤더로 보호 → period 정상 복원."""
        def build(ws):
            ws["A1"] = "(단위: 백만원)"
            ws["A2"], ws["B2"], ws["C2"] = "계정", 2024, 2025          # 정수 연도 헤더
            ws["A3"], ws["B3"], ws["C3"] = "매출", 1000, 1200
            ws["A4"], ws["B4"], ws["C4"] = "비용", 600, 700
        res = self._ingest_wb(build)
        rows = {(r.entity, r.period): r for r in res.tidy_rows}
        self.assertIn(("매출", "2024"), rows)        # 연도가 데이터로 둔갑 안 하고 period 복원
        self.assertEqual(rows[("매출", "2024")].value, 1000 * 1_000_000)

    def test_g12_year_range_data_value(self):
        """G12: 데이터 값이 우연히 연도범위(1900~2100)여도 헤더로 안 먹히고 값 유지.

        fuzz 회귀(dirty_129): 정수연도 보호가 셀 단위라 데이터 1932 를 연도로 오인 →
        그 행이 헤더로 흡수돼 데이터 손실. 행 맥락(_is_year_header_row)으로 수정."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "계정", 2024, 2025   # 정수 연도 헤더(행 전부 연도)
            ws["A2"], ws["B2"], ws["C2"] = "매출", 1932, 8831    # 1932=연도범위지만 데이터
            ws["A3"], ws["B3"], ws["C3"] = "비용", 2087, 600
        res = self._ingest_wb(build)
        rows = {(r.entity, r.period): r for r in res.tidy_rows}
        self.assertIn(("매출", "2024"), rows)        # 매출 행이 헤더로 안 먹힘
        self.assertEqual(rows[("매출", "2024")].value, 1932)   # 1932 가 값으로 유지

    def test_g13_meta_header_blank_gap_timeseries(self):
        """G13: 다층 메타(제목·발행일) + 텍스트 멀티헤더 + 빈행 + 시계열(시간=행라벨).

        실데이터(ONS 노동시장) 회귀: 헤더밴드가 텍스트 위주라 density 미달로 데이터블록과
        분리 → tidy 0(무음실패) 였던 것을 absorb_header_bands 병합으로 해소. 추가로
        ① _value_like 마커강화: 'rate (%)'·제목의 '-' 가 값으로 오인 안 됨,
        ② classify lookahead: 발행일 datetime(단발 메타)에서 헤더밴드가 조기종료 안 함,
        ③ 행라벨(기간) → period 승격(시계열 시간축 보존)."""
        import datetime as _dt

        def build(ws):
            ws["A1"] = "Table X: Labour Force - summary"   # 제목(하이픈 = 마커강화 검증)
            ws["A2"] = "Date of publication:"
            ws["B2"] = _dt.datetime(2020, 12, 15)          # 발행일 datetime(단발 메타 노이즈)
            # A3 빈행
            ws["B4"], ws["C4"], ws["D4"] = "Employment", "Unemployed", "rate (%)"  # 지표(텍스트)
            ws["A5"] = "id"
            ws["B5"], ws["C5"], ws["D5"] = "level", "level", "pct"
            ws["A6"], ws["B6"], ws["C6"], ws["D6"] = "Jan-Mar 2020", 100, 200, 63
            ws["A7"], ws["B7"], ws["C7"], ws["D7"] = "Feb-Apr 2020", 110, 210, 64
            ws["A8"], ws["B8"], ws["C8"], ws["D8"] = "Mar-May 2020", 120, 220, 65
        res = self._ingest_wb(build)
        self.assertTrue(res.tidy_rows, "tidy 0(무음실패) 아니어야 함")
        periods = {str(r.period) for r in res.tidy_rows}
        self.assertIn("Jan-Mar 2020", periods)             # ③ 시간 행라벨 → period 승격
        # ② 발행일 datetime 이 period 로 새지 않음(메타행은 헤더밴드로 흡수, 데이터 아님)
        self.assertFalse(any("2020-12-15" in p for p in periods))
        # value 보존 + ① 'rate (%)' 가 데이터로 오인되지 않아 100 이 살아있음(role 무관)
        vals = [r.value for r in res.tidy_rows if str(r.period) == "Jan-Mar 2020"]
        self.assertIn(100, vals)
        self.assertIn(63, vals)                            # 'rate (%)' 헤더 아래 값 보존


class TestDispatch(unittest.TestCase):
    def test_keyword_routing(self):
        self.assertEqual(dispatch("NPV IRR 투자 타당성").template, "investment_appraisal")
        self.assertEqual(dispatch("예실 변동 브리지").template, "variance")
        self.assertEqual(dispatch("MoM 추이").template, "period_trend")

    def test_column_signal(self):
        d = dispatch("월간 보고", columns=["계획", "실적", "항목"])
        self.assertEqual(d.template, "variance")

    def test_new_template_routing(self):
        # 신규 2종 — 일반 키워드(budget_build/unit_economics)보다 우선해야
        self.assertEqual(dispatch("코호트 잔존 NRR 분석").template, "cohort_retention")
        self.assertEqual(dispatch("리텐션 GRR").template, "cohort_retention")
        self.assertEqual(dispatch("인원 계획 인건비 증원").template, "headcount_plan")
        self.assertEqual(dispatch("fully-loaded headcount").template, "headcount_plan")
        # 기존 라우팅 회귀(unit_economics·budget_build 은 잔여 키워드 유지)
        self.assertEqual(dispatch("CAC LTV ARR").template, "unit_economics")
        self.assertEqual(dispatch("예산 수립").template, "budget_build")


class TestProfile(unittest.TestCase):
    """정제 마트 → SHAPE 추출. ⚠ 결정적 구조 픽스처(난수·실금액 아님)."""

    SEAS = [0.80, 0.77, 0.86, 0.92, 0.95, 0.99, 0.98, 0.98, 1.03, 1.11, 1.24, 1.37]

    def _make_mart(self, path):
        import csv
        rows = [["entity", "account", "period", "budget", "actual"]]
        for ent in ("E1", "E2"):
            for acc in ("Rev", "COGS"):
                for yi in range(2):
                    for mo in range(1, 13):
                        pi = yi * 12 + (mo - 1)
                        b = round(1000 * (1.02 ** pi) * self.SEAS[mo - 1])
                        a = round(b * 1.05)  # actual = budget×1.05 → 완전 상관
                        rows.append([ent, acc, "%d-%02d-01" % (2023 + yi, mo), b, a])
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            csv.writer(fh).writerows(rows)

    def test_axes_recovered(self):
        from fpna.profile import profile_table
        with tempfile.TemporaryDirectory() as tmp:
            csvp = os.path.join(tmp, "mart.csv")
            self._make_mart(csvp)
            spec = profile_table(csvp)
            cols = spec["tables"]["mart"]["columns"]
            self.assertEqual(cols["entity"]["type"], "choice")
            self.assertEqual(cols["entity"]["n"], 2)
            self.assertEqual(cols["period"]["type"], "date")
            self.assertEqual(cols["budget"]["type"], "measure")
            # 추세 ≈ 0.02 (입력), 시즌 12개·12월>1월
            self.assertAlmostEqual(cols["budget"]["trend"], 0.02, places=2)
            seas = cols["budget"]["seasonality"]
            self.assertEqual(len(seas), 12)
            self.assertGreater(seas[11], seas[0])
            # actual = budget×1.05 → 상관 ≈ 1.0
            self.assertEqual(cols["actual"].get("corr_with"), "budget")
            self.assertGreater(cols["actual"]["corr"], 0.99)

    def test_no_value_leak(self):
        from fpna.profile import profile_table, emit_yaml
        with tempfile.TemporaryDirectory() as tmp:
            csvp = os.path.join(tmp, "mart.csv")
            self._make_mart(csvp)
            text = emit_yaml(profile_table(csvp))
            self.assertIn("EDIT_ME_scale", text)          # base = 자리표시자
            self.assertNotIn("1000", text)                # 절대 금액 미유출
            self.assertNotIn("1050", text)


class TestCrypto(unittest.TestCase):
    def test_rfc8439_vectors(self):
        from fpna._chacha import test_vectors
        self.assertTrue(test_vectors())

    def test_roundtrip_unicode_newline(self):
        from fpna.crypto import encrypt_text, decrypt_text
        msg = "다국어 ünïcode 1,234 (단위:천원)\n줄바꿈\t탭 ✓"
        arm = encrypt_text("pw-암호", msg)
        self.assertEqual(decrypt_text("pw-암호", arm), msg)

    def test_wrong_passphrase_rejected(self):
        from fpna.crypto import encrypt_text, decrypt_text
        arm = encrypt_text("right-pw", "secret")
        with self.assertRaises(ValueError):
            decrypt_text("wrong-pw", arm)

    def test_tamper_rejected(self):
        import base64
        from fpna.crypto import encrypt_text, decrypt_text
        arm = encrypt_text("pw", "secret payload")
        blob = bytearray(base64.b64decode(arm))
        blob[-1] ^= 0x01  # ciphertext 1비트 변조
        tampered = base64.b64encode(bytes(blob)).decode()
        with self.assertRaises(ValueError):
            decrypt_text("pw", tampered)

    def test_mail_split_roundtrip(self):
        from fpna.crypto import encrypt_text, to_mail_text, decrypt_text
        msg = "line %d data 1234 한글\n" * 300
        arm = encrypt_text("pw", msg)
        parts = to_mail_text(arm, max_lines=5, msg_id="T")
        self.assertGreater(len(parts), 1)                  # 여러 통으로 분할
        for p in parts:
            self.assertLessEqual(p.count("\n") + 1, 5)     # 메일당 줄수 한정 준수
        combined = "\n".join(reversed(parts))              # 역순으로 붙여도
        self.assertEqual(decrypt_text("pw", combined), msg)  # 정렬·복원


class TestTemplatesQC(unittest.TestCase):
    def test_all_golden_pass_qc(self):
        for t in available():
            mod = get_template(t)
            data = mod.golden_sample()
            with tempfile.TemporaryDirectory() as tmp:
                out = os.path.join(tmp, "%s.xlsx" % t)
                res = render(t, data, out)
                self.assertTrue(res.qc.passed, "%s QC FAIL: %s" % (t, res.qc.summary()))
                self.assertTrue(res.saved)
                self.assertTrue(os.path.isfile(out))


class TestTieOutGates(unittest.TestCase):
    """핵심 보고(손익·이사회팩) tie-out 게이트 — plug 은폐·보드숫자≠모델숫자 차단.

    ⚠ 구조 더미만 변형(실금액 아님) — 게이트가 *불균형을 잡는지* 검증.
    """

    # --- pnl_3statement BS 균형(A=L+E) ---------------------------------
    def test_pnl_golden_is_linked_and_balanced(self):
        from fpna.templates import pnl_3statement as m
        data = m.golden_sample()
        self.assertTrue(data.linked, "골든은 3-statement 연결 모드여야 함")
        wb = m.build(data)
        rep = m.qc(wb, data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("BS 균형(A=L+E)", names)
        self.assertIn("RE roll(기초+NI-배당)", names)
        self.assertIn("CF 간접법 기말현금 == BS 현금", names)

    def test_pnl_bs_imbalance_fails_qc(self):
        """자산을 1 틀면(plug 은폐 시나리오) BS 균형 check 가 FAIL 해야."""
        from fpna.templates import pnl_3statement as m
        data = m.golden_sample()
        broken = replace(data, other_assets=data.other_assets + 1.0)
        wb = m.build(broken)
        rep = m.qc(wb, broken)
        self.assertFalse(rep.passed, "BS 불균형인데 QC 통과 — plug 은폐 미차단")
        bs = [(n, ok) for n, ok, _ in rep.checks if n == "BS 균형(A=L+E)"]
        self.assertEqual(bs, [("BS 균형(A=L+E)", False)])

    def test_pnl_re_roll_break_fails_qc(self):
        """기말 RE 가 roll 과 안 맞으면(배당 누락 시나리오) FAIL."""
        from fpna.templates import pnl_3statement as m
        data = m.golden_sample()
        # 배당을 늘리면 RE_end·equity·BS 가 전부 깨짐 → 최소 한 tie FAIL
        broken = replace(data, dividends=data.dividends + 50.0)
        rep = m.qc(m.build(broken), broken)
        self.assertFalse(rep.passed)

    # --- board_kpi_pack source tie-out ---------------------------------
    def test_board_golden_source_tie(self):
        from fpna.templates import board_kpi_pack as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("source tie-out(합)", names)
        self.assertIn("source tie-out(항목별)", names)

    def test_board_source_mismatch_fails_qc(self):
        """보드 표기값(actual) ≠ 출처(source) → source tie-out FAIL."""
        from fpna.templates import board_kpi_pack as m
        data = m.golden_sample()
        ks = list(data.kpis)
        ks[0] = replace(ks[0], actual=ks[0].source + 99.0)  # 보드숫자만 부풀림
        broken = replace(data, kpis=ks)
        rep = m.qc(m.build(broken), broken)
        self.assertFalse(rep.passed, "보드≠출처인데 QC 통과 — 신뢰붕괴 미차단")
        item = [(n, ok) for n, ok, _ in rep.checks if n == "source tie-out(항목별)"]
        self.assertEqual(item, [("source tie-out(항목별)", False)])

    # --- variance R9 시나리오정합 + 브리지합 tie ------------------------
    def test_variance_golden_scenario_and_bridge(self):
        from fpna.templates import variance as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R9 scenario_aligned", names)
        self.assertIn("브리지 합산 tie(Σ구간 == 양끝차)", names)

    def test_variance_scenario_misalignment_fails(self):
        """한 시나리오(Budget)에만 존재하는 키 → R9 FAIL(0 처리 금지)."""
        from fpna.templates import variance as m
        data = m.golden_sample()
        items = list(data.items)
        # plan 만 있고 actual 이 없는(=Budget-only) 항목 추가
        items.insert(-1, m.LineItem("신규부서", plan=50, actual=None))
        broken = replace(data, items=items)
        rep = m.qc(m.build(broken), broken)
        r9 = [(n, ok) for n, ok, _ in rep.checks if n == "R9 scenario_aligned"]
        self.assertEqual(r9, [("R9 scenario_aligned", False)],
                         "Budget-only 키인데 R9 통과 — silent default 미차단")

    def test_variance_bridge_break_fails(self):
        """총계가 항목 Δ 합과 안 맞으면(워터폴 미합치) 브리지 tie FAIL."""
        from fpna.templates import variance as m
        data = m.golden_sample()
        items = list(data.items)
        ti = next(i for i, it in enumerate(items) if it.is_total)
        items[ti] = replace(items[ti], actual=items[ti].actual + 33.0)  # 총계만 부풀림
        broken = replace(data, items=items)
        rep = m.qc(m.build(broken), broken)
        self.assertFalse(rep.passed, "워터폴 미합치인데 QC 통과")
        bt = [ok for n, ok, _ in rep.checks if n == "브리지 합산 tie(Σ구간 == 양끝차)"]
        self.assertEqual(bt, [False])

    # --- rolling_forecast R1 시간전수 + 컷오버 grain --------------------
    def test_rolling_golden_r1_and_cutover(self):
        from fpna.templates import rolling_forecast as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R1 time_ruler", names)
        self.assertIn("컷오버 grain(actual XOR forecast)", names)

    def test_rolling_period_gap_fails_r1(self):
        """캘린더 기간을 건너뛰면(P07 누락) R1 FAIL."""
        from fpna.templates import rolling_forecast as m
        data = m.golden_sample()
        # P07 을 빼서 갭 생성(periods/coords/series 모두 동기 제거)
        drop = 6  # index of P07
        periods = [p for i, p in enumerate(data.periods) if i != drop]
        coords = [c for i, c in enumerate(data.cal_coords) if i != drop]
        series = {k: [v for i, v in enumerate(vals) if i != drop]
                  for k, vals in data.series.items()}
        broken = replace(data, periods=periods, cal_coords=coords, series=series,
                         actual_until=5)
        rep = m.qc(m.build(broken), broken)
        r1 = [ok for n, ok, _ in rep.checks if n == "R1 time_ruler"]
        self.assertEqual(r1, [False], "기간 갭인데 R1 통과 — silent 갭 미차단")

    def test_rolling_cutover_overlap_fails(self):
        """같은 기간이 actual+forecast 양쪽 → 이중계상 게이트 FAIL."""
        from fpna.templates import rolling_forecast as m
        data = m.golden_sample()
        n = len(data.periods)
        # P06(idx5)을 actual·forecast 양쪽에 넣어 중첩
        a_idx = list(range(0, 6))
        f_idx = list(range(5, n))   # 5 가 양쪽
        broken = replace(data, actual_idx=a_idx, forecast_idx=f_idx)
        rep = m.qc(m.build(broken), broken)
        ov = [ok for n_, ok, _ in rep.checks
              if n_ == "컷오버 grain(actual XOR forecast)"]
        self.assertEqual(ov, [False], "중첩 기간인데 통과 — 이중계상 미차단")

    # --- cashflow_13w 13주 전수 + 주간 연속성 --------------------------
    def test_cashflow_golden_r1_and_continuity(self):
        from fpna.templates import cashflow_13w as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R1 13주 전수성", names)
        self.assertIn("주간 연속성(기말==익주기초)", names)

    def test_cashflow_short_weeks_fails_r1(self):
        """12주만(13주 미충족) → R1 전수성 FAIL."""
        from fpna.templates import cashflow_13w as m
        data = m.golden_sample()
        broken = replace(data, weeks=12, inflows=data.inflows[:12],
                         outflows=data.outflows[:12])
        rep = m.qc(m.build(broken), broken)
        r1 = [ok for n, ok, _ in rep.checks if n == "R1 13주 전수성"]
        self.assertEqual(r1, [False], "13주 미충족인데 통과")

    def test_cashflow_continuity_break_fails(self):
        """한 주 기초를 전주 기말과 다르게 덮어쓰면 연속성 단절 FAIL."""
        from fpna.templates import cashflow_13w as m
        data = m.golden_sample()
        opens, closes = data.week_balances()
        opens[5] = closes[4] + 777.0   # W6 기초를 전주 기말과 단절
        broken = replace(data, openings=opens)
        rep = m.qc(m.build(broken), broken)
        cont = [ok for n, ok, _ in rep.checks if n == "주간 연속성(기말==익주기초)"]
        self.assertEqual(cont, [False], "연속성 단절인데 통과 — 유동성 오판 미차단")


class TestCohortRetention(unittest.TestCase):
    """코호트 잔존 — GRR 단조·GRR≤NRR·R3 movement tie·R17 NA."""

    def test_golden_invariants(self):
        from fpna.templates import cohort_retention as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("GRR 단조 비증가(이탈 누적)", names)
        self.assertIn("GRR ≤ NRR(확장 비음수)", names)
        self.assertIn("R3 movement_tie", names)
        self.assertIn("R17 NA surfaced(0분모 은폐 금지)", names)

    def test_grr_monotone_break_fails(self):
        """이탈을 음수(=잔존 회복)로 만들면 GRR 단조 비증가 위배 FAIL."""
        from fpna.templates import cohort_retention as m
        data = m.golden_sample()
        cohorts = [replace(c) for c in data.cohorts]
        steps = list(cohorts[0].steps)
        # age2 의 churn 을 음수로 → 누적 이탈 감소 → GRR 상승(위배)
        steps[2] = replace(steps[2], churn=-200.0)
        cohorts[0] = replace(cohorts[0], steps=steps)
        broken = replace(data, cohorts=cohorts)
        rep = m.qc(m.build(broken), broken)
        mono = [ok for n, ok, _ in rep.checks if n == "GRR 단조 비증가(이탈 누적)"]
        self.assertEqual(mono, [False], "GRR 상승인데 통과 — 이탈누적 위배 미차단")

    def test_movement_tie_break_fails(self):
        """end_mrr 정의를 어기는 step 변형 시 R3 movement tie FAIL.

        ⚠ end_mrr 는 property 라 직접 못 깬다 → start_mrr 만 바꿔도 합은 유지되므로
        churn 을 음수로 바꿔 tie 가 아니라 단조에서 잡힘을 위 테스트가 담당.
        여기선 base 0 코호트로 R17 NA 표면화를 검증."""
        from fpna.templates import cohort_retention as m
        data = m.golden_sample()
        cohorts = [replace(c) for c in data.cohorts]
        # 시작 MRR 0 코호트 추가 → 모든 age 비율 NA(0분모, R17)
        zero = m.CohortLine("2024-99", [
            m.CohortStep(0, start_mrr=0.0),
            m.CohortStep(1, start_mrr=0.0),
        ])
        broken = replace(data, cohorts=cohorts + [zero])
        rep = m.qc(m.build(broken), broken)
        self.assertTrue(rep.passed, rep.summary())  # NA 는 정직 emit → 통과
        na = [d for n, ok, d in rep.checks if n == "R17 NA surfaced(0분모 은폐 금지)"]
        self.assertEqual(len(na), 1)


class TestHeadcountPlan(unittest.TestCase):
    """인원 계획 — R10 부서 roll-up tie·R3 loading tie·R1 시간 전수."""

    def test_golden_invariants(self):
        from fpna.templates import headcount_plan as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R10 dept_rollup_tie", names)
        self.assertIn("R3 loading_tie", names)
        self.assertIn("R1 time_ruler", names)

    def test_fully_loaded(self):
        """fully_loaded = base×(1+loading), 월 = annual/12."""
        from fpna.templates import headcount_plan as m
        ln = m.RosterLine("X", base_salary_annual=100.0, loading_rate=0.25)
        self.assertAlmostEqual(ln.fully_loaded_annual, 125.0)
        self.assertAlmostEqual(ln.fully_loaded_monthly, 125.0 / 12.0)

    def test_period_gap_fails_r1(self):
        """기간 벡터를 1기 짧게 주면 R1 전수성 FAIL."""
        from fpna.templates import headcount_plan as m
        data = m.golden_sample()
        lines = [replace(ln, headcount=ln.headcount[:11]) for ln in data.lines]
        broken = replace(data, lines=lines)  # end 는 그대로 → 12기 기대, 11기만 존재
        rep = m.qc(m.build(broken), broken)
        r1 = [ok for n, ok, _ in rep.checks if n == "R1 time_ruler"]
        self.assertEqual(r1, [False], "기간 갭인데 R1 통과")


class TestInvestmentAppraisalExt(unittest.TestCase):
    """investment_appraisal 보완 — MIRR·WACC·TV·tornado (기존 회귀 보존)."""

    def test_basic_golden_unchanged(self):
        """옵션 없는 기본 골든은 종전대로 통과(회귀)."""
        from fpna.templates import investment_appraisal as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())

    def test_full_golden_extras(self):
        from fpna.templates import investment_appraisal as m
        data = m.golden_sample_full()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("WACC 계산 가능", names)
        self.assertIn("MIRR 해 존재", names)
        self.assertIn("잔존가치(TV) 가산 tie", names)
        self.assertIn("토네이도 swing 정의", names)

    def test_terminal_value_added_to_last_cf(self):
        """TV 가 마지막 기 현금흐름에 정확히 가산되는지(이중가산·누락 차단)."""
        from fpna.templates import investment_appraisal as m
        data = replace(m.golden_sample(), terminal_value=200.0)
        eff = m._effective_cashflows(data)
        self.assertEqual(eff[-1], data.cashflows[-1] + 200.0)
        # NPV 가 TV 만큼 커진다(할인 후)
        npv_no_tv = finance.npv(0.10, data.cashflows)
        npv_tv = finance.npv(0.10, eff)
        self.assertGreater(npv_tv, npv_no_tv)

    def test_wacc_as_rate(self):
        """use_wacc_as_rate=True 면 할인율 = WACC."""
        from fpna.templates import investment_appraisal as m
        wb_ = m.WaccBuild(600, 400, 0.12, 0.05, 0.22)
        data = replace(m.golden_sample(), wacc_build=wb_, use_wacc_as_rate=True)
        self.assertAlmostEqual(m._effective_rate(data),
                               finance.wacc(600, 400, 0.12, 0.05, 0.22), places=8)


class TestPeriodTrendDepth(unittest.TestCase):
    """period_trend 깊이 — R1 전수·CAGR N 정합·계절지수 평균≈1."""

    def test_golden_invariants(self):
        from fpna.templates import period_trend as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R1 time_ruler", names)
        self.assertIn("CAGR N 정합(N=시점-1)", names)
        self.assertIn("계절지수 평균≈1", names)

    def test_period_gap_fails_r1(self):
        """캘린더 기간을 건너뛰면(P07 누락) R1 FAIL."""
        from fpna.templates import period_trend as m
        data = m.golden_sample()
        drop = 6
        periods = [p for i, p in enumerate(data.periods) if i != drop]
        coords = [c for i, c in enumerate(data.cal_coords) if i != drop]
        series = {k: [v for i, v in enumerate(vals) if i != drop]
                  for k, vals in data.series.items()}
        broken = replace(data, periods=periods, cal_coords=coords, series=series)
        rep = m.qc(m.build(broken), broken)
        r1 = [ok for n, ok, _ in rep.checks if n == "R1 time_ruler"]
        self.assertEqual(r1, [False], "기간 갭인데 R1 통과 — silent 갭 미차단")


class TestUnitEconDepth(unittest.TestCase):
    """unit_economics 깊이 — 할인 기여이익 LTV(특수해 정합 + 시간가치 보정)."""

    def test_golden_invariants(self):
        from fpna.templates import unit_economics as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("LTV(할인 기여이익) 계산", names)
        self.assertIn("할인 LTV < 무할인 LTV(시간가치 보정)", names)

    def test_no_discount_equals_simple(self):
        """할인율 0 이면 닫힌형 LTV == m/churn(현 단순식 특수해)."""
        from fpna.templates import unit_economics as m
        data = replace(m.golden_sample(), discount_monthly=0.0)
        ltv, _ = m.discounted_ltv(data)
        simple = data.arpu * data.gross_margin / data.churn_monthly
        self.assertAlmostEqual(ltv, simple, places=6)

    def test_discount_reduces_ltv(self):
        """할인율 > 0 이면 할인 LTV < 무할인 LTV(과대평가 보정)."""
        from fpna.templates import unit_economics as m
        d0 = replace(m.golden_sample(), discount_monthly=0.0)
        d1 = replace(m.golden_sample(), discount_monthly=0.02)
        ltv0, _ = m.discounted_ltv(d0)
        ltv1, _ = m.discounted_ltv(d1)
        self.assertLess(ltv1, ltv0)


class TestBudgetBuildDepth(unittest.TestCase):
    """budget_build 깊이 — R10 roll-up tie + ZBB/incremental 구분."""

    def test_golden_invariants(self):
        from fpna.templates import budget_build as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R10 dept_rollup_tie", names)
        self.assertIn("편성방식 유효(ZBB|incremental)", names)
        self.assertIn("incremental baseline 명시", names)

    def test_bad_method_fails(self):
        """편성방식이 ZBB/incremental 외면 FAIL."""
        from fpna.templates import budget_build as m
        data = m.golden_sample()
        depts = [replace(data.depts[0], method="freestyle")] + list(data.depts[1:])
        broken = replace(data, depts=depts)
        rep = m.qc(m.build(broken), broken)
        v = [ok for n, ok, _ in rep.checks if n == "편성방식 유효(ZBB|incremental)"]
        self.assertEqual(v, [False])

    def test_incremental_without_baseline_fails(self):
        """incremental 인데 전년예산(baseline) 누락 → FAIL."""
        from fpna.templates import budget_build as m
        data = m.golden_sample()
        depts = [replace(d, prior_budget=0.0) if d.method == "incremental" else d
                 for d in data.depts]
        broken = replace(data, depts=depts)
        rep = m.qc(m.build(broken), broken)
        v = [ok for n, ok, _ in rep.checks if n == "incremental baseline 명시"]
        self.assertEqual(v, [False])


class TestScenarioDepth(unittest.TestCase):
    """scenario_sensitivity 깊이 — R9 base=모델base + 동적 스위치."""

    def test_golden_invariants(self):
        from fpna.templates import scenario_sensitivity as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R9 base=모델base(시나리오 정합)", names)
        self.assertIn("시나리오 스위치 동적(케이스 분리)", names)

    def test_base_case_equals_model_base(self):
        """Base 케이스 결과 == base_outcome(모든 드라이버 base 값)."""
        from fpna.templates import scenario_sensitivity as m
        data = m.golden_sample()
        self.assertAlmostEqual(m._case_outcome(data, "Base"), data.base_outcome)


class TestRollingForecastBridge(unittest.TestCase):
    """rolling_forecast 깊이 — 버전 브리지 tie(ΣΔ == Σ현재−Σ직전)."""

    def test_golden_bridge_tie(self):
        from fpna.templates import rolling_forecast as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("버전 브리지 tie(ΣΔ == Σ현재−Σ직전)", names)

    def test_prior_length_mismatch_fails(self):
        """직전 전망 길이가 기간 수와 다르면 FAIL."""
        from fpna.templates import rolling_forecast as m
        data = m.golden_sample()
        prior = {k: v[:-1] for k, v in data.prior_series.items()}  # 1기 짧게
        broken = replace(data, prior_series=prior)
        rep = m.qc(m.build(broken), broken)
        v = [ok for n, ok, _ in rep.checks if n == "직전 전망 길이 == 기간 수"]
        self.assertEqual(v, [False])


class TestCashflowDirectMethod(unittest.TestCase):
    """cashflow_13w 깊이 — 직접법 분개 tie(Σ항목 == 집계)."""

    def test_golden_line_tie(self):
        from fpna.templates import cashflow_13w as m
        data = m.golden_sample()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("직접법 유입 분개 tie(Σ항목==계)", names)
        self.assertIn("직접법 유출 분개 tie(Σ항목==계)", names)

    def test_line_mismatch_fails(self):
        """분개 한 주를 어긋나게 만들면 tie FAIL(누락·중복 차단)."""
        from fpna.templates import cashflow_13w as m
        data = m.golden_sample()
        lines = {k: list(v) for k, v in data.inflow_lines.items()}
        first = next(iter(lines))
        lines[first][3] += 999.0  # W4 분개만 부풀림
        broken = replace(data, inflow_lines=lines)
        rep = m.qc(m.build(broken), broken)
        v = [ok for n, ok, _ in rep.checks if n == "직접법 유입 분개 tie(Σ항목==계)"]
        self.assertEqual(v, [False])


class TestInvestmentTornadoBase(unittest.TestCase):
    """investment_appraisal — R9 토네이도 base=모델 NPV(bracket)."""

    def test_full_golden_brackets_base(self):
        from fpna.templates import investment_appraisal as m
        data = m.golden_sample_full()
        rep = m.qc(m.build(data), data)
        self.assertTrue(rep.passed, rep.summary())
        names = [n for n, _, _ in rep.checks]
        self.assertIn("R9 토네이도 base=모델 NPV(bracket)", names)

    def test_stale_base_fails(self):
        """토네이도 범위가 모델 base NPV 를 bracket 안 하면 FAIL(stale base)."""
        from fpna.templates import investment_appraisal as m
        data = m.golden_sample_full()
        tor = list(data.tornado)
        # 한 변수의 low/high 를 base NPV 아래로 끌어내려 bracket 깨기
        tor[0] = replace(tor[0], npv_low=-9999.0, npv_high=-9000.0)
        broken = replace(data, tornado=tor)
        rep = m.qc(m.build(broken), broken)
        v = [ok for n, ok, _ in rep.checks if n == "R9 토네이도 base=모델 NPV(bracket)"]
        self.assertEqual(v, [False])


if __name__ == "__main__":
    unittest.main(verbosity=2)
