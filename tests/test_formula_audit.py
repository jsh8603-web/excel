"""
tests/test_formula_audit.py — 수식참조 계약 + 변이 검증 (자문 2026-06).

핵심 변이(teeth 증명): variance build 의 Δ 수식 칼럼을 *틀리게* 만들면 스파인이
반드시 FAIL 해야 한다. 기존 tautological qc 는 이 변이를 초록으로 통과시켰다.

실행: py -m unittest tests.test_formula_audit
"""
from __future__ import annotations

import unittest

import fpna._bootstrap  # noqa: F401

from fpna import formula_audit
from fpna.templates.base import QCReport


class CheckBinaryTest(unittest.TestCase):
    def test_correct(self):
        ok, _ = formula_audit.check_binary("=C5-B5", row=5, left_col=3, right_col=2)
        self.assertTrue(ok)

    def test_dollar_and_space_tolerant(self):
        ok, _ = formula_audit.check_binary("= $C$5 - $B$5", row=5, left_col=3, right_col=2)
        self.assertTrue(ok)

    def test_direction_flipped_fails(self):
        ok, _ = formula_audit.check_binary("=B5-C5", row=5, left_col=3, right_col=2)
        self.assertFalse(ok)

    def test_wrong_column_fails(self):
        ok, _ = formula_audit.check_binary("=D5-B5", row=5, left_col=3, right_col=2)
        self.assertFalse(ok)

    def test_wrong_row_fails(self):
        ok, _ = formula_audit.check_binary("=C6-B5", row=5, left_col=3, right_col=2)
        self.assertFalse(ok)


class VarianceFormulaGateTest(unittest.TestCase):
    """variance 를 정상/변이 빌드해 스파인 게이트의 검출력 확인."""

    def _run(self, mutate=None):
        from fpna.pipeline import _base_owned_gate
        from fpna.templates import variance
        data = variance.golden_sample()
        wb = variance.build(data)
        if mutate:
            mutate(wb)
        rep = QCReport("variance")
        _base_owned_gate(rep, wb, data, variance)
        return rep

    def test_clean_passes_formula_gate(self):
        rep = self._run()
        bad = [(n, d) for n, ok, d in rep.checks if not ok and "수식참조" in n]
        self.assertFalse(bad, "정상 산출이 수식참조 게이트에 걸림: %s" % bad)

    def test_flipped_delta_fails(self):
        """Δ 방향을 뒤집으면(=계획-실적) 게이트가 반드시 적색."""
        from openpyxl.utils import get_column_letter

        def flip(wb):
            m = wb._fpna_meta
            ws = wb.worksheets[0]
            pl = get_column_letter(m["plan_col"])
            ac = get_column_letter(m["act_col"])
            for r in range(m["data_start"], m["data_end"] + 1):
                c = ws.cell(row=r, column=m["var_col"])
                if isinstance(c.value, str) and c.value.startswith("="):
                    c.value = "=%s%d-%s%d" % (pl, r, ac, r)   # 뒤집음
        rep = self._run(flip)
        self.assertFalse(rep.passed, "뒤집힌 Δ 가 게이트를 통과(검출력 0)")
        self.assertTrue(any(not ok and "수식참조" in n for n, ok, _ in rep.checks))


if __name__ == "__main__":
    unittest.main()


class ContentTypeAndRatioTest(unittest.TestCase):
    def test_safe_ratio_formula(self):
        f = formula_audit.safe_ratio_formula("F52", "F60")
        self.assertIn("ISNUMBER(F60)", f)
        self.assertIn("F52/F60", f)

    def test_content_type_gate_flags_text_in_numeric_region(self):
        import openpyxl
        from fpna.templates.base import QCReport
        from fpna import layout_audit
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SUMMARY"
        ws.cell(row=13, column=3, value=100)
        ws.cell(row=14, column=3, value="AUG")     # 숫자영역 텍스트
        rep = QCReport("x")
        meta = {"numeric_regions": [("SUMMARY", 13, 14, 3, 3)]}
        ok = layout_audit.assert_cell_content_types(rep, wb, meta)
        self.assertFalse(ok)

    def test_content_type_gate_clean_passes(self):
        import openpyxl
        from fpna.templates.base import QCReport
        from fpna import layout_audit
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SUMMARY"
        ws.cell(row=13, column=3, value=100)
        ws.cell(row=14, column=3, value=200)
        rep = QCReport("x")
        meta = {"numeric_regions": [("SUMMARY", 13, 14, 3, 3)]}
        self.assertTrue(layout_audit.assert_cell_content_types(rep, wb, meta))
