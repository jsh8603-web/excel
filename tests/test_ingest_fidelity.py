"""
tests/test_ingest_fidelity.py — ingest 충실도 게이트 회귀(A0/A1/A4).

⚠ 합성 *구조* 픽스처 — 재무 수치는 의미 없는 더미(스케일/합계/좌표 로직 검증용).
실행: py -m unittest tests.test_ingest_fidelity

게이트:
- A0: 합계행 산술 불일치(스케일 정규화 後) → reject. 정합은 통과.
- A1: tidy 행에 출처 좌표(src_sheet/row/col) 부착 + per-cell coverage map
      (결측/중복 탐지) + 1:1 값 일치.
- A4: 선행0 ID('0001') → text 보존(계정코드/사번).
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.styles import Alignment


def _ingest(builder):
    """openpyxl 워크북 빌더 → 임시파일 → ingest 결과."""
    from fpna.ingest import ingest_workbook
    wb = openpyxl.Workbook()
    builder(wb.active)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
        path = fh.name
    try:
        wb.save(path)
        return ingest_workbook(path)
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


def _set(ws, coord, value, *, indent=0):
    c = ws[coord]
    c.value = value
    if indent:
        c.alignment = Alignment(indent=indent)
    return c


class TestA0SubtotalReject(unittest.TestCase):
    """A0: 합계행 산술 reject 승격(계층 부모==Σ자식)."""

    def test_subtotal_mismatch_rejected(self):
        """들여쓰기 부모(합계) != Σ자식 → reject + SUBTOTAL_ARITH_MISMATCH smell."""
        def build(ws):
            ws["A1"], ws["B1"] = "계정", "2024"
            # 합계(level0) = 자식 두 개 합과 어긋남(1000 != 300+600=900).
            _set(ws, "A2", "합계", indent=0)
            ws["B2"] = 1000
            _set(ws, "A3", "제품A", indent=1)
            ws["B3"] = 300
            _set(ws, "A4", "제품B", indent=1)
            ws["B4"] = 600
        res = _ingest(build)
        # 합계 행은 reject(tidy 에서 제외)
        ents = {r.entity for r in res.tidy_rows}
        self.assertNotIn("합계", ents)
        self.assertGreaterEqual(res.report.n_rejected, 1)
        kinds = {s["kind"] for s in res.smells}
        self.assertIn("SUBTOTAL_ARITH_MISMATCH", kinds)

    def test_subtotal_match_kept(self):
        """들여쓰기 부모(합계) == Σ자식(허용오차 내) → 통과(reject 없음)."""
        def build(ws):
            ws["A1"], ws["B1"] = "계정", "2024"
            _set(ws, "A2", "합계", indent=0)
            ws["B2"] = 900                       # 300+600 = 정합
            _set(ws, "A3", "제품A", indent=1)
            ws["B3"] = 300
            _set(ws, "A4", "제품B", indent=1)
            ws["B4"] = 600
        res = _ingest(build)
        kinds = {s["kind"] for s in res.smells}
        self.assertNotIn("SUBTOTAL_ARITH_MISMATCH", kinds)
        # 합계 행 보존(소계 role 이지만 tidy 에 존재)
        ents = {r.entity for r in res.tidy_rows}
        self.assertIn("합계", ents)

    def test_subtotal_match_after_scale_normalization(self):
        """★스케일 정규화 後 비교 — 합계와 자식이 같은 블록단위(백만)면 정합 유지."""
        def build(ws):
            ws["A1"] = "(단위: 백만원)"
            ws["A2"], ws["B2"] = "계정", "2024"
            _set(ws, "A3", "합계", indent=0)
            ws["B3"] = 900                       # 백만원 단위, 자식 합과 정합
            _set(ws, "A4", "제품A", indent=1)
            ws["B4"] = 300
            _set(ws, "A5", "제품B", indent=1)
            ws["B5"] = 600
        res = _ingest(build)
        kinds = {s["kind"] for s in res.smells}
        # 같은 스케일이라 base 환산 後에도 정합 → mismatch 없음
        self.assertNotIn("SUBTOTAL_ARITH_MISMATCH", kinds)
        rows = {r.entity: r for r in res.tidy_rows}
        self.assertEqual(rows["합계"].value, 900 * 1_000_000)

    def test_rounding_absorbed(self):
        """반올림 오차(허용오차 내)는 reject 안 함."""
        def build(ws):
            ws["A1"], ws["B1"] = "계정", "2024"
            _set(ws, "A2", "합계", indent=0)
            ws["B2"] = 1000
            _set(ws, "A3", "제품A", indent=1)
            ws["B3"] = 333
            _set(ws, "A4", "제품B", indent=1)
            ws["B4"] = 667                       # 333+667=1000 정확
        res = _ingest(build)
        self.assertNotIn("SUBTOTAL_ARITH_MISMATCH",
                         {s["kind"] for s in res.smells})


class TestA1Reconciliation(unittest.TestCase):
    """A1: provenance 좌표 + coverage map + 1:1 값검증."""

    def test_provenance_coords_attached(self):
        """모든 tidy 행이 src_sheet/src_row/src_col 보유."""
        def build(ws):
            ws.title = "손익"
            ws["A1"], ws["B1"], ws["C1"] = "계정", "2024", "2025"
            ws["A2"], ws["B2"], ws["C2"] = "매출", 100, 120
            ws["A3"], ws["B3"], ws["C3"] = "비용", 60, 70
        res = _ingest(build)
        self.assertTrue(res.tidy_rows)
        for r in res.tidy_rows:
            self.assertEqual(r.src_sheet, "손익")
            self.assertGreater(r.src_row, 0)
            self.assertGreater(r.src_col, 0)

    def test_coverage_full_clean(self):
        """정상 표 → 모든 data 셀 정확히 1회 덮임(missing/dup/mismatch 0)."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "계정", "2024", "2025"
            ws["A2"], ws["B2"], ws["C2"] = "매출", 100, 120
            ws["A3"], ws["B3"], ws["C3"] = "비용", 60, 70
        res = _ingest(build)
        self.assertEqual(len(res.recon), 1)
        rc = res.recon[0]
        self.assertEqual(rc.n_groundtruth, 4)        # 4 measure 셀
        self.assertEqual(len(rc.covered), 4)
        self.assertFalse(rc.missing)
        self.assertFalse(rc.duplicate)
        self.assertFalse(rc.value_mismatch)
        self.assertTrue(rc.ok)
        # recon smell 없음
        self.assertFalse({s["kind"] for s in res.smells if s["kind"].startswith("RECON")})

    def test_value_one_to_one_match(self):
        """tidy 값이 출처 셀 raw 와 1:1 일치(좌표→값)."""
        def build(ws):
            ws.title = "x"
            ws["A1"], ws["B1"] = "계정", "2024"
            ws["A2"], ws["B2"] = "매출", 12345
            ws["A3"], ws["B3"] = "비용", 6789
        res = _ingest(build)
        by_coord = {(r.src_row, r.src_col): r.value for r in res.tidy_rows}
        self.assertEqual(by_coord[(2, 2)], 12345)
        self.assertEqual(by_coord[(3, 2)], 6789)
        self.assertTrue(res.recon[0].ok)

    def test_missing_cell_detected(self):
        """data 셀이 tidy 에 누락되면 coverage map 이 missing 으로 잡는다.

        recon 의 독립 ground-truth 가 parse 와 다른 셀집합을 보는지 직접 검증:
        groundtruth_cells 로 GT 를 뽑고, tidy 에서 한 좌표를 빼 reconcile 하면
        그 좌표가 missing 으로 보고돼야 한다(coverage=0).
        """
        from fpna.ingest.reconcile import reconcile_sheet
        def build(ws):
            ws["A1"], ws["B1"] = "계정", "2024"
            ws["A2"], ws["B2"] = "매출", 100
            ws["A3"], ws["B3"] = "비용", 60
        res = _ingest(build)
        gt = {(2, 2), (3, 2)}                     # 두 measure 셀
        raw = {(2, 2): 100, (3, 2): 60}
        # tidy 에서 (3,2) 행을 의도적으로 누락
        partial = [r for r in res.tidy_rows if (r.src_row, r.src_col) != (3, 2)]
        rc = reconcile_sheet("s", gt, raw, partial)
        self.assertIn((3, 2), rc.missing)
        self.assertNotIn((3, 2), rc.covered)
        self.assertFalse(rc.ok)

    def test_duplicate_cell_detected(self):
        """한 출처셀이 tidy 행 2개로 중복 매핑되면 duplicate 로 잡는다."""
        from fpna.ingest.reconcile import reconcile_sheet
        from fpna.ingest.validate import TidyRow
        gt = {(2, 2)}
        raw = {(2, 2): 100}
        dup_rows = [
            TidyRow(entity="매출", period="2024", metric=None, value=100,
                    row_role="data", src_sheet="s", src_row=2, src_col=2),
            TidyRow(entity="매출b", period="2024", metric=None, value=100,
                    row_role="data", src_sheet="s", src_row=2, src_col=2),
        ]
        rc = reconcile_sheet("s", gt, raw, dup_rows)
        self.assertIn((2, 2), rc.duplicate)
        self.assertEqual(rc.duplicate[(2, 2)], 2)
        self.assertFalse(rc.ok)

    def test_value_mismatch_detected(self):
        """tidy 값이 원본 셀과 다르면 value_mismatch 로 잡는다(다값손실/오배치)."""
        from fpna.ingest.reconcile import reconcile_sheet
        from fpna.ingest.validate import TidyRow
        gt = {(2, 2)}
        raw = {(2, 2): 100}
        rows = [TidyRow(entity="매출", period="2024", metric=None, value=999,
                        row_role="data", src_sheet="s", src_row=2, src_col=2)]
        rc = reconcile_sheet("s", gt, raw, rows)
        self.assertTrue(rc.value_mismatch)
        self.assertEqual(rc.value_mismatch[0][0], (2, 2))
        self.assertFalse(rc.ok)

    def test_groundtruth_independent_of_parse(self):
        """ground-truth 스캐너는 소계/총계/헤더/연도를 measure 에서 제외(독립·관대)."""
        from fpna.ingest.cells import as_cells
        from fpna.ingest.reconcile import groundtruth_cells
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"], ws["B1"] = "계정", 2024        # bare 연도 = 헤더(제외)
        ws["A2"], ws["B2"] = "매출", 100
        ws["A3"], ws["B3"] = "합계", 100         # 소계 라벨 행(제외)
        cells = as_cells(ws, ws)
        gt = groundtruth_cells(cells)
        self.assertIn((2, 2), gt)                # 매출 measure
        self.assertNotIn((1, 2), gt)             # 연도 헤더
        self.assertNotIn((3, 2), gt)             # 합계 행

    def test_recon_summary_in_schema(self):
        """schema['reconciliation'] 에 시트별 coverage 요약이 박힌다."""
        def build(ws):
            ws["A1"], ws["B1"] = "계정", "2024"
            ws["A2"], ws["B2"] = "매출", 100
            ws["A3"], ws["B3"] = "비용", 60
        res = _ingest(build)
        recon = res.schema.get("reconciliation")
        self.assertTrue(recon)
        self.assertIn("n_groundtruth", recon[0])
        self.assertTrue(recon[0]["ok"])


class TestA4LeadingZeroId(unittest.TestCase):
    """A4: 선행0 ID(계정코드/사번) text 보존."""

    def test_leading_zero_preserved_as_text(self):
        """계정코드 열('0001'/'00123')이 entity 로 선행0 보존(int 변환 안 됨)."""
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "계정코드", "2023", "2024"
            ws["A2"], ws["B2"], ws["C2"] = "0001", 100, 110
            ws["A3"], ws["B3"], ws["C3"] = "00123", 200, 210
        res = _ingest(build)
        ents = {r.entity for r in res.tidy_rows}
        # 선행0 유실 시 '0001'→1, '00123'→123 (정수)로 깨짐.
        self.assertIn("0001", ents)
        self.assertIn("00123", ents)
        self.assertNotIn("1", ents)
        self.assertNotIn("123", ents)

    def test_normalize_unit_leading_zero(self):
        """normalize_value_ex 단위 — 선행0 ID 는 text, 일반 0/소수/정수는 숫자."""
        from fpna.ingest.normalize import normalize_value_ex
        self.assertEqual(normalize_value_ex("0001")[0], "0001")
        self.assertEqual(normalize_value_ex("007")[0], "007")
        self.assertEqual(normalize_value_ex("0")[0], 0)        # 진짜 0
        self.assertEqual(normalize_value_ex("0.5")[0], 0.5)    # 소수
        self.assertEqual(normalize_value_ex("10")[0], 10)      # 일반 정수
        # 선행0 ID 는 센티넬/스케일 아님
        v, sent, neg, scale = normalize_value_ex("0042")
        self.assertEqual(v, "0042")
        self.assertIsNone(sent)
        self.assertEqual(scale, 1)

    def test_regex_type_leading_zero_is_text(self):
        """선행0 ID 는 열 타입 추론에서 TEXT(수치 아님)."""
        from fpna.ingest.normalize import regex_type, infer_column_type
        self.assertEqual(regex_type("0001"), "TEXT")
        self.assertEqual(regex_type("100"), "NUM")
        self.assertEqual(infer_column_type(["0001", "0002", "0003"]), "TEXT")


if __name__ == "__main__":
    unittest.main()
