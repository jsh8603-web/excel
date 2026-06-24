#!/usr/bin/env python3
"""tests/test_design_audit.py — 디자인 게이트·edit_cell·restyle 회귀."""
import openpyxl
from openpyxl.styles import Alignment, Font

from fpna import design_audit, house_style as hs
from fpna.templates.base import QCReport


def test_standard_output_is_silent():
    wb = openpyxl.Workbook(); ws = wb.active
    hs.set_cell(ws, 1, 1, "Lease", role="label")
    hs.set_cell(ws, 1, 2, 1200, role="calc", number_format=hs.FMT_INT)
    rep = QCReport("t")
    assert design_audit.assert_design_standard(rep, wb) is True
    assert rep.checks[-1][1] is True  # ok


def test_decoration_hard_fails():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "*** 보고서 ***"; ws["A1"].font = Font(size=24)
    rep = QCReport("t")
    assert design_audit.assert_design_standard(rep, wb) is False


def test_number_left_align_reported():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["B5"] = 999; ws["B5"].alignment = Alignment(horizontal="left")
    f = design_audit.design_findings(wb)
    assert f["num_align"]


def test_edit_cell_preserves_alignment():
    wb = openpyxl.Workbook(); ws = wb.active
    hs.set_cell(ws, 1, 2, 1200, role="calc", number_format=hs.FMT_INT)
    hs.edit_cell(ws, "B1", 9999)
    assert ws["B1"].alignment.horizontal == "right"
    assert ws["B1"].value == 9999


def test_restyle_is_nondestructive():
    wb = openpyxl.Workbook(); ws = wb.active
    ws["A1"] = "*** x ***"
    ws["B5"] = 999; ws["B5"].alignment = Alignment(horizontal="left")
    ws["B6"] = "=SUM(B5:B5)"
    design_audit.restyle_inplace(wb)
    assert ws["B6"].value == "=SUM(B5:B5)"   # 수식 불변
    assert ws["B5"].value == 999             # 숫자 불변
    assert ws["B5"].alignment.horizontal == "right"
    assert ws["A1"].value == "x"             # 장식 제거
