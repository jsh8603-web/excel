"""
tests/test_conserve.py — T4 보존검증 N-version 스파인 + 변이 하니스.

실행: py -m unittest tests.test_conserve

자문 3R C4: "게이트가 실제로 잡는다는 증명 없는 게이트는 극장(theater)일 뿐."
변이 3종(끝행드롭=off-by-one / 부호1개반전 / 소계중복=이중계상)이 게이트를 trip
시키는지 확인 → 보존 게이트가 살아있음을 증명. + ConserveSpec 스윕 배선 + 모듈경계
독립성(ast) + fullCalcOnLoad + plausibility band.
"""
from __future__ import annotations

import ast
import dataclasses
import os
import tempfile
import unittest

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl import load_workbook

from fpna.conserve import ConserveSpec, eval_specs
from fpna.bands import in_band
from fpna.pipeline import run_report
from fpna.templates import get_template


# --------------------------------------------------------------------------- #
# ConserveSpec 선언형 스윕 단위                                                #
# --------------------------------------------------------------------------- #
class EvalSpecsTest(unittest.TestCase):
    def test_eval_basic(self):
        specs = [ConserveSpec("합", raw_sum_fn=lambda d: sum(d["xs"]), reported_key="tot")]
        got = eval_specs(specs, {"xs": [1, 2, 3]}, {"tot": 6})
        self.assertEqual(got, [("합", 6.0, 6, 0.5)])

    def test_eval_missing_key(self):
        specs = [ConserveSpec("합", raw_sum_fn=lambda d: 1.0, reported_key="없음")]
        name, lhs, rhs, tol = eval_specs(specs, {}, {})[0]
        self.assertIsNone(rhs)


class _FakeTemplate:
    """CONSERVE_SPECS 배선 테스트용 최소 템플릿(스파인 덕타이핑)."""
    TYPE = "_fake"
    CONSERVE_SPECS = [ConserveSpec("합계", raw_sum_fn=lambda d: sum(d["xs"]),
                                   reported_key="tot")]

    def __init__(self, reported):
        self._reported = reported

    def build(self, data, *, mode="create", base_path=None):
        wb = openpyxl.Workbook()
        wb.active["A1"] = self._reported
        wb._fpna_meta = {"tot": self._reported}
        return wb


class SpineConsumesSpecsTest(unittest.TestCase):
    def test_tie_passes(self):
        t = _FakeTemplate(reported=6)
        res = run_report(t, {"xs": [1, 2, 3]})
        self.assertTrue(res.qc.passed, res.qc.summary())

    def test_mismatch_blocks_save(self):
        t = _FakeTemplate(reported=999)        # 보고값 위조 → raw 합 6 과 불일치
        with tempfile.TemporaryDirectory() as d:
            res = run_report(t, {"xs": [1, 2, 3]}, out_path=os.path.join(d, "x.xlsx"))
        self.assertFalse(res.qc.passed)
        self.assertFalse(res.saved)            # 우회/오답 저장 차단


# --------------------------------------------------------------------------- #
# 변이 하니스 — 기존 conserves() 4종이 살아있는지(raw 변화에 trip)            #
# --------------------------------------------------------------------------- #
def _drop_last(contracts):
    return contracts[:-1]                       # off-by-one(끝행 드롭)


def _flip_one_sign(contracts):
    c0 = dataclasses.replace(contracts[0], amount_per_period=-contracts[0].amount_per_period)
    return [c0] + list(contracts[1:])           # 부호 1개 반전


def _dup_first(contracts):
    return list(contracts) + [contracts[0]]     # 소계 중복(이중계상)


class VariationHarnessTest(unittest.TestCase):
    """fc_maturity_wall: clean 으로 build → 변이된 raw 로 conserves 호출 시 불일치."""

    def setUp(self):
        self.mod = get_template("fc_maturity_wall")
        self.clean = self.mod.golden_sample()
        self.wb = self.mod.build(self.clean)     # 보고값(grand)은 clean 기준

    def _trips(self, mutated_contracts):
        mutated = dataclasses.replace(self.clean, contracts=mutated_contracts)
        items = self.mod.conserves(self.wb, mutated)   # raw=변이 / reported=clean
        return any(abs(raw - rep) > 1e-6 for _name, raw, rep in items)

    def test_drop_last_trips(self):
        self.assertTrue(self._trips(_drop_last(self.clean.contracts)))

    def test_flip_sign_trips(self):
        self.assertTrue(self._trips(_flip_one_sign(self.clean.contracts)))

    def test_dup_trips(self):
        self.assertTrue(self._trips(_dup_first(self.clean.contracts)))

    def test_clean_passes(self):
        """clean 은 통과해야(게이트가 무조건 trip 하는 가짜가 아님을 증명)."""
        items = self.mod.conserves(self.wb, self.clean)
        self.assertTrue(all(abs(raw - rep) <= 1e-6 for _n, raw, rep in items))


# --------------------------------------------------------------------------- #
# 모듈경계 독립성(ast) — conserve.py 가 build 패키지를 import 하지 않음        #
# --------------------------------------------------------------------------- #
class ModuleBoundaryTest(unittest.TestCase):
    def test_conserve_no_template_import(self):
        path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "fpna", "conserve.py")
        with open(path, encoding="utf-8") as fh:
            tree = ast.parse(fh.read())
        imported = []
        for node in ast.walk(tree):
            if isinstance(node, ast.ImportFrom) and node.module:
                imported.append(node.module)
            elif isinstance(node, ast.Import):
                imported += [a.name for a in node.names]
        bad = [m for m in imported if "templates" in m or m.endswith(".build")]
        self.assertEqual(bad, [], "conserve.py 가 build/templates 를 import(독립성 위배): %s" % bad)


# --------------------------------------------------------------------------- #
# fullCalcOnLoad(artifact-gap 완화) + band                                     #
# --------------------------------------------------------------------------- #
class ArtifactGuardTest(unittest.TestCase):
    def test_fullcalconload_set(self):
        mod = get_template("fc_maturity_wall")
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "x.xlsx")
            run_report(mod, mod.golden_sample(), out_path=p)
            wb = load_workbook(p)
            self.assertTrue(wb.calculation.fullCalcOnLoad)


class BandTest(unittest.TestCase):
    def test_rate_unit_confusion(self):
        self.assertFalse(in_band(5.0, "rate"))   # 500% = 단위혼동 의심
        self.assertTrue(in_band(0.05, "rate"))   # 5%
        self.assertTrue(in_band(123.0, "없는차원"))  # 미선언 = 보수 통과


if __name__ == "__main__":
    unittest.main()
