#!/usr/bin/env python3
"""
tools/styles_calibrate.py — 오프라인 1회 캘리브레이션(COM, 런타임 미사용).

DESIGN §6: openpyxl `cell.number_format`/`cell.font` (resolved) ↔ Excel-effective ↔
fpna.styles_interp (styles.xml applyX 해소) 3자를 대조해 "어느 속성이 충실히 읽히는가"
충실도맵을 1회 산출. 회사/집 무관하게 pywin32+Excel 있으면 실행. 런타임은 결과맵만 쓴다.

실행: py tools/styles_calibrate.py [fixture.xlsx]
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import fpna._bootstrap  # noqa: F401
import openpyxl
from fpna import styles_interp as si


def _make_fixture(path):
    """직접포맷(set_cell류) + named-style base 혼합 픽스처."""
    from fpna import house_style as hs
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "cal"
    hs.set_cell(ws, 2, 2, 1234, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, 3, 2, 0.25, role="input", number_format=hs.FMT_PCT1)
    hs.set_cell(ws, 4, 2, "라벨", role="label")
    wb.save(path)
    return [(2, 2), (3, 2), (4, 2)]


def calibrate(path):
    coords = _make_fixture(path) if not os.path.isfile(path) else None
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if coords is None:
        coords = [(c.row, c.column) for row in ws.iter_rows() for c in row if c.value is not None][:12]

    rows = []
    for (r, c) in coords:
        cell = ws.cell(r, c)
        eff = si.effective_of_cell(path, cell, ws.title)
        rows.append({
            "cell": cell.coordinate,
            "openpyxl_numfmt": cell.number_format,
            "interp_numfmt": eff.get("num_code"),
            "match_numfmt": cell.number_format == eff.get("num_code"),
            "interp_named": eff.get("named_style"),
        })

    com_rows = {}
    try:
        import win32com.client as win32
        xl = win32.gencache.EnsureDispatch("Excel.Application")
        xl.Visible = False; xl.DisplayAlerts = False
        wbx = xl.Workbooks.Open(os.path.abspath(path))
        wsx = wbx.Worksheets(1)
        for (r, c) in coords:
            rng = wsx.Cells(r, c)
            com_rows[(r, c)] = {"com_numfmt": rng.NumberFormat, "com_font": rng.Font.Name}
        wbx.Close(False); xl.Quit()
    except Exception as e:
        com_rows = {"_error": repr(e)[:80]}

    print("=" * 70)
    print("styles 캘리브레이션:", os.path.basename(path))
    print("=" * 70)
    for row in rows:
        print(row)
    print("\nCOM effective:", com_rows)
    nmatch = sum(1 for r in rows if r["match_numfmt"])
    print("\n충실도(numFmt openpyxl==interp): %d/%d" % (nmatch, len(rows)))
    print("→ 생성파일(set_cell 직접포맷)은 openpyxl-resolved==effective 기대. 불일치 시 styles_interp 사용.")
    return rows


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "out/zone/calib.xlsx"
    os.makedirs(os.path.dirname(p) or ".", exist_ok=True)
    calibrate(p)
