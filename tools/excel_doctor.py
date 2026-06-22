#!/usr/bin/env python3
"""
tools/excel_doctor.py — 이미 있는 .xlsx 를 한 번에 검사(+선택 수리)하는 로컬 도구.

무엇을 잡나
-----------
  1) 수식 에러 리터럴(#REF!/#DIV/0! 등)          [신뢰·자동수리 불가]
  2) 캐논 밖 number_format (house_style FMT_* 외)  [신뢰·--fix 로 수리]
  3) 같은 열 안 서식 불균형(outlier)              [휴리스틱·--fix 로 수리]
  4) 수치블록 cross-foot 의심(소계 ≠ 자식합)       [휴리스틱·자동수리 불가→재생성]

수리 경계(중요·정직)
--------------------
  ▶ 서식 문제(2·3)는 결정적으로 수리 가능 — house_style 캐논/열 다수서식으로 되돌린다.
  ▶ 값/수식 문제(1·4, "엉뚱한 칼럼" 류)는 **dead 파일만으로 자동수리 불가**다.
    올바른 값은 원천(소스 데이터)에서 다시 유도해야 하므로, 정답 경로는
    fpna.pipeline.run_report 로 **재생성**하는 것이다(그 경로엔 이제 formula_audit
    /layout_audit 게이트가 박혀 같은 버그를 재현 못 한다). doctor 는 이걸 탐지·보고만 한다.

사용
----
  py tools/excel_doctor.py <파일.xlsx>           # 검사만(리포트 + exit code)
  py tools/excel_doctor.py <파일.xlsx> --fix     # 서식만 수리 → <파일>.fixed.xlsx
  py tools/excel_doctor.py <파일.xlsx> --fix --inplace   # 원본 덮어쓰기(주의)

exit: 0 = 치명 이슈 없음, 1 = 발견(수리 못 한 값/수식 이슈 포함).
"""
from __future__ import annotations

import argparse
import os
import sys

# tools/ 에서 직접 실행 시 repo root 를 path 에 (import fpna 해소).
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import fpna._bootstrap  # noqa: F401

from openpyxl import load_workbook

from fpna import layout_audit
from fpna import formula_audit
from fpna.ingest.cells import ERROR_LITERALS


def _scan_error_cells(wb) -> list:
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERROR_LITERALS:
                    bad.append("%s!%s=%s" % (ws.title, c.coordinate, c.value))
    return bad


def _scan_disallowed_formats(wb) -> list:
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                if c.number_format not in layout_audit.CANON_FORMATS:
                    bad.append((ws.title, c.coordinate, c.number_format))
    return bad


def _scan_crossfoot(wb, *, abs_tol: float = 0.5) -> list:
    """휴리스틱 cross-foot: 굵게+상단테두리(=합계 관습) 숫자셀이 바로 위 연속 숫자블록의
    합과 다르면 의심 보고. 수식셀(문자열)은 건너뛴다(Excel 미계산이라 값 부재).
    """
    susp = []
    for ws in wb.worksheets:
        from collections import defaultdict
        colcells = defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    colcells[c.column].append((c.row, c.value, c))
        for col, items in colcells.items():
            items.sort()
            for i, (r, v, cell) in enumerate(items):
                is_totalish = bool(getattr(cell.font, "bold", False)) and \
                    getattr(cell.border, "top", None) is not None and \
                    getattr(cell.border.top, "style", None) is not None
                if not is_totalish or i == 0:
                    continue
                # 바로 위의 연속(행 인접) 숫자블록 합
                run = []
                pr = r
                for rr, vv, _cc in reversed(items[:i]):
                    if rr == pr - 1:
                        run.append(vv)
                        pr = rr
                    else:
                        break
                if len(run) >= 2 and abs(sum(run) - v) > abs_tol:
                    susp.append("%s!%s(=%g, 위 %d셀 합=%g)" %
                                (ws.title, cell.coordinate, v, len(run), sum(run)))
    return susp


def doctor(path: str, *, fix: bool = False, inplace: bool = False) -> int:
    path = os.path.abspath(path)
    if not os.path.isfile(path):
        print("파일 없음:", path)
        return 2
    wb = load_workbook(path)  # 수식=문자열 유지

    err = _scan_error_cells(wb)
    bad_fmt = _scan_disallowed_formats(wb)
    outliers = layout_audit.column_format_outliers(wb)
    crossfoot = _scan_crossfoot(wb)
    stray = layout_audit.text_in_numeric_columns(wb)
    unguarded = formula_audit.unguarded_divisions(wb)

    print("=" * 60)
    print("EXCEL DOCTOR:", os.path.basename(path))
    print("=" * 60)

    fatal = False

    print("\n[1] 수식 에러 리터럴: %s" % ("없음" if not err else "%d건" % len(err)))
    for x in err[:12]:
        print("    ✗", x)
    if err:
        fatal = True
        print("    → 자동수리 불가. 수식 참조 오류 — 소스에서 run_report 로 재생성 권장.")

    print("\n[2] 캐논 밖 number_format: %s" % ("없음" if not bad_fmt else "%d건" % len(bad_fmt)))
    for s, coord, nf in bad_fmt[:12]:
        print("    ✗ %s!%s = %r" % (s, coord, nf))

    print("\n[3] 열 내 서식 불균형(outlier): %s" %
          ("없음" if not outliers else "%d건" % len(outliers)))
    for o in outliers[:12]:
        print("    ✗ %s!%s = %r (열 다수서식 %r)" %
              (o["sheet"], o["coord"], o["found"], o["expected"]))

    print("\n[4] cross-foot 의심(소계≠자식합): %s  [자문성·exit 미반영]" %
          ("없음" if not crossfoot else "%d건" % len(crossfoot)))
    for x in crossfoot[:12]:
        print("    ⚠", x)
    if crossfoot:
        print("    → 휴리스틱 힌트. 부호규약(비용 차감)을 모르면 오탐 가능 — 단순합 전제.")
        print("      진짜 값 검증은 생성단계 게이트(conserve/formula_audit)가 소유.")

    print("\n[5] 숫자영역 텍스트 침투(헤더/주석 누수): %s" %
          ("없음" if not stray else "%d건" % len(stray)))
    for s in stray[:12]:
        print("    ✗ %s!%s = %r  [%s]" % (s["sheet"], s["coord"], s["value"], s["kind"]))
    if stray:
        fatal = True
        print("    → #VALUE! 의 상류 원인일 수 있음(분모 오염). 값 자동수리 불가 — 재생성.")

    print("\n[6] 가드 없는 나눗셈(#VALUE!/#DIV/0! 후보): %s  [자문성·exit 미반영]" %
          ("없음" if not unguarded else "%d건" % len(unguarded)))
    for u in unguarded[:12]:
        print("    ⚠ %s!%s = %s" % (u["sheet"], u["coord"], u["formula"]))
    if unguarded:
        print("    → CPU/CPP 류는 IF(ISNUMBER(분모)…) 가드 권장(formula_audit.safe_ratio_formula).")

    # ---- 수리(서식만) ----
    n_fixed = 0
    if fix and (bad_fmt or outliers):
        by_sheet = {ws.title: ws for ws in wb.worksheets}
        # (3) outlier → 열 다수서식으로
        for o in outliers:
            by_sheet[o["sheet"]][o["coord"]].number_format = o["expected"]
            n_fixed += 1
        # (2) 캐논 밖 → 같은 열 다수 캐논서식이 있으면 그걸로, 없으면 보수적 General 유지
        col_major = {}
        for o in outliers:
            col_major[(o["sheet"], o["col"])] = o["expected"]
        for s, coord, nf in bad_fmt:
            cell = by_sheet[s][coord]
            # 이미 outlier 수리로 캐논이 됐으면 skip
            if cell.number_format in layout_audit.CANON_FORMATS:
                continue
            ws = by_sheet[s]
            tgt = col_major.get((s, cell.column))
            if tgt:
                cell.number_format = tgt
                n_fixed += 1
        out = path if inplace else path.rsplit(".", 1)[0] + ".fixed.xlsx"
        wb.save(out)
        print("\n[FIX] 서식 %d건 수리 → %s" % (n_fixed, os.path.basename(out)))
        print("      (값/수식 이슈는 미수리 — 재생성 필요)")

    has_format_issue = bool(bad_fmt or outliers)
    print("\n" + "-" * 60)
    if not (fatal or has_format_issue):
        print("진단: 치명 이슈 없음 ✓")
        return 0
    if fix and not fatal and n_fixed:
        print("진단: 서식 이슈 수리 완료. 값/수식 치명 이슈 없음 ✓")
        return 0
    print("진단: 이슈 발견 — 위 항목 확인.")
    return 1


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="이미 있는 .xlsx 검사/수리")
    ap.add_argument("path")
    ap.add_argument("--fix", action="store_true", help="서식 이슈 자동수리")
    ap.add_argument("--inplace", action="store_true", help="원본 덮어쓰기(기본: .fixed.xlsx)")
    a = ap.parse_args(argv)
    return doctor(a.path, fix=a.fix, inplace=a.inplace)


if __name__ == "__main__":
    sys.exit(main())
