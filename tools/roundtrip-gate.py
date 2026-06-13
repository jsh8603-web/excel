#!/usr/bin/env python3
"""
roundtrip-gate.py — 워크북 변형이 '연결'을 끊지 않았는지 보증하는 무결성 게이트.

왜 필요한가(실측): openpyxl 라운드트립은 셀의 **수식 텍스트는 보존**하지만, 실제
외부링크 부품(xl/externalLinks/*)·차트·피벗·VBA 등은 **조용히 떨군다**. 수식만 보면
멀쩡해 보이는데 Excel 에선 끊긴 고아 참조가 된다. 그래서 셀 수준 + **zip 부품 수준**을
함께 지문화해 before↔after 를 비교한다. 신규 의존성 0(vendored openpyxl + stdlib).

쓰임:
  python3 tools/roundtrip-gate.py <before.xlsx> <after.xlsx>     # 직접 비교
  from roundtrip_gate import guard; guard(path, mutate_fn)        # 변형을 감싸 보증
Exit: 0 = 무결, 1 = 끊김(고아화) 검출.
"""
from __future__ import annotations
import sys, os, zipfile, argparse
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "vendor"))
from openpyxl import load_workbook  # noqa: E402

# 사라지면 '연결 끊김'을 의미하는 구조적 부품들
SIGNIFICANT = ("xl/externalLinks/", "xl/charts/", "xl/drawings/", "xl/pivotTables/",
               "xl/pivotCache/", "xl/calcChain.xml", "xl/vbaProject.bin", "xl/comments",
               # 외부 DB/쿼리 연결 — openpyxl 이 가장 잘 떨구는 부분
               "xl/connections.xml", "xl/queryTables/", "customXml/", "xl/model",
               # 다운스트림(Power BI/Tableau)이 바인딩하는 Excel Table
               "xl/tables/")
ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")

def _significant(part: str) -> bool:
    return any(part.startswith(p) or part == p for p in SIGNIFICANT)

def fingerprint(path: str) -> dict:
    parts = set(zipfile.ZipFile(path).namelist())
    fp = {"parts": parts, "sig_parts": {p for p in parts if _significant(p)},
          "formulas": {}, "names": set(), "sheets": [], "errors": set(),
          "tables": {}, "headers": {}}
    wb = load_workbook(path, data_only=False, keep_links=True)
    try:
        fp["sheets"] = list(wb.sheetnames)
        try: fp["names"] = set(wb.defined_names.keys())
        except Exception: fp["names"] = set(getattr(wb, "defined_names", []) or [])
        for ws in wb.worksheets:
            # 다운스트림 바인딩 계약: Excel Table(이름→범위) + 헤더행
            try:
                for tname, t in dict(ws.tables).items():
                    fp["tables"][f"{ws.title}!{tname}"] = getattr(t, "ref", None) or str(t)
            except Exception:
                pass
            try:
                fp["headers"][ws.title] = tuple(c.value for c in ws[1])
            except Exception:
                pass
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str):
                        if v.startswith("="):
                            fp["formulas"][f"{ws.title}!{c.coordinate}"] = v
                        if any(t in v for t in ERROR_TOKENS):
                            fp["errors"].add(f"{ws.title}!{c.coordinate}")
    finally:
        wb.close()
    return fp

def compare(before: dict, after: dict) -> list:
    issues = []
    dropped_sig = before["sig_parts"] - after["parts"]
    for p in sorted(dropped_sig):
        issues.append(f"구조 부품 소실(고아화): {p}")
    dropped_f = set(before["formulas"]) - set(after["formulas"])
    for k in sorted(dropped_f):
        issues.append(f"수식 소실: {k}  (was {before['formulas'][k]})")
    changed_f = [k for k in (set(before["formulas"]) & set(after["formulas"]))
                 if before["formulas"][k] != after["formulas"][k]]
    for k in sorted(changed_f):
        issues.append(f"수식 변경: {k}  {before['formulas'][k]} → {after['formulas'][k]}")
    for n in sorted(before["names"] - after["names"]):
        issues.append(f"정의된 이름 소실: {n}")
    for s in [x for x in before["sheets"] if x not in after["sheets"]]:
        issues.append(f"시트 소실(다운스트림 바인딩 깨짐): {s}")
    # 다운스트림(Power BI/Tableau) 바인딩 계약
    for t in sorted(set(before.get("tables", {})) - set(after.get("tables", {}))):
        issues.append(f"Excel Table 소실(바인딩 깨짐): {t}")
    for t in (set(before.get("tables", {})) & set(after.get("tables", {}))):
        if before["tables"][t] != after["tables"][t]:
            issues.append(f"Table 범위 변경: {t}  {before['tables'][t]} → {after['tables'][t]}")
    for sh, hdr in before.get("headers", {}).items():
        if sh in after.get("headers", {}) and after["headers"][sh] != hdr:
            issues.append(f"헤더 변경(컬럼 바인딩 깨짐): {sh}  {hdr} → {after['headers'][sh]}")
    for e in sorted(after["errors"] - before["errors"]):
        issues.append(f"신규 에러 셀: {e}")
    return issues

def guard(path: str, mutate_fn):
    """mutate_fn(path) 을 실행하되, 전후 무결성을 보증. 위반 시 RuntimeError."""
    before = fingerprint(path)
    mutate_fn(path)
    after = fingerprint(path)
    issues = compare(before, after)
    if issues:
        raise RuntimeError("라운드트립 무결성 위반:\n  - " + "\n  - ".join(issues))
    return True

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("before"); ap.add_argument("after")
    a = ap.parse_args()
    issues = compare(fingerprint(a.before), fingerprint(a.after))
    if not issues:
        print("ROUNDTRIP GATE PASS: 수식·외부링크·구조 부품·정의이름 모두 보존")
        return 0
    for i in issues:
        print(f"  ✗ {i}")
    print(f"ROUNDTRIP GATE FAIL: {len(issues)}건 (연결 끊김/고아화)")
    return 1

if __name__ == "__main__":
    sys.exit(main())
