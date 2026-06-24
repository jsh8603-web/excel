#!/usr/bin/env python3
"""tools/kosis_fetch.py — KOSIS OpenAPI 실데이터 → KOSIS portal export 포맷 .xlsx 렌더.

ingest 파이프라인 스트레스 테스트용. *데이터는 통계청 실값*, 레이아웃은 KOSIS 실제
export 의 누더기 구조(병합 2행 헤더 + 다기간 블록 + 좌측 다중 키컬럼 ditto + '-' 센티넬)를
충실 재현한다. 합성 수치 없음 — DT 는 KOSIS API 원값 그대로.

검증 보조 도구(tools/)라 stdlib + urllib 만 사용(런타임 fpna/ 와 무관).

usage:
  py tools/kosis_fetch.py <orgId> <tblId> <out.xlsx> [--periods N] [--prdse M|Y|Q|D]
"""
from __future__ import annotations

import json
import os
import sys
import urllib.parse
import urllib.request

import openpyxl

API = "https://kosis.kr/openapi/Param/statisticsParameterData.do"
# ⛔ API 키 하드코딩 금지(public 재전환 대비). 환경변수 KOSIS_API_KEY 로 주입.
#   회사/집 무관: `export KOSIS_API_KEY=...` (KOSIS 키는 base64 형태 그대로가 키).
KEY = os.environ.get("KOSIS_API_KEY", "")


def _get(params: dict) -> object:
    q = urllib.parse.urlencode(params, safe="")
    url = "%s?%s" % (API, q)
    with urllib.request.urlopen(url, timeout=40) as r:
        raw = r.read().decode("utf-8")
    return json.loads(raw)


def fetch(org_id: str, tbl_id: str, *, periods: int = 3, prdse: str = "M") -> list:
    """objL 차수를 모르므로 L1→L1L2→L1L2L3 순으로 시도, 첫 데이터 응답 채택."""
    base = {"method": "getList", "apiKey": KEY, "orgId": org_id, "tblId": tbl_id,
            "itmId": "ALL", "prdSe": prdse, "newEstPrdCnt": str(periods),
            "format": "json", "jsonVD": "Y"}
    last_err = None
    for objspec in ({"objL1": "ALL", "objL2": "ALL", "objL3": "ALL"},
                    {"objL1": "ALL", "objL2": "ALL"},
                    {"objL1": "ALL"}):
        p = dict(base, **objspec)
        try:
            out = _get(p)
        except Exception as e:                      # noqa: BLE001
            last_err = str(e); continue
        if isinstance(out, list) and out and "DT" in out[0]:
            return out
        if isinstance(out, dict):
            last_err = out.get("errMsg", str(out))
    raise RuntimeError("fetch 실패 %s/%s: %s" % (org_id, tbl_id, last_err))


def _fmt_period(prd: str) -> str:
    s = str(prd)
    if len(s) == 6 and s.isdigit():
        return "%s.%s" % (s[:4], s[4:])             # 202605 → 2026.05 (KOSIS 표기)
    return s


def _row_dims(rec: dict) -> list:
    """레코드에 존재하는 분류차수(C1_NM, C2_NM, ...) 순서대로."""
    dims = []
    for i in range(1, 6):
        if ("C%d_NM" % i) in rec:
            dims.append(i)
    return dims


def render(records: list, out_path: str, *, tbl_nm: str = "", tbl_id: str = "") -> dict:
    """KOSIS export 포맷으로 렌더. 반환 = 검증용 메타(행수/항목수/기간수/센티넬수)."""
    dims = _row_dims(records[0])
    periods = sorted({_fmt_period(r["PRD_DE"]) for r in records})
    # 항목: 첫 등장 순서 보존
    items, seen = [], set()
    for r in records:
        it = r.get("ITM_NM", "")
        if it not in seen:
            seen.add(it); items.append(it)
    units = {}
    for r in records:
        u = r.get("UNIT_NM", "") or ""
        units[u] = units.get(u, 0) + 1
    unit = max(units, key=units.get) if units else ""

    # 행키: 분류값 튜플(첫 등장 순서). 값 lookup = (rowkey, period, item)→DT
    rowkeys, rk_seen = [], set()
    lookup = {}
    for r in records:
        rk = tuple(r.get("C%d_NM" % i, "") for i in dims)
        if rk not in rk_seen:
            rk_seen.add(rk); rowkeys.append(rk)
        per = _fmt_period(r["PRD_DE"])
        lookup[(rk, per, r.get("ITM_NM", ""))] = r.get("DT")
    obj_names = [records[0].get("C%d_OBJ_NM" % i, "분류%d" % i) for i in dims]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "데이터"
    nkey = len(dims)

    # --- 헤더: row1 = 키 obj명(2행 병합) + 기간(항목 수만큼 병합) / row2 = 항목명 ---
    for ci, nm in enumerate(obj_names, start=1):
        ws.cell(1, ci, nm)
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)
    # 2행 헤더: row1 = 기간(항목 2+ 시 가로 병합), row2 = 항목명. 단일 항목도 동일 경로
    # (항목명은 항상 row2 에 박제 — KOSIS 실export 처럼 지표명이 헤더에 존재).
    col = nkey + 1
    for per in periods:
        start = col
        for it in items:
            lbl = "%s (%s)" % (it, unit) if unit else it
            ws.cell(2, col, lbl); col += 1
        ws.cell(1, start, per)
        if col - 1 > start:                          # 항목 2+ 일 때만 기간을 항목들 위로 병합
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)

    # --- 데이터행: 좌측 다중키 ditto(바깥 차원 반복 공백), 결측 '-' 센티넬 ---
    n_sent = 0
    prev = [None] * nkey
    rrow = 3
    for rk in rowkeys:
        for ci in range(nkey):
            val = rk[ci]
            # 바깥 차원(마지막 키 제외)은 직전과 같으면 ditto 공백
            if ci < nkey - 1 and val == prev[ci]:
                ws.cell(rrow, ci + 1, "")
            else:
                ws.cell(rrow, ci + 1, val)
                for k in range(ci + 1, nkey):
                    prev[k] = None                   # 바깥이 바뀌면 안쪽 ditto 리셋
        prev = list(rk)
        col = nkey + 1
        col_combo = [(per, it) for per in periods for it in items]
        for (per, it) in col_combo:
            dt = lookup.get((rk, per, it))
            if dt is None or dt == "":
                ws.cell(rrow, col, "-"); n_sent += 1
            else:
                try:
                    ws.cell(rrow, col, float(dt))
                except (TypeError, ValueError):
                    ws.cell(rrow, col, dt)
            col += 1
        rrow += 1

    # --- 메타정보 시트(KOSIS 실export 동봉) ---
    mw = wb.create_sheet("메타정보")
    mw.cell(1, 1, "○ 통계표ID"); mw.cell(1, 2, tbl_id)
    mw.cell(2, 1, "○ 통계표명"); mw.cell(2, 2, tbl_nm)
    mw.cell(3, 1, "○ 단위"); mw.cell(3, 2, unit)
    mw.cell(4, 1, "○ 자료출처"); mw.cell(4, 2, "KOSIS 국가통계포털")

    wb.save(out_path)
    return {"n_rows": len(rowkeys), "n_items": len(items), "n_periods": len(periods),
            "n_dims": nkey, "n_sentinel": n_sent, "unit": unit,
            "n_data_cells": len(rowkeys) * len(periods) * len(items)}


def main(argv):
    if not KEY:
        print("환경변수 KOSIS_API_KEY 미설정 — `export KOSIS_API_KEY=<키>` 후 재실행.",
              file=sys.stderr)
        return 2
    org_id, tbl_id, out_path = argv[1], argv[2], argv[3]
    periods, prdse = 3, "M"
    i = 4
    while i < len(argv):
        if argv[i] == "--periods":
            periods = int(argv[i + 1]); i += 2
        elif argv[i] == "--prdse":
            prdse = argv[i + 1]; i += 2
        else:
            i += 1
    recs = fetch(org_id, tbl_id, periods=periods, prdse=prdse)
    tbl_nm = recs[0].get("TBL_NM", "")
    meta = render(recs, out_path, tbl_nm=tbl_nm, tbl_id=tbl_id)
    print("OK %s | %s rec → %s" % (tbl_id, len(recs), out_path))
    print("   ", json.dumps(meta, ensure_ascii=False))


if __name__ == "__main__":
    sys.exit(main(sys.argv))
