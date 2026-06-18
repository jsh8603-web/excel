"""
tools/dirty_excel_gen.py — 누더기 엑셀 fuzz 생성기 (clean→dirty + oracle).

dev 전용(openpyxl — vendored 아님). fpna/ 런타임은 import 하지 않는다.

아이디어: 깨끗한 fact(entity×period→value)를 먼저 만들고, 누더기 패턴을 시드 기반으로
무작위 적용해 dirty.xlsx 를 만든다. 동시에 '복원 기대값'(expected_*.csv, oracle)을 박제한다.
→ ingest(dirty) 결과를 expected 와 비교하면 크래시뿐 아니라 '정확도'까지 검증된다.

누더기 패턴(시드별 무작위 on/off):
  단위행+스케일 / 정수 vs 문자열 연도헤더 / 소계행(볼드) / @표시형식 숫자 /
  헤더 각주마커(*) / 라벨 ditto(반복 빈칸)

CLI:
  python dirty_excel_gen.py --n 30 --out out/dirty_corpus   # dirty_{seed}.xlsx + _expected.csv 쌍 N개
"""
from __future__ import annotations

import argparse
import csv
import os
import random

import openpyxl
from openpyxl.styles import Font

_ACCOUNTS = ["매출", "비용", "이익", "자산", "부채", "자본", "현금"]
_YEARS = [2021, 2022, 2023, 2024, 2025]


def gen_one(seed: int, out_xlsx: str, out_expected: str) -> dict:
    """시드 1개 → dirty.xlsx + expected.csv. 적용한 누더기 옵션 dict 반환."""
    rng = random.Random(seed)
    accounts = _ACCOUNTS[:rng.randint(3, 6)]
    years = _YEARS[-rng.randint(2, 4):]

    # 누더기 옵션(시드별 무작위)
    opt = {
        "unit_scale": rng.random() < 0.6,      # 단위행 + 표시값 축소
        "int_year": rng.random() < 0.5,        # 정수 연도헤더(아니면 문자열)
        "subtotal": rng.random() < 0.4,        # 소계행(볼드)
        "at_format": rng.random() < 0.4,       # 일부 셀 @(텍스트)형식 숫자
        "footnote": rng.random() < 0.3,        # 헤더 각주마커
        "ditto": rng.random() < 0.3,           # 첫 열 라벨 일부 빈칸(상동)
    }
    scale = 1_000_000 if opt["unit_scale"] else 1

    wb = openpyxl.Workbook()
    ws = wb.active
    r = 1
    if opt["unit_scale"]:
        ws.cell(r, 1, "(단위: 백만원)")
        r += 1

    # 헤더행: 계정 | 연도들
    foot = "*" if opt["footnote"] else ""
    ws.cell(r, 1, "계정" + foot)
    for j, y in enumerate(years):
        ws.cell(r, 2 + j, (y if opt["int_year"] else str(y)))
    r += 1

    expected: list[tuple[str, str, float]] = []
    prev_acc = None
    for a in accounts:
        # ditto: 같은 계정 반복 시 라벨 빈칸(여기선 단일이라 첫 열 약하게 흉내)
        label = a
        if opt["ditto"] and prev_acc is not None and rng.random() < 0.3:
            label = None
        ws.cell(r, 1, label)
        prev_acc = a
        for j, y in enumerate(years):
            disp = rng.randint(100, 9999)         # 표시값(스케일 적용 전)
            val = disp * scale                    # 복원 기대 원본값
            cell = ws.cell(r, 2 + j, str(disp) if opt["at_format"] and rng.random() < 0.5 else disp)
            if opt["at_format"] and isinstance(cell.value, str):
                cell.number_format = "@"
            expected.append((a, str(y), float(val)))   # entity=계정, period=연도, value=원본
        r += 1

    if opt["subtotal"]:
        # 소계행(볼드) — oracle 에서는 제외(row_role=subtotal 로 분류될 것)
        ws.cell(r, 1, "소계").font = Font(bold=True)
        for j in range(len(years)):
            ws.cell(r, 2 + j, rng.randint(1000, 99999))
        r += 1

    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)), exist_ok=True)
    wb.save(out_xlsx)
    with open(out_expected, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["entity", "period", "value"])
        w.writerows(expected)
    return opt


def gen_corpus(n: int, out_dir: str, seed0: int = 1) -> list[dict]:
    os.makedirs(out_dir, exist_ok=True)
    metas = []
    for i in range(n):
        seed = seed0 + i
        xlsx = os.path.join(out_dir, "dirty_%03d.xlsx" % seed)
        exp = os.path.join(out_dir, "dirty_%03d_expected.csv" % seed)
        opt = gen_one(seed, xlsx, exp)
        metas.append({"seed": seed, "xlsx": xlsx, **opt})
    return metas


def _main(argv=None):
    p = argparse.ArgumentParser(description="누더기 엑셀 fuzz 생성기 (dev 전용)")
    p.add_argument("--n", type=int, default=20)
    p.add_argument("--out", default="out/dirty_corpus")
    p.add_argument("--seed0", type=int, default=1)
    args = p.parse_args(argv)
    metas = gen_corpus(args.n, args.out, args.seed0)
    on = lambda k: sum(1 for m in metas if m[k])  # noqa: E731
    print("생성: %d개 → %s" % (len(metas), args.out))
    print("패턴 분포: unit=%d int_year=%d subtotal=%d at_fmt=%d footnote=%d ditto=%d"
          % (on("unit_scale"), on("int_year"), on("subtotal"),
             on("at_format"), on("footnote"), on("ditto")))


if __name__ == "__main__":
    _main()
