#!/usr/bin/env python3
"""
format-in-place.py — 이미 연결된 워크북에 '서식만' 입히는 비파괴 패스.

원칙(실측 기반):
  - 로드는 수식 보존(data_only=False), keep_links=True.
  - **셀 값/수식을 절대 다시 쓰지 않는다.** number_format/font/fill/border/alignment 등 스타일 속성만.
  - 저장 후 roundtrip-gate 로 무결성 보증 → 위반 시 raise(원본 비변경은 호출측 책임).

중요한 한계(정직): openpyxl save 는 스타일만 만져도 외부링크 부품(xl/externalLinks/*)을 떨군다.
따라서 **외부링크가 있는 워크북은 이 패스가 게이트에서 거부**된다(고아화 직전 차단). 그런 파일은
탭 단일 워크북으로 통합하거나 읽기전용으로 두어야 한다. 탭간(intra-workbook) 전용 파일은 안전하게 통과.

쓰임: python3 tools/format-in-place.py <workbook.xlsx>
Exit: 0 = 서식 적용+무결, 1 = 무결성 위반(외부링크 등 → 통합 필요).
"""
from __future__ import annotations
import sys, os, argparse, shutil, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "vendor"))
from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import Font, Alignment  # noqa: E402

# roundtrip-gate 재사용
import importlib.util
_spec = importlib.util.spec_from_file_location("rtg", os.path.join(os.path.dirname(__file__), "roundtrip-gate.py"))
rtg = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(rtg)

NUM_FMT = "#,##0;(#,##0)"   # 천단위 콤마 + 괄호 음수(한국 회계 관행)

def apply_house_style(path: str) -> None:
    """스타일 속성만 변경. cell.value 는 절대 건드리지 않는다."""
    wb = load_workbook(path, data_only=False, keep_links=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    # 숫자(수식 결과 포함)에 회계 서식 — 단, 값은 재기록하지 않음
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        c.number_format = NUM_FMT
                        c.alignment = Alignment(horizontal="right")
                    # 헤더행(1행) 굵게 — 텍스트만
                    if c.row == 1 and isinstance(v, str):
                        c.font = Font(bold=True)
                        c.alignment = Alignment(horizontal="left")
        wb.save(path)            # ← 여기서 openpyxl 이 외부링크 부품을 떨굴 수 있음(게이트가 잡음)
    finally:
        wb.close()

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--in-place", action="store_true",
                    help="원본을 직접 수정(기본은 .styled.xlsx 사본에 적용)")
    a = ap.parse_args()
    target = a.workbook
    if not a.in_place:
        target = os.path.splitext(a.workbook)[0] + ".styled.xlsx"
        shutil.copy(a.workbook, target)
    try:
        rtg.guard(target, apply_house_style)   # 서식 적용 + 무결성 보증
    except RuntimeError as e:
        print(f"서식 적용 거부 — {e}")
        print("→ 외부링크가 있는 워크북이면 단일 워크북으로 통합 후 재시도하거나, 읽기전용으로 두세요.")
        if not a.in_place and os.path.exists(target):
            os.remove(target)   # 고아 산출물 남기지 않음
        return 1
    print(f"서식 적용 + 무결성 통과 → {target}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
