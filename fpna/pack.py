"""
fpna.pack — 다중 exhibit 팩(연동 워크페이퍼) 오케스트레이터.

R2(FAST Control 시트)·R7(PFRAM 다중시트 연동)·R8(Doubletalk ledger 대안) 배선.
report.py(SheetSpec/build_report)가 *임의 builder* 다중시트라면, pack 은 *완전 템플릿
exhibit*(_MODULES 키)을 시트당 1장으로 묶는다. 연동은 셀참조가 아니라 **공유 facts** —
각 exhibit 가 단일 가정집합(공유 calendar/dims/assumptions)에서 build 되고, 시트 간
정합은 ConserveSpec(공유 facts dict source, build 호출 0)으로 스파인이 강제한다.

핵심:
  - 각 exhibit.build() 는 독립 wb 를 만든다 → `_graft_sheet`로 단일 pack wb 에 합본.
  - ★openpyxl 제약(정직 박제): 차트/조건부서식은 cross-wb Reference 가 깨져 graft 에서
    누락될 수 있다. 값·수식·서식·병합·너비는 보존. 수식 정확성은 fullCalcOnLoad +
    tools/verify_xlsx(COM) 가 담당(B 와 동일 경계).
  - 첫 시트 = Control(Index): exhibit↔시트 맵 + cross ties + 모델체크(A=L+E·현금 tie)를
    메모리값 OK/XX(조건부 초록/빨강)로 표면화.
  - build_pack 은 pipeline.run_report 스파인을 경유 → receipt 없이는 저장 불가(우회 차단).
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import finance, house_style as hs
from fpna.pipeline import run_report, RunResult
from fpna.templates import get_template


# --------------------------------------------------------------------------- #
# 선언                                                                         #
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class ExhibitSpec:
    """팩 구성 장표 1개. type_name=_MODULES 키, data=해당 템플릿 INPUT."""
    type_name: str
    data: object
    title: str = ""             # 시트 제목(없으면 type_name)
    section: str = "detail"


@dataclass(frozen=True)
class ModelCheck:
    """Control 모델체크 1개 — 공유 facts 의 lhs == rhs(±tol). 메모리값 OK/XX."""
    name: str
    lhs_key: str
    rhs_key: str
    tol: float = 0.5


@dataclass(frozen=True)
class PackSpec:
    """다중 exhibit 팩 선언(frozen)."""
    name: str
    title: str
    exhibits: tuple                              # tuple[ExhibitSpec]
    shared_facts: dict = field(default_factory=dict)   # cross tie/모델체크 source(평탄 dict)
    cross_ties: tuple = ()                       # tuple[ConserveSpec] over shared_facts
    model_checks: tuple = ()                     # tuple[ModelCheck]
    subtitle: str = ""
    as_of: str = ""
    unit: str = "₩"
    ledger_mode: bool = False                    # (옵션) Doubletalk 원장 posting — 기본 off


# --------------------------------------------------------------------------- #
# graft — cross-wb 시트 합본                                                   #
# --------------------------------------------------------------------------- #
def _graft_sheet(src_ws, dst_ws) -> int:
    """src_ws 의 셀값·스타일·병합·너비·freeze 를 dst_ws 로 복사. 차트/조건부서식 제외.

    반환 = 복사한 셀 수(비어있지 않거나 스타일 있는). 차트 누락은 호출측이 집계.
    """
    n = 0
    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            d = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                d.font = copy(cell.font)
                d.fill = copy(cell.fill)
                d.border = copy(cell.border)
                d.alignment = copy(cell.alignment)
                d.number_format = cell.number_format
                d.protection = copy(cell.protection)
            n += 1
    for mc in list(src_ws.merged_cells.ranges):
        dst_ws.merge_cells(str(mc))
    for key, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dst_ws.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[key].height = dim.height
    dst_ws.freeze_panes = src_ws.freeze_panes
    try:
        dst_ws.sheet_view.zoomScale = src_ws.sheet_view.zoomScale
    except Exception:
        pass
    return n


# --------------------------------------------------------------------------- #
# 조립                                                                         #
# --------------------------------------------------------------------------- #
def _eval_model_checks(spec: PackSpec) -> list:
    """공유 facts 로 모델체크 평가 → [(name, ok, lhs, rhs, detail), ...]."""
    out = []
    f = spec.shared_facts
    for mc in spec.model_checks:
        lhs = f.get(mc.lhs_key)
        rhs = f.get(mc.rhs_key)
        if lhs is None or rhs is None:
            out.append((mc.name, False, lhs, rhs,
                        "facts 부재(%s/%s)" % (mc.lhs_key, mc.rhs_key)))
        else:
            ok = abs(float(lhs) - float(rhs)) <= mc.tol
            out.append((mc.name, ok, lhs, rhs,
                        "" if ok else "Δ=%.6g" % (float(lhs) - float(rhs))))
    return out


def _build_control(ws, spec: PackSpec, sheet_titles: list, checks: list,
                   chart_drops: int) -> None:
    """Control(Index) 시트: exhibit 맵 + 모델체크 OK/XX + cross ties 목록."""
    last_col = 5
    hs.set_widths(ws, {1: 28, 2: 22, 3: 16, 4: 16, 5: 12})
    r = hs.report_frame(ws, spec.title, subtitle=spec.subtitle or "Control / Index",
                        unit=spec.unit, as_of=spec.as_of, last_col=last_col)
    ws.sheet_properties.tabColor = "2E5A87"

    # exhibit 맵
    r = hs.section_header(ws, r, "장표 구성 (Exhibits)", last_col=last_col)
    for j, h in enumerate(("장표 유형", "시트명", "구분"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j <= 2 else hs.CENTER)
    r += 1
    for ex, title in zip(spec.exhibits, sheet_titles):
        hs.set_cell(ws, r, 1, ex.type_name, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, "='%s'!A1" % title, role="link", align=hs.LEFT)
        hs.set_cell(ws, r, 3, ex.section, role="soft", align=hs.CENTER)
        r += 1

    # 모델체크 (메모리값 OK/XX)
    r = hs.section_header(ws, r + 1, "모델 체크 (Model Checks)", last_col=last_col)
    for j, h in enumerate(("점검", "LHS", "RHS", "결과"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1
    for name, ok, lhs, rhs, detail in checks:
        hs.set_cell(ws, r, 1, name, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, lhs if lhs is not None else "—", role="calc",
                    number_format=hs.FMT_INT, align=hs.CENTER)
        hs.set_cell(ws, r, 3, rhs if rhs is not None else "—", role="calc",
                    number_format=hs.FMT_INT, align=hs.CENTER)
        hs.set_cell(ws, r, 4, "OK" if ok else "XX", role="total", align=hs.CENTER)
        ws.cell(row=r, column=4).font = hs.font(hs.POS_FG if ok else hs.NEG_FG, bold=True)
        if detail:
            hs.set_cell(ws, r, 5, detail, role="soft", align=hs.LEFT)
        r += 1

    # cross ties 목록(스파인이 강제 — 여기선 선언 표시)
    if spec.cross_ties:
        r = hs.section_header(ws, r + 1, "크로스시트 정합 (Cross Ties — 스파인 강제)",
                              last_col=last_col)
        for ct in spec.cross_ties:
            hs.set_cell(ws, r, 1, "• " + ct.name, role="soft", align=hs.LEFT)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
            r += 1

    note = "graft 차트/조건부서식 누락 %d (openpyxl 제약 — 값·수식 보존, COM/verify_xlsx 로 보완)" % chart_drops
    hs.report_footer(ws, r + 1, source="공유 가정집합 · pack.PackSpec",
                     note=note, prepared_by="FP&A", last_col=last_col)


def _assemble(spec: PackSpec) -> openpyxl.Workbook:
    """exhibit 들을 graft 합본 + Control 시트 구성. wb._fpna_meta 에 facts·검증 노출."""
    wb = openpyxl.Workbook()
    ctrl = wb.active
    ctrl.title = hs.safe_sheet_title("Control")

    sheet_titles: list = []
    facts_by_sheet: dict = {}
    chart_drops = 0
    used_titles: set = set()
    for ex in spec.exhibits:
        mod = get_template(ex.type_name)
        sub = mod.build(ex.data)
        src = sub.active
        base = hs.safe_sheet_title(ex.title or ex.type_name)
        title = base
        k = 2
        while title in used_titles:                  # 31자 잘림 충돌 회피
            title = hs.safe_sheet_title("%s_%d" % (base[:28], k))
            k += 1
        used_titles.add(title)
        dst = wb.create_sheet(title)
        _graft_sheet(src, dst)
        chart_drops += len(getattr(src, "_charts", []))
        facts_by_sheet[ex.type_name] = getattr(sub, "_fpna_meta", {}) or {}
        sheet_titles.append(title)

    checks = _eval_model_checks(spec)
    _build_control(ctrl, spec, sheet_titles, checks, chart_drops)

    wb._fpna_meta = {
        "shared_facts": dict(spec.shared_facts),
        "facts_by_sheet": facts_by_sheet,
        "model_checks": checks,
        "chart_drops": chart_drops,
        # cross_ties reported_key 가 평탄 shared_facts 를 읽도록 meta 에도 평탄값 심음
        **{k: v for k, v in spec.shared_facts.items()},
    }
    wb.calculation.fullCalcOnLoad = True
    return wb


# --------------------------------------------------------------------------- #
# 스파인 어댑터 (ReportTemplate 덕타이핑)                                       #
# --------------------------------------------------------------------------- #
class _PackRunner:
    """PackSpec → run_report 가 받는 template 객체. cross_ties = CONSERVE_SPECS 노출."""

    def __init__(self, spec: PackSpec):
        self.spec = spec
        self.TYPE = "pack:" + spec.name
        self.CONSERVE_SPECS = list(spec.cross_ties)

    def build(self, data, *, mode: str = "create", base_path=None):
        return _assemble(self.spec)

    def qc(self, wb, data):
        from fpna.templates.base import QCReport, qc_no_formula_errors
        rep = QCReport(self.TYPE)
        qc_no_formula_errors(wb, rep)
        # 모델체크 전수 통과(메모리값) — BS 항등·현금 tie 등
        for name, ok, lhs, rhs, detail in wb._fpna_meta.get("model_checks", []):
            rep.add("모델체크:%s" % name, ok, detail)
        rep.add("Control 시트 존재", wb.worksheets[0].title.startswith("Control"), "")
        rep.add("exhibit 시트 수", len(wb.worksheets) - 1 == len(self.spec.exhibits),
                "" if len(wb.worksheets) - 1 == len(self.spec.exhibits)
                else "시트 %d ≠ exhibit %d" % (len(wb.worksheets) - 1, len(self.spec.exhibits)))
        return rep


def build_pack(spec: PackSpec, *, out_path: str | None = None,
               force: bool = False) -> RunResult:
    """팩 빌드 → run_report 스파인 경유(receipt 없이 저장 불가). RunResult 반환."""
    runner = _PackRunner(spec)
    return run_report(runner, spec, out_path=out_path, force=force)


__all__ = ["ExhibitSpec", "ModelCheck", "PackSpec", "build_pack",
           "_graft_sheet", "_assemble", "_PackRunner"]
