#!/usr/bin/env python3
"""
fpna/design_audit.py — 디자인 표준(FAST/ICAEW/Macabacus) 준수 게이트 + 비파괴 정규화.

layout_audit(서식 캐논·content-type)와 짝을 이루는 *디자인* 계층. house_style 가 생성 측
SSOT 라면, 여기는 그 표준을 *어겼는지* 잡는 검사 측이다. 토큰(허용 폰트·역할 색)은
fpna.house_style 에서 직접 읽어 SSOT 와 정합(드리프트 없음).

무엇을 잡나(보수적·저오탐)
  · 장식문자(별표/마크다운 강조/장식선) — fpna 스타일은 간결 라벨(FAST 단순성)
  · 숫자 셀 좌/가운데 정렬 — 숫자는 우측정렬(FAST/Macabacus)
  · 비표준/과대 폰트 — house_style ALLOWED_SIZES/SIZE_TITLE 밖(ICAEW cell styles)
  · 헤더 근처(상단 5행) 장문 설명 — 주석은 전용 notes/cover 로(ICAEW 구조)

게이트(assert_design_standard)는 layout_audit 과 같은 QCReport 패턴. 기본은 장식문자만
hard-fail(가장 명확한 위반), 나머지는 보고. restyle_inplace 는 *서식만* 비파괴 정규화.
"""
from __future__ import annotations

import re

from fpna import house_style as hs
from fpna.templates.base import QCReport

_DECOR = re.compile(r"(\*\*.+\*\*|\*{3,}|[★☆■◆▶●]{1,}\s*\S|^\s*[\*\-=]{3,})")
_DECOR_STRIP = re.compile(r"^\s*[\*★☆■◆▶●=\-_]{1,}\s*|\s*[\*★☆■◆▶●=\-_]{1,}\s*$")
_MD_EMPH = re.compile(r"\*\*(.+?)\*\*")


def _isnum(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def design_findings(wb) -> dict:
    """디자인 위반을 범주별로 수집. 반환 {decoration, num_align, font, annotation}."""
    title_max = float(hs.SIZE_TITLE)
    allowed = set(hs.ALLOWED_SIZES)
    f = {"decoration": [], "num_align": [], "font": [], "annotation": []}
    for ws in wb.worksheets:
        long_top = 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and not _is_formula(v) and _DECOR.search(v):
                    f["decoration"].append("%s!%s=%r" % (ws.title, c.coordinate, v[:30]))
                if _isnum(v) and c.alignment and c.alignment.horizontal in ("left", "center"):
                    f["num_align"].append("%s!%s(%s)" % (ws.title, c.coordinate, c.alignment.horizontal))
                if c.font and c.font.size and v not in (None, ""):
                    if c.font.size > title_max:
                        f["font"].append("%s!%s %gpt>제목%g" % (ws.title, c.coordinate, c.font.size, title_max))
                    elif c.font.size not in allowed:
                        f["font"].append("%s!%s %gpt(비표준)" % (ws.title, c.coordinate, c.font.size))
                if isinstance(v, str) and len(v) > 150 and c.row <= 5:
                    long_top += 1
        if long_top:
            f["annotation"].append("%s 상단 장문 %d개(전용 notes 로 분리)" % (ws.title, long_top))
    return f


def zone_findings(wb, contract) -> dict:
    """영역(zone) 위반 수집 — 한 시트 내 정형블록 한정(좌표 free, 마커 2트랙 기반).

    반환 {resolved_drift, unsealed, unknown_block}. 마커 없는 시트는 빈 결과(혼합 아님).
      · resolved_drift: 블록 셀의 실제 서식 ≠ 계약 스펙(태그 불신·재계산). role=set_cell role.
      · unsealed: 마커 bounding box 안 데이터셀인데 row/col band 미커버(미경유 삽입·구멍).
    """
    from fpna import design_zones as dz
    out = {"resolved_drift": [], "unsealed": [], "unknown_block": []}
    blocks = contract.get("blocks", {})
    for ws in wb.worksheets:
        cells, row_map, col_map = dz.resolve_blocks(ws, contract)
        _, _, anchor = dz.read_band_maps(ws)
        if anchor is None:
            continue
        for (r, c), (rb, cb, bid) in cells.items():
            spec = blocks.get(bid, {}).get("spec")
            if spec is None:
                continue                              # 스펙 대상 아닌 band(label 등) — freehand 관용
            if dz.canon(dz.resolved(ws.cell(r, c))) != dz.canon(spec):
                out["resolved_drift"].append((ws.title, ws.cell(r, c).coordinate, bid, cb))
        # unsealed: 마커 quadrant 안 데이터셀인데 row/col band 중 *한쪽만* 커버(=미경유 침입/구멍).
        # 둘 다 커버=정상 셀, 둘 다 미커버=진짜 freehand → 관용.
        ar, ac = anchor
        for rr in range(ar + 1, ws.max_row + 1):
            for cc in range(ac + 1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                if cell.value in (None, ""):
                    continue
                inr, inc = (rr in row_map), (cc in col_map)
                if inr != inc:                        # XOR = 한 축만 블록 territory
                    out["unsealed"].append((ws.title, cell.coordinate))
    return out


def assert_design_standard(rep: QCReport, wb, *, fail_on_decoration: bool = True,
                           contract=None,
                           name: str = "디자인 표준(FAST/ICAEW/Macabacus)") -> bool:
    """디자인 위반을 QCReport 에 기록. 장식문자는 hard-fail(기본), 나머지는 보고.

    house_style 토큰을 직접 읽으므로 표준대로 생성된 산출(set_cell/brand_header)엔 침묵한다.
    contract 주면(혼합 시트) 영역별 strict — resolved_drift·unsealed 가 추가 hard-fail.
    """
    f = design_findings(wb)
    decor = f["decoration"]
    rest = f["num_align"] + f["font"] + f["annotation"]
    ok = not decor if fail_on_decoration else True
    detail = ""
    if decor:
        detail = "장식문자 %d건: %s" % (len(decor), ", ".join(decor[:6]))
    if rest:
        detail = (detail + " | " if detail else "") + "정렬/폰트/주석 %d건: %s" % (len(rest), ", ".join(rest[:6]))
    rep.add(name, ok, detail)
    if contract is not None:
        z = zone_findings(wb, contract)
        zhard = z["resolved_drift"] + z["unsealed"] + z["unknown_block"]
        zok = not zhard
        zdetail = ""
        if z["resolved_drift"]:
            zdetail = "drift %d: %s" % (len(z["resolved_drift"]), z["resolved_drift"][:4])
        if z["unsealed"]:
            zdetail = (zdetail + " | " if zdetail else "") + "unsealed %d: %s" % (len(z["unsealed"]), z["unsealed"][:4])
        rep.add("영역 디자인(strict zone)", zok, zdetail)
        ok = ok and zok
    return ok


# --------------------------------------------------------------------------- #
# 비파괴 정규화 — 외부 입수 .xlsx 를 표준으로(서식만, 값·수식 불변)               #
# --------------------------------------------------------------------------- #
def _snap_size(s):
    allowed = sorted(hs.ALLOWED_SIZES)
    if s in hs.ALLOWED_SIZES:
        return s
    if s > hs.SIZE_TITLE:
        return hs.SIZE_TITLE
    return min(allowed, key=lambda a: abs(a - s))


def _clean_label(text):
    t = _MD_EMPH.sub(r"\1", text)
    prev = None
    while prev != t:
        prev = t
        t = _DECOR_STRIP.sub("", t)
    return t.strip()


def restyle_zone(wb, contract) -> list:
    """혼합 시트의 strict 블록 수선 — resolved_drift 셀에 set_cell(role) 재적용(override 제거,
    값 불변). unsealed 는 자동수정 안 하고 flag 반환(인간 판정). 비파괴: 값/수식 불변 단언."""
    from fpna import house_style as hs
    z = zone_findings(wb, contract)
    blocks = contract.get("blocks", {})
    actions, before = [], {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if _isnum(c.value) or _is_formula(c.value) or isinstance(c.value, str):
                    before[(ws.title, c.coordinate)] = c.value
    for (title, coord, bid, cb) in z["resolved_drift"]:
        ws = wb[title]
        role = blocks.get(bid, {}).get("role", "calc")
        cell = ws[coord]
        hs.set_cell(ws, cell.row, cell.column, cell.value, role=role)   # 룩 재적용, 값 보존
        actions.append(("retag", title, coord, role))
    for (title, coord) in z["unsealed"]:
        actions.append(("flag_unsealed", title, coord))
    for ws in wb.worksheets:                              # 비파괴 단언
        for row in ws.iter_rows():
            for c in row:
                k = (ws.title, c.coordinate)
                if k in before and c.value != before[k]:
                    raise AssertionError("restyle_zone 가 값을 바꿈: %s %r→%r" % (k, before[k], c.value))
    return actions


def restyle_inplace(wb) -> list:
    """워크북을 house_style 표준으로 *서식만* 정규화(숫자·수식 값 불변). 변경 목록 반환.
    숫자 좌/가운데→우측, 비표준 폰트→스케일 스냅, 라벨 장식문자 제거."""
    from openpyxl.styles import Font
    before = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if _isnum(c.value) or _is_formula(c.value):
                    before[(ws.title, c.coordinate)] = c.value
    changes = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if _isnum(v) and c.alignment and c.alignment.horizontal in ("left", "center"):
                    c.alignment = hs.RIGHT
                    changes.append("%s!%s 정렬→우측" % (ws.title, c.coordinate))
                if c.font and c.font.size and c.font.size not in hs.ALLOWED_SIZES and v not in (None, ""):
                    old = c.font.size
                    c.font = Font(name=c.font.name, size=_snap_size(old), bold=c.font.bold,
                                  italic=c.font.italic, color=c.font.color)
                    changes.append("%s!%s 폰트 %g→%g" % (ws.title, c.coordinate, old, _snap_size(old)))
                if isinstance(v, str) and not _is_formula(v):
                    cleaned = _clean_label(v)
                    if cleaned and cleaned != v:
                        c.value = cleaned
                        changes.append("%s!%s 라벨 정리" % (ws.title, c.coordinate))
    # 불변 검증(값/수식)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                k = (ws.title, c.coordinate)
                if k in before and c.value != before[k]:
                    raise AssertionError("restyle 가 값을 바꿈: %s %r→%r" % (k, before[k], c.value))
    return changes
