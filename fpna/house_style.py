"""
fpna.house_style — 맥킨지/회계법인 룩 단일 SSOT.

모든 템플릿 빌더·렌더러는 색·폰트·숫자서식·보더를 *직접* 지정하지 않고
이 모듈의 상수와 헬퍼만 사용한다. 룩을 바꾸려면 여기 한 곳만 고친다.

관찰한 공통 규약(CFI / Macabacus / Vertex42)을 토큰화한 것이며,
외부 템플릿 파일은 커밋하지 않는다(개념만 재현).

핵심 규약
---------
- 숫자: 음수 괄호 `#,##0;(#,##0)`, % `0.0%`. 단위는 헤더에 명시(₩mn 등).
  백만 단위 스케일은 `FMT_INT_MN`(원값 두고 `,,` 로 표시만 ÷1e6).
- 무채색 본문 + 단일 액센트 1색.
- 입력셀=파랑 글씨, 계산셀=검정 글씨(IB 4색 관례), 링크셀=초록, 외부참조=빨강.
- gridlines off, 본문 9~10pt sans(맑은 고딕/Calibri).
- 항목 좌측·숫자 우측·들여쓰기 계층. 합계는 상단 단일 보더.
- tie-out check 셀은 ≠0(또는 |x|>tol)이면 적색 강조(`check_cell`, CellIsRule).
- variance bridge = 누적 막대 + base 시리즈 투명(fill=none) 트릭.

표준 레이아웃 규칙 (report_frame / report_footer)
------------------------------------------------
파이프라인 등록 테이블은 아래 골격으로 통일한다(레이아웃 일관성 = 신뢰성).

  ① 제목 블록 (title / subtitle)
  ② 메타 헤더 (단위 · 통화 · 회계기준 · 기준일 — 숫자 해석 모호성 제거)
  ③ 본문 (표 / 차트)
  ④ _RECON 대사 (completeness / accuracy / cutoff — view_contract.recon_block)
  ⑤ 출처 footer (Source / Note / Prepared by — 감사 추적성)
  ⑥ page_setup (반복 헤더 · 가로폭 맞춤 · 페이지/날짜 푸터)

  + gridlines off · freeze_panes(헤더 고정).

`report_frame` 이 ①②③의 머리(①②+style_sheet)와 freeze 를 일괄 적용하고
본문 시작 행을 반환한다. ④_RECON 은 템플릿별 대사 의미가 달라 본문에서 직접
그리며, `report_footer` 가 ⑤⑥을 묶어 마감한다.
"""
from __future__ import annotations

import fpna._bootstrap  # noqa: F401  (vendor/ 주입)

from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.formatting.rule import (CellIsRule, ColorScaleRule, DataBarRule,
                                      FormulaRule, IconSetRule)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ==========================================================================
# [핵심 디자인 코드 ①] 디자인 토큰 — 테마(팔레트) + 타이포(타이틀≠본문)
# ==========================================================================
# 빅4/컨설팅 산출물의 룩을 "토큰"으로 박는다. 색·폰트를 템플릿이 직접 쓰지 않고
# 이 토큰만 쓰므로, apply_theme() 한 줄로 29개 템플릿이 동시에 재스킨된다.
#
# 레퍼런스(브랜드 가이드/관례, 색값만 차용):
#   - Deloitte: green #86BC24 + black (brandpalettes/colorcodeshub).
#   - 타이포 관례(McKinsey/KPMG): "타이틀 폰트 ≠ 본문 폰트". 본문은 거의 항상
#     Arial/Calibri 계열. EY Interstate 처럼 사내 전용폰트는 클라이언트 PC 에
#     없어 깨진다(=우리 폐쇄망 제약과 동일) → 확정 보유 폰트만 쓴다.
#   - 회사 PC 확정 보유: "맑은 고딕"(한글) · "Calibri"/"Calibri Light"(라틴, Office 기본).
#
# 규칙:
#   - 타이틀 = 영문 + Calibri Light(가벼운 헤드라인). 본문 = Calibri/맑은 고딕 10pt.
#   - 무채색 본문 + 단일 브랜드 액센트 1색. 액센트는 '구분/강조'에만(남발 금지).
#   - 헤더는 진한 솔리드밴드 대신 '옅은밴드 + 액센트 하단룰'(현대 회계법인 표 룩).

# ---- 테마 정의(팔레트). 색은 ARGB hex(앞 2자리 alpha 생략 시 openpyxl FF 보정) ----
THEMES: dict[str, dict] = {
    # 기본: 차분한 무채색 + 딥틸 액센트. 어느 회사든 무난한 '컨설팅 뉴트럴'.
    "graphite": {
        "INK": "111418", "INK_SOFT": "5B616B",
        "ACCENT": "12404A", "ACCENT_SOFT": "BBD3D6", "ACCENT_DEEP": "0B2A31",
        "EYEBROW_FG": "12404A", "MARK": "12404A",
        "HEADER_BAND": "EEF2F3", "HEADER_FG": "111418",
        "BAND": "F6F8F8",
        "POS_FG": "1F7A1F", "NEG_FG": "C0392B",
        "INPUT_FG": "0B66C2", "LINK_FG": "1F7A1F",
        "TITLE_FONT": "Calibri Light", "BODY_FONT": "맑은 고딕",
        "BODY_FONT_LATIN": "Calibri",
    },
    # 딜로이트풍: Deloitte Green + 블랙. (사내/내부용. 외부배포 시 브랜드 충돌 주의)
    "deloitte": {
        "INK": "0F0B0B", "INK_SOFT": "53565A",
        "ACCENT": "86BC24", "ACCENT_SOFT": "DCEBBD", "ACCENT_DEEP": "046A38",
        "EYEBROW_FG": "046A38", "MARK": "86BC24",
        "HEADER_BAND": "F1F1F1", "HEADER_FG": "0F0B0B",
        "BAND": "F7F7F7",
        "POS_FG": "046A38", "NEG_FG": "DA291C",
        "INPUT_FG": "0076A8", "LINK_FG": "046A38",
        "TITLE_FONT": "Calibri Light", "BODY_FONT": "맑은 고딕",
        "BODY_FONT_LATIN": "Calibri",
    },
    # 네이비: 보수적 금융/IB 톤.
    "navy": {
        "INK": "0F1B2D", "INK_SOFT": "586071",
        "ACCENT": "1F3A5F", "ACCENT_SOFT": "C3CFE0", "ACCENT_DEEP": "0C1F38",
        "EYEBROW_FG": "1F3A5F", "MARK": "1F3A5F",
        "HEADER_BAND": "EDF1F6", "HEADER_FG": "0F1B2D",
        "BAND": "F5F8FB",
        "POS_FG": "1F7A1F", "NEG_FG": "B3261E",
        "INPUT_FG": "0B66C2", "LINK_FG": "1F7A1F",
        "TITLE_FONT": "Calibri Light", "BODY_FONT": "맑은 고딕",
        "BODY_FONT_LATIN": "Calibri",
    },
}
DEFAULT_THEME = "graphite"

# ---- 테마 비의존 중립 토큰(룰 회색은 어느 테마든 동일하게 차분) ----
WHITE = "FFFFFF"
RULE = "D8DCE0"          # 가는 구분선(hairline) 회색
RULE_STRONG = "AAB0B7"   # 합계 보더 진회색

# ---- 타이포 스케일(크기는 테마 무관 고정) ----
SIZE_BODY = 10
SIZE_SMALL = 9
SIZE_EYEBROW = 9         # eyebrow/kicker(대문자 영문, 트래킹)
SIZE_TITLE = 20          # 영문 헤드라인(Calibri Light, 큼직하게)
SIZE_SUBTITLE = 10.5     # 한글 부제
SIZE_SECTION = 11        # 섹션 헤더

# ---- 아래 토큰들은 apply_theme() 가 채운다(placeholder; import 시 1회 호출) ----
INK = INK_SOFT = ACCENT = ACCENT_SOFT = ACCENT_DEEP = ""
EYEBROW_FG = MARK = HEADER_BAND = HEADER_FG = HEADER_BG = BAND = ""
POS_FG = NEG_FG = INPUT_FG = LINK_FG = CALC_FG = ""
TITLE_FONT = BODY_FONT = FONT_NAME = FONT_NAME_LATIN = ""

# 2. 폰트 placeholder 종료 — 실제 값은 파일 하단 apply_theme(DEFAULT_THEME) 에서 주입

# --------------------------------------------------------------------------
# 3. 숫자 서식 코드 (Excel 범용)
# --------------------------------------------------------------------------
FMT_INT = "#,##0;(#,##0)"
FMT_INT_DASH = "#,##0;(#,##0);\"-\""       # 0을 대시로
FMT_NUM1 = "#,##0.0;(#,##0.0)"
FMT_NUM2 = "#,##0.00;(#,##0.00)"
FMT_PCT1 = "0.0%;(0.0%)"
FMT_PCT2 = "0.00%;(0.00%)"
FMT_PCT0 = "0%;(0%)"
FMT_MULT = '0.0"x"'                          # 배수(LTV/CAC 등)
FMT_DATE = "yyyy-mm-dd"
FMT_MONTH = "yyyy-mm"
FMT_MONEY_MN = '#,##0;(#,##0)'               # 단위는 헤더 표기(₩mn)
# 셀 안에서 백만 단위 스케일(원값 그대로 두고 표시만 ÷1e6). 헤더에 (mn) 표기 병행.
FMT_INT_MN = "#,##0,,;(#,##0,,)"
FMT_NUM1_MN = "#,##0.0,,;(#,##0.0,,)"
# KPI 신호 화살표(증감을 색+▲▼로). 양=초록▲ / 음=빨강▼ / 0=무채색.
FMT_KPI_ARROW = '[Green]"▲"0.0%;[Red]"▼"0.0%;0.0%'
# 빨강 음수(MR/대시보드 옵션). IB 정통은 검정 괄호(FMT_INT)가 기본, 빨강은 선택.
FMT_INT_RED = "#,##0;[Red](#,##0)"
FMT_NUM1_RED = "#,##0.0;[Red](#,##0.0)"

# --------------------------------------------------------------------------
# 4. 정렬
# --------------------------------------------------------------------------
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)


def indent(level: int) -> Alignment:
    """들여쓰기 계층용 좌측 정렬(level 만큼 indent)."""
    return Alignment(horizontal="left", vertical="center", indent=max(0, level))


# --------------------------------------------------------------------------
# 5. 보더 (합계 상단 단일 보더가 핵심 IB 관례)
# --------------------------------------------------------------------------
_thin = Side(style="thin", color=RULE)
_med = Side(style="medium", color=RULE_STRONG)
_dbl = Side(style="double", color=RULE_STRONG)

BORDER_NONE = Border()
BORDER_TOP = Border(top=_thin)                       # 소계 상단
BORDER_TOP_STRONG = Border(top=_med)                 # 합계 상단
BORDER_TOP_BOTTOM = Border(top=_thin, bottom=_thin)
BORDER_TOTAL = Border(top=_med, bottom=_dbl)         # 최종 합계(위 single·아래 double)
BORDER_BOTTOM = Border(bottom=_thin)


# --------------------------------------------------------------------------
# 6. Font 팩토리 + apply_theme(테마 주입)
# --------------------------------------------------------------------------
def font(color: str | None = None, *, bold: bool = False, size: float = SIZE_BODY,
         italic: bool = False, name: str | None = None) -> Font:
    """본문 폰트(기본 BODY_FONT=맑은 고딕). 색 미지정 시 활성 테마 INK."""
    return Font(name=name or BODY_FONT, size=size, bold=bold, italic=italic,
                color=color or INK)


def title_font(*, size: float = SIZE_TITLE, bold: bool = False,
               color: str | None = None) -> Font:
    """타이틀 폰트(TITLE_FONT=Calibri Light). 영문 헤드라인 전용."""
    return Font(name=TITLE_FONT, size=size, bold=bold, color=color or INK)


FILL_NONE = PatternFill(fill_type=None)

# 파생 폰트/필 placeholder — apply_theme 가 채운다.
F_BODY = F_BODY_BOLD = F_INPUT = F_CALC = F_LINK = F_SOFT = None
F_TITLE = F_SUBTITLE = F_SECTION = F_HEADER = F_TOTAL = F_EYEBROW = None
FILL_HEADER = FILL_BAND = FILL_ACCENT = None

_ACTIVE_THEME = None


def apply_theme(name: str = DEFAULT_THEME) -> None:
    """테마 토큰을 모듈 전역 + 파생 Font/Fill 에 주입한다.

    템플릿은 hs.ACCENT / hs.F_TITLE / hs.FILL_HEADER 등만 쓰므로, 이 함수 한 번이면
    29개 템플릿 룩이 동시에 바뀐다. import 시 1회 자동 호출(아래). 런타임에 다시
    부르면(예: hs.apply_theme("deloitte")) 이후 빌드부터 적용된다.
    """
    global _ACTIVE_THEME
    t = THEMES.get(name)
    if t is None:
        raise KeyError("unknown theme %r (있음: %s)" % (name, ", ".join(THEMES)))
    g = globals()
    # 팔레트/타이포 토큰 주입
    for k, v in t.items():
        g[k] = v
    g["FONT_NAME"] = t["BODY_FONT"]            # 하위호환 별칭
    g["FONT_NAME_LATIN"] = t["BODY_FONT_LATIN"]
    g["CALC_FG"] = t["INK"]
    g["HEADER_BG"] = t["HEADER_BAND"]          # 하위호환(과거 'BG' 명칭)
    # 파생 Font (활성 토큰 기준 재생성)
    g["F_BODY"] = font()
    g["F_BODY_BOLD"] = font(bold=True)
    g["F_INPUT"] = font(g["INPUT_FG"])
    g["F_CALC"] = font(g["CALC_FG"])
    g["F_LINK"] = font(g["LINK_FG"])
    g["F_SOFT"] = font(g["INK_SOFT"], size=SIZE_SMALL)
    g["F_TITLE"] = title_font(size=SIZE_TITLE)                       # 영문 헤드라인
    g["F_SUBTITLE"] = font(g["INK_SOFT"], size=SIZE_SUBTITLE)        # 한글 부제
    g["F_SECTION"] = font(g["ACCENT"], bold=True, size=SIZE_SECTION)
    g["F_HEADER"] = font(g["HEADER_FG"], bold=True, size=SIZE_SMALL)
    g["F_TOTAL"] = font(g["INK"], bold=True)
    g["F_EYEBROW"] = font(g["EYEBROW_FG"], bold=True, size=SIZE_EYEBROW,
                          name=t["BODY_FONT_LATIN"])
    # 파생 Fill
    g["FILL_HEADER"] = PatternFill("solid", fgColor=g["HEADER_BAND"])
    g["FILL_BAND"] = PatternFill("solid", fgColor=g["BAND"])
    g["FILL_ACCENT"] = PatternFill("solid", fgColor=g["ACCENT"])
    _ACTIVE_THEME = name


import os as _os
apply_theme(_os.environ.get("FPNA_THEME", DEFAULT_THEME)
            if _os.environ.get("FPNA_THEME") in THEMES else DEFAULT_THEME)


# --------------------------------------------------------------------------
# 7. 셀 역할 헬퍼 — 빌더는 이걸로 셀 의미(입력/계산/링크)를 표현
# --------------------------------------------------------------------------
def set_cell(ws: Worksheet, row: int, col: int, value=None, *,
             role: str = "calc", number_format: str | None = None,
             align: Alignment | None = None, bold: bool = False,
             border: Border | None = None, fill: PatternFill | None = None,
             size: int = SIZE_BODY):
    """단일 셀에 값+서식 일괄 적용. role ∈ {calc,input,link,label,header,total}.

    재무 관례상 셀의 '역할'을 색으로 구분(입력 파랑·계산 검정·링크 초록).
    """
    cell = ws.cell(row=row, column=col, value=value)
    role_font = {
        "calc": font(CALC_FG, bold=bold, size=size),
        "input": font(INPUT_FG, bold=bold, size=size),
        "link": font(LINK_FG, bold=bold, size=size),
        "label": font(INK, bold=bold, size=size),
        "soft": font(INK_SOFT, bold=bold, size=size),
        "header": F_HEADER,
        "total": font(INK, bold=True, size=size),
    }.get(role, font(CALC_FG, bold=bold, size=size))
    cell.font = role_font
    if role == "header":
        cell.fill = FILL_HEADER
        cell.alignment = align or CENTER
        # 현대 회계법인 표 헤더 = 옅은밴드 + 액센트 '하단룰'(진한 솔리드밴드 대신).
        cell.border = Border(bottom=Side(style="medium", color=ACCENT))
    else:
        cell.alignment = align or (LEFT if role in ("label", "soft") else RIGHT)
    if fill is not None:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if border is not None:
        cell.border = border
    return cell


# --------------------------------------------------------------------------
# 8. 시트 외관 (gridlines off, 여백, freeze, 열폭)
# --------------------------------------------------------------------------
def style_sheet(ws: Worksheet, *, freeze: str | None = "A2",
                zoom: int = 100) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    if freeze:
        ws.freeze_panes = freeze
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True if ws.sheet_properties.pageSetUpPr else None


def set_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    """{열번호: 폭} 일괄 적용."""
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ==========================================================================
# [핵심 디자인 코드 ②] 타이틀 규칙 — brand_header
# ==========================================================================
# 빅4/컨설팅 표지 관례를 코드로 박은 타이틀 블록:
#   ① eyebrow(kicker)  = 대문자 영문 워크스트림, 작게·트래킹·액센트색 (제목 위 1줄)
#   ② headline         = 영문 타이틀(Calibri Light, 큼직). "타이틀은 영문" 규칙.
#   ③ subtitle         = 한글 설명(보조 회색)
#   ④ accent rule      = 제목 아래 액센트 가로룰(브랜드 시그니처) + 좌상단 짧은 마크룰
# 영문 헤드라인은 (a) 명시 title_en (b) 제목의 "(English)" 괄호 추출 (c) 원제목 순.

import re as _re

# 영문 헤드라인이 없을 때 쓰는 기본 eyebrow.
EYEBROW_DEFAULT = "FP&A REPORTING"

_EN_PAREN = _re.compile(r"\(([A-Za-z0-9][A-Za-z0-9 &/\-\.\+]*)\)\s*$")


def _resolve_headline(title: str, title_en: str | None) -> tuple[str, str | None]:
    """(headline_en, subtitle_kr) 결정.

    - title_en 명시 → 그걸 헤드라인, 원제목 전체를 부제로.
    - 제목 끝 "(English)" 괄호 → 괄호 영문이 헤드라인, 괄호 제거한 한글이 부제.
    - 둘 다 없으면 → 원제목이 헤드라인(부제 None).
    """
    if title_en:
        return title_en, title
    m = _EN_PAREN.search(title or "")
    if m:
        en = m.group(1).strip()
        kr = (title[:m.start()]).strip(" -—·")
        return en, (kr or None)
    return title, None


def _track(text: str, gap: str = "\u2009") -> str:
    """eyebrow 트래킹(자간) — 글자 사이 thin space. 대문자 영문에만 권장."""
    return gap.join(list(text))


def brand_header(ws: Worksheet, title: str, subtitle: str = "", *,
                 title_en: str | None = None, eyebrow: str | None = None,
                 row: int = 1, last_col: int = 6) -> int:
    """브랜드 타이틀 블록을 그리고 본문 시작 행을 반환한다(title_block 대체).

    한글 제목을 넘겨도 괄호 영문을 헤드라인으로 끌어올리고, 한글은 부제로 내린다.
    """
    headline, kr_sub = _resolve_headline(title, title_en)
    sub = subtitle or kr_sub or ""

    # ① 좌상단 짧은 액센트 마크룰(A:B 상단 thick) — 브랜드 dot 의 절제된 변형
    for col in (1, 2):
        ws.cell(row=row, column=col).border = Border(
            top=Side(style="thick", color=MARK))
    # ① eyebrow
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    ce = ws.cell(row=row, column=1, value=_track(eyebrow or EYEBROW_DEFAULT))
    ce.font = F_EYEBROW
    ce.alignment = LEFT
    ws.row_dimensions[row].height = 16
    r = row + 1
    # ② 영문 헤드라인
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    ch = ws.cell(row=r, column=1, value=headline)
    ch.font = F_TITLE
    ch.alignment = LEFT
    ws.row_dimensions[r].height = 28
    r += 1
    # ③ 한글 부제
    if sub:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        cs = ws.cell(row=r, column=1, value=sub)
        cs.font = F_SUBTITLE
        cs.alignment = LEFT
        r += 1
    # ④ 제목 아래 액센트 가로룰(시그니처)
    for col in range(1, last_col + 1):
        ws.cell(row=r, column=col).border = Border(
            bottom=Side(style="medium", color=ACCENT))
    ws.row_dimensions[r].height = 4
    r += 1
    return r + 1  # 한 줄 띄움


def title_block(ws: Worksheet, title: str, subtitle: str = "",
                *, row: int = 1, last_col: int = 6) -> int:
    """하위호환 shim — brand_header 로 위임(기존 호출부 무수정 재스킨)."""
    return brand_header(ws, title, subtitle, row=row, last_col=last_col)


def section_header(ws: Worksheet, row: int, text: str, last_col: int = 6) -> int:
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SECTION
    c.alignment = LEFT
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).border = BORDER_BOTTOM
    return row + 1


def safe_sheet_title(name: str) -> str:
    """Excel 시트명 제약: ≤31자, 금지문자 []:*?/\\ 제거."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    name = name.strip() or "Sheet"
    return name[:31]


# --------------------------------------------------------------------------
# 9. 워터폴(variance bridge) — 누적막대 + base 투명 트릭
# --------------------------------------------------------------------------
def add_waterfall(ws: Worksheet, *, anchor: str, data_min_row: int,
                  data_max_row: int, base_col: int, value_col: int,
                  cat_col: int, title: str = "Bridge",
                  height: float = 8.0, width: float = 18.0) -> BarChart:
    """워터폴 차트를 stacked bar 로 그린다.

    레이아웃 전제(빌더가 보조 컬럼을 미리 채워둠):
      cat_col   : 카테고리 라벨
      base_col  : 투명 받침(floating bar 의 바닥 높이)
      value_col : 실제 표시 막대 높이(증감분/기둥)

    base 시리즈는 fill=none + line=none 으로 투명 처리해 막대가 떠 있게 만든다.
    """
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    chart.height = height
    chart.width = width
    chart.gapWidth = 40

    cats = Reference(ws, min_col=cat_col, min_row=data_min_row, max_row=data_max_row)

    base_ref = Reference(ws, min_col=base_col, min_row=data_min_row - 1,
                         max_row=data_max_row)
    val_ref = Reference(ws, min_col=value_col, min_row=data_min_row - 1,
                        max_row=data_max_row)
    chart.add_data(base_ref, titles_from_data=True)
    chart.add_data(val_ref, titles_from_data=True)
    chart.set_categories(cats)

    # series[0] = base → 투명
    base_series: Series = chart.series[0]
    base_series.graphicalProperties.noFill = True
    base_series.graphicalProperties.line.noFill = True

    # series[1] = value → 액센트 채움
    val_series: Series = chart.series[1]
    val_series.graphicalProperties.solidFill = ACCENT

    chart.legend = None
    ws.add_chart(chart, anchor)
    return chart


def add_line_chart(ws: Worksheet, *, anchor: str, data_min_col: int,
                   data_max_col: int, data_min_row: int, data_max_row: int,
                   cat_col: int, title: str = "", height: float = 7.5,
                   width: float = 16.0) -> LineChart:
    """라인 차트(추이/포캐스트). data 영역 1행 위에 시리즈명 가정."""
    ch = LineChart()
    ch.title = title or None
    ch.height = height
    ch.width = width
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col,
                     min_row=data_min_row - 1, max_row=data_max_row)
    cats = Reference(ws, min_col=cat_col, min_row=data_min_row, max_row=data_max_row)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    for s in ch.series:
        s.smooth = False
    ws.add_chart(ch, anchor)
    return ch


def add_bar_chart(ws: Worksheet, *, anchor: str, data_min_col: int,
                  data_max_col: int, data_min_row: int, data_max_row: int,
                  cat_col: int, title: str = "", height: float = 7.5,
                  width: float = 16.0, stacked: bool = False) -> BarChart:
    ch = BarChart()
    ch.type = "col"
    if stacked:
        ch.grouping = "stacked"
        ch.overlap = 100
    ch.title = title or None
    ch.height = height
    ch.width = width
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col,
                     min_row=data_min_row - 1, max_row=data_max_row)
    cats = Reference(ws, min_col=cat_col, min_row=data_min_row, max_row=data_max_row)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, anchor)
    return ch


def write_matrix(ws: Worksheet, top: int, left: int, headers: list[str],
                 rows: list[tuple], *, value_fmt: str = FMT_INT,
                 label_width: bool = True) -> int:
    """간단 표 작성 헬퍼. headers[0]=라벨열. rows=(label, v1, v2, ...).

    숫자 셀은 value_fmt, 라벨은 좌측. 마지막 행이 합계면 호출측이 보더 처리.
    반환: 표 다음 행번호.
    """
    for j, h in enumerate(headers):
        set_cell(ws, top, left + j, h, role="header",
                 align=LEFT if j == 0 else CENTER)
    r = top + 1
    for row in rows:
        set_cell(ws, r, left, row[0], role="label", align=LEFT)
        for j, v in enumerate(row[1:], start=1):
            is_pct = isinstance(v, float) and -2 < v < 2 and value_fmt == FMT_PCT1
            set_cell(ws, r, left + j, v, role="calc",
                     number_format=value_fmt)
        r += 1
    return r


# --------------------------------------------------------------------------
# 10. 정직성 헤더/푸터 — 단위·통화·기준일 명시 + 출처·주석
# --------------------------------------------------------------------------
def meta_header(ws: Worksheet, row: int, *, unit: str = "", currency: str = "",
                period_basis: str = "", as_of: str = "", last_col: int = 6) -> int:
    """단위·통화·기준일 메타 헤더 1줄(우측 정렬, 보조 회색).

    회계법인/IB 산출물의 정직성 규약: 표 상단에 스케일·통화·회계기준·기준일을
    명시해 숫자 해석의 모호성을 제거한다. 예) "(KRW mn) · FY ending Dec 31 ·
    as-of 2026-05-31". 빈 항목은 생략. 다음 시작 행을 반환한다.
    """
    parts: list[str] = []
    money = " ".join(p for p in (currency, unit) if p).strip()
    if money:
        parts.append("(%s)" % money)
    if period_basis:
        parts.append(period_basis)
    if as_of:
        parts.append("as-of %s" % as_of)
    text = "  ·  ".join(parts)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SOFT
    c.alignment = RIGHT
    return row + 1


def source_footer(ws: Worksheet, row: int, *, source: str = "", note: str = "",
                  prepared_by: str = "", last_col: int = 6) -> int:
    """출처·주석·작성자 푸터 블록(보조 회색, 좌측 wrap).

    각 라인은 "Source: ...", "Note: ...", "Prepared by: ..." 형태로 병합 표기.
    감사 추적성(누가·무엇을 근거로)을 산출물에 박제한다. 다음 행을 반환한다.
    """
    r = row
    for label, val in (("Source", source), ("Note", note),
                       ("Prepared by", prepared_by)):
        if not val:
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        c = ws.cell(row=r, column=1, value="%s: %s" % (label, val))
        c.font = F_SOFT
        c.alignment = LEFT_WRAP
        r += 1
    return r


# --------------------------------------------------------------------------
# 10b. 표준 레이아웃 골격 — report_frame / report_footer
# --------------------------------------------------------------------------
# 파이프라인 등록 템플릿(variance / fc_* 등)이 제각각 그리던 머리·꼬리 골격을
# 한 곳으로 통일한다. 호출측 build 는 frame→[본문]→footer 순서만 지키면
# 모든 산출물의 룩(제목블록·메타헤더·gridlines off·출처footer·인쇄설정)이 일관된다.
#
#   row = hs.report_frame(ws, title, subtitle=..., unit="₩mn", as_of="2026-05-31",
#                         last_col=6)
#   ... 본문(표/차트/_RECON)을 row 부터 그림. nxt = 본문 다음 행 ...
#   hs.report_footer(ws, nxt, source="GL export", prepared_by="FP&A", last_col=6)
#
# _RECON(completeness/accuracy/cutoff) 대사 블록은 view_contract.recon_block 로
# 본문에서 그리므로 frame 이 떠안지 않는다(템플릿별 대사 의미가 달라 호출측 책임).
def report_frame(ws: Worksheet, title: str, *, subtitle: str = "",
                 unit: str = "", currency: str = "", period_basis: str = "",
                 as_of: str = "", last_col: int = 6,
                 freeze=True, freeze_col: str = "A", zoom: int = 100) -> int:
    """표준 머리 골격을 일괄 적용하고 본문 시작 행을 반환한다.

    title_block → (unit/currency/period_basis/as_of 중 하나라도 있으면) meta_header
    → style_sheet(gridlines off · freeze) 순으로 적용한다.

    freeze 인자:
      - True  → 본문 헤더 다음 행(첫 데이터 행)에서 freeze. 고정 열은 freeze_col
                ("A"=행만 고정, "B"=좌측 1열도 고정 — 라벨열 보존용).
      - False → freeze 안 함.
      - str   → 명시 좌표("B7" 등)로 freeze(완전 수동).

    본문은 반환행(=헤더 행)부터 그리고, 마감은 report_footer 로 한다. set_widths 는
    표 폭이 템플릿마다 달라 호출측이 별도로 지정한다(frame 은 골격만 통일).
    """
    r = title_block(ws, title, subtitle, last_col=last_col)
    if any((unit, currency, period_basis, as_of)):
        r = meta_header(ws, r, unit=unit, currency=currency,
                        period_basis=period_basis, as_of=as_of, last_col=last_col)
    # 본문 헤더 행은 반환행(r). freeze 는 그 다음 행(첫 데이터 행) 기준.
    if isinstance(freeze, str):
        fz = freeze
    elif freeze:
        fz = "%s%d" % (freeze_col, r + 1)
    else:
        fz = None
    style_sheet(ws, freeze=fz, zoom=zoom)
    return r


def report_footer(ws: Worksheet, row: int, *, source: str = "", note: str = "",
                  prepared_by: str = "", last_col: int = 6,
                  title_rows: str | None = "1:2",
                  print_area: str | None = None,
                  footer_text: str | None = None) -> int:
    """표준 꼬리 골격 — source_footer + page_setup_report 를 묶어 마감한다.

    출처·주석·작성자 푸터를 그린 뒤 인쇄 설정(반복 헤더·가로폭 맞춤·페이지 푸터)을
    적용한다. 출처가 모두 비어 있으면 푸터 줄은 생략하되 page_setup 은 적용한다.
    반환 = 푸터 다음 행.
    """
    nxt = source_footer(ws, row, source=source, note=note,
                        prepared_by=prepared_by, last_col=last_col)
    page_setup_report(ws, title_rows=title_rows, print_area=print_area,
                      footer=footer_text)
    return nxt


# --------------------------------------------------------------------------
# 11. 조건부 서식 — openpyxl native rule (heatmap/databar/iconset)
# --------------------------------------------------------------------------
def apply_heatmap(ws: Worksheet, cell_range: str, *, low=POS_FG, mid=WHITE,
                  high=NEG_FG, three_scale: bool = True) -> None:
    """ColorScale 히트맵. 기본은 3색(낮음 초록 → 중간 흰 → 높음 빨강).

    비용/리스크성 매트릭스에 적합. 부호 반대 표현이 필요하면 low/high 교체.
    two-scale 가 필요하면 three_scale=False (min→max 2색).
    """
    if three_scale:
        rule = ColorScaleRule(
            start_type="min", start_color=low,
            mid_type="percentile", mid_value=50, mid_color=mid,
            end_type="max", end_color=high)
    else:
        rule = ColorScaleRule(start_type="min", start_color=low,
                              end_type="max", end_color=high)
    ws.conditional_formatting.add(cell_range, rule)


def apply_databar(ws: Worksheet, cell_range: str, *, color: str = ACCENT_SOFT,
                  show_value: bool = True) -> None:
    """DataBar(셀 내 막대). 규모 비교용. 기본 색 = 액센트 옅은 톤."""
    rule = DataBarRule(start_type="min", end_type="max", color=color,
                       showValue=show_value, minLength=None, maxLength=None)
    ws.conditional_formatting.add(cell_range, rule)


def apply_iconset(ws: Worksheet, cell_range: str, *,
                  icon_style: str = "3Arrows", reverse: bool = False,
                  show_value: bool = True) -> None:
    """IconSet(화살표/신호등 아이콘). KPI 방향·상태 신호화.

    icon_style 예: '3Arrows'(▲▶▼), '3TrafficLights1'(신호등),
    '3Symbols'(✓!✕), '4Arrows', '5Arrows'. reverse=True 면 임계 반전(비용성).
    임계값(values)은 icon 개수(접두 숫자)에서 균등분할로 자동 생성.
    """
    try:
        n = int(icon_style[0])
    except (ValueError, IndexError):
        n = 3
    n = max(3, min(5, n))
    values = [round(i * 100 / n) for i in range(n)]   # 균등 percent 임계
    rule = IconSetRule(icon_style=icon_style, type="percent",
                       values=values, showValue=show_value, reverse=reverse)
    ws.conditional_formatting.add(cell_range, rule)


def apply_zebra(ws: Worksheet, cell_range: str, *, fill: str = BAND) -> None:
    """짝수 행 음영밴딩(FormulaRule MOD(ROW(),2)=0). 차트 줄무늬 충돌 회피용.

    테이블 스타일 대신 조건부 서식으로 banding 을 깔아 가독성 확보.
    """
    rule = FormulaRule(formula=["MOD(ROW(),2)=0"], stopIfTrue=False,
                       fill=PatternFill("solid", fgColor=fill))
    ws.conditional_formatting.add(cell_range, rule)


# --------------------------------------------------------------------------
# 12. 페이지 setup — 인쇄 반복헤더·맞춤폭·푸터
# --------------------------------------------------------------------------
def page_setup_report(ws: Worksheet, *, title_rows: str | None = "1:2",
                      print_area: str | None = None, fit_width: int = 1,
                      fit_height: int = 0, footer: str | None = None) -> None:
    """회계법인 룩 인쇄 설정: 반복 헤더 행·가로폭 맞춤·푸터(페이지/날짜).

    - title_rows: 페이지마다 반복 인쇄할 헤더 행("1:2" 형식). None 이면 미설정.
    - print_area: 인쇄 영역("A1:H40"). None 이면 미설정.
    - fit_width/fit_height: fitToPage. height=0 = 행 길이 제한 없음(세로 분할 허용).
    - footer: oddFooter 문자열. 기본 = "&P / &N  ·  &D"(페이지 N/총 · 날짜).
    """
    ps = ws.page_setup
    ps.orientation = "landscape"
    ps.fitToWidth = fit_width
    ps.fitToHeight = fit_height
    if ws.sheet_properties.pageSetUpPr is not None:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    else:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    if title_rows:
        ws.print_title_rows = title_rows
    if print_area:
        ws.print_area = print_area
    ws.oddFooter.center.text = footer if footer is not None else "&P / &N  ·  &D"


# --------------------------------------------------------------------------
# 13. check / tie-out 셀 — ≠0(또는 |x|>tol)이면 적색 강조
# --------------------------------------------------------------------------
def check_cell(ws: Worksheet, row: int, col: int, formula: str, *, tol: float = 0,
               number_format: str = FMT_INT, label: str | None = None,
               label_col: int | None = None):
    """tie-out 체크 셀. 수식 결과가 허용오차 tol 을 벗어나면 적색 배경 강조.

    View Contract tie-out(BS 균형·브리지 합·소스 일치)의 시각화. 정상(=0)이면
    무채색, 깨지면 한눈에 빨강. 라벨을 주면 label_col(기본 col-1)에 함께 기입.
    """
    f = formula if formula.startswith("=") else "=" + formula
    cell = ws.cell(row=row, column=col, value=f)
    cell.font = F_CALC
    cell.alignment = RIGHT
    cell.number_format = number_format
    if label:
        lc = label_col if label_col is not None else max(1, col - 1)
        lcell = ws.cell(row=row, column=lc, value=label)
        lcell.font = F_SOFT
        lcell.alignment = LEFT
    coord = "%s%d" % (get_column_letter(col), row)
    bad_fill = PatternFill("solid", fgColor=NEG_FG)
    bad_font = font(WHITE, bold=True)
    if tol and tol > 0:
        # |x| > tol 이면 강조 → 두 단측(>tol, <-tol) 규칙
        for op, fm in (("greaterThan", [str(tol)]), ("lessThan", [str(-tol)])):
            ws.conditional_formatting.add(
                coord, CellIsRule(operator=op, formula=fm, stopIfTrue=False,
                                  fill=bad_fill, font=bad_font))
    else:
        ws.conditional_formatting.add(
            coord, CellIsRule(operator="notEqual", formula=["0"],
                              stopIfTrue=False, fill=bad_fill, font=bad_font))
    return cell


__all__ = [name for name in dir() if not name.startswith("_")]


# ==========================================================================
# [핵심 디자인 코드 ③] 표 스타일 — style_table_block
# ==========================================================================
# 헤더/합계는 set_cell(role=...) 가 처리하므로, 여기선 '표 전체'에 일관된
# 가독성 레이어(zebra 밴딩 · 외곽 hairline · 합계 상단 액센트룰)를 한 번에 입힌다.
# 결정적(openpyxl 직접 fill — 조건부서식 roundtrip 손실 회피).
def style_table_block(ws: Worksheet, *, header_row: int, first_row: int,
                      last_row: int, left: int, right: int,
                      zebra: bool = True, total_row: int | None = None) -> None:
    """표 영역에 zebra 밴딩 + 헤더 하단 액센트룰 + (옵션)합계 상단 룰을 적용.

    - header_row: 헤더행. 각 셀에 액센트 하단 medium 룰(set_cell 과 동일 톤) 재확인.
    - first_row..last_row: 본문. zebra=True 면 짝수번째 데이터행에 옅은 BAND fill.
    - total_row: 주면 그 행 상단에 액센트 medium 룰(합계 강조).
    파스텔 밴딩은 '값 없는 셀'도 칠해 표 경계를 또렷하게 만든다(회계법인 표 관례).
    """
    accent_rule = Side(style="medium", color=ACCENT)
    for col in range(left, right + 1):
        ws.cell(row=header_row, column=col).border = Border(bottom=accent_rule)
    if zebra:
        band = PatternFill("solid", fgColor=BAND)
        for i, r in enumerate(range(first_row, last_row + 1)):
            if i % 2 == 1:                      # 짝수번째 데이터행만 밴딩
                for col in range(left, right + 1):
                    cell = ws.cell(row=r, column=col)
                    if cell.fill is None or cell.fill.fgColor is None \
                            or cell.fill.patternType is None:
                        cell.fill = band
    if total_row is not None:
        for col in range(left, right + 1):
            cur = ws.cell(row=total_row, column=col).border
            ws.cell(row=total_row, column=col).border = Border(
                top=accent_rule, bottom=(cur.bottom if cur else None))


__all__ = [n for n in dir() if not n.startswith("_")]
