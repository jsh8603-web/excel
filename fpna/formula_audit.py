"""
fpna.formula_audit — 수식 참조 계약 ("엉뚱한 칼럼에서 뺀다" 의 해독제).

배경(독립 리뷰 2026-06): variance 등 템플릿은 셀에 Excel 수식을 직접 기입한다
(예: Δ = "=C{r}-B{r}"). 그런데 QC 는 finance.variance(actual,plan)=actual-plan 을
재계산해 (actual-plan) 과 대조 → **tautology**(같은 식 양변). 렌더된 수식이 실제로
어느 칼럼을 어느 순서로 빼는지는 한 번도 검사하지 않았다. 그래서 build 가 칼럼을
잘못 짚거나(=D-B) 방향을 뒤집어도(=B-C) QC 가 초록.

이 모듈은 **렌더된 수식 문자열을 파싱해 참조 칼럼을 meta 선언 의도와 대조**한다.
값 재계산(N-version)과 직교하는 *형태* 검증이라, 부호 맞고 칼럼만 틀린 침묵형도 잡는다.

선언 방식(템플릿이 _fpna_meta 에 심음):
  meta["formula_checks"] = [
    {"sheet": "Variance", "region": (r0, r1, var_col),
     "op": "-", "left": act_col, "right": plan_col,
     "name": "Δ=실적-계획 칼럼참조"},
    ...
  ]
region = (시작행, 끝행, 대상열). 그 범위의 각 셀 수식이
  =<left_col><행><op><right_col><행>
구조(행=자기행, 좌우열=선언값)인지 검증. 스파인(_base_owned_gate)이 자동 스윕한다.
"""
from __future__ import annotations

import re

import fpna._bootstrap  # noqa: F401

from openpyxl.utils import get_column_letter

from fpna.templates.base import QCReport

# 셀 참조 토큰: 선택적 $, 열문자, 선택적 $, 행번호.
_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


def parse_refs(formula: str) -> list:
    """수식 문자열에서 (열문자, 행번호) 참조를 출현 순서대로 추출.

    문자열 리터럴 안의 내용은 본 구현 범위 밖(단순 변동표 수식 전제). IF/ABS 등
    함수가 끼면 첫 두 참조가 피연산자가 아닐 수 있으니, 단순 2항식 검증에만 쓴다.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    return [(c, int(r)) for c, r in _REF_RE.findall(formula)]


def check_binary(formula: str, *, row: int, left_col: int, right_col: int,
                 op: str = "-") -> tuple[bool, str]:
    """수식이 정확히 '=<left><row><op><right><row>' 인지 검증.

    반환 (ok, detail). 공백·$ 는 무시하고 정규화 비교한다.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return False, "수식 아님(%r)" % formula
    norm = formula.replace("$", "").replace(" ", "").upper()
    want = "=%s%d%s%s%d" % (get_column_letter(left_col), row, op,
                            get_column_letter(right_col), row)
    if norm == want.upper():
        return True, ""
    return False, "기대 %s ≠ 실제 %s" % (want, formula)


def run_meta_checks(rep: QCReport, wb, meta: dict) -> None:
    """meta['formula_checks'] 선언을 스윕해 region 의 수식 참조를 강제.

    스파인이 전 템플릿에 호출 — 선언 없으면 no-op. 선언했는데 region 셀이 비거나
    수식이 아니면 fail(침묵 통과 금지).
    """
    checks = (meta or {}).get("formula_checks")
    if not checks:
        return
    by_sheet = {ws.title: ws for ws in wb.worksheets}
    for chk in checks:
        sheet = chk.get("sheet")
        r0, r1, col = chk["region"]
        op = chk.get("op", "-")
        left, right = chk["left"], chk["right"]
        name = chk.get("name", "수식참조 %s!col%d" % (sheet, col))
        ws = by_sheet.get(sheet)
        if ws is None:
            rep.add(name, False, "시트 없음: %s" % sheet)
            continue
        bad: list = []
        for r in range(r0, r1 + 1):
            cell = ws.cell(row=r, column=col)
            if cell.value is None:
                continue
            ok, detail = check_binary(cell.value, row=r, left_col=left,
                                      right_col=right, op=op)
            if not ok:
                bad.append("%s%d(%s)" % (get_column_letter(col), r, detail))
        rep.add(name, not bad,
                "" if not bad else "참조 불일치 %d건: %s" % (len(bad), "; ".join(bad[:6])))


# --------------------------------------------------------------------------- #
# 비율 안전(CPU/CPP) — #VALUE!/#DIV/0! 구조적 차단                              #
#   (W26 스크린샷: FCPU/FCPP 가 분모 오염으로 #VALUE!. 가드 없는 '/' 가 원인.)   #
# --------------------------------------------------------------------------- #
def safe_ratio_formula(num_ref: str, den_ref: str, *, na: str = '"NA"') -> str:
    """분모가 숫자 아님/0 이면 NA, 아니면 num/den 인 Excel 수식 문자열.

    CPU = 금액/물량 같은 비율 셀은 *반드시* 이걸로 쓴다. 분모 셀에 텍스트(주석)나
    공란이 와도 #VALUE!/#DIV/0! 대신 깨끗한 "NA" 를 표시 → 화면 오염 차단.
    view_contract.ratio_or_na(파이썬 측 NA)와 같은 철학의 수식 버전.
    """
    return '=IF(OR(NOT(ISNUMBER(%s)),%s=0),%s,%s/%s)' % (
        den_ref, den_ref, na, num_ref, den_ref)


def unguarded_divisions(wb) -> list:
    """가드(IFERROR/ISNUMBER/IF) 없이 '/' 를 쓴 수식 셀을 탐지(dead 파일 휴리스틱).

    분모가 텍스트/공란이면 #VALUE!/#DIV/0! 가 날 후보. 반환 [{sheet,coord,formula}].
    ⚠ 휴리스틱(advisory) — 분모가 늘 숫자라 안전한 경우도 잡을 수 있다.
    """
    out: list = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=") and "/" in v):
                    continue
                up = v.upper()
                if "IFERROR" in up or "ISNUMBER" in up or "IFNA" in up:
                    continue
                out.append({"sheet": ws.title, "coord": c.coordinate, "formula": v})
    return out


__all__ = ["parse_refs", "check_binary", "run_meta_checks",
           "safe_ratio_formula", "unguarded_divisions"]
