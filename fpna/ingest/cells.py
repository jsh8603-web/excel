"""
fpna.ingest.cells — 시트를 "셀 1개 = 1행" 평면 테이블로 환원(as_cells).

unpivotr 의 핵심 통찰: 모든 정형화는 grid 가 아니라 셀 좌표 평면 위에서 돈다.
값/타입/서식/병합 anchor 를 한 dataclass(Cell)에 담아 이후 단계가 소비한다.

순수 파이썬(openpyxl + stdlib). 결정적.
"""
from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

from openpyxl.worksheet.worksheet import Worksheet

# 셀 클래스(분류 결과)
BLANK = "blank"
DATA = "data"
HEADER = "header"
LABEL = "label"
CORNER = "corner"

# data_type
T_BLANK = "blank"
T_NUMERIC = "numeric"
T_CHARACTER = "character"
T_DATE = "date"
T_ERROR = "error"
T_LOGICAL = "logical"

# Excel 에러 센티넬
ERROR_LITERALS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?",
                  "#NULL!", "#NUM!", "#SPILL!", "#CALC!", "#GETTING_DATA"}


@dataclass
class Fmt:
    bold: bool = False
    italic: bool = False
    indent: int = 0
    align_h: str | None = None
    align_v: str | None = None
    font_color: str | None = None
    number_format: str = "General"
    has_top_border: bool = False
    has_bottom_border: bool = False


@dataclass
class Cell:
    row: int
    col: int
    value: object = None
    formula: str | None = None          # data_only=False 로 읽은 수식 문자열(있으면)
    data_type: str = T_BLANK
    is_blank: bool = True
    fmt: Fmt = field(default_factory=Fmt)
    merge_anchor: tuple[int, int] | None = None   # 병합영역 좌상단 (없으면 None)
    cls: str = BLANK                    # 분류 결과(classify 단계에서 채움)


def _classify_type(v) -> str:
    if v is None:
        return T_BLANK
    if isinstance(v, bool):
        return T_LOGICAL
    if isinstance(v, (int, float)):
        return T_NUMERIC
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return T_DATE
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return T_BLANK
        if s in ERROR_LITERALS:
            return T_ERROR
        return T_CHARACTER
    return T_CHARACTER


def _color_hex(font) -> str | None:
    try:
        c = font.color
        if c is None:
            return None
        rgb = getattr(c, "rgb", None)
        if isinstance(rgb, str) and len(rgb) >= 6:
            return rgb[-6:].upper()
    except Exception:
        pass
    return None


def _read_fmt(cell) -> Fmt:
    f = Fmt()
    try:
        font = cell.font
        f.bold = bool(font.bold)
        f.italic = bool(font.italic)
        f.font_color = _color_hex(font)
    except Exception:
        pass
    try:
        al = cell.alignment
        f.indent = int(al.indent or 0)
        f.align_h = al.horizontal
        f.align_v = al.vertical
    except Exception:
        pass
    try:
        f.number_format = cell.number_format or "General"
    except Exception:
        pass
    try:
        b = cell.border
        f.has_top_border = bool(b.top and b.top.style)
        f.has_bottom_border = bool(b.bottom and b.bottom.style)
    except Exception:
        pass
    return f


def build_merge_index(ws: Worksheet) -> dict[tuple[int, int], tuple[int, int]]:
    """(row,col) -> 병합영역 좌상단 anchor(row,col)."""
    idx: dict[tuple[int, int], tuple[int, int]] = {}
    for rng in ws.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                idx[(r, c)] = anchor
    return idx


# as_cells 안전 상한: 실데이터 경계 면적이 이걸 넘으면 명시적 에러로 차단(조용한 OOM/행 방지).
MAX_INGEST_CELLS = 3_000_000


def _content_extent(ws: Worksheet) -> tuple[int, int]:
    """시트에서 '실제 값/수식이 있는' 셀의 최대 (row, col).

    openpyxl 의 sparse `_cells` 를 직접 보아 멀리 떨어진 '잔여 서식'(값 없는 빈칸)은 제외한다.
    → max_row/max_column 이 잔여 서식 한 칸 때문에 부풀려져도 실데이터 경계만 잡는다.
    `_cells` 부재(read_only 시트)면 (0,0) → 호출측이 max_row/col 로 폴백.
    """
    store = getattr(ws, "_cells", None)
    if not store:
        return 0, 0
    mr = mc = 0
    for (r, c), cell in store.items():
        v = cell.value
        if v is None or (isinstance(v, str) and v.strip() == ""):
            continue
        if r > mr:
            mr = r
        if c > mc:
            mc = c
    return mr, mc


def as_cells(ws_formula: Worksheet, ws_value: Worksheet | None = None) -> list[Cell]:
    """워크시트를 Cell 리스트로 평면화.

    ws_formula : data_only=False 로 연 시트(수식 문자열 보유)
    ws_value   : data_only=True 로 연 같은 시트(캐시 계산값 보유, 선택)

    값은 ws_value(있으면) 우선, 수식은 ws_formula 에서 채운다.
    openpyxl 한계상 ws_value 의 캐시는 'Excel 이 마지막 저장한 값' 이므로
    미개봉 파일은 None 일 수 있다(상위에서 경고).

    ⚠ 경계는 부풀려진 bbox(max_row×max_col)가 아니라 '실데이터 경계'(_content_extent
    + 값셀 + 병합영역)로 clamp 한다 — 잔여 서식 한 칸 때문에 수백만 빈 Cell 을 만들어
    메모리가 터지는(=조용히 멈추는) 사고를 막는다. 경계 내부의 빈칸은 병합전파·ditto fill 에
    필요하므로 그대로 유지한다.
    """
    merged = build_merge_index(ws_formula)
    cells: list[Cell] = []
    bbox_row = ws_formula.max_row or 0
    bbox_col = ws_formula.max_column or 0

    val_grid = {}
    if ws_value is not None:
        vstore = getattr(ws_value, "_cells", None)
        if vstore:                              # sparse 우선(부풀린 bbox 전수순회 회피)
            for (r, c), cell in vstore.items():
                v = cell.value
                if v is not None:
                    val_grid[(r, c)] = v
        else:                                   # read_only 등 _cells 부재 → 종전 방식 폴백
            for r in range(1, (ws_value.max_row or 0) + 1):
                for c in range(1, (ws_value.max_column or 0) + 1):
                    v = ws_value.cell(row=r, column=c).value
                    if v is not None:
                        val_grid[(r, c)] = v

    # 실데이터 경계 = 값/수식 보유 셀 ∪ 값셀 ∪ 병합영역(잔여 서식 빈칸은 제외).
    fr, fc = _content_extent(ws_formula)
    max_row = max(fr, max((r for (r, _) in val_grid), default=0),
                  max((r for (r, _) in merged), default=0))
    max_col = max(fc, max((c for (_, c) in val_grid), default=0),
                  max((c for (_, c) in merged), default=0))
    if max_row == 0:                            # sparse 폴백 실패 → 종전 bbox 사용
        max_row = bbox_row
    if max_col == 0:
        max_col = bbox_col
    if bbox_row:                                # 안전: 부풀린 bbox 를 절대 넘지 않음
        max_row = min(max_row, bbox_row)
    if bbox_col:
        max_col = min(max_col, bbox_col)

    area = max_row * max_col
    if area > MAX_INGEST_CELLS:                 # 거대 시트 → 조용한 멈춤 대신 명시 차단
        raise ValueError(
            "시트가 너무 큼: 실데이터 경계 %d행 × %d열 = %d셀 (상한 %d). "
            "--sheet 로 단일 시트를 지정하거나 데이터 범위를 줄여 다시 시도하세요."
            % (max_row, max_col, area, MAX_INGEST_CELLS)
        )

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            fcell = ws_formula.cell(row=r, column=c)
            raw = fcell.value
            formula = None
            value = raw
            if isinstance(raw, str) and raw.startswith("="):
                formula = raw
                value = val_grid.get((r, c))  # 캐시값(없으면 None)
            elif (r, c) in val_grid:
                value = val_grid[(r, c)]

            dt = _classify_type(value)
            is_blank = value is None or (isinstance(value, str) and value.strip() == "")
            cells.append(Cell(
                row=r, col=c, value=value, formula=formula,
                data_type=dt, is_blank=is_blank,
                fmt=_read_fmt(fcell),
                merge_anchor=merged.get((r, c)),
            ))
    return cells


def cells_by_pos(cells: list[Cell]) -> dict[tuple[int, int], Cell]:
    return {(c.row, c.col): c for c in cells}


__all__ = [
    "Cell", "Fmt", "as_cells", "build_merge_index", "cells_by_pos",
    "BLANK", "DATA", "HEADER", "LABEL", "CORNER",
    "T_BLANK", "T_NUMERIC", "T_CHARACTER", "T_DATE", "T_ERROR", "T_LOGICAL",
    "ERROR_LITERALS",
]
