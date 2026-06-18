"""
fpna.ingest.pipeline — 누더기 엑셀 → tidy 변환 오케스트레이터.

단계(리서치 보강 반영, 결정적·재현가능):
  1) load: data_only False/True 두 번 열어 값+수식 동시 수집(as_cells)
  2) detect_blocks: connected-component + density 로 다중 표 분리
  2.5) 비데이터 행 격리: 제목/단위/각주
  3) unmerge_fill + classify + 헤더 밴드
  4) behead 언피벗 → long
  4.5) 소계/합계 행 플래그(이중집계 방지)
  5) 센티넬·단위·스케일 정규화
  6) dataclass 스키마 검증 → reject 리포트
  7) 수식 스멜 스캔
  8) 산출: tidy.csv(utf-8-sig) + schema.json + smell_report.md

회사 PC: `py main.py ingest <파일.xlsx>` 로 호출.
"""
from __future__ import annotations

import csv
import datetime as _dt
import json
import os
import re
from dataclasses import asdict, dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from .cells import as_cells, T_NUMERIC, T_DATE, ERROR_LITERALS
from .detect import (detect_blocks, cells_in_block, strip_title_rows,
                     strip_footnote_rows, strip_repeated_header_rows,
                     label_is_subtotal, subtotal_signal_score, Block, UNIT_RE)
from .headers import (unpivot_block, unmerge_fill, LongRow,
                      classify_cells, fill_down_ditto, no_header_suspect)
from .normalize import (normalize_value_ex, parse_unit_label, infer_column_type,
                        scale_for_unit, strip_footnote_marker)
from .validate import TidyRow, validate_rows, scan_formula_smells, Smell
from .reconcile import groundtruth_cells, reconcile_sheet, recon_to_smells

_PERIOD_RE = re.compile(r"(\d{4}\s*[-/.년]?\s*(\d{1,2})?\s*(월|분기|Q|H)?|\dQ|[1-4]분기|상반기|하반기)",
                        re.IGNORECASE)


@dataclass
class IngestResult:
    tidy_rows: list = field(default_factory=list)
    schema: dict = field(default_factory=dict)
    smells: list = field(default_factory=list)
    report: object = None
    n_blocks: int = 0
    recon: list = field(default_factory=list)   # A1: 시트별 ReconReport


def _looks_like_period(v) -> bool:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return True
    if v is None:
        return False
    return bool(_PERIOD_RE.search(str(v)))


def _map_long_to_tidy(longs: list[LongRow], unit: str | None, sheet: str,
                      smells: list, *, block_scale: int = 1) -> list[TidyRow]:
    """LongRow.attrs(hdr_c*/hdr_r*) → (entity, period, metric, value) 휴리스틱 매핑.

    - period: period 패턴인 헤더값 우선(보통 열헤더).
    - metric: period 아닌 열헤더 또는 가장 안쪽 행헤더.
    - entity: 가장 바깥 행헤더.
    나머지 헤더는 metric 에 ' > ' 로 결합 보존.

    무음 손상 방어:
      ① 단위 스케일 환산 — 셀내 접미(셀) > 블록 단위(블록) > 1 우선순위로 base(원) 환산.
         열 단위로 셀 스케일 2종+ 혼재 → smell SCALE_HETEROGENEOUS.
      ③ 오류값 보존 — ERROR_LITERALS → value=null + raw_value 보존 + smell ERROR_CELL.
         ⛔ 0/NaN coerce 금지.

    block_scale: 블록 자체 단위행에서 산출한 base 환산 곱수(시트 폴백 단위는
                 라벨로만 보존하고 환산엔 쓰지 않는다 — 무단위 블록 오염 방지).
    """
    out: list[TidyRow] = []
    # 열별 셀-스케일 집계(혼재 탐지용). key = (entity? no) → metric+col 식별 위해 src_col 사용.
    col_cell_scales: dict[int, set] = {}
    # MetaCollector 흡수: 열별 표시서식 집계 → text-as-num(@서식+숫자) / mixed-format 탐지
    col_number_formats: dict[int, set] = {}
    col_text_as_num: dict[int, int] = {}

    for lr in longs:
        col_vals = [(k, v) for k, v in lr.attrs.items() if k.startswith("hdr_c")]
        row_vals = [(k, v) for k, v in lr.attrs.items() if k.startswith("hdr_r")]

        period = None
        metric_parts: list[str] = []
        for _, v in col_vals:
            if v is None:
                continue
            if period is None and _looks_like_period(v):
                period = str(v)
            else:
                metric_parts.append(str(v))

        entity = None
        row_labels = [str(v) for _, v in row_vals if v is not None]
        if row_labels:
            entity = row_labels[0]
            if len(row_labels) > 1:
                metric_parts = row_labels[1:] + metric_parts
        metric = " > ".join(metric_parts) if metric_parts else None

        coord = "R%dC%d" % (lr.row, lr.col)

        # ③ 오류값 보존: data_type=='e' 또는 ERROR_LITERALS → null + 원본보존 + smell.
        if isinstance(lr.value, str) and lr.value.strip() in ERROR_LITERALS:
            kind = lr.value.strip()
            smells.append(Smell(coord, "ERROR_CELL", "엑셀 오류값 보존(null 처리): %s" % kind))
            role = "data"
            label_for_role = " ".join(filter(None, [entity, metric or ""]))
            if label_is_subtotal(label_for_role):
                role = "subtotal"
            out.append(TidyRow(
                entity=entity, period=period, metric=metric,
                value=None, unit=unit, row_role=role, level=lr.level,
                src_sheet=sheet, src_row=lr.row, src_col=lr.col,
                scale_applied=1, scale_source="none",
                raw_value=kind, flags="ERROR_CELL",
            ))
            continue

        # ① 정규화 + 셀내 접미 스케일 노출
        num, sentinel, _neg, cell_scale = normalize_value_ex(lr.value)

        # 퍼센트는 스케일 환산 대상 아님(0.85 를 1000배 하면 안 됨).
        is_pct = isinstance(lr.value, str) and lr.value.strip().endswith("%")

        # 환산 곱수 결정: 셀 > 블록 > 1
        scale = 1
        scale_source = "none"
        is_num = sentinel is None and isinstance(num, (int, float)) and not isinstance(num, bool)
        if is_num and not is_pct:
            if cell_scale > 1:
                scale = cell_scale
                scale_source = "cell"
            elif block_scale > 1:
                scale = block_scale
                scale_source = "block"
            col_cell_scales.setdefault(lr.col, set()).add(cell_scale)

        # MetaCollector 흡수: 열별 표시서식 집계(에러셀은 위에서 continue 됨)
        _nf = lr.number_format or "General"
        col_number_formats.setdefault(lr.col, set()).add(_nf)
        if _nf == "@" and is_num:        # 텍스트서식인데 숫자 내용 → SUM 누락 위험
            col_text_as_num[lr.col] = col_text_as_num.get(lr.col, 0) + 1

        raw_value = None
        value = None if sentinel is not None else num
        if is_num and scale > 1:
            raw_value = num                 # 환산 전 원본 보존
            value = num * scale

        flags = ""
        # G5 색 음수: 빨강폰트 양수 → 음수 보정 + SIGN_FROM_COLOR.
        #   이미 음수표기(괄호/△)면 중복 부호화 금지(_neg 가 True면 건너뜀).
        if (lr.cell_red and is_num and not _neg
                and isinstance(value, (int, float)) and value > 0):
            if raw_value is None:
                raw_value = value
            value = -value
            flags = "SIGN_FROM_COLOR"
            smells.append(Smell(coord, "SIGN_FROM_COLOR",
                                "빨강폰트 음수 보정: %s → %s" % (raw_value, value)))

        # G3 계층 레벨: lr.level(indent+선행공백). G5 볼드 신호 보존.
        role = "data"
        label_for_role = " ".join(filter(None, [entity, metric or ""]))
        if label_is_subtotal(label_for_role):
            role = "subtotal"
        out.append(TidyRow(
            entity=entity, period=period, metric=metric,
            value=value, unit=unit, row_role=role, level=lr.level,
            src_sheet=sheet, src_row=lr.row, src_col=lr.col,
            scale_applied=scale, scale_source=scale_source,
            raw_value=raw_value, flags=flags,
        ))
        # 후처리 신호 보관(볼드 → G5 산술 교집합).
        out[-1]._g5_bold = lr.label_bold        # type: ignore[attr-defined]
        out[-1]._g5_label = label_for_role       # type: ignore[attr-defined]

    # ① 열 단위 셀-스케일 혼재(2종+) → SCALE_HETEROGENEOUS smell
    for col, scs in sorted(col_cell_scales.items()):
        distinct = {s for s in scs if s > 1}
        # 셀 스케일이 2종 이상이거나, 일부 셀만 스케일 가짐(1 과 >1 혼재)
        if len(distinct) >= 2 or (distinct and (1 in scs)):
            smells.append(Smell(
                "C%d" % col, "SCALE_HETEROGENEOUS",
                "한 열에 셀 스케일 혼재: %s" % sorted(scs)))

    # MetaCollector 흡수: text-as-num(@서식+숫자) — Excel SUM 에서 무음 누락되는 사고
    for col, cnt in sorted(col_text_as_num.items()):
        smells.append(Smell("C%d" % col, "TEXT_AS_NUM_SUSPECT",
            "텍스트서식(@) 셀에 숫자 내용 %d개 — Excel SUM 누락 위험" % cnt))
    # mixed number_format(General/@ 제외 2종+) — 복붙 오염 지문
    for col, fmts in sorted(col_number_formats.items()):
        distinct = sorted({f for f in fmts if f and f not in ("General", "@")})
        if len(distinct) >= 2:
            smells.append(Smell("C%d" % col, "MIXED_NUMBER_FORMAT",
                "한 열에 표시서식 %d종 혼재: %s" % (len(distinct), ", ".join(distinct[:4]))))

    _apply_subtotal_and_hierarchy(out, smells)
    return out


_ARITH_TOL = 1e-6
# A0: 합계행 산술 reject 허용오차(반올림 흡수). 상대+절대 둘 다 둔다 —
#   재무 반올림(천/백만 절사) 누적분을 흡수하되 명백 불일치만 reject.
#   ★전제: value 는 이미 base(원) 스케일 환산 後 → 천원/원 혼재 오탐 없음.
_RECONCILE_REL_TOL = 0.005   # 0.5% 상대
_RECONCILE_ABS_TOL = 1.0     # base(원) 1 절대(반올림 1단위 흡수)


def _arith_close(a: float, b: float) -> bool:
    """합계 정합 판정(상대 0.5% 또는 절대 1 이내)."""
    return abs(a - b) <= max(_RECONCILE_ABS_TOL, _RECONCILE_REL_TOL * (abs(a) + abs(b)))


def _apply_subtotal_and_hierarchy(rows: list[TidyRow], smells: list) -> None:
    """G5 산술 소계 교집합 + G3 부모==Σ자식 검증 + A0 합계 reject(후처리, 결정적).

    - G5: (label·bold·arith) 3신호 중 score≥2 → row_role='subtotal'.
      arith = 같은 (period,metric) 그룹에서 그 행 값 == 나머지 형제 합.
    - G3: 같은 (period,metric) 그룹에서 부모(level L)가 직후 자식(level>L) 합과
      일치하면 PARENT_EQ_CHILDREN_SUM 플래그(이중집계 방지 안전망).
    - A0: 라벨이 명시 합계/소계인 행이 자식 합과 명백히 어긋나면(허용오차 밖)
      SUBTOTAL_ARITH_MISMATCH 플래그 → validate 에서 reject(원본 합계오류 차단).
      ★스케일 정규화 後 비교(value 가 base 환산값) — 천원/원 혼재 오탐 회피.
    """
    def num(r):
        return r.value if isinstance(r.value, (int, float)) and not isinstance(r.value, bool) else None

    # (period, metric) 그룹 — 같은 열 슬롯에서 형제 비교.
    groups: dict[tuple, list[TidyRow]] = {}
    for r in rows:
        groups.setdefault((r.period, r.metric), []).append(r)

    for _key, grp in groups.items():
        grp_sorted = sorted(grp, key=lambda r: r.src_row)
        # --- G5 산술 소계: 후보 == 나머지 형제(소계 아님) 합 ---
        numeric = [r for r in grp_sorted if num(r) is not None]
        # ⚠ label_is_subtotal 는 행당 1회만 계산한다. 형제 합도 그룹당 1회만 구하고
        #   각 행에서 자기 값만 차감 → O(G²)(큰 시트 폭주) 를 O(G) 로 낮춘다.
        non_sub = [o for o in numeric
                   if not label_is_subtotal(getattr(o, "_g5_label", ""))]
        non_sub_sum = sum(num(o) for o in non_sub)
        non_sub_ids = {id(o) for o in non_sub}
        for r in grp_sorted:
            v = num(r)
            bold = getattr(r, "_g5_bold", False)
            label = getattr(r, "_g5_label", "")
            arith = False
            if v is not None and len(numeric) >= 2:
                # 형제 = 자신 외, 소계 라벨 아닌 나머지 numeric (합 1회분에서 자기 차감).
                r_in = id(r) in non_sub_ids
                sibs_sum = non_sub_sum - (v if r_in else 0.0)
                sibs_n = len(non_sub) - (1 if r_in else 0)
                if sibs_n >= 1 and abs(v - sibs_sum) <= _ARITH_TOL * (abs(v) + 1):
                    arith = True
            score, _sig = subtotal_signal_score(label, bold=bold, arith_match=arith)
            if score >= 2:
                if r.row_role != "subtotal":
                    r.row_role = "subtotal"
                if "SUBTOTAL_COLOR_BOLD" not in r.flags:
                    r.flags = (r.flags + ";SUBTOTAL_COLOR_BOLD").strip(";")
                    smells.append(Smell(
                        "R%dC%d" % (r.src_row, r.src_col), "SUBTOTAL_DETECTED",
                        "소계 신호 score=%d (label=%s bold=%s arith=%s)"
                        % (score, _sig["label"], bold, arith)))

        # --- G3 부모==Σ자식: level 기반 ---
        for i, parent in enumerate(grp_sorted):
            pv = num(parent)
            if pv is None:
                continue
            children = []
            for child in grp_sorted[i + 1:]:
                if child.level > parent.level:
                    if num(child) is not None:
                        children.append(child)
                elif child.level <= parent.level:
                    break  # 같은/상위 레벨 만나면 형제 경계
            if len(children) >= 2:
                csum = sum(num(c) for c in children)
                if abs(pv - csum) <= _ARITH_TOL * (abs(pv) + 1):
                    parent.flags = (parent.flags + ";PARENT_EQ_CHILDREN_SUM").strip(";")
                    smells.append(Smell(
                        "R%dC%d" % (parent.src_row, parent.src_col),
                        "PARENT_EQ_CHILDREN_SUM",
                        "부모 %s == Σ자식(%d) — 계층 합 정합" % (pv, len(children))))

        # --- A0 합계행 산술 reject: 명시 합계/소계 라벨이 자식 합과 명백 불일치 ---
        #   level 계층(부모 들여쓰기 + 자식) 우선, 없으면 평면 형제합으로 폴백.
        #   허용오차(반올림) 밖이면 SUBTOTAL_ARITH_MISMATCH → validate reject.
        _flag_subtotal_arith_mismatch(grp_sorted, num, smells)


def _flag_subtotal_arith_mismatch(grp_sorted: list, num, smells: list) -> None:
    """A0: 같은 (period,metric) 그룹에서 명시 합계/소계 행의 산술 불일치 탐지.

    ★보수적: **계층(level) 부모-자식 관계가 명시된 경우만** 비교한다. 들여쓰기로
    부모(합계/소계 라벨)가 자식보다 얕은 level 이고, 아래에 더 깊은 자식이 ≥2 개
    이어질 때만 'Σ자식' 을 산출해 대조. 평면 형제합 추정은 무관 행을 합산해
    오탐(정상 표를 reject)을 내므로 쓰지 않는다.

    후보 = label_is_subtotal 또는 row_role=='subtotal'. 부모값 != Σ자식(허용오차
    밖)이면 SUBTOTAL_ARITH_MISMATCH 플래그 → validate_rows reject(원본 합계오류 차단).
    ⚠ 명백 불일치만(허용오차 안 = 침묵). 계층 미형성이면 무동작(reject 안 함).
    """
    for idx, cand in enumerate(grp_sorted):
        cv = num(cand)
        if cv is None:
            continue
        is_cand = (cand.row_role == "subtotal"
                   or label_is_subtotal(getattr(cand, "_g5_label", "")))
        if not is_cand:
            continue
        # 계층 자식: 후보 아래, 더 깊은 level, 같은/얕은 level 만나면 형제 경계.
        children = []
        for child in grp_sorted[idx + 1:]:
            if child.level > cand.level:
                if num(child) is not None:
                    children.append(child)
            else:
                break
        if len(children) < 2:
            continue  # 계층 미형성 — 비교 근거 없음(오탐 회피, reject 보류)
        ssum = sum(num(c) for c in children)
        if _arith_close(cv, ssum):
            continue  # 정합(반올림 흡수) — 침묵
        if "SUBTOTAL_ARITH_MISMATCH" not in (cand.flags or ""):
            cand.flags = (cand.flags + ";SUBTOTAL_ARITH_MISMATCH").strip(";")
            smells.append(Smell(
                "R%dC%d" % (cand.src_row, cand.src_col),
                "SUBTOTAL_ARITH_MISMATCH",
                "합계 %s != Σ계층자식(%d)=%s — 원본 산술 오류(reject)"
                % (cv, len(children), ssum)))


def ingest_workbook(path: str, *, sheet: str | None = None) -> IngestResult:
    """엑셀 파일을 tidy long 으로 변환."""
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True, read_only=False)

    sheets = [sheet] if sheet else wb_f.sheetnames
    all_tidy: list[TidyRow] = []
    all_smells = []
    schema_blocks = []
    n_blocks = 0

    recon_reports: list = []
    for sn in sheets:
        ws_f = wb_f[sn]
        ws_v = wb_v[sn]
        cells = as_cells(ws_f, ws_v)
        all_smells.extend([asdict_smell(s, sn) for s in scan_formula_smells(cells)])

        # A1: ground-truth 좌표 = 정규화/병합전파 전, parse 와 독립한 단순 스캔.
        #   unmerge_fill 이 cells 를 in-place 변형하기 전에 좌표집합 + raw값맵을 떠야
        #   순환(같은 블록탐지 재사용)·오염(병합전파 후 값) 둘 다 회피.
        sheet_gt = groundtruth_cells(cells)
        sheet_raw_values = {(c.row, c.col): c.value for c in cells if not c.is_blank}

        # 병합셀 값을 영역 내 전파(헤더 해소·블록탐지·제목격리 정확도↑)
        unmerge_fill(cells)
        # 시트 단위 라벨 폴백: '(단위: ...)' 를 시트 전체에서 1회 스캔
        sheet_unit = None
        for c in cells:
            if isinstance(c.value, str):
                m = UNIT_RE.search(c.value)
                if m:
                    sheet_unit = m.group(1).strip()
                    break

        blocks = detect_blocks(cells)
        for bi, b in enumerate(blocks):
            n_blocks += 1
            bc = cells_in_block(cells, b)
            bc, _titles, unit_meta = strip_title_rows(bc, b)
            bc, _foots = strip_footnote_rows(bc, b)
            block_unit = unit_meta.get("unit")          # 블록 자체 선언 단위
            unit = block_unit or sheet_unit             # 라벨(표시)용 — 시트 폴백 포함
            # ⚠ 환산 곱수는 블록 자체 단위에서만(시트 폴백 단위는 무단위 블록을
            #   오염시키므로 값 환산에 쓰지 않는다 — 라벨로만 보존).
            block_scale = scale_for_unit(block_unit)
            # 블록 bbox 재계산(격리 후)
            if not bc:
                continue
            rows = [c.row for c in bc]
            cols = [c.col for c in bc]
            b2 = Block(min(rows), max(rows), min(cols), max(cols))

            # ② ditto fill-down: 좌측 라벨 밴드의 빈칸(상동)을 위→아래로 충전.
            #    classify 로 좌헤더 폭 산출 → 그 열들만 단방향 fill.
            unmerge_fill(bc)
            _top, left_cols = classify_cells(bc, b2)
            ditto_cols = list(range(b2.min_col, b2.min_col + left_cols))
            filled = fill_down_ditto(bc, ditto_cols, (b2.min_row, b2.max_row))

            block_smells: list = []

            # ② no-header 의심(MetaCollector GuessHeaderRange 흡수): 헤더행이 데이터와
            #    타입 유사 + 숫자 포함 → 헤더 없는 표일 수 있어 첫 행이 헤더로 먹힐 위험.
            #    오탐 위험이 있어 격하(정제변경)는 안 하고 경고만 — 사용자가 판단.
            if no_header_suspect(bc, b2, _top):
                block_smells.append(Smell(
                    "R%dC%d" % (b2.min_row, b2.min_col), "NO_HEADER_SUSPECT",
                    "헤더행이 데이터와 타입 유사+숫자 포함 — 헤더 없는 표 의심(첫 행 손실 가능)"))

            # G8 반복헤더: 페이지브레이크로 재삽입된 헤더행(첫 헤더와 동일) 제거.
            header_rows = list(range(b2.min_row, b2.min_row + max(_top, 1)))
            bc, dropped_hdr = strip_repeated_header_rows(bc, b2, header_rows)
            if dropped_hdr:
                # 격리 후 bbox 재계산
                rr = [c.row for c in bc]
                cc = [c.col for c in bc]
                b2 = Block(min(rr), max(rr), min(cc), max(cc))
                block_smells.append(Smell(
                    "R%dC%d" % (header_rows[0], b2.min_col),
                    "REPEATED_HEADER_DROPPED",
                    "표 중간 반복헤더 %d행 제거: %s" % (len(dropped_hdr), dropped_hdr)))

            # G7 각주마커 DUP_KEY: 헤더밴드에서 마커 제거 후 같은 키 충돌 탐지.
            _detect_dup_keys(bc, header_rows, block_smells)

            longs = unpivot_block(bc, b2)
            tidy = _map_long_to_tidy(longs, unit, sn, block_smells,
                                     block_scale=block_scale)
            # ditto 로 채워진 src 라벨이 기여한 행에 DITTO_FILLED 플래그
            if filled:
                filled_rows = {r for (r, _c) in filled}
                for tr in tidy:
                    if tr.src_row in filled_rows:
                        tr.flags = (tr.flags + ";DITTO_FILLED").strip(";")
                block_smells.append(Smell(
                    "R%dC%d" % (b2.min_row, b2.min_col), "DITTO_FILLED",
                    "카테고리 빈칸 %d개 위→아래 충전" % len(filled)))
            all_smells.extend([asdict_smell(s, sn) for s in block_smells])
            all_tidy.extend(tidy)
            schema_blocks.append({
                "sheet": sn, "block": bi,
                "range": "R%dC%d:R%dC%d" % (b.min_row, b.min_col, b.max_row, b.max_col),
                "unit": unit, "unit_scale": block_scale,
                "n_long_rows": len(tidy),
            })

        # A1: 시트 단위 per-cell reconciliation(충실도 게이트). GT/raw 는 unmerge 前
        #   snapshot, tidy 는 이 시트 소속 행만. 결과는 smell 로 노출(reject 와 독립).
        sheet_tidy = [t for t in all_tidy if t.src_sheet == sn]
        recon = reconcile_sheet(sn, sheet_gt, sheet_raw_values, sheet_tidy)
        recon_reports.append(recon)
        all_smells.extend([asdict_smell(s, sn) for s in recon_to_smells(recon)])

    kept, rep = validate_rows(all_tidy, numeric_metric=False)

    # 컬럼 타입 추론(스키마 문서화용)
    schema = {
        "columns": {
            "entity": "TEXT", "period": "TEXT",
            "metric": "TEXT",
            "value": infer_column_type([r.value for r in kept]),
            "unit": "TEXT", "row_role": "TEXT", "level": "NUM",
            "src_sheet": "TEXT", "src_row": "NUM", "src_col": "NUM",
            "scale_applied": "NUM", "scale_source": "TEXT",
            "raw_value": "TEXT", "flags": "TEXT",
        },
        "blocks": schema_blocks,
        "n_rows": len(kept),
        "n_rejected": rep.n_rejected,
        # A1: 충실도(per-cell reconciliation) 요약 — 시트별 coverage/mismatch.
        "reconciliation": [{
            "sheet": rc.sheet,
            "n_groundtruth": rc.n_groundtruth,
            "n_covered": len(rc.covered),
            "n_missing": len(rc.missing),
            "n_duplicate": len(rc.duplicate),
            "n_value_mismatch": len(rc.value_mismatch),
            "ok": rc.ok,
        } for rc in recon_reports],
        "generated_by": "fpna.ingest.pipeline",
    }
    return IngestResult(tidy_rows=kept, schema=schema, smells=all_smells,
                        report=rep, n_blocks=n_blocks, recon=recon_reports)


def _detect_dup_keys(block_cells: list, header_rows: list, smells: list) -> None:
    """G7: 헤더밴드 각주마커 제거 후 동일 행에서 같은 논리키 충돌 → DUP_KEY smell.

    마커 제거 전엔 '매출¹'/'매출²' 로 구별되던 두 열헤더가 같은 '매출' 로 붕괴하면
    언피벗 시 metric 키가 합쳐져 값이 섞일 위험 → 표면화.
    """
    hr = set(header_rows)
    for r in sorted(hr):
        seen: dict[str, int] = {}
        for c in sorted([c for c in block_cells if c.row == r and not c.is_blank],
                        key=lambda x: x.col):
            if not isinstance(c.value, str):
                continue
            key, stripped = strip_footnote_marker(c.value)
            if not stripped:
                continue
            if key in seen and seen[key] != c.col:
                smells.append(Smell(
                    "R%dC%d" % (r, c.col), "DUP_KEY",
                    "각주마커 제거 후 키 충돌: '%s'(C%d, C%d)"
                    % (key, seen[key], c.col)))
            else:
                seen.setdefault(key, c.col)


def asdict_smell(s, sheet) -> dict:
    return {"sheet": sheet, "cell": s.cell, "kind": s.kind, "detail": s.detail}


# --------------------------------------------------------------------------
# 산출 writer (utf-8-sig: 한글 Excel 더블클릭 호환)
# --------------------------------------------------------------------------
def write_tidy_csv(rows: list[TidyRow], path: str) -> None:
    cols = ["entity", "period", "metric", "value", "unit",
            "row_role", "level", "src_sheet", "src_row", "src_col",
            "scale_applied", "scale_source", "raw_value", "flags"]
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(cols)
        for r in rows:
            d = asdict(r)
            w.writerow([d[c] for c in cols])


def write_schema_json(schema: dict, path: str) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(schema, fh, ensure_ascii=False, indent=2, default=str)


def write_smell_report(smells: list[dict], path: str) -> None:
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("# 수식 스멜 리포트\n\n")
        if not smells:
            fh.write("스멜 없음.\n")
            return
        by_kind: dict[str, list] = {}
        for s in smells:
            by_kind.setdefault(s["kind"], []).append(s)
        for kind, items in sorted(by_kind.items()):
            fh.write("## %s (%d건)\n\n" % (kind, len(items)))
            for s in items[:50]:
                fh.write("- `%s` %s — %s\n" % (s["sheet"], s["cell"], s["detail"]))
            if len(items) > 50:
                fh.write("- … 외 %d건\n" % (len(items) - 50))
            fh.write("\n")


def run_ingest(path: str, out_dir: str, *, sheet: str | None = None) -> IngestResult:
    """파일 정형화 + 3종 산출 기록. out_dir 에 tidy.csv/schema.json/smell_report.md."""
    os.makedirs(out_dir, exist_ok=True)
    res = ingest_workbook(path, sheet=sheet)
    write_tidy_csv(res.tidy_rows, os.path.join(out_dir, "tidy.csv"))
    write_schema_json(res.schema, os.path.join(out_dir, "schema.json"))
    write_smell_report(res.smells, os.path.join(out_dir, "smell_report.md"))
    return res


__all__ = ["IngestResult", "ingest_workbook", "run_ingest",
           "write_tidy_csv", "write_schema_json", "write_smell_report"]
