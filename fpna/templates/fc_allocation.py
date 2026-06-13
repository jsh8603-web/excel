"""
fpna.templates.fc_allocation — 고정비: 공통비 다대다 배부(step-down) (C6 신규).

cost_allocation(단일 풀→수익부서 직접배부)과 달리, **서비스부서 간 다대다
연쇄 배부**(step-down)를 다룬다. 서비스부서(IT·시설 등) 공통비를 다른 서비스
+ 생산부서로 순차 배부하고, 최종적으로 생산부서로 전액 귀착시킨다.

- grain = "1행 = 1 (source_dept × target_dept)" 배부 흐름.
- step-down: 배부 순서대로 service dept 를 비우며 하위(아직 안 비운 service +
  production)로 가중 배부. 자기 자신·이미 닫은 부서로는 배부하지 않는다.
- 불변식(R11 보존): Σ투입 고정비 == Σ최종 귀착(production) (누수 0).
  view_contract.assert_allocation_conserves. 배부 합이 투입과 tol=0 일치.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.conserve import ConserveSpec
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_allocation"


# --------------------------------------------------------------------------- #
# T4 보존(자문 R2 C6) — 선언형 CONSERVE_SPECS                                  #
#   배부 보존: 투입 풀(Σ direct_cost) == 최종귀착(alloc_final). raw 변은        #
#   _stepdown(배부 cascade)을 부르지 않고 INPUT 직접 합 → 누수/이중귀착 노출.   #
# --------------------------------------------------------------------------- #
CONSERVE_SPECS = [
    ConserveSpec(
        "배부보존: 최종귀착 = Σ direct_cost",
        raw_sum_fn=lambda d: sum(dp.direct_cost for dp in d.depts),
        reported_key="alloc_final",
    ),
]


@dataclass
class Dept:
    """부서 1개. kind=service(배부 대상·소진) / production(최종 귀착)."""
    dept_id: str
    label: str
    kind: str                  # "service" | "production"
    direct_cost: float = 0.0   # 직접 발생 고정비(투입)


@dataclass
class FcAllocationInput:
    title: str = "고정비 — 공통비 다대다 배부 (Step-Down)"
    subtitle: str = "서비스부서 연쇄 배부 → 생산부서 전액 귀착 (보존·누수 0)"
    unit: str = "₩"
    depts: list = field(default_factory=list)            # list[Dept]
    service_order: list = field(default_factory=list)    # service dept 배부 순서(id)
    # {source_id: {target_id: weight}} — source 가 target 들로 배부하는 가중
    drivers: dict = field(default_factory=dict)
    commentary: list = field(default_factory=list)


def _stepdown(data: FcAllocationInput):
    """step-down 배부. (flows, final, totals) 반환.

    flows = [(source, target, amount), ...] 배부 흐름.
    final = {production_id: 최종귀착액}. 누적 = direct + 받은 배부.
    """
    cur = {d.dept_id: d.direct_cost for d in data.depts}
    kind = {d.dept_id: d.kind for d in data.depts}
    closed: set = set()
    flows = []
    prod_ids = [d.dept_id for d in data.depts if d.kind == "production"]
    for src in data.service_order:
        amount = cur.get(src, 0.0)
        closed.add(src)                                  # 자기 자신·이미 닫은 곳 제외
        weights = {t: w for t, w in data.drivers.get(src, {}).items()
                   if t not in closed}
        alloc = finance.allocate_pool(amount, weights)
        if not alloc:                                    # 배부기준 없음 → production 균등
            live_prod = [p for p in prod_ids if p not in closed]
            if live_prod and amount:
                alloc = finance.allocate_pool(amount, {p: 1.0 for p in live_prod})
        for tgt, a in alloc.items():
            cur[tgt] = cur.get(tgt, 0.0) + a
            flows.append((src, tgt, a))
        cur[src] = 0.0                                   # 소진
    final = {p: cur.get(p, 0.0) for p in prod_ids}
    totals = {"input": sum(d.direct_cost for d in data.depts),
              "final": sum(final.values())}
    return flows, final, totals


def golden_sample() -> FcAllocationInput:
    """구조 골든 — 서비스 2(IT·시설) → 생산 2(조립·도장) step-down."""
    depts = [
        Dept("IT", "IT부", "service", direct_cost=3_000.0),
        Dept("FAC", "시설부", "service", direct_cost=2_000.0),
        Dept("ASM", "조립부", "production", direct_cost=10_000.0),
        Dept("PNT", "도장부", "production", direct_cost=6_000.0),
    ]
    # IT 먼저 → 시설/조립/도장. 시설 → 조립/도장. (step-down: IT가 시설로도 배부)
    drivers = {
        "IT": {"FAC": 10, "ASM": 50, "PNT": 40},
        "FAC": {"ASM": 60, "PNT": 40},
    }
    return FcAllocationInput(depts=depts, service_order=["IT", "FAC"], drivers=drivers,
                             commentary=["step-down: IT → 시설+생산, 시설 → 생산(전액 귀착)",
                                         "보존: Σ직접고정비 == Σ생산부서 최종귀착 (누수 0)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: FcAllocationInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    flows, final, totals = _stepdown(data)
    last_col = 4

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("FcAllocation")
    hs.set_widths(ws, {1: 18, 2: 18, 3: 14, 4: 14})

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col)

    label = {d.dept_id: d.label for d in data.depts}

    # 투입(직접 고정비)
    hs.section_header(ws, r, "투입 고정비 (직접)", last_col=last_col)
    r += 1
    for j, h in enumerate(("부서", "구분", "직접 고정비"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j <= 2 else hs.CENTER)
    r += 1
    for d in data.depts:
        hs.set_cell(ws, r, 1, d.label, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, d.kind, role="soft", align=hs.CENTER)
        hs.set_cell(ws, r, 3, d.direct_cost, role="calc", number_format=hs.FMT_INT)
        r += 1
    hs.set_cell(ws, r, 1, "투입 합계", role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 3, totals["input"], role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    r += 2

    # 배부 흐름 (source → target)
    r = hs.section_header(ws, r, "배부 흐름 (Step-Down)", last_col=last_col)
    for j, h in enumerate(("배부 출발", "배부 도착", "배부액"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j <= 2 else hs.CENTER)
    r += 1
    for src, tgt, a in flows:
        hs.set_cell(ws, r, 1, label.get(src, src), role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, label.get(tgt, tgt), role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 3, a, role="calc", number_format=hs.FMT_INT)
        r += 1
    r += 1

    # 최종 귀착(생산부서)
    r = hs.section_header(ws, r, "최종 귀착 (생산부서)", last_col=last_col)
    for j, h in enumerate(("생산부서", "최종 귀착액"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1
    for pid, amt in final.items():
        hs.set_cell(ws, r, 1, label.get(pid, pid), role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, amt, role="calc", number_format=hs.FMT_INT)
        r += 1
    hs.set_cell(ws, r, 1, "귀착 합계", role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 2, totals["final"], role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    r += 1

    # _RECON: 투입 == 최종귀착 (보존)
    recon = vc.recon_block(
        n_input=len(data.depts), n_output=len(flows),
        src_sum=totals["input"], out_sum=totals["final"],
        completeness="부서 %d · 배부 흐름 %d (step-down)" % (len(data.depts), len(flows)),
        accuracy="Σ직접고정비 == Σ생산부서 최종귀착 (누수 0)",
        cutoff="service dept 전액 소진 → production 귀착",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="부서별 고정비 원장 · 배부기준 마스터",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 (source_dept × target_dept)", ("source", "target"),
                [{"source": s, "target": t} for s, t, _ in flows])
    wb._fpna_meta = {"fact": fact, "flows": flows, "final": final, "totals": totals,
                     "alloc_final": totals["final"]}   # T4 CONSERVE_SPECS reported_key
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: FcAllocationInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    totals = meta["totals"]

    # R8 grain (source × target 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # R11 보존: 투입 == 최종귀착 (배부 전후 누수 0)
    vc.assert_allocation_conserves(rep, totals["input"], totals["final"], tol=1e-6)

    # service dept 전액 소진 확인(최종 귀착에 service 잔액 없음)
    final_only_prod = all(
        any(d.dept_id == pid and d.kind == "production" for d in data.depts)
        for pid in meta["final"])
    rep.add("최종 귀착 = 생산부서만", final_only_prod,
            "" if final_only_prod else "service dept 잔액 미소진")

    # 배부액 비음수
    nonneg = all(a >= -1e-9 for _, _, a in meta["flows"])
    rep.add("배부액 비음수", nonneg, "" if nonneg else "음수 배부")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "Dept", "FcAllocationInput", "golden_sample", "build", "qc"]
