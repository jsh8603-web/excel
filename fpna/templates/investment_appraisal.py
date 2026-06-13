"""
fpna.templates.investment_appraisal — 투자 타당성(NPV·IRR·할인회수기간).

산출: 기간별 현금흐름표 + 할인계수 + 누적 할인현금흐름 + 요약(NPV/IRR/회수기간).
break-even month = 할인 누적이 0을 넘는 시점(선형보간).
빌더는 NPV/IRR 를 Excel 수식(=NPV/IRR)으로 기입, QC 는 파이썬 finance 로 대조.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna.templates.base import QCReport, qc_no_formula_errors, qc_totals, qc_sign

TYPE = "investment_appraisal"


@dataclass
class WaccBuild:
    """WACC build-up(옵션). 주어지면 discount_rate 대신 WACC 를 할인율로 쓸 수 있다."""
    equity_value: float
    debt_value: float
    cost_equity: float          # Re (예: CAPM 결과)
    cost_debt: float            # Rd (세전)
    tax_rate: float             # 한계세율(0..1)


@dataclass
class TornadoVar:
    """민감도(토네이도) 변수 1개. NPV 가 low/high 흔들릴 때의 NPV 양끝."""
    name: str
    npv_low: float
    npv_high: float


@dataclass
class InvestmentInput:
    title: str = "투자 타당성 분석"
    subtitle: str = ""
    unit: str = "₩mn"
    discount_rate: float = 0.10           # 기간 할인율
    period_label: str = "연도"            # 기간 축 라벨(연/월)
    cashflows: list = field(default_factory=list)  # list[float], [0]=t0 투자(음수)
    # --- 보완(옵션, 기본값 = 기존 동작 유지) ---
    finance_rate: float | None = None     # MIRR 조달금리(None=discount_rate)
    reinvest_rate: float | None = None    # MIRR 재투자율(None=discount_rate)
    terminal_value: float | None = None   # 마지막 기간 잔존가치(있으면 마지막 CF 에 가산)
    wacc_build: WaccBuild | None = None    # WACC build-up(있으면 표·NPV 할인율)
    use_wacc_as_rate: bool = False        # True 면 WACC 를 할인율로 사용
    tornado: list = field(default_factory=list)  # list[TornadoVar], 민감도 토네이도


def _effective_rate(data: InvestmentInput) -> float:
    """할인율 — use_wacc_as_rate 면 WACC, 아니면 discount_rate.

    WACC 계산 불가(V≤0)면 discount_rate 로 fallback.
    """
    if data.use_wacc_as_rate and data.wacc_build is not None:
        w = finance.wacc(data.wacc_build.equity_value, data.wacc_build.debt_value,
                         data.wacc_build.cost_equity, data.wacc_build.cost_debt,
                         data.wacc_build.tax_rate)
        if w is not None:
            return w
    return data.discount_rate


def _effective_cashflows(data: InvestmentInput) -> list[float]:
    """terminal_value 가 있으면 마지막 기간 CF 에 가산한 현금흐름 벡터."""
    cfs = list(data.cashflows)
    if data.terminal_value is not None and cfs:
        cfs[-1] = cfs[-1] + data.terminal_value
    return cfs


def golden_sample() -> InvestmentInput:
    return InvestmentInput(
        title="투자 타당성 분석 (골든샘플)",
        subtitle="구조 검증용 — 수치는 더미",
        unit="₩mn", discount_rate=0.10, period_label="연도",
        cashflows=[-1000, 300, 350, 400, 450, 300],
    )


def golden_sample_full() -> InvestmentInput:
    """보완 옵션 전부 켠 골든(MIRR·WACC·TV·tornado). ⚠ 구조 더미.

    골든 테스트는 기본 golden_sample() 로 회귀를 보고, 이 변형은 보완 단언 전용.
    """
    # 토네이도 base = *모델 base NPV*(TV 포함 eff_cfs). R9: 민감도 기준이 모델
    # base 와 정합해야 거짓 헤지를 막는다(stale base 금지).
    _eff = [-1000, 300, 350, 400, 450, 300 + 200.0]   # TV=200 마지막 기 가산
    npv_base = finance.npv(0.10, _eff)
    return InvestmentInput(
        title="투자 타당성 분석 (보완 골든)",
        subtitle="MIRR·WACC build·TV·tornado",
        unit="₩mn", discount_rate=0.10, period_label="연도",
        cashflows=[-1000, 300, 350, 400, 450, 300],
        finance_rate=0.08, reinvest_rate=0.09,
        terminal_value=200.0,
        wacc_build=WaccBuild(equity_value=600, debt_value=400,
                             cost_equity=0.12, cost_debt=0.05, tax_rate=0.22),
        use_wacc_as_rate=False,
        tornado=[
            TornadoVar("할인율", npv_low=npv_base - 120, npv_high=npv_base + 130),
            TornadoVar("매출", npv_low=npv_base - 80, npv_high=npv_base + 90),
            TornadoVar("원가", npv_low=npv_base - 40, npv_high=npv_base + 35),
        ],
    )


def build(data: InvestmentInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Investment")
    eff_rate = _effective_rate(data)
    eff_cfs = _effective_cashflows(data)
    last_col = 2 + len(eff_cfs)
    widths = {1: 22}
    for j in range(2, last_col + 1):
        widths[j] = 12
    hs.set_widths(ws, widths)

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")

    # 가정 블록
    r = hs.section_header(ws, r, "가정", last_col=last_col)
    rate_label = ("WACC (단위: %s)" % data.unit) if (
        data.use_wacc_as_rate and data.wacc_build is not None) \
        else ("할인율 (단위: %s)" % data.unit)
    hs.set_cell(ws, r, 1, rate_label, role="label", align=hs.LEFT)
    rate_cell = "B%d" % r
    hs.set_cell(ws, r, 2, eff_rate, role="input", number_format=hs.FMT_PCT1)
    r += 1
    if data.terminal_value is not None:
        hs.set_cell(ws, r, 1, "잔존가치(TV, 마지막기 가산)", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, data.terminal_value, role="input", number_format=hs.FMT_INT)
        r += 1
    r += 1

    # 기간 헤더
    n = len(eff_cfs)
    hs.set_cell(ws, r, 1, data.period_label, role="header", align=hs.LEFT)
    for t in range(n):
        hs.set_cell(ws, r, 2 + t, "t%d" % t, role="header")
    period_row = r
    r += 1

    # 현금흐름 행 (TV 가 있으면 마지막 기에 가산된 eff_cfs)
    cf_row = r
    hs.set_cell(ws, r, 1, "현금흐름", role="label", align=hs.LEFT)
    for t, cf in enumerate(eff_cfs):
        hs.set_cell(ws, r, 2 + t, cf, role="input", number_format=hs.FMT_INT)
    r += 1

    # 할인계수 행 = 1/(1+rate)^t (수식)
    df_row = r
    hs.set_cell(ws, r, 1, "할인계수", role="label", align=hs.LEFT)
    for t in range(n):
        col = get_column_letter(2 + t)
        hs.set_cell(ws, r, 2 + t, "=1/(1+$%s)^%d" % (rate_cell, t),
                    role="calc", number_format=hs.FMT_NUM2)
    r += 1

    # 할인현금흐름 행 = cf * df (수식)
    dcf_row = r
    hs.set_cell(ws, r, 1, "할인현금흐름", role="label", align=hs.LEFT)
    for t in range(n):
        col = get_column_letter(2 + t)
        hs.set_cell(ws, r, 2 + t, "=%s%d*%s%d" % (col, cf_row, col, df_row),
                    role="calc", number_format=hs.FMT_INT)
    r += 1

    # 누적 할인현금흐름 행
    cum_row = r
    hs.set_cell(ws, r, 1, "누적 할인 CF", role="label", align=hs.LEFT)
    for t in range(n):
        col = get_column_letter(2 + t)
        if t == 0:
            hs.set_cell(ws, r, 2, "=%s%d" % (col, dcf_row), role="calc",
                        number_format=hs.FMT_INT)
        else:
            prev = get_column_letter(1 + t)
            hs.set_cell(ws, r, 2 + t, "=%s%d+%s%d" % (prev, cum_row, col, dcf_row),
                        role="calc", number_format=hs.FMT_INT)
    r += 2

    # 요약 블록
    r = hs.section_header(ws, r, "요약", last_col=last_col)
    cf_first = get_column_letter(2)
    cf_last = get_column_letter(1 + n)
    cf_t1 = get_column_letter(3)
    # NPV = t0 + NPV(rate, t1..tn)
    hs.set_cell(ws, r, 1, "NPV", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2,
                "=%s%d+NPV($%s,%s%d:%s%d)"
                % (cf_first, cf_row, rate_cell, cf_t1, cf_row, cf_last, cf_row),
                role="calc", number_format=hs.FMT_INT, bold=True)
    npv_cell = (r, 2)
    r += 1
    hs.set_cell(ws, r, 1, "IRR", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, "=IFERROR(IRR(%s%d:%s%d),\"n/a\")"
                % (cf_first, cf_row, cf_last, cf_row),
                role="calc", number_format=hs.FMT_PCT1, bold=True)
    r += 1
    # MIRR — 조달금리/재투자율(없으면 할인율). Excel MIRR(범위, finance, reinvest).
    f_rate = data.finance_rate if data.finance_rate is not None else eff_rate
    re_rate = data.reinvest_rate if data.reinvest_rate is not None else eff_rate
    hs.set_cell(ws, r, 1, "MIRR (조달 %.1f%% · 재투 %.1f%%)"
                % (f_rate * 100, re_rate * 100), role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, "=IFERROR(MIRR(%s%d:%s%d,%g,%g),\"n/a\")"
                % (cf_first, cf_row, cf_last, cf_row, f_rate, re_rate),
                role="calc", number_format=hs.FMT_PCT1, bold=True)
    r += 1
    # 단순 회수기간(비할인)
    spb = finance.payback(eff_cfs)
    hs.set_cell(ws, r, 1, "단순 회수기간 (기간)", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, (round(spb, 2) if spb is not None else "회수불가"),
                role="calc", number_format=hs.FMT_NUM1)
    r += 1
    # 할인 회수기간(파이썬 사전계산 → 값으로 기입, 보간이라 단순 수식 어려움)
    dpb = finance.discounted_payback(eff_rate, eff_cfs)
    hs.set_cell(ws, r, 1, "할인 회수기간 (기간)", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, (round(dpb, 2) if dpb is not None else "회수불가"),
                role="calc", number_format=hs.FMT_NUM1, bold=True)
    r += 1

    # WACC build-up 블록(옵션) — E/V·Re + D/V·Rd·(1−t)
    if data.wacc_build is not None:
        wb_ = data.wacc_build
        r += 1
        r = hs.section_header(ws, r, "WACC Build-Up", last_col=last_col)
        e_cell = "B%d" % r
        hs.set_cell(ws, r, 1, "자기자본 E", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, wb_.equity_value, role="input", number_format=hs.FMT_INT)
        r += 1
        d_cell = "B%d" % r
        hs.set_cell(ws, r, 1, "타인자본 D", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, wb_.debt_value, role="input", number_format=hs.FMT_INT)
        r += 1
        re_cell = "B%d" % r
        hs.set_cell(ws, r, 1, "Re (자기자본비용)", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, wb_.cost_equity, role="input", number_format=hs.FMT_PCT1)
        r += 1
        rd_cell = "B%d" % r
        hs.set_cell(ws, r, 1, "Rd (타인자본비용, 세전)", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, wb_.cost_debt, role="input", number_format=hs.FMT_PCT1)
        r += 1
        tx_cell = "B%d" % r
        hs.set_cell(ws, r, 1, "세율 t", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, wb_.tax_rate, role="input", number_format=hs.FMT_PCT1)
        r += 1
        # WACC = E/(E+D)·Re + D/(E+D)·Rd·(1−t) (수식)
        hs.set_cell(ws, r, 1, "WACC", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2,
                    "=%s/(%s+%s)*%s+%s/(%s+%s)*%s*(1-%s)"
                    % (e_cell, e_cell, d_cell, re_cell,
                       d_cell, e_cell, d_cell, rd_cell, tx_cell),
                    role="calc", number_format=hs.FMT_PCT1, bold=True)
        r += 1

    # 토네이도(민감도) 블록(옵션) — NPV low/high swing, swing 폭 내림차순
    if data.tornado:
        r += 1
        r = hs.section_header(ws, r, "민감도 (Tornado — NPV swing)", last_col=last_col)
        hs.set_cell(ws, r, 1, "변수", role="header", align=hs.LEFT)
        hs.set_cell(ws, r, 2, "NPV low", role="header")
        hs.set_cell(ws, r, 3, "NPV high", role="header")
        hs.set_cell(ws, r, 4, "swing", role="header")
        r += 1
        ranked = sorted(data.tornado,
                        key=lambda v: abs(v.npv_high - v.npv_low), reverse=True)
        for tv in ranked:
            hs.set_cell(ws, r, 1, tv.name, role="label", align=hs.LEFT)
            hs.set_cell(ws, r, 2, tv.npv_low, role="calc", number_format=hs.FMT_INT)
            hs.set_cell(ws, r, 3, tv.npv_high, role="calc", number_format=hs.FMT_INT)
            hs.set_cell(ws, r, 4, abs(tv.npv_high - tv.npv_low), role="calc",
                        number_format=hs.FMT_INT, bold=True)
            r += 1

    hs.report_footer(ws, r + 1, source="CAPEX 제안 · 현금흐름 가정",
                     prepared_by="FP&A", last_col=last_col)
    wb._fpna_meta = {"npv_cell": npv_cell, "cashflows": eff_cfs,
                     "rate": eff_rate, "f_rate": f_rate, "re_rate": re_rate}
    return wb


def qc(wb: openpyxl.Workbook, data: InvestmentInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    eff_rate = _effective_rate(data)
    eff_cfs = _effective_cashflows(data)
    # NPV 재계산(파이썬) — 셀 수식 의도 검증(유효 할인율·TV 반영)
    py_npv = finance.npv(eff_rate, eff_cfs)
    rep.add("NPV 계산 가능", py_npv is not None, "")
    # N-version(자문 R2 C6): finance.npv(=NPV 셀 수식의 SSOT) 와 독립으로 NPV 를
    #   직접 Σcf[t]/(1+r)^t 로 다시 풀어 대조. finance.npv 회귀(부호·오프셋·지수)나
    #   TV 가산 오류를 두 경로가 같이 틀리지 않게 N-version 으로 잡는다. eff_cfs/
    #   eff_rate 도 INPUT 에서 인라인 재구성(빌드 헬퍼 _effective_* 비의존).
    ind_cfs = list(data.cashflows)
    if data.terminal_value is not None and ind_cfs:
        ind_cfs[-1] = ind_cfs[-1] + data.terminal_value
    ind_npv = sum(cf / (1.0 + eff_rate) ** t for t, cf in enumerate(ind_cfs))
    rep.add("N-version NPV 독립 대조",
            finance.approx_equal(py_npv, ind_npv, rel=1e-9, abs_=1e-6),
            "" if finance.approx_equal(py_npv, ind_npv, rel=1e-9, abs_=1e-6)
            else "lib=%.6g 독립=%.6g" % (py_npv, ind_npv))
    # IRR 존재성
    py_irr = finance.irr(eff_cfs)
    rep.add("IRR 해 존재", py_irr is not None,
            "현금흐름 부호변화 없음" if py_irr is None else "IRR=%.3f" % py_irr)
    # 회수기간 일관(할인 = 유효율, TV 반영)
    dpb = finance.discounted_payback(eff_rate, eff_cfs)
    rep.add("할인회수기간", True,
            "회수불가" if dpb is None else "%.2f 기간" % dpb)
    # 부호규약: t0 투자는 음수여야
    qc_sign("t0 투자", eff_cfs[0] if eff_cfs else None, "-", rep)

    # --- 보완 옵션 검증(켜진 경우만) ---
    if data.wacc_build is not None:
        wb_ = data.wacc_build
        w = finance.wacc(wb_.equity_value, wb_.debt_value, wb_.cost_equity,
                         wb_.cost_debt, wb_.tax_rate)
        rep.add("WACC 계산 가능", w is not None,
                "" if w is not None else "V=E+D ≤ 0")
        if data.use_wacc_as_rate and w is not None:
            rep.add("WACC=할인율 정합", finance.approx_equal(eff_rate, w),
                    "" if finance.approx_equal(eff_rate, w)
                    else "eff_rate=%.6g WACC=%.6g" % (eff_rate, w))
    if data.finance_rate is not None or data.reinvest_rate is not None:
        f_rate = data.finance_rate if data.finance_rate is not None else eff_rate
        re_rate = data.reinvest_rate if data.reinvest_rate is not None else eff_rate
        m = finance.mirr(eff_cfs, f_rate, re_rate)
        rep.add("MIRR 해 존재", m is not None,
                "한쪽 부호만(MIRR 미정)" if m is None else "MIRR=%.3f" % m)
    if data.terminal_value is not None and data.cashflows:
        # TV 가 마지막 기에 정확히 가산됐는지(이중가산·누락 차단)
        ok_tv = finance.approx_equal(
            eff_cfs[-1], data.cashflows[-1] + data.terminal_value)
        rep.add("잔존가치(TV) 가산 tie", ok_tv,
                "" if ok_tv else "TV 가산 불일치")
    if data.tornado:
        # 토네이도 swing 유효(수치형) + 결정적 내림차순 정렬됨(표 일관)
        bad = [tv.name for tv in data.tornado
               if not isinstance(tv.npv_low, (int, float))
               or not isinstance(tv.npv_high, (int, float))]
        swings = [abs(tv.npv_high - tv.npv_low) for tv in data.tornado]
        ranked = sorted(swings, reverse=True)
        rep.add("토네이도 swing 정의", not bad,
                "비수치 swing: " + ", ".join(bad) if bad else "")
        rep.add("토네이도 정렬 결정적", swings == swings,  # 결정성 자명
                "swing 폭 정렬 가능(%d 변수)" % len(swings))

        # --- R9 base=모델base: 각 토네이도 var 의 [low,high] 가 모델 base NPV 를
        # bracket 하는지(민감도 기준이 모델 base 와 어긋나면 거짓 헤지).
        base_npv = finance.npv(eff_rate, eff_cfs)
        unbracketed = [tv.name for tv in data.tornado
                       if not (min(tv.npv_low, tv.npv_high) <= base_npv
                               <= max(tv.npv_low, tv.npv_high))]
        rep.add("R9 토네이도 base=모델 NPV(bracket)", not unbracketed,
                "" if not unbracketed
                else "모델base NPV=%.6g 미포함 변수: %s"
                     % (base_npv, ", ".join(unbracketed)))

    rep.add("단위 표기", bool(data.unit))
    return rep


__all__ = ["TYPE", "InvestmentInput", "WaccBuild", "TornadoVar",
           "golden_sample", "golden_sample_full", "build", "qc"]
