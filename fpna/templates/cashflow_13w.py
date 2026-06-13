"""fpna.templates.cashflow_13w — 13주 단기 현금흐름/유동성."""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401
import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna.templates.base import QCReport, qc_no_formula_errors
from fpna.view_contract import assert_tie_out

TYPE = "cashflow_13w"


@dataclass
class CashInput:
    title: str = "13주 현금흐름"
    subtitle: str = ""
    unit: str = "₩mn"
    opening: float = 0.0
    weeks: int = 13
    inflows: list = field(default_factory=list)    # 주별 유입
    outflows: list = field(default_factory=list)   # 주별 유출
    # --- 주간 연속성(선택) ------------------------------------------------
    # 주별 명시 기초잔액. 주면 각 주 기초 == 직전주 기말 인지 검증(연속성 tie).
    # 비우면 build 와 동일하게 opening 에서 순현금 누적으로 파생(자기 tie).
    openings: list = field(default_factory=list)   # list[float], 길이 = weeks

    def week_balances(self):
        """(openings, closings) 주별 잔액 체인. openings 명시 없으면 연속 파생."""
        opens, closes = [], []
        bal = self.opening
        for w in range(self.weeks):
            ob = self.openings[w] if (self.openings and w < len(self.openings)) else bal
            cb = ob + self.inflows[w] - self.outflows[w]
            opens.append(ob)
            closes.append(cb)
            bal = cb
        return opens, closes


def golden_sample() -> CashInput:
    return CashInput(
        title="13주 현금흐름 (골든샘플)", subtitle="구조 검증용 — 더미", unit="₩mn",
        opening=500, weeks=13,
        inflows=[120] * 13, outflows=[100, 110, 130, 90, 100, 140, 95, 100, 120, 110, 100, 130, 90],
    )


def build(data: CashInput, *, mode="create", base_path=None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = hs.safe_sheet_title("Cash13W")
    n = data.weeks; last_col = 1 + n
    hs.style_sheet(ws, freeze="B6")
    hs.set_widths(ws, {1: 18, **{j: 9 for j in range(2, last_col + 1)}})
    r = hs.title_block(ws, data.title, data.subtitle, last_col=last_col)

    hs.set_cell(ws, r, 1, "항목 (단위: %s)" % data.unit, role="header", align=hs.LEFT)
    for w in range(n):
        hs.set_cell(ws, r, 2 + w, "W%d" % (w + 1), role="header")
    r += 1
    in_row = r
    hs.set_cell(ws, r, 1, "유입", role="label", align=hs.LEFT)
    for w in range(n):
        hs.set_cell(ws, r, 2 + w, data.inflows[w], role="input", number_format=hs.FMT_INT)
    r += 1
    out_row = r
    hs.set_cell(ws, r, 1, "유출", role="label", align=hs.LEFT)
    for w in range(n):
        hs.set_cell(ws, r, 2 + w, data.outflows[w], role="input", number_format=hs.FMT_INT)
    r += 1
    net_row = r
    hs.set_cell(ws, r, 1, "순현금", role="label", align=hs.LEFT)
    for w in range(n):
        col = get_column_letter(2 + w)
        hs.set_cell(ws, r, 2 + w, "=%s%d-%s%d" % (col, in_row, col, out_row),
                    role="calc", number_format=hs.FMT_INT)
    r += 1
    bal_row = r
    hs.set_cell(ws, r, 1, "기말 잔액", role="label", align=hs.LEFT, bold=True)
    for w in range(n):
        col = get_column_letter(2 + w)
        if w == 0:
            hs.set_cell(ws, r, 2, "=%g+%s%d" % (data.opening, col, net_row),
                        role="calc", number_format=hs.FMT_INT, bold=True)
        else:
            prev = get_column_letter(1 + w)
            hs.set_cell(ws, r, 2 + w, "=%s%d+%s%d" % (prev, bal_row, col, net_row),
                        role="calc", number_format=hs.FMT_INT, bold=True)
    r += 2
    hs.add_line_chart(ws, anchor="A%d" % r, data_min_col=2, data_max_col=last_col,
                      data_min_row=bal_row, data_max_row=bal_row, cat_col=1,
                      title="주간 잔액(유동성)")
    return wb


def qc(wb, data: CashInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    ok_len = len(data.inflows) == data.weeks == len(data.outflows)
    rep.add("주 수 일치", ok_len, "" if ok_len else "inflow/outflow 길이≠weeks")

    # --- R1 13주 전수성: 주차가 W1..W{weeks} 연속(결측·중복 없음) ----------
    # 주를 건너뛰면(행 누락) 유동성 갭을 못 본다. weeks==13 + 데이터 길이 정합.
    full_13w = (data.weeks == 13) and ok_len
    rep.add("R1 13주 전수성", full_13w,
            "" if full_13w else "13주 미충족 또는 inflow/outflow 길이 불일치(weeks=%d)"
            % data.weeks)

    # --- 주간 연속성 tie-out: 각 주 기초 == 직전주 기말 ---------------------
    # 연속성이 깨지면(누군가 기초를 하드코딩/덮어씀) 잔액 체인이 거짓말 →
    # 유동성 오판. opening 명시 모드에서 강제로 잡는다.
    opens, closes = data.week_balances()
    cont_break = []
    for w in range(1, data.weeks):
        if abs(opens[w] - closes[w - 1]) > 1e-9:
            cont_break.append("W%d(기초=%.0f≠전주기말=%.0f)"
                              % (w + 1, opens[w], closes[w - 1]))
    rep.add("주간 연속성(기말==익주기초)", not cont_break,
            "" if not cont_break else "연속성 단절: " + ", ".join(cont_break[:6]))
    # 첫 주 기초 == opening tie
    if opens:
        assert_tie_out(rep, opens[0], data.opening, tol=1e-9,
                       name="첫 주 기초 == opening")

    # 파이썬 잔액 시뮬 → 최소잔액 음수 경고 (연속 체인 기준)
    min_bal = min([data.opening] + closes) if closes else data.opening
    rep.add("유동성(최소잔액≥0)", min_bal >= 0, "최소잔액=%.0f" % min_bal)
    rep.add("단위 표기", bool(data.unit))
    return rep


__all__ = ["TYPE", "CashInput", "golden_sample", "build", "qc"]
