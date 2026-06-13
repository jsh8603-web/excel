"""
fpna.templates.fc_depreciation_schedule — 고정비: 자산별 감가상각 스케줄.

자산 × 기간 **전수** 매트릭스(결측 기간도 행/열 유지, R1)로 월 상각을 펼치고,
스케줄 합을 GL 상각비와 대사(R11)한다. 차이는 일회성/조정 노출 신호.

- grain = "1행 = 1 자산 × 1 기간" (R8).
- 표시 = wide(자산 행 × 기간 열), 내부 = tidy Fact.
- 상각액은 fpna.finance.depreciation_schedule(정액법) 재계산.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Asset, Fact, DOMAIN_LABEL
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_depreciation_schedule"


@dataclass
class FixedCostDeprInput:
    title: str = "고정비 — 감가상각 스케줄"
    subtitle: str = "자산별 월 상각 (정액법) · GL 대사"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)          # (fy, period) inclusive
    end: tuple = (2024, 12)
    assets: list = field(default_factory=list)          # list[dims.Asset]
    gl_dep_by_period: dict = field(default_factory=dict)  # {"_total": GL 상각비합} 대사용
    commentary: list = field(default_factory=list)

    # --- C11 보완 (dims.Asset 은 frozen·타작업자 소유 → 자산번호 키 평행맵으로 부여) ---
    # 부분월 proration: {asset_no: 0<factor<=1}  (1차월 일할/반월). 미지정=1.0(전월).
    first_period_factor: dict = field(default_factory=dict)
    # 자산처분: {asset_no: (fy, period)} 처분월부터 D&A 중단·장부가 고정.
    disposals: dict = field(default_factory=dict)
    # C11 처분손익: {asset_no: 처분가}. 처분손익 = 처분가 − NBV(처분시점 장부가).
    #   양(+)=처분이익 / 음(−)=처분손실. disposals 와 짝(처분월 없으면 무시).
    disposal_prices: dict = field(default_factory=dict)
    # 손상(impairment): {asset_no: ((fy, period), carrying_after)} base-reset 이벤트.
    impairments: dict = field(default_factory=dict)
    # R12 활성 window: {asset_no: ((start_fy,p), (end_fy,p) | None)} — 없으면
    #   in_service ~ 표시끝(또는 disposal 직전)으로 파생. None=evergreen(표시끝까지).
    active_window: dict = field(default_factory=dict)
    # 활성 window 내인데 GL 계상이 누락된 (asset, period) 강제 주입(테스트/실데이터 결측 표현).
    #   {asset_no: [period_label, ...]} — 해당 셀을 NO_DATA 로 비우고 MISSING_ACCRUAL emit.
    missing_accruals: dict = field(default_factory=dict)


# --------------------------------------------------------------------------- #
# 공유 헬퍼 (build/qc/golden 공통)                                            #
# --------------------------------------------------------------------------- #
def _ordinal(fyp: tuple) -> int:
    return fyp[0] * 12 + (fyp[1] - 1)


def _active_window_ords(inp: FixedCostDeprInput, a, first_ord: int, last_ord: int):
    """자산 a 의 R12 활성 window 를 (start_ord, end_ord) ordinal 로 반환.

    명시 active_window 우선. 없으면 in_service ~ (disposal 직전 또는 표시끝).
    evergreen(end=None) = 표시끝까지.
    """
    win = inp.active_window.get(a.asset_no)
    if win is not None:
        ws, we = win
        s = _ordinal(ws)
        e = last_ord if we is None else _ordinal(we)
        return max(s, first_ord), min(e, last_ord)
    s = _ordinal(a.in_service) if a.in_service else first_ord
    disp = inp.disposals.get(a.asset_no)
    e = (_ordinal(disp) - 1) if disp is not None else last_ord
    return max(s, first_ord), min(e, last_ord)


def _build_fact(inp: FixedCostDeprInput):
    """캘린더·기간·tidy Fact 를 (cal, periods, fact) 로 반환. 자산×기간 전수.

    C11: 부분월 proration / 처분(D&A 중단) / 손상(base-reset) 을 확장 스케줄로 반영.
    missing_accruals 로 지정된 (asset, period) 셀은 dep=None(NO_DATA) 로 비운다(R12 결측).
    """
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    periods = cal.periods(inp.start, inp.end)
    first_ord, last_ord = periods[0].ordinal, periods[-1].ordinal
    rows: list[dict] = []
    for a in inp.assets:
        in_ord = _ordinal(a.in_service) if a.in_service else first_ord
        span = last_ord - in_ord + 1
        n = max(span, a.life_months, 0)
        # 처분/손상 이벤트의 0-base 인덱스(가동 시점 in_ord 기준)
        disp = inp.disposals.get(a.asset_no)
        disp_idx = (_ordinal(disp) - in_ord) if disp is not None else None
        imp = inp.impairments.get(a.asset_no)
        imp_idx = imp_to = None
        if imp is not None:
            (imp_fyp, imp_to) = imp
            imp_idx = _ordinal(imp_fyp) - in_ord
        sch = finance.depreciation_schedule_ext(
            a.acq_cost, a.salvage, a.life_months, n_periods=n, start_index=0,
            first_period_factor=inp.first_period_factor.get(a.asset_no, 1.0),
            disposal_index=disp_idx, impair_index=imp_idx, impair_to=imp_to)
        miss = set(inp.missing_accruals.get(a.asset_no, ()))
        for p in periods:
            offset = p.ordinal - in_ord
            if offset < 0:                       # 가동 전 — 미계상(전수 유지, dep=0)
                opening, dep, closing = a.acq_cost, 0.0, a.acq_cost
            elif offset < len(sch):
                opening, dep, closing = sch[offset]
            else:                                # 내용연수 종료 후
                opening, dep, closing = a.salvage, 0.0, a.salvage
            if p.label in miss:                  # GL 결측(R12): 값 비움 → MISSING_ACCRUAL
                dep = None
            rows.append({
                "asset_no": a.asset_no, "period": p.label, "domain": a.domain,
                "opening": opening, "dep": dep, "closing": closing,
            })
    fact = Fact("1행 = 1 자산 × 1 기간", ("asset_no", "period"), rows)
    return cal, periods, fact


def golden_sample() -> FixedCostDeprInput:
    """재무 의미 없는 구조 골든. 자산 3건(분야·가동시점·잔존가 다양)."""
    assets = [
        Asset("V-001", "리스금융A", "vehicle",
              acq_cost=36_000, life_months=36, salvage=0.0, in_service=(2024, 1)),
        Asset("P-001", "빌딩임대B", "property",
              acq_cost=120_000, life_months=120, salvage=0.0, in_service=(2024, 1)),
        Asset("F-001", "설비공급C", "fixed_parts",
              acq_cost=24_000, life_months=24, salvage=2_400, in_service=(2024, 4)),
    ]
    inp = FixedCostDeprInput(assets=assets, start=(2024, 1), end=(2024, 12))
    # GL 상각비 = 스케줄 총합으로 세팅(R11 대사 PASS 시연)
    _, _, fact = _build_fact(inp)
    inp.gl_dep_by_period = {"_total": sum(r["dep"] for r in fact.rows)}
    return inp


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def _expected_grid(data: FixedCostDeprInput, periods) -> set:
    """R12 활성 window 전수 (asset_no, period_label) 기대 격자."""
    first_ord, last_ord = periods[0].ordinal, periods[-1].ordinal
    by_ord = {p.ordinal: p.label for p in periods}
    grid: set = set()
    for a in data.assets:
        s, e = _active_window_ords(data, a, first_ord, last_ord)
        for o in range(s, e + 1):
            grid.add(((a.asset_no,), by_ord[o]))
    return grid


def _disposal_gain_loss(data: FixedCostDeprInput, periods, fact) -> list[dict]:
    """C11 처분손익 = 처분가 − NBV(처분시점 장부가).

    NBV = 처분 period 의 opening(D&A 중단으로 closing=opening 고정). 양=이익/음=손실.
    disposals + disposal_prices 짝이 있는 자산만.
    """
    by_ap = {(r["asset_no"], r["period"]): r for r in fact.rows}
    by_label = {p.ordinal: p.label for p in periods}
    out = []
    for a in data.assets:
        disp = data.disposals.get(a.asset_no)
        if disp is None or a.asset_no not in data.disposal_prices:
            continue
        disp_ord = disp[0] * 12 + (disp[1] - 1)
        lbl = by_label.get(disp_ord)
        nbv = by_ap[(a.asset_no, lbl)]["opening"] if lbl and (a.asset_no, lbl) in by_ap \
            else a.acq_cost
        price = data.disposal_prices[a.asset_no]
        out.append({"asset_no": a.asset_no, "period": lbl, "nbv": nbv,
                    "price": price, "gain_loss": price - nbv})
    return out


def build(data: FixedCostDeprInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, fact = _build_fact(data)
    nP = len(periods)
    last_col = 1 + nP + 1                       # 자산열 + 기간열들 + 합계열

    # R12 재발성: 활성 window 내 (asset, period) 행 중 dep 미계상(None) → MISSING_ACCRUAL.
    #   1층 emit: ledger 1행 + 시트 NO_DATA 셀(surfaced flag). 저장 막지 않음(자문 R3).
    ledger = vc.AnomalyLedger()
    present_filled = {((r["asset_no"],), r["period"])
                      for r in fact.rows if r["dep"] is not None}
    for keyt, per in sorted(_expected_grid(data, periods) - present_filled):
        ledger.add(grain=keyt, period=per, anomaly_type="MISSING_ACCRUAL",
                   detail="활성 window 내 GL 상각 미계상")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Depreciation")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    header_row = r
    hs.set_widths(ws, {1: 24, last_col: 14})
    for c in range(2, last_col):
        ws.column_dimensions[get_column_letter(c)].width = 11

    # 헤더
    hs.set_cell(ws, r, 1, "자산 (분야)", role="header", align=hs.LEFT)
    for j, p in enumerate(periods, start=2):
        hs.set_cell(ws, r, j, p.label, role="header")
    hs.set_cell(ws, r, last_col, "합계", role="header")
    r += 1

    data_start = r
    # 자산별 행 (dep 값 = 정액법 재계산값, calc). dep=None 셀은 NO_DATA 로 노출(은폐 금지).
    by_asset: dict[str, dict[str, float]] = {}
    for row in fact.rows:
        by_asset.setdefault(row["asset_no"], {})[row["period"]] = row["dep"]
    surfaced = 0
    for a in data.assets:
        label = "%s (%s)" % (a.asset_no, DOMAIN_LABEL.get(a.domain, a.domain))
        hs.set_cell(ws, r, 1, label, role="label", align=hs.LEFT)
        for j, p in enumerate(periods, start=2):
            v = by_asset[a.asset_no][p.label]
            if v is None:                        # R12 결측 노출 = surfaced flag
                hs.set_cell(ws, r, j, "NO_DATA", role="soft", align=hs.CENTER)
                surfaced += 1
            else:
                hs.set_cell(ws, r, j, v, role="calc", number_format=hs.FMT_INT_DASH)
        # 합계 = NO_DATA 셀 제외 SUM (NO_DATA 텍스트는 SUM 에 무시됨)
        c0, c1 = get_column_letter(2), get_column_letter(last_col - 1)
        hs.set_cell(ws, r, last_col, "=SUM(%s%d:%s%d)" % (c0, r, c1, r),
                    role="calc", number_format=hs.FMT_INT, bold=True)
        r += 1
    data_end = r - 1

    # 합계행 (기간별 SUM)
    hs.set_cell(ws, r, 1, "합계", role="total", align=hs.LEFT)
    for j in range(2, last_col + 1):
        cl = get_column_letter(j)
        hs.set_cell(ws, r, j, "=SUM(%s%d:%s%d)" % (cl, data_start, cl, data_end),
                    role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    total_row = r

    # _RECON 블록 (None=결측 셀은 합계 제외)
    master_total = sum(rrow["dep"] for rrow in fact.rows if rrow["dep"] is not None)
    gl_total = data.gl_dep_by_period.get("_total")
    recon = vc.recon_block(
        n_input=len(data.assets), n_output=len(fact.rows),
        src_sum=master_total, out_sum=(gl_total if gl_total is not None else master_total),
        completeness="자산 %d × 기간 %d 전수 (결측 기간 dep=0 유지)" % (len(data.assets), nP),
        accuracy="정액법 재계산 대조 (fpna.finance)",
        cutoff="기간 cutoff = 캘린더 기간말",
    )
    rec_top = total_row + 2
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    # Anomaly ledger 노출 (R12 MISSING_ACCRUAL) — 2층 emit. 행이 시트에 보여야 보존 성립.
    if len(ledger):
        nxt = hs.section_header(ws, nxt + 1, "이상치 대장 (Anomaly Ledger)", last_col=last_col)
        for j, h in enumerate(("자산", "기간", "유형", "비고"), start=1):
            hs.set_cell(ws, nxt, j, h, role="header", align=hs.LEFT)
        nxt += 1
        for row in ledger.rows:
            hs.set_cell(ws, nxt, 1, row["grain"][0], role="label", align=hs.LEFT)
            hs.set_cell(ws, nxt, 2, row["period"], role="soft", align=hs.LEFT)
            hs.set_cell(ws, nxt, 3, row["anomaly_type"], role="soft", align=hs.LEFT)
            hs.set_cell(ws, nxt, 4, row["detail"], role="soft", align=hs.LEFT)
            nxt += 1

    # C11 처분손익 (disposal gain/loss = 처분가 − NBV)
    disposal_gl = _disposal_gain_loss(data, periods, fact)
    if disposal_gl:
        nxt = hs.section_header(ws, nxt + 1, "처분손익 (Disposal Gain/Loss)", last_col=last_col)
        for j, h in enumerate(("자산", "처분월", "NBV(장부가)", "처분가", "처분손익"), start=1):
            hs.set_cell(ws, nxt, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
        nxt += 1
        for g in disposal_gl:
            hs.set_cell(ws, nxt, 1, g["asset_no"], role="label", align=hs.LEFT)
            hs.set_cell(ws, nxt, 2, g["period"], role="soft", align=hs.CENTER)
            hs.set_cell(ws, nxt, 3, g["nbv"], role="calc", number_format=hs.FMT_INT)
            hs.set_cell(ws, nxt, 4, g["price"], role="input", number_format=hs.FMT_INT)
            hs.set_cell(ws, nxt, 5, g["gain_loss"], role="calc", number_format=hs.FMT_INT)
            # 손실(음)=빨강 / 이익(양)=초록
            col = hs.POS_FG if g["gain_loss"] >= 0 else hs.NEG_FG
            ws.cell(row=nxt, column=5).font = hs.font(col, bold=True)
            nxt += 1

    # 코멘터리
    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="고정자산 대장 · GL 상각",
                     prepared_by="FP&A", last_col=last_col)
    wb._fpna_meta = {"cal": cal, "periods": periods, "fact": fact,
                     "master_total": master_total, "gl_total": gl_total,
                     "anomaly_ledger": ledger, "surfaced_flags": surfaced,
                     "disposal_gl": disposal_gl}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: FixedCostDeprInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal, periods, fact = meta["cal"], meta["periods"], meta["fact"]

    # R8 grain + R1 시간축 전수 + R7 no_silent_drop
    vc.assert_grain(rep, fact)
    vc.assert_time_ruler(rep, fact, cal, data.start, data.end, period_key="period")
    vc.assert_no_silent_drop(rep, fact, expected_n=len(data.assets) * len(periods))

    # 자산별 총 dep 재계산 대조 (가동~표시끝이 내용연수 이상이면 취득가-잔존)
    #   처분/손상/결측이 있으면 전체상각 가정이 깨지므로 그 자산은 대조 skip(메모).
    touched = set(data.disposals) | set(data.impairments) | set(data.missing_accruals)
    for a in data.assets:
        in_ord = _ordinal(a.in_service) if a.in_service else periods[0].ordinal
        elapsed = periods[-1].ordinal - in_ord + 1
        got = sum(r["dep"] for r in fact.rows
                  if r["asset_no"] == a.asset_no and r["dep"] is not None)
        if a.asset_no in touched:
            continue                              # 이벤트 자산 — 전체상각 대조 부적격
        if elapsed >= a.life_months:
            # N-version 폐형식: 전수명 도달 비이벤트 자산은 Σ기간상각 == 취득−잔존.
            #   got 은 fact(빌드 산출), expected 는 INPUT 폐형식 — 독립 두 경로 대조.
            expected = a.acq_cost - a.salvage
            ok = finance.approx_equal(got, expected, rel=1e-9, abs_=1e-6)
            rep.add("자산상각합:%s" % a.asset_no, ok,
                    "" if ok else "계산=%.6g 기대=%.6g" % (got, expected))
        else:
            # 부분활성(전수명 미도달) 자산: Σdep == 월정액 × 활성기간수 폐형식 대조.
            #   단 가동 1차월 일할(first_period_factor)·처분·손상은 이 단순항등을
            #   깨므로 전수월·비이벤트만 적용. monthly 는 finance 정액(빌드 ext 와
            #   다른 경로). 가동월이 표시구간 안이면(부분 1차월) skip(메모).
            monthly = finance.straight_line_depreciation(
                a.acq_cost, a.salvage, a.life_months)
            cnt = sum(1 for r in fact.rows
                      if r["asset_no"] == a.asset_no and r["dep"] is not None)
            # 단순 정액×기간수가 깨지는 경우 skip: (a) 가동월이 표시 시작보다 뒤
            #   (1차월 일할) (b) first_period_factor≠1(명시 일할). 둘 다 1차월 부분
            #   상각 → 폐형식 부등. 처분/손상/결측은 위 touched 에서 이미 제외.
            ff = data.first_period_factor.get(a.asset_no, 1.0)
            partial_first = (in_ord > periods[0].ordinal) or (ff != 1.0)
            if not partial_first and cnt > 0:
                exp_partial = monthly * cnt
                ok = finance.approx_equal(got, exp_partial, rel=1e-9, abs_=1e-6)
                rep.add("N-version 부분상각:%s(정액×기간)" % a.asset_no, ok,
                        "" if ok else "계산=%.6g 기대=%.6g(월%.6g×%d)"
                        % (got, exp_partial, monthly, cnt))

    # R12 재발성: 활성 window 의 (asset, period) 가 모두 *계상*됐는지(결측=MISSING_ACCRUAL).
    #   build 의 ledger 를 신선 ledger 로 재도출해 동일 결과인지 교차검증(은폐 불가).
    filled = Fact("1행 = 1 자산 × 1 기간", ("asset_no", "period"),
                  [r for r in fact.rows if r["dep"] is not None])
    fresh = vc.AnomalyLedger()
    vc.assert_recurrence(rep, filled, _expected_grid(data, periods), fresh,
                         period_key="period", key_keys=["asset_no"])
    led = meta["anomaly_ledger"]
    rep.add("R12 ledger 일치", len(fresh) == len(led),
            "" if len(fresh) == len(led) else "build=%d qc재도출=%d" % (len(led), len(fresh)))

    # Anomaly 2층 보존: |ledger| == 시트 surfaced flag (은폐 시 passed=False).
    vc.assert_anomaly_conserved(rep, led, meta["surfaced_flags"])

    # R11 master ↔ GL 대사
    if meta["gl_total"] is not None:
        vc.assert_master_to_gl(rep, meta["master_total"], meta["gl_total"])

    # C11 처분손익 재계산 대조 (처분가 − NBV)
    gl_ok = True
    for g in meta.get("disposal_gl", []):
        if abs(g["gain_loss"] - (g["price"] - g["nbv"])) > 1e-6:
            gl_ok = False
    if meta.get("disposal_gl"):
        rep.add("C11 처분손익 = 처분가 − NBV", gl_ok, "" if gl_ok else "처분손익 재계산 불일치")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "FixedCostDeprInput", "golden_sample", "build", "qc"]
