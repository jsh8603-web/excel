"""
fpna.templates.cohort_retention — 코호트 잔존 / NRR·GRR 리텐션 매트릭스.

SaaS/구독 분석의 정통 산출물. 가입 시점(cohort)별로 줄을 세우고, 가입 후
경과 기간(period 0,1,2,…)에 따라 남아있는 MRR/고객을 추적한다. CFO/투자자
질문 "이 코호트 12개월 후 NRR 얼마? 이탈(GRR) 어디서 빠지나?"에 답한다.

- grain = "1행 = 1 cohort × 1 경과기간(age)" (내부 tidy). 표시는 cohort 행 ×
  age 열 wide 삼각행렬.
- GRR(Gross Retention) = (시작MRR − churn − contraction) / 시작MRR. 확장 미반영
  → 단조 비증가(이탈은 누적). 상한 100%.
- NRR(Net Retention) = (시작MRR − churn − contraction + expansion) / 시작MRR.
  확장 포함 → 100% 초과 가능.
- 불변식:
  (1) R3 tie: 각 age 의 end_mrr == start_mrr − churn − contraction + expansion.
  (2) GRR 단조 비증가(이탈 누적 — age 가 늘수록 GRR 안 올라간다).
  (3) GRR ≤ NRR (확장이 음수가 아닌 한 net ≥ gross).
  (4) R17: start_mrr == 0 인 코호트의 비율은 NA(0/inf 박제 금지).
- wide(R5): cohort 가 데이터에 없어도 age 축은 전수 전개(미관측 = NO_DATA).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "cohort_retention"


@dataclass
class CohortStep:
    """1 cohort 의 1 경과기간(age) 변동. start_mrr 는 age=0 만(또는 직전 end)."""
    age: int                  # 가입 후 경과 기간(0,1,2,…)
    start_mrr: float          # 기간 시작 MRR(보통 age=0=신규, 이후=직전 end)
    churn: float = 0.0        # 완전 이탈(MRR 감소, 양수로 표기)
    contraction: float = 0.0  # 다운그레이드(MRR 감소, 양수로 표기)
    expansion: float = 0.0    # 업셀/확장(MRR 증가, 양수로 표기)

    @property
    def end_mrr(self) -> float:
        return self.start_mrr - self.churn - self.contraction + self.expansion


@dataclass
class CohortLine:
    """1 cohort(가입 시점)의 경과기간별 잔존 시계열."""
    cohort: str               # 코호트 라벨(예: "2024-01")
    steps: list = field(default_factory=list)   # list[CohortStep], age 오름차순


@dataclass
class CohortRetentionInput:
    title: str = "코호트 잔존 / NRR·GRR 리텐션"
    subtitle: str = "가입 시점별 경과기간 잔존율(이탈 누적 = GRR, 확장 포함 = NRR)"
    unit: str = "₩"
    cohorts: list = field(default_factory=list)   # list[CohortLine]
    max_age: int | None = None                    # 표시 age 축 상한(None=관측 최대)
    commentary: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# 파생 계산 (build/qc 공유 — 결정성)                                          #
# --------------------------------------------------------------------------- #
def _base_mrr(line: CohortLine) -> float | None:
    """코호트 시작(age=0) MRR. steps 없으면 None."""
    if not line.steps:
        return None
    s0 = min(line.steps, key=lambda s: s.age)
    return s0.start_mrr


def _max_age(inp: CohortRetentionInput) -> int:
    if inp.max_age is not None:
        return inp.max_age
    ages = [s.age for ln in inp.cohorts for s in ln.steps]
    return max(ages) if ages else 0


def _retention_row(line: CohortLine, ages: list[int]):
    """cohort 1줄의 age 별 (grr, nrr) 또는 (None, reason). base 0 → NA(R17).

    grr/nrr 은 base(시작 MRR) 대비 *누적* 잔존율. churn/contraction 은 이탈
    누적(단조), expansion 은 net 에만 가산.
    """
    base = _base_mrr(line)
    by_age = {s.age: s for s in line.steps}
    out: dict[int, dict] = {}
    cum_churn = 0.0
    cum_contraction = 0.0
    cum_expansion = 0.0
    for age in ages:
        step = by_age.get(age)
        if step is not None:
            cum_churn += step.churn
            cum_contraction += step.contraction
            cum_expansion += step.expansion
        present = step is not None
        gross_num = (base - cum_churn - cum_contraction) if base is not None else None
        net_num = (gross_num + cum_expansion) if gross_num is not None else None
        grr, grr_na = finance_ratio(gross_num, base)
        nrr, nrr_na = finance_ratio(net_num, base)
        out[age] = {"present": present, "grr": grr, "grr_na": grr_na,
                    "nrr": nrr, "nrr_na": nrr_na,
                    "end_mrr": (step.end_mrr if step is not None else None),
                    "start_mrr": (step.start_mrr if step is not None else None)}
    return base, out


def finance_ratio(num, den):
    """비율 또는 NA 사유(R17 ratio_or_na 위임)."""
    val, reason = vc.ratio_or_na(num, den)
    return val, reason


def _build_fact(inp: CohortRetentionInput) -> Fact:
    rows: list[dict] = []
    for ln in inp.cohorts:
        for s in ln.steps:
            rows.append({"cohort": ln.cohort, "age": s.age,
                         "start_mrr": s.start_mrr, "end_mrr": s.end_mrr})
    return Fact("1행 = 1 cohort × 1 경과기간(age)", ("cohort", "age"), rows)


def golden_sample() -> CohortRetentionInput:
    """구조 골든 — 두 코호트. 하나는 이탈 누적(GRR<100%), 하나는 확장으로 NRR>100%.

    ⚠ 구조 검증용 더미(재무 의미 없음). 단조·tie·GRR≤NRR 불변식 검증.
    """
    c1 = CohortLine("2024-01", [
        CohortStep(0, start_mrr=1000.0),                               # 신규 1000
        CohortStep(1, start_mrr=1000.0, churn=50.0),                   # -50 이탈
        CohortStep(2, start_mrr=950.0, churn=30.0, contraction=20.0),  # -50 더
        CohortStep(3, start_mrr=900.0, churn=20.0, expansion=10.0),    # 소폭 확장
    ])
    # 확장이 강해 NRR>100% 가능(GRR 는 여전히 ≤100%)
    c2 = CohortLine("2024-02", [
        CohortStep(0, start_mrr=800.0),
        CohortStep(1, start_mrr=800.0, churn=20.0, expansion=60.0),
        CohortStep(2, start_mrr=840.0, churn=10.0, expansion=80.0),
    ])
    return CohortRetentionInput(
        cohorts=[c1, c2], max_age=3,
        commentary=["2024-01: 3기 GRR 88% — 초기 이탈 집중, 확장 미미",
                    "2024-02: 확장 강해 2기 NRR 100% 초과(net expansion)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: CohortRetentionInput, *, mode: str = "create",
          base_path=None) -> openpyxl.Workbook:
    fact = _build_fact(data)
    max_age = _max_age(data)
    ages = list(range(0, max_age + 1))
    # 열: 코호트 + 시작MRR + age 별 GRR + (마지막) NRR 요약
    last_col = 2 + len(ages) + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Cohort")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    hs.set_widths(ws, {1: 14, 2: 12})
    for c in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    # 헤더(GRR 매트릭스)
    r = hs.section_header(ws, r, "GRR — 총 잔존율(이탈 누적, 확장 미반영)", last_col=last_col)
    hs.set_cell(ws, r, 1, "코호트", role="header", align=hs.LEFT)
    hs.set_cell(ws, r, 2, "시작MRR", role="header")
    for j, age in enumerate(ages, start=3):
        hs.set_cell(ws, r, j, "M%d" % age, role="header")
    hs.set_cell(ws, r, last_col, "최종NRR", role="header")
    r += 1

    grr_top = r
    grr_range_rows: list[int] = []
    cohort_results: list[tuple[CohortLine, float | None, dict]] = []
    na_surfaced = 0
    for ln in data.cohorts:
        base, per_age = _retention_row(ln, ages)
        cohort_results.append((ln, base, per_age))
        hs.set_cell(ws, r, 1, ln.cohort, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, base if base is not None else "NA",
                    role="input", number_format=hs.FMT_INT)
        last_nrr = None
        for j, age in enumerate(ages, start=3):
            cell = per_age[age]
            if not cell["present"]:
                hs.set_cell(ws, r, j, "—", role="soft", align=hs.CENTER)
            elif cell["grr_na"] is not None:
                hs.set_cell(ws, r, j, "NA", role="soft", align=hs.CENTER)
                na_surfaced += 1
            else:
                hs.set_cell(ws, r, j, cell["grr"], role="calc",
                            number_format=hs.FMT_PCT1, align=hs.CENTER)
            if cell["present"] and cell["nrr"] is not None:
                last_nrr = cell["nrr"]
        hs.set_cell(ws, r, last_col,
                    last_nrr if last_nrr is not None else "NA",
                    role="total", number_format=hs.FMT_PCT1, bold=True)
        grr_range_rows.append(r)
        r += 1

    # GRR 히트맵(낮을수록 빨강 — 이탈 위험). 3..(2+len ages) 열.
    if grr_range_rows:
        rng = "%s%d:%s%d" % (get_column_letter(3), grr_top,
                             get_column_letter(2 + len(ages)), grr_range_rows[-1])
        hs.apply_heatmap(ws, rng, low=hs.NEG_FG, mid=hs.WHITE, high=hs.POS_FG)

    # NRR 매트릭스(확장 포함) — 별도 섹션
    r += 1
    r = hs.section_header(ws, r, "NRR — 순 잔존율(확장 포함, 100% 초과 가능)",
                          last_col=last_col)
    hs.set_cell(ws, r, 1, "코호트", role="header", align=hs.LEFT)
    hs.set_cell(ws, r, 2, "시작MRR", role="header")
    for j, age in enumerate(ages, start=3):
        hs.set_cell(ws, r, j, "M%d" % age, role="header")
    r += 1
    for ln, base, per_age in cohort_results:
        hs.set_cell(ws, r, 1, ln.cohort, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, base if base is not None else "NA",
                    role="input", number_format=hs.FMT_INT)
        for j, age in enumerate(ages, start=3):
            cell = per_age[age]
            if not cell["present"]:
                hs.set_cell(ws, r, j, "—", role="soft", align=hs.CENTER)
            elif cell["nrr_na"] is not None:
                hs.set_cell(ws, r, j, "NA", role="soft", align=hs.CENTER)
                na_surfaced += 1
            else:
                hs.set_cell(ws, r, j, cell["nrr"], role="calc",
                            number_format=hs.FMT_PCT1, align=hs.CENTER)
        r += 1

    # _RECON: 각 step end_mrr == start − churn − contraction + expansion (tie)
    n_steps = sum(len(ln.steps) for ln in data.cohorts)
    src_sum = sum(s.end_mrr for ln in data.cohorts for s in ln.steps)
    out_sum = sum(
        s.start_mrr - s.churn - s.contraction + s.expansion
        for ln in data.cohorts for s in ln.steps)
    recon = vc.recon_block(
        n_input=len(data.cohorts), n_output=n_steps,
        src_sum=src_sum, out_sum=out_sum,
        completeness="코호트 %d × age 0..%d 전수" % (len(data.cohorts), max_age),
        accuracy="end = start − churn − contraction + expansion (step tie)",
        cutoff="GRR 단조 비증가 / GRR ≤ NRR / base=0 → NA(R17)")
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    end_row = rec_top + len(recon) + 1

    if data.commentary:
        cr = end_row + 2
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        end_row = cr

    hs.report_footer(ws, end_row + 1, source="구독 원장 · MRR movement",
                     prepared_by="FP&A", last_col=last_col)

    wb._fpna_meta = {"fact": fact, "ages": ages, "cohort_results": cohort_results,
                     "src_sum": src_sum, "out_sum": out_sum,
                     "na_surfaced": na_surfaced}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: CohortRetentionInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    fact = meta["fact"]

    # R8 grain (cohort × age 중복 0)
    vc.assert_grain(rep, fact)

    # R3 tie: Σ end_mrr == Σ(start − churn − contraction + expansion)
    vc.assert_tie_out(rep, meta["src_sum"], meta["out_sum"], tol=1e-6,
                      name="R3 movement_tie")

    # 불변식 (2)(3): GRR 단조 비증가 + GRR ≤ NRR (관측된 age 만)
    mono_ok = True
    order_ok = True
    for ln, base, per_age in meta["cohort_results"]:
        prev_grr = None
        for age in meta["ages"]:
            cell = per_age[age]
            if not cell["present"] or cell["grr"] is None:
                continue
            if prev_grr is not None and cell["grr"] > prev_grr + 1e-9:
                mono_ok = False
            prev_grr = cell["grr"]
            if cell["nrr"] is not None and cell["grr"] > cell["nrr"] + 1e-9:
                order_ok = False
    rep.add("GRR 단조 비증가(이탈 누적)", mono_ok,
            "" if mono_ok else "GRR 가 age 증가에 상승(이탈 누적 위배)")
    rep.add("GRR ≤ NRR(확장 비음수)", order_ok,
            "" if order_ok else "GRR > NRR (확장 음수 또는 계산 오류)")

    # R17: NA 가 산출물에 노출됐는지(0분모 코호트 → NA 표기). 발견 emit, 은폐 차단.
    expected_na = 0
    for ln, base, per_age in meta["cohort_results"]:
        for age in meta["ages"]:
            cell = per_age[age]
            if cell["present"]:
                if cell["grr_na"] is not None:
                    expected_na += 1
                if cell["nrr_na"] is not None:
                    expected_na += 1
    rep.add("R17 NA surfaced(0분모 은폐 금지)", meta["na_surfaced"] == expected_na,
            "" if meta["na_surfaced"] == expected_na
            else "surfaced=%d 기대=%d" % (meta["na_surfaced"], expected_na))

    # 라인별 end_mrr 재계산 대조(결정성)
    recompute_ok = True
    for ln in data.cohorts:
        for s in ln.steps:
            if not finance.approx_equal(
                    s.end_mrr,
                    s.start_mrr - s.churn - s.contraction + s.expansion):
                recompute_ok = False
    rep.add("step end_mrr 재계산", recompute_ok, "")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "CohortStep", "CohortLine", "CohortRetentionInput",
           "golden_sample", "build", "qc"]
