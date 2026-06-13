"""
fpna.templates.headcount_plan — 인원 계획 → fully-loaded 인건비 → opex.

roster(부서별 인원·기본급)에 부담률(4대보험·복리후생·상여 등 loading)을 얹어
"진짜 회사가 부담하는" 1인당 비용(fully-loaded)을 구하고, 기간(월)에 걸쳐
부서별 총 인건비를 계획한다. CFO 질문 "내년 인건비 얼마, 증원 N명이면 opex
얼마 늘어?"에 답한다.

- grain = "1행 = 1 부서 × 1 기간" (내부 tidy). 표시는 부서 행 × 기간 열 wide.
- fully-loaded cost = base_salary × (1 + loading_rate). 1인당 월부담 = annual / 12.
- 부서 월 인건비 = headcount(기간) × fully_loaded_monthly.
- 불변식:
  (1) R10 tie: Σ부서별 인건비 == 총계(전 기간·전 부서 합).
  (2) R3 tie: Σ fully_loaded == Σ(base + loading) — loading 누수 0.
  (3) headcount ≥ 0(음수 인원 금지).
- ramp(중도 입사)는 기간별 headcount 벡터로 표현(전수 — 빈 기간 0 명시).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "headcount_plan"


@dataclass
class RosterLine:
    """1 부서(또는 직급) 의 인원 계획 라인."""
    dept: str                          # 부서/조직 라벨
    grade: str = ""                    # 직급(옵션 — 표시용)
    base_salary_annual: float = 0.0    # 1인 연 기본급
    loading_rate: float = 0.0          # 부담률(4대보험+복리후생+상여), 예 0.25 = 25%
    headcount: list = field(default_factory=list)  # 기간순 인원(period 와 1:1)

    @property
    def fully_loaded_annual(self) -> float:
        return self.base_salary_annual * (1.0 + self.loading_rate)

    @property
    def fully_loaded_monthly(self) -> float:
        return self.fully_loaded_annual / 12.0


@dataclass
class HeadcountPlanInput:
    title: str = "인원 계획 / Fully-Loaded 인건비"
    subtitle: str = "roster → 부담률 적용 → 부서·기간별 인건비 opex"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)
    end: tuple = (2024, 12)
    lines: list = field(default_factory=list)   # list[RosterLine]
    commentary: list = field(default_factory=list)


def _periods(inp: HeadcountPlanInput):
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    return cal, cal.periods(inp.start, inp.end)


def _build_fact(inp: HeadcountPlanInput):
    """tidy Fact (dept × period). 부서별 월 인건비 = hc × fully_loaded_monthly."""
    cal, periods = _periods(inp)
    rows: list[dict] = []
    for ln in inp.lines:
        flm = ln.fully_loaded_monthly
        for p, hc in zip(periods, ln.headcount):
            rows.append({"dept": ln.dept, "period": p.label,
                         "headcount": hc, "cost": hc * flm})
    fact = Fact("1행 = 1 부서 × 1 기간", ("dept", "period"), rows)
    return cal, periods, fact


def golden_sample() -> HeadcountPlanInput:
    """구조 골든 — 두 부서. 하나는 평탄, 하나는 중도 증원(ramp).

    ⚠ 구조 검증용 더미(재무 의미 없음). R10 부서 roll-up tie 검증.
    """
    lines = [
        # 영업: 5명 평탄, 연봉 60(천 단위 구조 더미), 부담률 25%
        RosterLine("영업", "", base_salary_annual=60_000.0, loading_rate=0.25,
                   headcount=[5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]),
        # 개발: 3명 → 7월부터 5명 증원(ramp), 연봉 80, 부담률 30%
        RosterLine("개발", "", base_salary_annual=80_000.0, loading_rate=0.30,
                   headcount=[3, 3, 3, 3, 3, 3, 5, 5, 5, 5, 5, 5]),
    ]
    return HeadcountPlanInput(
        lines=lines, start=(2024, 1), end=(2024, 12),
        commentary=["개발 7월 +2명 증원 — 하반기 인건비 step up",
                    "부담률: 영업 25% / 개발 30%(상여·복리후생 차등)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: HeadcountPlanInput, *, mode: str = "create",
          base_path=None) -> openpyxl.Workbook:
    cal, periods, fact = _build_fact(data)
    nP = len(periods)
    # 열: 부서 + [연봉 · 부담률 · FL월] 가정 + 기간별 인건비 + 연계
    assum_cols = ("연기본급", "부담률", "FL월/인")
    last_col = 1 + len(assum_cols) + nP + 1   # +1 = 연계(합)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Headcount")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    hs.set_widths(ws, {1: 14, 2: 12, 3: 9, 4: 11})
    for c in range(5, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    # 헤더
    hs.set_cell(ws, r, 1, "부서", role="header", align=hs.LEFT)
    for k, h in enumerate(assum_cols):
        hs.set_cell(ws, r, 2 + k, h, role="header")
    base_period_col = 2 + len(assum_cols)
    for j, p in enumerate(periods):
        hs.set_cell(ws, r, base_period_col + j, p.label, role="header")
    hs.set_cell(ws, r, last_col, "연계", role="header")
    r += 1

    # 부서별 인건비(기간 열) + 가정
    dept_period_cost: dict[str, list] = {}
    dept_year_total: dict[str, float] = {}
    period_total = [0.0] * nP
    data_top = r
    for ln in data.lines:
        flm = ln.fully_loaded_monthly
        hs.set_cell(ws, r, 1, ln.dept + (" / " + ln.grade if ln.grade else ""),
                    role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, ln.base_salary_annual, role="input", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 3, ln.loading_rate, role="input", number_format=hs.FMT_PCT0)
        hs.set_cell(ws, r, 4, flm, role="calc", number_format=hs.FMT_INT)
        costs: list[float] = []
        yr = 0.0
        for j, hc in enumerate(ln.headcount):
            c = hc * flm
            costs.append(c)
            yr += c
            period_total[j] += c
            hs.set_cell(ws, r, base_period_col + j, c, role="calc",
                        number_format=hs.FMT_INT)
        hs.set_cell(ws, r, last_col, yr, role="total", number_format=hs.FMT_INT, bold=True)
        dept_period_cost[ln.dept] = costs
        dept_year_total[ln.dept] = yr
        r += 1

    # 총계행(전 부서 합) — R10 tie 의 시각화
    total_row = r
    hs.set_cell(ws, r, 1, "총계", role="total", align=hs.LEFT)
    grand = 0.0
    for j in range(nP):
        hs.set_cell(ws, r, base_period_col + j, period_total[j], role="total",
                    number_format=hs.FMT_INT)
        grand += period_total[j]
    hs.set_cell(ws, r, last_col, grand, role="total", number_format=hs.FMT_INT, bold=True)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG

    # 인원수 추이 막대(헤드카운트, 부서 합) — 보조 시각화는 생략(표 우선).

    # _RECON: Σ부서별 == 총계(R10) + Σfully_loaded == Σ(base+loading)
    fl_sum = sum(ln.fully_loaded_annual for ln in data.lines)
    base_plus_loading = sum(ln.base_salary_annual * (1.0 + ln.loading_rate)
                            for ln in data.lines)
    recon = vc.recon_block(
        n_input=len(data.lines), n_output=len(fact.rows),
        src_sum=sum(dept_year_total.values()), out_sum=grand,
        completeness="부서 %d × 기간 %d 전수" % (len(data.lines), nP),
        accuracy="fully_loaded = base × (1+loading) / 부서합 = 총계",
        cutoff="기간별 인건비 = headcount × FL월")
    rec_top = total_row + 2
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    end_row = rec_top + len(recon) + 1

    if data.commentary:
        cr = end_row + 2
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        end_row = cr

    hs.report_footer(ws, end_row + 1, source="HR roster · 인건비 가정",
                     prepared_by="FP&A", last_col=last_col)

    wb._fpna_meta = {
        "cal": cal, "periods": periods, "fact": fact,
        "dept_period_cost": dept_period_cost, "dept_year_total": dept_year_total,
        "period_total": period_total, "grand": grand,
        "fl_sum": fl_sum, "base_plus_loading": base_plus_loading,
    }
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# T2 바인딩 (from_tidy) + T4 보존 (conserves) — module-level                   #
# --------------------------------------------------------------------------- #
GRAIN = ("dept", "period")                     # 1행 = 1 부서 × 1 기간
REQUIRED = ("lines",)
UNIT_POLICY = {"lines.base_salary_annual": float, "lines.loading_rate": float}


def from_tidy(rows) -> HeadcountPlanInput:
    """tidy rows(dept × period) → HeadcountPlanInput. 부서별 headcount 벡터 재조립.

    행 컬럼: dept, period, headcount, [base_salary_annual, loading_rate, grade].
    부서 가정(연봉·부담률·직급)은 그 부서 행에서 동일하다고 보고 첫 행에서 취득.
    headcount 는 period 정렬(라벨 오름차순)로 벡터화 — 시간축 순서 보존.
    """
    from fpna.binding import _coerce
    from itertools import groupby
    keyf = lambda r: _coerce(r.get("dept"), str)
    srt = sorted(rows, key=lambda r: (keyf(r), str(r.get("period"))))
    lines = []
    for dept, grp in groupby(srt, key=keyf):
        grp = list(grp)
        first = grp[0]
        hc = [_coerce(r.get("headcount"), int) or 0 for r in grp]
        lines.append(RosterLine(
            dept=dept,
            grade=_coerce(first.get("grade"), str) or "",
            base_salary_annual=_coerce(first.get("base_salary_annual"), float) or 0.0,
            loading_rate=_coerce(first.get("loading_rate"), float) or 0.0,
            headcount=hc,
        ))
    return HeadcountPlanInput(lines=lines)


def conserves(wb, data):
    """T4 보존: INPUT 직접 산술로 총 인건비 독립 재합산 == build 보고 grand.

    ⛔ build 의 _build_fact 재호출 금지. RosterLine 의 fully_loaded_monthly 는
    dataclass property(build 헬퍼 아님)라 INPUT 고유 산술로 간주 — 독립 경로.
    부서 월 인건비 = Σ(headcount × FL월). 전 부서·전 기간 합.
    """
    raw = 0.0
    for ln in data.lines:                        # INPUT 직접 순회(독립 경로)
        flm = ln.base_salary_annual * (1.0 + ln.loading_rate) / 12.0
        for hc in ln.headcount:
            raw += hc * flm
    return [("총인건비", raw, wb._fpna_meta["grand"])]


def qc(wb: openpyxl.Workbook, data: HeadcountPlanInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal, periods, fact = meta["cal"], meta["periods"], meta["fact"]

    # R8 grain + R1 시간축 전수
    vc.assert_grain(rep, fact)
    vc.assert_time_ruler(rep, fact, cal, data.start, data.end, period_key="period")
    vc.assert_no_silent_drop(rep, fact, expected_n=len(data.lines) * len(periods))

    # R10 tie: Σ부서 연합 == 총계(grand)
    dept_sum = sum(meta["dept_year_total"].values())
    vc.assert_tie_out(rep, dept_sum, meta["grand"], tol=1e-6, name="R10 dept_rollup_tie")

    # R3 tie: Σ fully_loaded == Σ(base + loading) — loading 누수 0
    vc.assert_tie_out(rep, meta["fl_sum"], meta["base_plus_loading"], tol=1e-6,
                      name="R3 loading_tie")

    # 부서·기간 인건비 재계산 대조(결정성)
    recompute_ok = True
    for ln in data.lines:
        flm = ln.fully_loaded_monthly
        for j, hc in enumerate(ln.headcount):
            if not finance.approx_equal(meta["dept_period_cost"][ln.dept][j], hc * flm):
                recompute_ok = False
    rep.add("부서·기간 인건비 재계산", recompute_ok, "")

    # headcount ≥ 0 (음수 인원 금지)
    neg = [(ln.dept, j) for ln in data.lines
           for j, hc in enumerate(ln.headcount) if hc < 0]
    rep.add("headcount ≥ 0", not neg,
            "" if not neg else "음수 인원: " + ", ".join("%s@%d" % t for t in neg[:6]))

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "RosterLine", "HeadcountPlanInput",
           "golden_sample", "build", "qc",
           "GRAIN", "REQUIRED", "UNIT_POLICY", "from_tidy", "conserves"]
