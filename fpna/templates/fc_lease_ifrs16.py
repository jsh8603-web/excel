"""
fpna.templates.fc_lease_ifrs16 — 고정비: K-IFRS 1116(IFRS 16) 리스 (C6 신규).

리스를 자본화(사용권자산·리스부채)해 상각·이자비용을 인식한다. CFO 질문
"리스 자본화 후 매기 이자/상각비, 부채 잔액 추이, rent-free(무상기간) 정액화는?"

- grain = "1행 = 1 리스(계약) × 1 기간".
- 리스부채 = Σ 미래 리스료 PV(증분차입이자율 할인). 매기 이자=기초부채×rate,
  부채상각=지급−이자, 기말=기초+이자−지급. 사용권자산=부채초기(+선급−인센티브
  +초기직접원가), 정액상각. rent-free=지급 0 이나 비용은 정액 인식.
- 불변식(R3 tie): Σ부채상각 == 리스부채 초기측정 / Σ자산상각 == 사용권자산 초기.
  기말부채(마지막) ≈ 0, 기말자산(마지막) ≈ 0. 잔액 chain 연속.
- dims.Contract 를 lease subtype 으로 받는다(recurrence=monthly, amount_per_period
  = 리스료). rent_free_periods 로 무상기간 지정(해당 기간 지급 0).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Contract, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_lease_ifrs16"


@dataclass
class LeaseTerms:
    """리스 1건의 IFRS 16 측정 파라미터(Contract 보강)."""
    contract: Contract                    # lease subtype (recurrence=monthly 가정)
    discount_rate: float                  # 기간 증분차입이자율(월)
    term_periods: int                     # 리스기간(개월)
    rent_free_periods: tuple = ()         # 무상기간 인덱스(0-base) — 지급 0
    initial_direct_costs: float = 0.0
    prepaid: float = 0.0
    incentives: float = 0.0


@dataclass
class LeaseIfrs16Input:
    title: str = "고정비 — 리스 자본화 (K-IFRS 1116)"
    subtitle: str = "사용권자산 상각 + 리스부채 이자 · rent-free 정액화"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)
    leases: list = field(default_factory=list)        # list[LeaseTerms]
    commentary: list = field(default_factory=list)


def _payments(lt: LeaseTerms) -> list[float]:
    """리스기간 지급 벡터. rent-free 기간은 0(무상)."""
    rf = set(lt.rent_free_periods)
    amt = lt.contract.amount_per_period
    return [0.0 if t in rf else amt for t in range(lt.term_periods)]


def _schedules(data: LeaseIfrs16Input):
    """리스별 finance.lease_schedule + 표시 기간(전체 최대 term) 정렬."""
    cal = AccountingCalendar(fiscal_year_start_month=data.fy_start_month)
    max_term = max((lt.term_periods for lt in data.leases), default=0)
    # 표시 기간 = start 부터 max_term 개월
    s_ord = data.start[0] * 12 + (data.start[1] - 1)
    e_ord = s_ord + max(max_term - 1, 0)
    periods = cal.periods((s_ord // 12, s_ord % 12 + 1), (e_ord // 12, e_ord % 12 + 1))
    sched = {}
    for lt in data.leases:
        rows = finance.lease_schedule(
            _payments(lt), lt.discount_rate,
            initial_direct_costs=lt.initial_direct_costs,
            prepaid=lt.prepaid, incentives=lt.incentives)
        sched[lt.contract.contract_id] = rows
    return cal, periods, sched


def golden_sample() -> LeaseIfrs16Input:
    """구조 골든 — 일반 리스 1건 + rent-free 2개월 리스 1건."""
    import datetime as _dt
    leases = [
        LeaseTerms(
            contract=Contract("LSE-01", "6010", "빌딩임대A", _dt.date(2024, 1, 1),
                              _dt.date(2025, 12, 31), "monthly", 1_000.0),
            discount_rate=0.004, term_periods=24),
        LeaseTerms(
            contract=Contract("LSE-02", "6010", "빌딩임대B(무상2M)", _dt.date(2024, 1, 1),
                              _dt.date(2025, 12, 31), "monthly", 800.0),
            discount_rate=0.004, term_periods=24, rent_free_periods=(0, 1)),
    ]
    return LeaseIfrs16Input(leases=leases, start=(2024, 1),
                            commentary=["LSE-02 첫 2개월 rent-free(지급 0) — 비용은 정액 인식",
                                        "리스부채 = 미래 리스료 PV(증분차입이자율 할인)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: LeaseIfrs16Input, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, sched = _schedules(data)
    nP = len(periods)
    last_col = 1 + nP + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("LeaseIFRS16")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    hs.set_widths(ws, {1: 26, last_col: 12})
    for c in range(2, last_col):
        ws.column_dimensions[get_column_letter(c)].width = 9

    hs.set_cell(ws, r, 1, "리스 / 흐름", role="header", align=hs.LEFT)
    for j, p in enumerate(periods, start=2):
        hs.set_cell(ws, r, j, p.label, role="header")
    hs.set_cell(ws, r, last_col, "합계", role="header")
    r += 1

    fact_rows = []
    flows = (("지급액", "payment", "input"), ("이자", "interest", "calc"),
             ("부채상각(−)", "principal", "calc"), ("기말 리스부채", "closing_liab", "total"),
             ("사용권자산 상각", "rou_amort", "calc"), ("기말 사용권자산", "rou_close", "total"))
    totals = {"interest": 0.0, "principal": 0.0, "rou_amort": 0.0, "payment": 0.0}
    liab0_sum = rou0_sum = 0.0
    for lt in data.leases:
        rows = sched[lt.contract.contract_id]
        cid = lt.contract.contract_id
        hs.section_header(ws, r, "%s [%s]" % (lt.contract.counterparty, cid), last_col=last_col)
        r += 1
        if rows:
            liab0_sum += rows[0].opening_liab
            rou0_sum += rows[0].rou_open
        for fname, key, role in flows:
            hs.set_cell(ws, r, 1, fname, role="label", align=hs.LEFT)
            for j, p in enumerate(periods, start=2):
                idx = j - 2
                if idx < len(rows):
                    val = getattr(rows[idx], key)
                else:
                    val = 0.0
                hs.set_cell(ws, r, j, val, role=role, number_format=hs.FMT_INT_DASH)
            if key in totals:
                ssum = sum(getattr(rr, key) for rr in rows)
                hs.set_cell(ws, r, last_col, ssum, role="calc",
                            number_format=hs.FMT_INT, bold=True)
                totals[key] += ssum
            r += 1
        # tidy fact (리스×기간)
        for idx, rr in enumerate(rows):
            if idx < nP:
                fact_rows.append({"lease_id": cid, "period": periods[idx].label})
        r += 1

    # _RECON: Σ부채상각 == 리스부채 초기 / Σ자산상각 == 사용권자산 초기
    recon = vc.recon_block(
        n_input=len(data.leases), n_output=len(fact_rows),
        src_sum=liab0_sum, out_sum=totals["principal"],
        completeness="리스 %d 전수 (term 별 전 기간)" % len(data.leases),
        accuracy="Σ부채상각 == 리스부채 초기측정 (지급=이자+원금)",
        cutoff="리스부채 = 미래 리스료 PV(증분차입이자율)",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    rr2 = rec_top + len(recon) + 1
    hs.set_cell(ws, rr2, 1, "사용권자산 초기 / Σ자산상각", role="label", align=hs.LEFT)
    hs.set_cell(ws, rr2, 2, rou0_sum, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, rr2, 3, totals["rou_amort"], role="calc", number_format=hs.FMT_INT)
    nxt = rr2 + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="리스 계약서 · 증분차입이자율",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 리스 × 1 기간", ("lease_id", "period"), fact_rows)
    wb._fpna_meta = {"cal": cal, "periods": periods, "fact": fact, "sched": sched,
                     "liab0_sum": liab0_sum, "rou0_sum": rou0_sum, "totals": totals}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: LeaseIfrs16Input) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    totals = meta["totals"]

    # R8 grain (lease × period 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # R3 tie: Σ부채상각 == 리스부채 초기측정
    vc.assert_tie_out(rep, meta["liab0_sum"], totals["principal"], tol=1e-6,
                      name="R3 liability_amort_tie")
    # R3 tie: Σ사용권자산상각 == 사용권자산 초기
    vc.assert_tie_out(rep, meta["rou0_sum"], totals["rou_amort"], tol=1e-6,
                      name="R3 rou_amort_tie")

    # 부채 chain: closing(t)=opening(t+1) + closing 식 정합 + 마지막 ≈ 0
    chain_ok = eq_ok = end_ok = True
    for cid, rows in meta["sched"].items():
        for i, rr in enumerate(rows):
            if abs((rr.opening_liab + rr.interest - rr.payment) - rr.closing_liab) > 1e-6:
                eq_ok = False
            if i + 1 < len(rows) and abs(rr.closing_liab - rows[i + 1].opening_liab) > 1e-6:
                chain_ok = False
        if rows and abs(rows[-1].closing_liab) > 1e-6:
            end_ok = False
        if rows and abs(rows[-1].rou_close) > 1e-6:
            end_ok = False
    rep.add("부채 chain 연속", chain_ok, "" if chain_ok else "chain 단절")
    rep.add("부채식 정합(기말=기초+이자−지급)", eq_ok, "" if eq_ok else "부채식 불일치")
    rep.add("기말 부채/자산 ≈ 0(상각 완료)", end_ok, "" if end_ok else "잔액 미소진")

    # rent-free 정액화: 무상기간도 사용권자산 상각은 정액(지급=0 이나 비용 인식)
    rf_ok = True
    for lt in data.leases:
        rows = meta["sched"][lt.contract.contract_id]
        for t in lt.rent_free_periods:
            if t < len(rows):
                if rows[t].payment != 0.0:
                    rf_ok = False
                if rows[t].rou_amort <= 0.0:        # 무상이어도 정액 상각(비용 인식)
                    rf_ok = False
    rep.add("rent-free 정액화(지급0·상각>0)", rf_ok, "" if rf_ok else "무상기간 상각 미인식")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "LeaseTerms", "LeaseIfrs16Input", "golden_sample", "build", "qc"]
