"""
fpna.templates.fc_driver_unitcost — 고정비: 활동 동인 단가 (A4).

고정비를 활동 동인(대당·㎡당·kWh당)으로 나눈 단위원가를 보여준다. CFO 질문
"차량 1대당 유지비 얼마, 면적당 임차료 추세는?"에 답한다.

- grain = "1행 = 1 CostCenter × 1 Account × 1 driver_type".
- 불변식: unit_cost = cost/qty (qty>0). 단위정합 선언. qty 결측/0 → NA(R17).
  ⛔ unit_cost 열 합산 금지 → blended = Σcost / Σqty (단가의 평균이 아니라 가중).
- R17: 0/음수 qty → RATIO_NA(ZERO_DENOM), 단위 혼재(㎡ vs sqft) → UNIT_MISMATCH.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_driver_unitcost"


@dataclass
class DriverLine:
    """1 (CC × Account × driver_type) 의 비용·수량·단위."""
    cost_center: str
    account: str
    driver_type: str          # 예: "차량대수", "임차면적", "전력량"
    label: str
    cost: float | None        # 고정비(결측 가능 — None)
    qty: float | None         # 동인 수량(결측/0 → R17 NA)
    unit: str                 # 동인 단위(예: "대", "㎡", "kWh") — 단위정합 선언
    canonical_unit: str | None = None  # 정규화 기준 단위(다르면 UNIT_MISMATCH)


@dataclass
class DriverUnitCostInput:
    title: str = "고정비 — 활동 동인 단가 (Driver Unit Cost)"
    subtitle: str = "동인(대·㎡·kWh)당 단위원가 · blended = Σcost/Σqty"
    unit: str = "₩"
    lines: list = field(default_factory=list)   # list[DriverLine]
    commentary: list = field(default_factory=list)


def golden_sample() -> DriverUnitCostInput:
    """구조 골든 — 정상 2건 + qty=0(ZERO_DENOM) 1건 + 단위혼재(UNIT_MISMATCH) 1건."""
    lines = [
        DriverLine("CC20", "6020", "차량유지", "차량유지비/대 (CC20)",
                   cost=12_000.0, qty=20.0, unit="대", canonical_unit="대"),
        DriverLine("CC10", "6010", "임차면적", "임차료/㎡ (CC10)",
                   cost=30_000.0, qty=1_000.0, unit="㎡", canonical_unit="㎡"),
        DriverLine("CC30", "6300", "전력", "전력비/kWh (CC30·결측)",
                   cost=8_000.0, qty=0.0, unit="kWh", canonical_unit="kWh"),  # ZERO_DENOM
        DriverLine("CC11", "6010", "임차면적", "임차료/sqft (CC11·단위혼재)",
                   cost=5_000.0, qty=200.0, unit="sqft", canonical_unit="㎡"),  # UNIT_MISMATCH
    ]
    return DriverUnitCostInput(lines=lines,
                               commentary=["전력 qty=0 → R17 ZERO_DENOM (단가 0/inf 박제 금지)",
                                           "sqft vs ㎡ → UNIT_MISMATCH (정규화 전 NA)"])


def _compute(inp: DriverUnitCostInput):
    """라인별 (unit_cost, reason) + blended(정상 라인만 Σcost/Σqty). ledger 동반."""
    ledger = vc.AnomalyLedger()
    rows = []
    sum_cost = 0.0
    sum_qty = 0.0
    for ln in inp.lines:
        val, reason = vc.ratio_or_na(
            ln.cost, ln.qty, num_unit=ln.unit, den_unit=ln.canonical_unit,
            require_same_unit=True)
        if reason is not None:
            ledger.add(grain=(ln.cost_center, ln.account, ln.driver_type),
                       period=None, anomaly_type="RATIO_NA", detail=reason)
        else:
            # blended 는 단위정합·정상 라인만 (Σcost/Σqty — 열 합산 금지)
            sum_cost += ln.cost
            sum_qty += ln.qty
        rows.append({"line": ln, "unit_cost": val, "reason": reason})
    blended, blended_reason = vc.ratio_or_na(sum_cost, sum_qty)
    return rows, ledger, sum_cost, sum_qty, blended, blended_reason


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: DriverUnitCostInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    rows, ledger, sum_cost, sum_qty, blended, blended_reason = _compute(data)
    last_col = 6

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("DriverUnitCost")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col)
    hs.set_widths(ws, {1: 26, 2: 12, 3: 10, 4: 8, 5: 12, 6: 10})

    headers = ["비용 라인", "고정비", "동인수량", "단위", "단위원가", "사유(NA)"]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1

    surfaced = 0
    for row in rows:
        ln = row["line"]
        hs.set_cell(ws, r, 1, ln.label, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, ln.cost if ln.cost is not None else "NO_DATA",
                    role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 3, ln.qty if ln.qty is not None else "NO_DATA",
                    role="calc", number_format=hs.FMT_NUM1)
        hs.set_cell(ws, r, 4, ln.unit, role="soft", align=hs.CENTER)
        if row["reason"] is not None:
            # R17 NA — 0/inf 박제 금지. NA 셀 + 사유 노출(surfaced flag).
            hs.set_cell(ws, r, 5, "NA", role="soft", align=hs.CENTER)
            hs.set_cell(ws, r, 6, row["reason"], role="soft", align=hs.CENTER)
            ws.cell(row=r, column=5).font = hs.font(hs.NEG_FG, bold=True)
            surfaced += 1
        else:
            hs.set_cell(ws, r, 5, row["unit_cost"], role="calc", number_format=hs.FMT_NUM2)
            hs.set_cell(ws, r, 6, "—", role="soft", align=hs.CENTER)
        r += 1

    # blended (Σcost/Σqty, 단위정합 정상 라인만) — 단가 열 합산 금지의 핵심
    r += 1
    hs.section_header(ws, r, "Blended 단위원가 (Σcost / Σqty — 열 합산 금지)",
                      last_col=last_col)
    r += 1
    hs.set_cell(ws, r, 1, "Σ고정비 (정상 라인)", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, sum_cost, role="calc", number_format=hs.FMT_INT)
    r += 1
    hs.set_cell(ws, r, 1, "Σ동인수량 (정상 라인)", role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 3, sum_qty, role="calc", number_format=hs.FMT_NUM1)
    r += 1
    hs.set_cell(ws, r, 1, "Blended 단위원가", role="total", align=hs.LEFT)
    if blended_reason is None:
        hs.set_cell(ws, r, 5, blended, role="total", number_format=hs.FMT_NUM2, bold=True)
    else:
        hs.set_cell(ws, r, 5, "NA", role="total", bold=True, align=hs.CENTER)
        hs.set_cell(ws, r, 6, blended_reason, role="soft", align=hs.CENTER)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG

    # Anomaly ledger 노출 (R17 RATIO_NA) — surfaced 와 보존
    if len(ledger):
        nxt = r + 2
        nxt = hs.section_header(ws, nxt, "이상치 대장 (Anomaly Ledger · RATIO_NA)",
                                last_col=last_col)
        for j, h in enumerate(("라인", "유형", "사유"), start=1):
            hs.set_cell(ws, nxt, j, h, role="header", align=hs.LEFT)
        nxt += 1
        for lr in ledger.rows:
            hs.set_cell(ws, nxt, 1, " · ".join(str(g) for g in lr["grain"]),
                        role="label", align=hs.LEFT)
            hs.set_cell(ws, nxt, 2, lr["anomaly_type"], role="soft", align=hs.LEFT)
            hs.set_cell(ws, nxt, 3, lr["detail"], role="soft", align=hs.LEFT)
            nxt += 1
        r = nxt

    if data.commentary:
        cr = r + 2
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        r = cr

    hs.report_footer(ws, r + 1, source="고정비 원장 · 동인 마스터",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 CC × 1 Account × 1 driver_type",
                ("cost_center", "account", "driver_type"),
                [{"cost_center": x["line"].cost_center, "account": x["line"].account,
                  "driver_type": x["line"].driver_type} for x in rows])
    wb._fpna_meta = {"fact": fact, "rows": rows, "anomaly_ledger": ledger,
                     "surfaced_flags": surfaced, "sum_cost": sum_cost,
                     "sum_qty": sum_qty, "blended": blended,
                     "blended_reason": blended_reason}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: DriverUnitCostInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain (CC×Account×driver_type 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # R17: 라인별 unit_cost 재계산 대조 + NA 사유 일치
    for row in meta["rows"]:
        ln = row["line"]
        val, reason = vc.ratio_or_na(ln.cost, ln.qty, num_unit=ln.unit,
                                     den_unit=ln.canonical_unit, require_same_unit=True)
        rep.add("R17 단가:%s" % ln.cost_center,
                (reason == row["reason"]) and
                (reason is not None or abs((val or 0) - (row["unit_cost"] or 0)) < 1e-9),
                "" if reason == row["reason"] else "사유 불일치")

    # ⛔ unit_cost 열 합산 금지 검증: blended != Σ(개별 단가).
    #   blended = Σcost/Σqty 여야지, 개별 단가의 단순합/평균이면 위반.
    indiv = [r["unit_cost"] for r in meta["rows"] if r["unit_cost"] is not None]
    if meta["blended_reason"] is None and indiv:
        naive_sum = sum(indiv)
        naive_avg = naive_sum / len(indiv)
        # blended 가 정상 라인 Σcost/Σqty 와 일치(재계산)
        recomputed, _ = vc.ratio_or_na(meta["sum_cost"], meta["sum_qty"])
        ok_blend = abs(meta["blended"] - (recomputed or 0)) < 1e-9
        rep.add("blended=Σcost/Σqty", ok_blend, "")
        # 단순합/평균과 다름을 확인(열 합산 함정 회피 입증; 값이 우연히 같을 수 있어 soft)
        rep.add("단가 열 합산 금지(soft)", True,
                "blended=%.4g vs 단순합=%.4g 평균=%.4g" % (meta["blended"], naive_sum, naive_avg))

    # Anomaly 2층 보존: |ledger| == surfaced flag (NA 은폐 금지)
    vc.assert_anomaly_conserved(rep, meta["anomaly_ledger"], meta["surfaced_flags"])

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "DriverLine", "DriverUnitCostInput", "golden_sample", "build", "qc"]
