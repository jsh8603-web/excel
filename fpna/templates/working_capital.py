"""
fpna.templates.working_capital — 운전자본 DSO/DPO/DIO 드라이버 (C3 빠진 템플릿).

CFO 질문 "매출채권 회수일(DSO)·재고 회전일(DIO)·매입채무 지급일(DPO) 추세,
운전자본(AR+재고−AP)이 현금을 얼마나 묶나(CCC)?"에 답한다.

- grain = "1행 = 1 기간".
- 드라이버(R17 비율, 0분모 → NA):
    DSO = AR / revenue × days   /  DIO = Inventory / cogs × days
    DPO = AP / cogs × days      /  CCC = DSO + DIO − DPO
- 불변식: working_capital = AR + Inventory − AP (정의 항등, R3 tie).
- 결측/0분모 → RATIO_NA(ledger emit, surfaced 보존). 침묵 0 박제 금지.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.conserve import ConserveSpec
from fpna.dims import AccountingCalendar, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "working_capital"


# --------------------------------------------------------------------------- #
# T4 보존(자문 R2 C6) — 선언형 CONSERVE_SPECS                                  #
#   정의 항등: ΣWC == Σ(AR + 재고 − AP). raw 변은 INPUT(rows_in)에서 직접 합 → #
#   build 의 _compute(period 매핑·결측 처리)을 부르지 않는 N-version. 어느 기간 #
#   WC 가 누락/덧셈오류면 sum_wc 와 불일치로 trip.                              #
# --------------------------------------------------------------------------- #
CONSERVE_SPECS = [
    ConserveSpec(
        "ΣWC = Σ(AR + 재고 − AP)",
        raw_sum_fn=lambda d: sum(p.ar + p.inventory - p.ap for p in d.rows_in),
        reported_key="sum_wc",
    ),
]


@dataclass
class WcPeriod:
    """1 기간의 운전자본 구성요소 + 손익(드라이버 분모)."""
    period: tuple              # (fy, period)
    ar: float                  # 매출채권
    inventory: float           # 재고
    ap: float                  # 매입채무
    revenue: float | None      # 매출(DSO 분모) — 결측 가능
    cogs: float | None         # 매출원가(DIO/DPO 분모)


@dataclass
class WorkingCapitalInput:
    title: str = "운전자본 드라이버 (Working Capital)"
    subtitle: str = "DSO / DIO / DPO · CCC · WC = AR + 재고 − AP"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)
    end: tuple = (2024, 6)
    days_in_period: float = 30.0      # 기간 일수(월=30 단순화)
    rows_in: list = field(default_factory=list)   # list[WcPeriod]
    commentary: list = field(default_factory=list)


def _compute(data: WorkingCapitalInput):
    """기간별 WC + DSO/DIO/DPO/CCC. ledger 동반(R17 NA). 메타 반환."""
    cal = AccountingCalendar(fiscal_year_start_month=data.fy_start_month)
    periods = cal.periods(data.start, data.end)
    # 입력을 period label 로 매핑
    in_by_label = {}
    for w in data.rows_in:
        lbl = cal.period(*w.period).label
        in_by_label[lbl] = w
    ledger = vc.AnomalyLedger()
    rows = []
    surfaced = 0
    days = data.days_in_period
    for p in periods:
        w = in_by_label.get(p.label)
        if w is None:                      # 결측 기간 — 전수 유지(NO_DATA)
            rows.append({"period": p.label, "missing": True})
            continue
        wc = w.ar + w.inventory - w.ap
        dso, r_dso = vc.ratio_or_na(w.ar * days, w.revenue)
        dio, r_dio = vc.ratio_or_na(w.inventory * days, w.cogs)
        dpo, r_dpo = vc.ratio_or_na(w.ap * days, w.cogs)
        for reason, drv in ((r_dso, "DSO"), (r_dio, "DIO"), (r_dpo, "DPO")):
            if reason is not None:
                ledger.add(grain=(p.label, drv), period=p.label,
                           anomaly_type="RATIO_NA", detail=reason)
                surfaced += 1
        ccc = None
        if None not in (dso, dio, dpo):
            ccc = dso + dio - dpo
        rows.append({"period": p.label, "missing": False, "ar": w.ar,
                     "inventory": w.inventory, "ap": w.ap, "wc": wc,
                     "dso": dso, "dio": dio, "dpo": dpo, "ccc": ccc})
    return cal, periods, rows, ledger, surfaced


def golden_sample() -> WorkingCapitalInput:
    """구조 골든 — 정상 5기간 + cogs 결측 1기간(DIO/DPO ZERO_DENOM)."""
    rows = [
        WcPeriod((2024, 1), ar=3_000.0, inventory=2_000.0, ap=1_500.0,
                 revenue=10_000.0, cogs=6_000.0),
        WcPeriod((2024, 2), ar=3_200.0, inventory=2_100.0, ap=1_600.0,
                 revenue=10_500.0, cogs=6_200.0),
        WcPeriod((2024, 3), ar=3_100.0, inventory=2_300.0, ap=1_550.0,
                 revenue=10_200.0, cogs=6_100.0),
        WcPeriod((2024, 4), ar=3_300.0, inventory=2_200.0, ap=1_700.0,
                 revenue=11_000.0, cogs=6_500.0),
        WcPeriod((2024, 5), ar=3_400.0, inventory=2_400.0, ap=1_800.0,
                 revenue=11_200.0, cogs=6_600.0),
        # 6월: cogs=0 → DIO/DPO ZERO_DENOM (R17 NA)
        WcPeriod((2024, 6), ar=3_500.0, inventory=2_500.0, ap=1_900.0,
                 revenue=11_500.0, cogs=0.0),
    ]
    return WorkingCapitalInput(rows_in=rows, start=(2024, 1), end=(2024, 6),
                               commentary=["6월 cogs=0 → DIO/DPO R17 ZERO_DENOM (0/inf 박제 금지)",
                                           "CCC = DSO + DIO − DPO (현금 전환 사이클)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: WorkingCapitalInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, rows, ledger, surfaced = _compute(data)
    last_col = 9

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("WorkingCapital")
    hs.set_widths(ws, {1: 12, 2: 11, 3: 11, 4: 11, 5: 13, 6: 9, 7: 9, 8: 9, 9: 9})

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col)

    headers = ["기간", "매출채권", "재고", "매입채무", "운전자본", "DSO", "DIO", "DPO", "CCC"]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1

    sum_ar = sum_inv = sum_ap = sum_wc = 0.0
    for row in rows:
        hs.set_cell(ws, r, 1, row["period"], role="label", align=hs.LEFT)
        if row["missing"]:
            for j in range(2, last_col + 1):
                hs.set_cell(ws, r, j, "NO_DATA", role="soft", align=hs.CENTER)
            r += 1
            continue
        hs.set_cell(ws, r, 2, row["ar"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 3, row["inventory"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 4, row["ap"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 5, row["wc"], role="total", number_format=hs.FMT_INT)
        for j, key in ((6, "dso"), (7, "dio"), (8, "dpo"), (9, "ccc")):
            v = row[key]
            if v is None:
                hs.set_cell(ws, r, j, "NA", role="soft", align=hs.CENTER)
                ws.cell(row=r, column=j).font = hs.font(hs.NEG_FG, bold=True)
            else:
                hs.set_cell(ws, r, j, v, role="calc", number_format=hs.FMT_NUM1)
        sum_ar += row["ar"]; sum_inv += row["inventory"]
        sum_ap += row["ap"]; sum_wc += row["wc"]
        r += 1

    # _RECON: ΣWC == ΣAR + Σ재고 − ΣAP (정의 항등)
    recon = vc.recon_block(
        n_input=len(data.rows_in), n_output=len(periods),
        src_sum=sum_wc, out_sum=sum_ar + sum_inv - sum_ap,
        completeness="기간 %d 전수 (결측 NO_DATA 유지)" % len(periods),
        accuracy="WC = AR + 재고 − AP (정의 항등)",
        cutoff="드라이버 분모(매출/원가) 결측 → R17 NA",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    # Anomaly ledger 노출 (R17 RATIO_NA)
    if len(ledger):
        nxt = hs.section_header(ws, nxt + 1, "이상치 대장 (Anomaly Ledger · RATIO_NA)",
                                last_col=last_col)
        for j, h in enumerate(("기간·드라이버", "유형", "사유"), start=1):
            hs.set_cell(ws, nxt, j, h, role="header", align=hs.LEFT)
        nxt += 1
        for lr in ledger.rows:
            hs.set_cell(ws, nxt, 1, " · ".join(str(g) for g in lr["grain"]),
                        role="label", align=hs.LEFT)
            hs.set_cell(ws, nxt, 2, lr["anomaly_type"], role="soft", align=hs.LEFT)
            hs.set_cell(ws, nxt, 3, lr["detail"], role="soft", align=hs.LEFT)
            nxt += 1

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="BS 보조원장 · 손익 (AR/재고/AP · 매출/원가)",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 기간", ("period",),
                [{"period": row["period"]} for row in rows])
    wb._fpna_meta = {"cal": cal, "periods": periods, "fact": fact, "rows": rows,
                     "anomaly_ledger": ledger, "surfaced_flags": surfaced,
                     "sum_ar": sum_ar, "sum_inv": sum_inv, "sum_ap": sum_ap,
                     "sum_wc": sum_wc}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: WorkingCapitalInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal = meta["cal"]

    # R8 grain + R1 시간축 전수
    vc.assert_grain(rep, meta["fact"])
    vc.assert_time_ruler(rep, meta["fact"], cal, data.start, data.end, period_key="period")

    # R3 정의 항등: ΣWC == ΣAR + Σ재고 − ΣAP
    vc.assert_tie_out(rep, meta["sum_wc"],
                      meta["sum_ar"] + meta["sum_inv"] - meta["sum_ap"], tol=1e-6,
                      name="R3 wc_identity")

    # 행별 WC 정의 재계산 + CCC = DSO+DIO−DPO 정합
    wc_ok = ccc_ok = True
    for row in meta["rows"]:
        if row.get("missing"):
            continue
        if abs(row["wc"] - (row["ar"] + row["inventory"] - row["ap"])) > 1e-6:
            wc_ok = False
        if row["ccc"] is not None:
            if abs(row["ccc"] - (row["dso"] + row["dio"] - row["dpo"])) > 1e-6:
                ccc_ok = False
    rep.add("WC = AR+재고−AP (행별)", wc_ok, "" if wc_ok else "WC 정의 불일치")
    rep.add("CCC = DSO+DIO−DPO", ccc_ok, "" if ccc_ok else "CCC 정합 불일치")

    # Anomaly 2층 보존: |ledger| == surfaced flag (NA 은폐 금지)
    vc.assert_anomaly_conserved(rep, meta["anomaly_ledger"], meta["surfaced_flags"])

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "WcPeriod", "WorkingCapitalInput", "golden_sample", "build", "qc"]
