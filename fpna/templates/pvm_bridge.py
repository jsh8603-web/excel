"""
fpna.templates.pvm_bridge — Price / Volume / Mix 브리지 (C3 빠진 템플릿).

매출(또는 비용) 변동을 단가(Price)·물량(Volume)·믹스(Mix) 3효과로 분해하는
워터폴. CFO 질문 "매출 +Δ 중 가격 인상분 얼마, 물량 증가분 얼마, 제품 구성
변화(mix)분 얼마?"에 답한다.

- grain = "1행 = 1 product(또는 segment)".
- 분해(Horngren rate/volume + mix):
    ΔPrice  = Σ (P1 − P0) · Q1                (단가 효과, Q1 가중)
    ΔVolume = Σ P0 · (Q0_share·ΣQ1 − Q0)      (총량 효과, 구성 불변)
    ΔMix    = Σ P0 · (Q1 − Q0_share·ΣQ1)      (구성 변화 효과)
  ΔPrice + ΔVolume + ΔMix == ΔTotal (잔차 0, R3 tie). volume/mix 합 == P0 기준
  Σ(Q1−Q0)·P0 이므로 합산 시 항등.
- 불변식: Σ분해 == ΔTotal(=ΣP1Q1 − ΣP0Q0), tol=0.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "pvm_bridge"


@dataclass
class PvmLine:
    """1 product 의 기준(0)/대상(1) 단가·물량."""
    product: str
    label: str
    p0: float            # 기준 단가
    q0: float            # 기준 물량
    p1: float            # 대상 단가
    q1: float            # 대상 물량


@dataclass
class PvmBridgeInput:
    title: str = "Price / Volume / Mix 브리지"
    subtitle: str = "매출 변동 = 단가효과 + 물량효과 + 믹스효과 (Σ분해 = ΔTotal)"
    unit: str = "₩"
    base_label: str = "기준(P0·Q0)"
    end_label: str = "대상(P1·Q1)"
    lines: list = field(default_factory=list)        # list[PvmLine]
    commentary: list = field(default_factory=list)


def _decompose(inp: PvmBridgeInput):
    """라인별·합계 PVM 분해. 반환 dict(메타) — build/qc 공통.

    Volume/Mix 분리(구성 효과 격리):
      total_q0, total_q1 = Σq0, Σq1
      share0_i = q0_i / total_q0  (기준 구성비)
      vol_i  = P0_i · (share0_i·total_q1 − q0_i)   # 구성 불변 가정 총량 변화
      mix_i  = P0_i · (q1_i − share0_i·total_q1)    # 구성 변화
      price_i = (P1_i − P0_i) · q1_i
    Σ(vol+mix) = Σ P0·(q1−q0) (구성항 상쇄) → ΔTotal 의 물량부와 항등.
    """
    total_q0 = sum(ln.q0 for ln in inp.lines)
    total_q1 = sum(ln.q1 for ln in inp.lines)
    price = vol = mix = 0.0
    per_line = []
    base = sum(ln.p0 * ln.q0 for ln in inp.lines)
    end = sum(ln.p1 * ln.q1 for ln in inp.lines)
    for ln in inp.lines:
        share0 = (ln.q0 / total_q0) if total_q0 else 0.0
        expected_q1 = share0 * total_q1            # 구성 불변 시 기대 물량
        pe = (ln.p1 - ln.p0) * ln.q1
        ve = ln.p0 * (expected_q1 - ln.q0)
        me = ln.p0 * (ln.q1 - expected_q1)
        price += pe
        vol += ve
        mix += me
        per_line.append({"product": ln.product, "label": ln.label,
                         "price": pe, "volume": ve, "mix": me})
    decomposed = price + vol + mix
    total = end - base
    return {"base": base, "end": end, "total": total,
            "price": price, "volume": vol, "mix": mix,
            "decomposed": decomposed, "residual": total - decomposed,
            "per_line": per_line}


def golden_sample() -> PvmBridgeInput:
    """구조 골든 — 3 product, 단가↑·물량↑·믹스 이동이 섞여 Σ분해=ΔTotal."""
    lines = [
        PvmLine("A", "제품 A (고가)", p0=100.0, q0=50.0, p1=110.0, q1=60.0),
        PvmLine("B", "제품 B (중가)", p0=60.0, q0=80.0, p1=60.0, q1=70.0),
        PvmLine("C", "제품 C (저가)", p0=30.0, q0=40.0, p1=32.0, q1=55.0),
    ]
    return PvmBridgeInput(lines=lines,
                          commentary=["A 단가 +10% + 물량 +20% = price·volume 동시 기여",
                                      "C 물량 +37.5% = 저가 비중 상승(mix 음의 압력)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: PvmBridgeInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    m = _decompose(data)
    last_col = 4

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("PVM")
    hs.set_widths(ws, {1: 26, 2: 12, 3: 12, 4: 12})

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col)

    # 워터폴 보조영역(cat / base 투명받침 / 표시값 / 누적)
    hs.section_header(ws, r, "변동 브리지 (Bridge)", last_col=last_col)
    r += 1
    for j, h in enumerate(("구간 / 효과", "base", "표시값", "누적"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1

    data_start = r
    cum = m["base"]
    hs.set_cell(ws, r, 1, data.base_label, role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, 0, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 3, m["base"], role="input", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 4, cum, role="calc", number_format=hs.FMT_INT)
    r += 1
    for name, amt in (("단가효과 (Price)", m["price"]),
                      ("물량효과 (Volume)", m["volume"]),
                      ("믹스효과 (Mix)", m["mix"])):
        base = cum if amt >= 0 else cum + amt
        cum += amt
        hs.set_cell(ws, r, 1, name, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, base, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 3, abs(amt), role="input", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 4, cum, role="calc", number_format=hs.FMT_INT)
        r += 1
    hs.set_cell(ws, r, 1, data.end_label, role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 2, 0, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 3, cum, role="total", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 4, cum, role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    data_end = r

    anchor = "%s%d" % (get_column_letter(last_col + 1), data_start - 1)
    hs.add_waterfall(ws, anchor=anchor, data_min_row=data_start, data_max_row=data_end,
                     base_col=2, value_col=3, cat_col=1, title="PVM 브리지")

    # 제품별 분해 표 (price/volume/mix per line)
    pr = data_end + 2
    pr = hs.section_header(ws, pr, "제품별 분해 (Price / Volume / Mix)", last_col=last_col)
    for j, h in enumerate(("제품", "Price", "Volume", "Mix"), start=1):
        hs.set_cell(ws, pr, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    pr += 1
    for pl in m["per_line"]:
        hs.set_cell(ws, pr, 1, pl["label"], role="label", align=hs.LEFT)
        hs.set_cell(ws, pr, 2, pl["price"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, pr, 3, pl["volume"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, pr, 4, pl["mix"], role="calc", number_format=hs.FMT_INT)
        pr += 1
    hs.set_cell(ws, pr, 1, "합계", role="total", align=hs.LEFT)
    hs.set_cell(ws, pr, 2, m["price"], role="total", number_format=hs.FMT_INT)
    hs.set_cell(ws, pr, 3, m["volume"], role="total", number_format=hs.FMT_INT)
    hs.set_cell(ws, pr, 4, m["mix"], role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=pr, column=j).border = hs.BORDER_TOP_STRONG

    # _RECON
    recon = vc.recon_block(
        n_input=len(data.lines), n_output=len(data.lines),
        src_sum=m["total"], out_sum=m["decomposed"],
        completeness="제품 %d 전수 분해 (price·volume·mix)" % len(data.lines),
        accuracy="ΔPrice+ΔVolume+ΔMix == ΔTotal (잔차 0)",
        cutoff="%s → %s" % (data.base_label, data.end_label),
    )
    rec_top = pr + 2
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="매출 원장 · 제품별 단가·물량",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 product", ("product",),
                [{"product": ln.product} for ln in data.lines])
    wb._fpna_meta = {"fact": fact, **m}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: PvmBridgeInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain (product 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # R3 tie: ΔPrice+ΔVolume+ΔMix == ΔTotal (잔차 0)
    vc.assert_tie_out(rep, meta["total"], meta["decomposed"], tol=1e-6,
                      name="R3 pvm_tie")
    rep.add("잔차 0(완전분해)", abs(meta["residual"]) < 1e-6,
            "" if abs(meta["residual"]) < 1e-6 else "잔차=%.6g" % meta["residual"])

    # 재계산 대조 (build 와 동일 _decompose)
    m2 = _decompose(data)
    same = all(abs(meta[k] - m2[k]) < 1e-6 for k in ("price", "volume", "mix", "total"))
    rep.add("분해 재계산 대조", same, "" if same else "재계산 불일치")

    # 워터폴 차트 1개 존재
    rep.add("워터폴 차트", len(wb.active._charts) >= 1,
            "" if wb.active._charts else "차트 없음")
    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "PvmLine", "PvmBridgeInput", "golden_sample", "build", "qc"]
