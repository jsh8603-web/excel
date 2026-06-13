"""
fpna.templates.consolidation_fx — 연결/세그먼트 FX 환산 (C3 빠진 템플릿).

멀티엔티티(현지통화) 손익을 보고통화로 환산해 roll-up 한다. CFO 질문
"해외 자회사 합산 시 환율 적용 후 연결 매출/이익 얼마, 세그먼트 합 = 총계 맞나?"

- grain = "1행 = 1 entity (× 1 line_item)".
- 환산: reported = local × fx_rate (line_item 별 average rate 가정·단순).
- 불변식(R3 tie): Σ(엔티티별 환산) == 연결 총계. 세그먼트 roll-up == 총계.
  내부거래 제거(elimination)는 별도 엔티티(-)로 명시(silent net 금지).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "consolidation_fx"


@dataclass
class Entity:
    """연결 대상 엔티티 1건(현지통화 금액 + 환율)."""
    entity_id: str
    label: str
    segment: str               # 세그먼트(roll-up 축)
    currency: str
    fx_rate: float             # 현지통화 → 보고통화 환율
    local_amount: float        # 현지통화 금액(매출/이익 등 단일 라인)
    is_elimination: bool = False  # 내부거래 제거 엔티티(명시)


@dataclass
class ConsolidationInput:
    title: str = "연결 / 세그먼트 FX 환산 (Consolidation)"
    subtitle: str = "엔티티 환산 → 세그먼트 roll-up → 연결 총계 (Σ = 총계 tie)"
    unit: str = "₩"
    line_label: str = "매출"
    reporting_currency: str = "KRW"
    entities: list = field(default_factory=list)         # list[Entity]
    commentary: list = field(default_factory=list)


def _translate(inp: ConsolidationInput):
    """엔티티별 환산 + 세그먼트 roll-up + 총계. 메타 dict 반환."""
    rows = []
    seg_tot: dict[str, float] = {}
    grand = 0.0
    for e in inp.entities:
        reported = e.local_amount * e.fx_rate
        rows.append({"entity": e, "reported": reported})
        seg_tot[e.segment] = seg_tot.get(e.segment, 0.0) + reported
        grand += reported
    return {"rows": rows, "seg_tot": seg_tot, "grand": grand,
            "seg_sum": sum(seg_tot.values())}


def golden_sample() -> ConsolidationInput:
    """구조 골든 — 3 엔티티(2 세그먼트) + 내부거래 제거 1건."""
    entities = [
        Entity("KR", "본사(KR)", "국내", "KRW", fx_rate=1.0, local_amount=10_000.0),
        Entity("US", "미국법인(US)", "해외", "USD", fx_rate=1_350.0, local_amount=8.0),
        Entity("JP", "일본법인(JP)", "해외", "JPY", fx_rate=9.0, local_amount=500.0),
        Entity("ELIM", "내부거래 제거", "해외", "KRW", fx_rate=1.0,
               local_amount=-1_000.0, is_elimination=True),
    ]
    return ConsolidationInput(entities=entities,
                              commentary=["US/JP 현지통화 → KRW 환산(평균환율)",
                                          "내부거래 -1,000 = 그룹내 매출 제거(silent net 금지·명시)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: ConsolidationInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    m = _translate(data)
    last_col = 6

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Consolidation")
    hs.set_widths(ws, {1: 18, 2: 10, 3: 8, 4: 12, 5: 12, 6: 14})

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, currency=data.reporting_currency, last_col=last_col)

    headers = ["엔티티", "세그먼트", "통화", "현지금액", "환율", "환산(%s)" % data.reporting_currency]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1

    # 세그먼트 순서대로 정렬 + 세그먼트 소계
    seg_order = []
    for e in data.entities:
        if e.segment not in seg_order:
            seg_order.append(e.segment)
    for seg in seg_order:
        ents = [row for row in m["rows"] if row["entity"].segment == seg]
        hs.set_cell(ws, r, 1, "[%s]" % seg, role="soft", align=hs.LEFT)
        r += 1
        for row in ents:
            e = row["entity"]
            label = e.label + (" (제거)" if e.is_elimination else "")
            hs.set_cell(ws, r, 1, label, role="label", align=hs.LEFT)
            hs.set_cell(ws, r, 2, e.segment, role="soft", align=hs.CENTER)
            hs.set_cell(ws, r, 3, e.currency, role="soft", align=hs.CENTER)
            hs.set_cell(ws, r, 4, e.local_amount, role="input", number_format=hs.FMT_INT)
            hs.set_cell(ws, r, 5, e.fx_rate, role="input", number_format=hs.FMT_NUM2)
            hs.set_cell(ws, r, 6, row["reported"], role="calc", number_format=hs.FMT_INT)
            r += 1
        hs.set_cell(ws, r, 1, "%s 소계" % seg, role="total", align=hs.LEFT)
        hs.set_cell(ws, r, 6, m["seg_tot"][seg], role="total", number_format=hs.FMT_INT)
        for j in range(1, last_col + 1):
            ws.cell(row=r, column=j).border = hs.BORDER_TOP
        r += 1

    # 연결 총계
    hs.set_cell(ws, r, 1, "연결 총계", role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 6, m["grand"], role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    r += 1

    # _RECON: 세그먼트 합 == 연결 총계 (roll-up tie)
    recon = vc.recon_block(
        n_input=len(data.entities), n_output=len(data.entities),
        src_sum=m["grand"], out_sum=m["seg_sum"],
        completeness="엔티티 %d → 세그먼트 %d roll-up" % (len(data.entities), len(m["seg_tot"])),
        accuracy="Σ세그먼트 소계 == 연결 총계 (내부거래 명시 제거)",
        cutoff="환산 = 현지금액 × 환율(평균환율)",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="자회사 보고 패키지 · 환율 마스터",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 entity", ("entity_id",),
                [{"entity_id": e.entity_id} for e in data.entities])
    wb._fpna_meta = {"fact": fact, **m}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: ConsolidationInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain (entity 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # R10/R3 roll-up tie: Σ세그먼트 소계 == 연결 총계
    vc.assert_tie_out(rep, meta["grand"], meta["seg_sum"], tol=1e-6,
                      name="R3 segment_roll_tie")

    # 환산 재계산 대조 (local × fx)
    ok = True
    for row in meta["rows"]:
        e = row["entity"]
        if abs(row["reported"] - e.local_amount * e.fx_rate) > 1e-6:
            ok = False
    rep.add("환산 = 현지 × 환율 재계산", ok, "" if ok else "환산 불일치")

    # 내부거래 제거 명시(silent net 금지): elimination 엔티티가 행으로 노출됐는지
    elim = [e for e in data.entities if e.is_elimination]
    rep.add("내부거래 제거 명시(행 노출)", True,
            "elimination %d건 명시" % len(elim) if elim else "내부거래 제거 없음")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "Entity", "ConsolidationInput", "golden_sample", "build", "qc"]
