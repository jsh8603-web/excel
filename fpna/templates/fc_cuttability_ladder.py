"""
fpna.templates.fc_cuttability_ladder — 고정비: 절감 가능성 사다리 (A3).

약정을 time-to-exit(해지가능 시점)로 줄 세워 "지금 끊을 수 있는 비용 vs 잠긴 비용"을
보여준다. CFO 질문 "비상시 3개월 내 줄일 수 있는 고정비 얼마?"에 답한다.

- grain = "1행 = 1 약정(Contract)".
- 등급(rung)은 dims.cuttability_rung — **Contract 속성(약정·만기·notice)이 主**.
  레퍼런스(차용): Horngren committed↔discretionary + ZBB Pyhrr(HBR 1970 Nov-Dec, 확인)
  decision-package. 함정 회피 근거 = ABJ 2003 JAR 41(1):47-63(확인).
  stickiness(ABJ 비대칭)는 ⛔ 단일신호 금지 — 보조 메모로만 부착(등급 안 바꿈).
  ⛔ stickiness 는 집계 추정치 → 라인 단위 인과 해석 금지(ecological fallacy).
- 불변식: earliest_exit = max(report+notice, break) (dims 가 계산), time-to-exit
  단조 정렬, Σ티어별 연환산 == Σ전체 연환산(R3 tie).
- 엣지: notice 미정의 → locked(committed), penalty>절감 → net 음수 표기.
"""
from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import (AccountingCalendar, Contract, Fact, cuttability_rung,
                       CUTTABILITY_RUNGS)
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_cuttability_ladder"

# 연환산 계수
_PER_YEAR = {"monthly": 12, "quarterly": 4, "annual": 1, "one_time": 1}
# 등급 표시 순서(절감 쉬운 → 어려운). committed 가 가장 잠김.
_RUNG_ORDER = ("discretionary", "semi_discretionary", "contractual_locked", "committed")
_RUNG_LABEL = {
    "discretionary": "재량 (즉시 절감)",
    "semi_discretionary": "준재량 (단기 협상)",
    "contractual_locked": "계약잔존 (notice·위약)",
    "committed": "약정고정 (해지 전 불가)",
}


@dataclass
class CuttabilityTerm:
    """약정 1건 + 해지 파라미터(notice/break/penalty) + 보조 stickiness."""
    contract: Contract
    notice_months: int = 0
    break_month: int | None = None
    annual_saving: float = 0.0       # 해지 시 연 절감액(보통 연환산액)
    exit_penalty: float = 0.0        # 해지 위약금(net = saving - penalty)
    stickiness: float | None = None  # ABJ 비대칭(보조 신호) — 등급 판정 X


@dataclass
class CuttabilityInput:
    title: str = "고정비 — 절감 가능성 사다리 (Cuttability Ladder)"
    subtitle: str = "약정별 해지가능 시점(time-to-exit)·절감 vs 위약"
    unit: str = "₩"
    fy_start_month: int = 1
    report_period: tuple = (2024, 1)    # 기준 시점
    terms: list = field(default_factory=list)   # list[CuttabilityTerm]
    commentary: list = field(default_factory=list)


def _annualized(c: Contract) -> float:
    return c.amount_per_period * _PER_YEAR.get(c.recurrence, 1)


def _rows(inp: CuttabilityInput):
    """(as_of, rows) — 약정별 등급·exit·절감. rows: dict per term."""
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    as_of = cal.period(*inp.report_period).cutoff_date
    out = []
    for t in inp.terms:
        res = cuttability_rung(t.contract, as_of=as_of,
                               notice_months=t.notice_months, break_month=t.break_month,
                               stickiness=t.stickiness)
        ann = _annualized(t.contract)
        saving = t.annual_saving or ann
        out.append({
            "contract_id": t.contract.contract_id,
            "counterparty": t.contract.counterparty,
            "annualized": ann, "rung": res["rung"],
            "exit_m": res["earliest_exit_m"],
            "saving": saving, "penalty": t.exit_penalty,
            "net_saving": saving - t.exit_penalty,
            "stickiness_note": res["stickiness_note"],
            "drivers": "; ".join(res["drivers"]),
        })
    return cal, as_of, out


def golden_sample() -> CuttabilityInput:
    """구조 골든 — 4등급이 모두 나오도록 약정 4건."""
    terms = [
        # 만료 임박 + 즉시 해지가능(notice 0, break 0) → discretionary
        CuttabilityTerm(
            Contract("M-01", "6210", "마케팅SaaS", _dt.date(2023, 1, 1),
                     _dt.date(2024, 1, 31), "monthly", 200.0),
            notice_months=0, break_month=0, annual_saving=2_400.0, exit_penalty=0.0,
            stickiness=-0.1),
        # 단기 notice 2M → semi_discretionary
        CuttabilityTerm(
            Contract("V-01", "6020", "차량리스", _dt.date(2023, 1, 1),
                     _dt.date(2026, 6, 30), "monthly", 600.0),
            notice_months=2, annual_saving=7_200.0, exit_penalty=1_000.0,
            stickiness=0.2),
        # 잔여 ~8M(≤12) notice 없음 → contractual_locked
        CuttabilityTerm(
            Contract("L-01", "6010", "사무실임차", _dt.date(2023, 1, 1),
                     _dt.date(2024, 9, 30), "monthly", 1_000.0),
            notice_months=0, annual_saving=12_000.0, exit_penalty=15_000.0,  # penalty>saving → net 음수
            stickiness=0.5),
        # evergreen + notice 미정의 → committed(locked)
        CuttabilityTerm(
            Contract("U-01", "6300", "전력공급", _dt.date(2023, 1, 1),
                     None, "monthly", 800.0),
            notice_months=0, annual_saving=9_600.0, exit_penalty=0.0,
            stickiness=0.8),
    ]
    return CuttabilityInput(report_period=(2024, 1), terms=terms,
                            commentary=["L-01 위약금>연절감 → net 음수(해지 비경제)",
                                        "U-01 evergreen+notice 미정의 → committed(locked)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: CuttabilityInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, as_of, rows = _rows(data)
    last_col = 7

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Cuttability")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, as_of=as_of.isoformat(), last_col=last_col)
    hs.set_widths(ws, {1: 10, 2: 16, 3: 12, 4: 10, 5: 12, 6: 12, 7: 12})

    headers = ["약정", "거래처", "연환산", "해지(월)", "연절감", "위약금", "순절감"]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j <= 2 else hs.CENTER)
    r += 1

    # 등급 순서(절감 쉬운 → 어려운) → time-to-exit 단조 정렬(None=∞ 맨 뒤)
    rung_idx = {g: i for i, g in enumerate(_RUNG_ORDER)}

    def _sort_key(x):
        return (rung_idx.get(x["rung"], 9),
                x["exit_m"] if x["exit_m"] is not None else 9_999)

    rows_sorted = sorted(rows, key=_sort_key)

    tier_tot: dict[str, float] = {}
    cur = None
    for row in rows_sorted:
        if row["rung"] != cur:
            cur = row["rung"]
            hs.section_header(ws, r, "[%s] %s" % (cur, _RUNG_LABEL.get(cur, cur)),
                              last_col=last_col)
            r += 1
        hs.set_cell(ws, r, 1, row["contract_id"], role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, row["counterparty"], role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 3, row["annualized"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 4, row["exit_m"] if row["exit_m"] is not None else "locked",
                    role="calc", number_format=hs.FMT_INT, align=hs.CENTER)
        hs.set_cell(ws, r, 5, row["saving"], role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 6, row["penalty"], role="calc", number_format=hs.FMT_INT)
        # 순절감 음수면 빨강 표기(penalty>saving)
        net_cell = hs.set_cell(ws, r, 7, row["net_saving"], role="calc",
                               number_format=hs.FMT_INT)
        if row["net_saving"] < 0:
            net_cell.font = hs.font(hs.NEG_FG, bold=True)
        tier_tot[row["rung"]] = tier_tot.get(row["rung"], 0.0) + row["annualized"]
        # stickiness 보조 메모(soft, 다음 행)
        if row["stickiness_note"]:
            hs.set_cell(ws, r, 2, row["stickiness_note"], role="soft", align=hs.LEFT)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
            # 메모는 같은 행 우측을 덮지 않도록 별도 행으로 분리
        r += 1

    # 티어별 연환산 소계 + 총계
    r += 1
    hs.section_header(ws, r, "등급별 연환산 소계 (절감 쉬운 → 어려운)", last_col=last_col)
    r += 1
    grand = 0.0
    for g in _RUNG_ORDER:
        if g in tier_tot:
            hs.set_cell(ws, r, 1, _RUNG_LABEL.get(g, g), role="label", align=hs.LEFT)
            hs.set_cell(ws, r, 3, tier_tot[g], role="calc", number_format=hs.FMT_INT)
            grand += tier_tot[g]
            r += 1
    hs.set_cell(ws, r, 1, "총 고정비(연환산)", role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 3, grand, role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG

    if data.commentary:
        cr = r + 2
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        r = cr

    hs.report_footer(ws, r + 1, source="약정 등록부 · 해지 조건표",
                     prepared_by="FP&A", last_col=last_col)
    # 내부 tidy fact (grain = 1 약정)
    fact = Fact("1행 = 1 약정", ("contract_id",),
                [{"contract_id": x["contract_id"], "rung": x["rung"],
                  "exit_m": x["exit_m"], "annualized": x["annualized"]} for x in rows])
    wb._fpna_meta = {"fact": fact, "rows": rows, "rows_sorted": rows_sorted,
                     "tier_tot": tier_tot, "grand": grand,
                     "active_total": sum(x["annualized"] for x in rows)}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: CuttabilityInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain (약정 1건 = 1행, 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # A3 tie: Σ티어 연환산 == Σ전체 active 연환산
    vc.assert_tie_out(rep, meta["active_total"], meta["grand"], tol=1e-6,
                      name="A3 tier_tie")

    # time-to-exit 단조 정렬 검증(등급 순 + exit 오름차순; locked=∞ 맨 뒤)
    rung_idx = {g: i for i, g in enumerate(_RUNG_ORDER)}
    keys = [(rung_idx.get(x["rung"], 9),
             x["exit_m"] if x["exit_m"] is not None else 9_999)
            for x in meta["rows_sorted"]]
    mono = all(keys[i] <= keys[i + 1] for i in range(len(keys) - 1))
    rep.add("time-to-exit 단조정렬", mono, "" if mono else "정렬 깨짐")

    # 등급 enum 유효성
    bad = [x["contract_id"] for x in meta["rows"] if x["rung"] not in CUTTABILITY_RUNGS]
    rep.add("등급 enum 유효", not bad, "" if not bad else "잘못된 rung: " + ", ".join(bad))

    # ⛔ 단일신호 금지 준수 검증: stickiness 가 등급을 바꾸지 않았는지.
    #   stickiness 를 제거하고 재판정해도 동일 등급이어야(보조 신호 원칙).
    cal = AccountingCalendar(fiscal_year_start_month=data.fy_start_month)
    as_of = cal.period(*data.report_period).cutoff_date
    drift = []
    for t in data.terms:
        from fpna.dims import cuttability_rung as _cr
        with_st = _cr(t.contract, as_of=as_of, notice_months=t.notice_months,
                      break_month=t.break_month, stickiness=t.stickiness)["rung"]
        without_st = _cr(t.contract, as_of=as_of, notice_months=t.notice_months,
                         break_month=t.break_month, stickiness=None)["rung"]
        if with_st != without_st:
            drift.append(t.contract.contract_id)
    rep.add("단일신호 금지(stickiness 보조)", not drift,
            "" if not drift else "stickiness 가 등급 변경: " + ", ".join(drift))

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "CuttabilityTerm", "CuttabilityInput", "golden_sample", "build", "qc"]
