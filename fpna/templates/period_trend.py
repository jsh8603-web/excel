"""fpna.templates.period_trend — 기간별 추이(MoM/QoQ/YoY).

깊이(C4): TTM/LTM 롤링합(직전 12기 합, 기간 부족 시 NA) · CAGR(N 기간 정합 —
12기 미만이면 연환산 불가 명시) · 계절지수(평균=1 정규화) 컬럼.
게이트(C5): R1 시간축 전수성(cal_coords 주면 캘린더 연속 ruler 검증).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401
import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna.dims import AccountingCalendar, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors
from fpna.view_contract import assert_time_ruler

TYPE = "period_trend"


@dataclass
class TrendInput:
    title: str = "기간별 추이"
    subtitle: str = ""
    unit: str = "₩mn"
    periods: list = field(default_factory=list)   # ["2025-01", ...]
    series: dict = field(default_factory=dict)     # {metric: [값,...]}
    # --- R1 시간축 전수성(선택) -------------------------------------------
    # cal_coords 를 주면 periods 가 캘린더 연속 ruler 인지 검증(R1). 각 원소 =
    # (fy, period). start/end 사이 결측 기간이 있으면 R1 FAIL. 비우면 skip(라벨 모드).
    cal_coords: list = field(default_factory=list)   # list[(fy, period)]
    fiscal_year_start_month: int = 1
    # --- TTM/계절 윈도(선택) ----------------------------------------------
    ttm_window: int = 12        # 롤링합 윈도(12=TTM/LTM). period 가 분기면 4 등.

    @property
    def has_calendar(self) -> bool:
        return bool(self.cal_coords) and len(self.cal_coords) == len(self.periods)


def golden_sample() -> TrendInput:
    # 캘린더 연속 12기 — R1 전수성 + TTM(12기) + 계절지수 검증 가능.
    # period 라벨은 캘린더 ruler 와 일치해야 R1 통과(FY2025-P01..P12).
    coords = [(2025, p) for p in range(1, 13)]
    periods = ["FY2025-P%02d" % p for p in range(1, 13)]
    return TrendInput(
        title="기간별 추이 (골든샘플)", subtitle="구조 검증용 — 더미", unit="₩mn",
        periods=periods,
        series={"매출": [100, 110, 121, 133, 120, 130, 140, 150, 160, 170, 180, 200],
                "영업이익": [10, 12, 15, 18, 14, 16, 18, 20, 22, 24, 26, 30]},
        cal_coords=coords,
    )


def _first_series(data: TrendInput):
    """(metric, values) — 첫 시리즈(TTM/CAGR/계절 기준)."""
    if not data.series:
        return None, []
    k = next(iter(data.series))
    return k, list(data.series[k])


def build(data: TrendInput, *, mode="create", base_path=None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = hs.safe_sheet_title("Trend")
    n = len(data.periods)
    last_col = 1 + n
    hs.set_widths(ws, {1: 18, **{j: 11 for j in range(2, last_col + 1)}})
    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")

    hs.set_cell(ws, r, 1, "지표 (단위: %s)" % data.unit, role="header", align=hs.LEFT)
    for j, p in enumerate(data.periods, start=2):
        hs.set_cell(ws, r, j, p, role="header")
    r += 1
    data_start = r
    for metric, vals in data.series.items():
        hs.set_cell(ws, r, 1, metric, role="label", align=hs.LEFT)
        for j, v in enumerate(vals, start=2):
            hs.set_cell(ws, r, j, v, role="input", number_format=hs.FMT_INT)
        r += 1
    data_end = r - 1

    # MoM% 블록(첫 시리즈 기준)
    r += 1
    r = hs.section_header(ws, r, "MoM 증감률 (첫 지표)", last_col=last_col)
    hs.set_cell(ws, r, 1, "MoM%", role="label", align=hs.LEFT)
    for t in range(n):
        col = get_column_letter(2 + t)
        if t == 0:
            hs.set_cell(ws, r, 2, "", role="calc")
        else:
            prev = get_column_letter(1 + t)
            hs.set_cell(ws, r, 2 + t,
                        "=IF(%s%d=0,\"\",%s%d/%s%d-1)" % (prev, data_start, col, data_start, prev, data_start),
                        role="calc", number_format=hs.FMT_PCT1)
    r += 2

    # --- TTM/LTM 롤링합(첫 지표) — 직전 W기 합. 기간 부족하면 "" (NA, 0 박제 금지)
    w = data.ttm_window
    r = hs.section_header(ws, r, "TTM/LTM 롤링합 (직전 %d기, 첫 지표)" % w, last_col=last_col)
    hs.set_cell(ws, r, 1, "TTM(%d)" % w, role="label", align=hs.LEFT)
    for t in range(n):
        if t + 1 < w:
            hs.set_cell(ws, r, 2 + t, "", role="calc")     # 윈도 부족 → NA
        else:
            lo = get_column_letter(2 + t - (w - 1))
            hi = get_column_letter(2 + t)
            hs.set_cell(ws, r, 2 + t, "=SUM(%s%d:%s%d)" % (lo, data_start, hi, data_start),
                        role="calc", number_format=hs.FMT_INT, bold=(t == n - 1))
    ttm_row = r
    r += 2

    # --- 계절지수(평균=1) — 첫 지표. 12기(또는 윈도)≥1주기 있을 때만 의미.
    _, fvals = _first_series(data)
    seas = finance.seasonal_indices(fvals, season_len=w) if fvals else []
    r = hs.section_header(ws, r, "계절지수 (평균=1, 첫 지표)", last_col=last_col)
    hs.set_cell(ws, r, 1, "Seasonal idx", role="label", align=hs.LEFT)
    for t in range(n):
        sv = seas[t] if t < len(seas) else 1.0
        hs.set_cell(ws, r, 2 + t, sv, role="calc", number_format=hs.FMT_NUM2)
    r += 2

    # --- CAGR 요약(N 정합) — begin→end, N=관측기간-1(연환산은 period 가 연일 때만)
    _, sv = _first_series(data)
    r = hs.section_header(ws, r, "CAGR 요약 (첫 지표)", last_col=last_col)
    if sv and len(sv) >= 2 and sv[0] > 0:
        n_steps = len(sv) - 1
        g = finance.cagr(sv[0], sv[-1], n_steps)
        hs.set_cell(ws, r, 1, "기간성장률(%d기간, period당)" % n_steps,
                    role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, g if g is not None else "n/a",
                    role="calc", number_format=hs.FMT_PCT1, bold=True)
    else:
        hs.set_cell(ws, r, 1, "기간성장률", role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, "n/a(시점<2 또는 시작≤0)", role="calc")
    r += 2

    hs.add_line_chart(ws, anchor="A%d" % r, data_min_col=2, data_max_col=last_col,
                      data_min_row=data_start, data_max_row=data_end, cat_col=1,
                      title="추이")
    hs.report_footer(ws, r + 16, source="실적 시계열(기간별)",
                     prepared_by="FP&A", last_col=last_col)
    return wb


def qc(wb, data: TrendInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    lens = {len(v) for v in data.series.values()}
    rep.add("기간-시리즈 길이 일치", lens <= {len(data.periods)},
            "" if lens <= {len(data.periods)} else "길이 불일치")

    # --- R1 시간축 전수성: periods 가 캘린더 연속 ruler 인지(silent 갭 차단) ---
    # 갭이 있으면 허위 추세·잘못된 CAGR N(분모 오류). cal_coords 모드에서 강제.
    if data.has_calendar:
        cal = AccountingCalendar(fiscal_year_start_month=data.fiscal_year_start_month)
        fact = Fact("1행 = 1 period", ("period",),
                    [{"period": lbl} for lbl in data.periods])
        assert_time_ruler(rep, fact, cal, data.cal_coords[0], data.cal_coords[-1],
                          period_key="period")
    else:
        rep.add("R1 time_ruler", True, "cal_coords 미입력 — 레거시 라벨 모드(skip)")

    # --- CAGR N 정합: 연환산 단정 금지(period 단위 성장률만, N=시점-1) ---------
    _, sv = _first_series(data)
    if sv and len(sv) >= 2:
        n_steps = len(sv) - 1
        ok_n = (n_steps == len(data.periods) - 1)
        rep.add("CAGR N 정합(N=시점-1)", ok_n,
                "" if ok_n else "N=%d ≠ 시점-1=%d" % (n_steps, len(data.periods) - 1))

    # --- 계절지수 평균 ≈ 1(정규화 정합) ------------------------------------
    _, fvals = _first_series(data)
    seas = finance.seasonal_indices(fvals, season_len=data.ttm_window) if fvals else []
    if seas:
        avg = sum(seas) / len(seas)
        rep.add("계절지수 평균≈1", finance.approx_equal(avg, 1.0, rel=1e-6, abs_=1e-6),
                "" if finance.approx_equal(avg, 1.0, rel=1e-6, abs_=1e-6)
                else "평균=%.6g" % avg)

    rep.add("단위 표기", bool(data.unit))
    return rep


__all__ = ["TYPE", "TrendInput", "golden_sample", "build", "qc"]
