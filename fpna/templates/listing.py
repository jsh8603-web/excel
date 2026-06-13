"""
fpna.templates.listing — 범용 정리표 (문자 식별자 위주 데이터의 착지점).

measure(수치)가 없거나 적은 실무 데이터(식별자·범주·상태 등 문자 컬럼 다수)를 *계산
없이* 회계법인 룩으로 정리한다. 28종 분석표가 measure 시계열을 전제하는 것과 달리,
listing 은 "임의 컬럼을 보기 좋게 나열 + (선택)그룹 소계 + 총계"가 전부다.

검증 성격 = 데이터 충실도(fidelity): 분석 계산이 없으므로 "입력 ↔ 산출 숫자 일치"를
강제한다 — (a) 행 누락·중복 0 (b) 그룹 소계 == 구성요소 합 (c) 총계 == 전체 합.
⛔ 도메인 고유어 하드코딩 0. 골든 = 무의미 구조 더미.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.conserve import ConserveSpec
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "listing"


# T4 보존: 총계(meta) == Σ(입력 행의 첫 number_col). raw 변은 INPUT 직접(build 호출 0).
CONSERVE_SPECS = [
    ConserveSpec(
        "정리표 총계 보존(Σ입력 == 보고 총계)",
        raw_sum_fn=lambda d: sum(_num(r.get(d.number_cols[0])) for r in d.rows)
        if d.number_cols else 0.0,
        reported_key="grand_total",
        tol=0.5,
    ),
]


@dataclass
class ListingInput:
    title: str = "데이터 정리표"
    subtitle: str = ""
    unit: str = ""
    headers: list = field(default_factory=list)       # 표시 컬럼 순서(없으면 rows[0] 키)
    rows: list = field(default_factory=list)          # list[dict]
    number_cols: list = field(default_factory=list)   # 숫자 컬럼(우측정렬·합계 대상)
    group_by: str = ""                                # 소계 기준 dimension(빈=소계 없음)
    show_total: bool = True
    commentary: list = field(default_factory=list)


def _num(v) -> float:
    """표시값 → 합산용 float. 비숫자/결측 = 0(누락은 행 보존으로 별도 검증)."""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace(",", "")
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    try:
        x = float(s)
        return -x if neg else x
    except (TypeError, ValueError):
        return 0.0


def _headers(data: ListingInput) -> list:
    if data.headers:
        return list(data.headers)
    return list(data.rows[0].keys()) if data.rows else []


def golden_sample() -> ListingInput:
    """구조 더미 — 식별자 + 범주 + 수치 1개(재무 의미 없음). 그룹 소계 시연."""
    rows = [
        {"코드": "A-001", "구역": "1구역", "상태": "정상", "수량": 12},
        {"코드": "A-002", "구역": "1구역", "상태": "보류", "수량": 8},
        {"코드": "A-003", "구역": "2구역", "상태": "정상", "수량": 20},
        {"코드": "A-004", "구역": "2구역", "상태": "정상", "수량": 15},
    ]
    return ListingInput(
        title="데이터 정리표 (골든)", subtitle="구조 검증용 더미 — 식별자·범주·수치",
        headers=["코드", "구역", "상태", "수량"], rows=rows,
        number_cols=["수량"], group_by="구역", show_total=True,
        commentary=["measure 가 적은 식별자 위주 데이터의 기본 레이아웃"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def _grouped(data: ListingInput):
    """group_by 기준 정렬된 (group_key, [row,...]) 리스트. group_by 없으면 단일 그룹."""
    if not data.group_by:
        return [(None, list(data.rows))]
    order: list = []
    buckets: dict = {}
    for r in data.rows:
        k = r.get(data.group_by)
        if k not in buckets:
            buckets[k] = []
            order.append(k)
        buckets[k].append(r)
    return [(k, buckets[k]) for k in order]


def build(data: ListingInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    headers = _headers(data)
    last_col = max(1, len(headers))
    numset = set(data.number_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Listing")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    # 폭: 첫 열 넓게, 숫자열 보통
    widths = {1: 22}
    for j, h in enumerate(headers, start=1):
        widths.setdefault(j, 13 if h in numset else 16)
    hs.set_widths(ws, widths)

    # 헤더
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header",
                    align=hs.CENTER if h in numset else hs.LEFT)
    r += 1
    data_start = r

    subtotals: list = []          # [(group_key, {col: sum})]
    col_totals: dict = {c: 0.0 for c in data.number_cols}
    emitted = 0
    for gkey, grp in _grouped(data):
        gsum = {c: 0.0 for c in data.number_cols}
        for row in grp:
            for j, h in enumerate(headers, start=1):
                v = row.get(h)
                if h in numset:
                    hs.set_cell(ws, r, j, v, role="calc", number_format=hs.FMT_INT_DASH,
                                align=hs.RIGHT)
                    gsum[h] += _num(v)
                else:
                    hs.set_cell(ws, r, j, v, role="label", align=hs.LEFT)
            emitted += 1
            r += 1
        # 그룹 소계
        if data.group_by and data.number_cols:
            hs.set_cell(ws, r, 1, "  소계 (%s)" % gkey, role="total", align=hs.LEFT)
            for j, h in enumerate(headers, start=1):
                if h in numset:
                    hs.set_cell(ws, r, j, gsum[h], role="total", number_format=hs.FMT_INT,
                                align=hs.RIGHT, bold=True)
            for j in range(1, last_col + 1):
                ws.cell(row=r, column=j).border = hs.BORDER_TOP
            subtotals.append((gkey, gsum))
            for c in data.number_cols:
                col_totals[c] += gsum[c]
            r += 1
        else:
            for c in data.number_cols:
                col_totals[c] += gsum[c]
    data_end = r - 1

    # zebra(데이터 영역) — 가독성
    if data_end >= data_start and last_col >= 1:
        rng = "A%d:%s%d" % (data_start, get_column_letter(last_col), data_end)
        try:
            hs.apply_zebra(ws, rng)
        except Exception:
            pass

    # 총계
    grand_total = 0.0
    if data.show_total and data.number_cols:
        hs.set_cell(ws, r, 1, "총계", role="total", align=hs.LEFT)
        for j, h in enumerate(headers, start=1):
            if h in numset:
                hs.set_cell(ws, r, j, col_totals[h], role="total",
                            number_format=hs.FMT_INT, align=hs.RIGHT, bold=True)
        for j in range(1, last_col + 1):
            ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
        grand_total = col_totals[data.number_cols[0]]
        r += 1

    if data.commentary:
        cr = r + 1
        cr = hs.section_header(ws, cr, "비고", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        r = cr

    hs.report_footer(ws, r + 1, source="입력 데이터 정리(계산 없음)",
                     prepared_by="FP&A", last_col=last_col)
    wb._fpna_meta = {
        "n_rows_in": len(data.rows), "n_emitted": emitted,
        "subtotals": subtotals, "col_totals": col_totals,
        "grand_total": grand_total, "headers": headers,
    }
    return wb


# --------------------------------------------------------------------------- #
# qc — 데이터 충실도(fidelity) 검증                                            #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: ListingInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # (a) no_silent_drop: 입력 행수 == 산출 행수 (누락·중복 0)
    ok_rows = meta["n_emitted"] == meta["n_rows_in"]
    rep.add("행 보존(no_silent_drop)", ok_rows,
            "" if ok_rows else "입력=%d 산출=%d" % (meta["n_rows_in"], meta["n_emitted"]))

    # (b) 그룹 소계 == 구성요소 합 (number_cols)
    if data.group_by and data.number_cols:
        sub_ok = True
        for gkey, grp in _grouped(data):
            for c in data.number_cols:
                want = sum(_num(row.get(c)) for row in grp)
                got = dict(meta["subtotals"]).get(gkey, {}).get(c)
                if got is None or abs(got - want) > 0.5:
                    sub_ok = False
        rep.add("그룹 소계 == 구성요소 합", sub_ok, "" if sub_ok else "소계 불일치")

    # (c) 총계 == 전체 합 (각 number_col)
    if data.number_cols:
        tot_ok = True
        for c in data.number_cols:
            want = sum(_num(r.get(c)) for r in data.rows)
            if abs(meta["col_totals"][c] - want) > 0.5:
                tot_ok = False
        rep.add("열 총계 == 전체 합", tot_ok, "" if tot_ok else "총계 불일치")

    # 헤더 비어있지 않음
    rep.add("헤더 존재", bool(meta["headers"]), "" if meta["headers"] else "headers/rows 비어있음")
    return rep


__all__ = ["TYPE", "ListingInput", "CONSERVE_SPECS", "golden_sample", "build", "qc"]
