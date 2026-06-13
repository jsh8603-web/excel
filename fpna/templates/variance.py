"""
fpna.templates.variance — 예실(Plan vs Actual) 변동 분석.

산출: 변동표(Plan/Actual/Δ/Δ%) + 워터폴 브리지 + 코멘터리.
빌더는 셀에 Excel 수식(=Actual-Plan 등)을 직접 기입 → 회사 PC에서 라이브 모델.
QC 는 파이썬에서 Δ 를 재계산해 셀 의도와 대조.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna.templates.base import (QCReport, qc_no_formula_errors, qc_totals, qc_sign)
from fpna.view_contract import assert_scenario_aligned, assert_tie_out

TYPE = "variance"


@dataclass
class LineItem:
    name: str
    plan: float
    actual: float
    cost_nature: bool = False   # True=비용성(증가가 악화). 부호규약 플래그.
    level: int = 0              # 들여쓰기 계층
    is_total: bool = False
    key: str = ""               # 모집단 키(예: cost center). 빈 값이면 name 으로 fallback.

    @property
    def pop_key(self) -> str:
        """R9 모집단 정렬용 키. 명시 key 없으면 항목명 사용."""
        return self.key or self.name


@dataclass
class VarianceInput:
    title: str = "예실 변동 분석"
    subtitle: str = ""
    unit: str = "₩mn"
    period: str = ""
    items: list = field(default_factory=list)   # list[LineItem]
    commentary: list = field(default_factory=list)  # list[str]


def golden_sample() -> VarianceInput:
    """재무 의미 없는 구조 골든(빌더·QC 회귀용)."""
    return VarianceInput(
        title="예실 변동 분석 (골든샘플)",
        subtitle="구조 검증용 — 수치는 더미",
        unit="₩mn", period="2025-06",
        items=[
            LineItem("매출", plan=1000, actual=1120),
            LineItem("매출원가", plan=600, actual=640, cost_nature=True),
            LineItem("판관비", plan=200, actual=190, cost_nature=True),
            LineItem("영업이익", plan=200, actual=290, is_total=True),
        ],
        commentary=[
            "매출 +120 (물량 효과 추정)",
            "매출원가 +40 악화 (단가 상승)",
            "판관비 -10 개선",
        ],
    )


def build(data: VarianceInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Variance")
    last_col = 5
    hs.style_sheet(ws, freeze="A6")
    hs.set_widths(ws, {1: 28, 2: 14, 3: 14, 4: 14, 5: 12})

    r = hs.title_block(ws, data.title,
                       (data.subtitle + ("  ·  " + data.period if data.period else ""))
                       .strip(" ·"), last_col=last_col)

    # 헤더 행
    headers = ["항목 (단위: %s)" % data.unit, "계획", "실적", "Δ", "Δ%"]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header",
                    align=hs.LEFT if j == 1 else hs.CENTER)
    header_row = r
    r += 1

    data_start = r
    plan_col, act_col, var_col, pct_col = 2, 3, 4, 5
    total_rows = []
    for it in data.items:
        hs.set_cell(ws, r, 1, it.name, role="label", align=hs.indent(it.level),
                    bold=it.is_total)
        hs.set_cell(ws, r, plan_col, it.plan, role="input", number_format=hs.FMT_INT,
                    bold=it.is_total)
        hs.set_cell(ws, r, act_col, it.actual, role="input", number_format=hs.FMT_INT,
                    bold=it.is_total)
        # Δ = 실적-계획 (수식)
        pl, ac = get_column_letter(plan_col), get_column_letter(act_col)
        hs.set_cell(ws, r, var_col, "=%s%d-%s%d" % (ac, r, pl, r),
                    role="calc", number_format=hs.FMT_INT, bold=it.is_total)
        # Δ% = Δ/|계획|
        va = get_column_letter(var_col)
        hs.set_cell(ws, r, pct_col,
                    "=IF(%s%d=0,\"\",%s%d/ABS(%s%d))" % (pl, r, va, r, pl, r),
                    role="calc", number_format=hs.FMT_PCT1, bold=it.is_total)
        if it.is_total:
            for j in range(1, last_col + 1):
                ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
            total_rows.append(r)
        r += 1
    data_end = r - 1

    # 워터폴 브리지용 보조 영역(차트 데이터) — 항목별 Δ 누적
    bridge_top = r + 1
    hs.section_header(ws, bridge_top, "변동 브리지 (Bridge)", last_col=last_col)
    br = bridge_top + 1
    # 보조 컬럼: cat / base(투명) / value
    hs.set_cell(ws, br, 1, "구간", role="header", align=hs.LEFT)
    hs.set_cell(ws, br, 2, "base", role="header")
    hs.set_cell(ws, br, 3, "값", role="header")
    br += 1
    bridge_data_start = br
    cum = 0.0
    # 시작점 = 계획 합(총계 항목이 있으면 그 계획, 없으면 매출 계획)
    base_total = next((it for it in data.items if it.is_total), data.items[0])
    hs.set_cell(ws, br, 1, "계획", role="label", align=hs.LEFT)
    hs.set_cell(ws, br, 2, 0, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, br, 3, base_total.plan, role="calc", number_format=hs.FMT_INT)
    cum = base_total.plan
    br += 1
    for it in data.items:
        if it.is_total:
            continue
        if it.actual is None or it.plan is None:
            continue  # 미정렬 시나리오 키 — R9 가 잡는다(브리지 누적 제외)
        delta = (it.actual - it.plan)
        signed = -delta if it.cost_nature else delta  # 비용은 부호 반전(이익 기여 기준)
        base = cum if signed >= 0 else cum + signed
        hs.set_cell(ws, br, 1, it.name, role="label", align=hs.LEFT)
        hs.set_cell(ws, br, 2, base, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, br, 3, abs(signed), role="calc", number_format=hs.FMT_INT)
        cum += signed
        br += 1
    hs.set_cell(ws, br, 1, "실적", role="label", align=hs.LEFT)
    hs.set_cell(ws, br, 2, 0, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, br, 3, cum, role="calc", number_format=hs.FMT_INT)
    bridge_data_end = br

    anchor = "%s%d" % (get_column_letter(last_col + 1), bridge_top)
    hs.add_waterfall(ws, anchor=anchor, data_min_row=bridge_data_start,
                     data_max_row=bridge_data_end, base_col=2, value_col=3,
                     cat_col=1, title="예실 브리지")

    # 코멘터리
    if data.commentary:
        cr = bridge_data_end + 2
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1

    # 메타(QC 가 읽을 좌표 기록)
    wb._fpna_meta = {
        "data_start": data_start, "data_end": data_end,
        "plan_col": plan_col, "act_col": act_col, "var_col": var_col,
        "items": data.items,
    }
    return wb


def qc(wb: openpyxl.Workbook, data: VarianceInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    nontotal = [it for it in data.items if not it.is_total]
    totals = [it for it in data.items if it.is_total]

    # --- R9 시나리오 정합: Actual/Budget 가 같은 모집단(키)인지 -----------
    # variance = Scenario 축의 차이(R9)여야 한다. 한쪽 시나리오에만 존재하는
    # 키를 0 으로 버리면(silent default) 변동요인 누락·이중계상의 토양.
    # 각 LineItem 은 plan(=Budget)·actual(=Actual) 두 값을 가지므로
    # "값이 실제로 존재(None 아님)"하는 키 집합을 두 시나리오로 본다.
    # ⚠ R9 를 먼저 돌려 None(미정렬) 키를 노출한 뒤, 산술 검증은 None 을 건넌다.
    actual_keys = [it.pop_key for it in nontotal if it.actual is not None]
    budget_keys = [it.pop_key for it in nontotal if it.plan is not None]
    assert_scenario_aligned(rep, actual_keys, budget_keys)

    # Δ 재계산 대조(파이썬) — 셀 수식은 Excel 가 계산하므로 의도 검증
    for it in data.items:
        if it.actual is None or it.plan is None:
            continue  # R9 가 이미 미정렬로 잡음
        expected = finance.variance(it.actual, it.plan)
        qc_totals("Δ:%s" % it.name, expected, it.actual - it.plan, rep)
    # 총계 항목 = 하위(비총계) 합과 정합한지(부호규약 고려는 생략, 단순 합)
    if totals and nontotal and all(it.actual is not None for it in nontotal):
        # 영업이익 = 매출 - 비용성 항목 합 (cost_nature=True 는 차감)
        calc = sum((-it.actual if it.cost_nature else it.actual) for it in nontotal)
        qc_totals("총계(실적)", calc, totals[0].actual, rep)

    # --- 브리지 합산 tie: 워터폴 component 합 == 양끝(계획→실적) 차 --------
    # build 의 브리지는 계획합에서 항목별 signed Δ 를 누적해 실적합에 닿는다.
    # Σ(signed Δ) == (실적 시작점 − 계획 시작점) 이어야 워터폴이 양끝과 합치.
    # 미합치 = 변동요인 중복/누락(워터폴이 거짓말).
    if totals and nontotal and all(
            (it.actual is not None and it.plan is not None) for it in nontotal):
        base_total = next((it for it in data.items if it.is_total), data.items[0])
        bridge_sum = sum((-(it.actual - it.plan) if it.cost_nature
                          else (it.actual - it.plan)) for it in nontotal)
        endpoint_diff = base_total.actual - base_total.plan
        assert_tie_out(rep, bridge_sum, endpoint_diff, tol=0.0,
                       name="브리지 합산 tie(Σ구간 == 양끝차)")

    # 단위/포맷 일관: unit 문자열 존재
    rep.add("단위 표기", bool(data.unit), "unit 비어있음" if not data.unit else "")
    return rep


__all__ = ["TYPE", "LineItem", "VarianceInput", "golden_sample", "build", "qc"]
