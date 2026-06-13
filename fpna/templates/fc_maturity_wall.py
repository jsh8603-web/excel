"""
fpna.templates.fc_maturity_wall — 고정비: 약정 만기 도래(maturity wall).

CEO/CFO 질문 "12개월 내 만료/갱신 도래 고정비 얼마, 갱신 인상 리스크는?"에 답한다.
계약을 잔여 기간 bucket(≤12 / 13-24 / 25-36 / >36M)으로 줄 세우고, evergreen(만기
없음)은 별도 rolling 으로 격리한다.

- grain = "1행 = 기준시점 1 계약".
- 연환산액 = amount_per_period × (연 발생 횟수). monthly×12 / quarterly×4 / annual×1.
- 불변식(R3 tie): Σ(버킷별 연환산) == Σ(전체 active 연환산). end≠None 은 정확히 1버킷.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Contract
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_maturity_wall"

# 연 발생 횟수(연환산 계수)
_PER_YEAR = {"monthly": 12, "quarterly": 4, "annual": 1, "one_time": 1}
# 잔여기간(월) → bucket
_BUCKETS = ("≤12M", "13-24M", "25-36M", ">36M")
_ROLLING = "rolling(evergreen)"


@dataclass
class MaturityWallInput:
    title: str = "고정비 — 약정 만기 도래 (Maturity Wall)"
    subtitle: str = "잔여 기간 버킷별 연환산 약정"
    unit: str = "₩"
    fy_start_month: int = 1
    report_period: tuple = (2024, 1)        # 기준 시점 (fy, period)
    contracts: list = field(default_factory=list)   # list[dims.Contract]
    commentary: list = field(default_factory=list)


def _bucket(remaining_m: int | None) -> str:
    if remaining_m is None:
        return _ROLLING
    if remaining_m <= 12:
        return "≤12M"
    if remaining_m <= 24:
        return "13-24M"
    if remaining_m <= 36:
        return "25-36M"
    return ">36M"


def _annualized(c: Contract) -> float:
    return c.amount_per_period * _PER_YEAR.get(c.recurrence, 1)


def _rows(inp: MaturityWallInput):
    """(cal, ref_cutoff, rows) — active 계약만. rows: dict per 계약."""
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    ref = cal.period(*inp.report_period)
    ref_cd = ref.cutoff_date
    out = []
    for c in inp.contracts:
        if c.status != "active":
            continue
        if c.end_date is None:
            remaining = None                       # evergreen → rolling
        else:
            remaining = (c.end_date.year - ref_cd.year) * 12 + (c.end_date.month - ref_cd.month)
            if remaining < 0:
                continue                           # 이미 만료
        out.append({
            "contract_id": c.contract_id, "counterparty": c.counterparty,
            "account_id": c.account_id, "remaining_m": remaining,
            "expiry": (c.end_date.isoformat() if c.end_date else "evergreen"),
            "annualized": _annualized(c), "bucket": _bucket(remaining),
        })
    return cal, ref_cd, out


def golden_sample() -> MaturityWallInput:
    import datetime as _dt
    contracts = [
        Contract("L-01", "6010", "빌딩임대A", _dt.date(2023, 1, 1), _dt.date(2024, 9, 30),
                 "monthly", 1_000.0),                       # 잔여 8M → ≤12M
        Contract("L-02", "6010", "빌딩임대B", _dt.date(2023, 1, 1), _dt.date(2026, 6, 30),
                 "monthly", 800.0),                         # 잔여 ~29M → 25-36M
        Contract("V-01", "6020", "차량리스C", _dt.date(2023, 1, 1), _dt.date(2025, 6, 30),
                 "quarterly", 600.0),                       # 잔여 ~17M → 13-24M
        Contract("S-01", "6030", "SaaS-D", _dt.date(2023, 1, 1), None,
                 "monthly", 200.0),                         # evergreen → rolling
    ]
    return MaturityWallInput(report_period=(2024, 1), contracts=contracts,
                             commentary=["L-01 8개월 내 만료 — 갱신 인상 리스크 점검"])


def build(data: MaturityWallInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, ref_cd, rows = _rows(data)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("MaturityWall")
    last_col = 6

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, as_of=ref_cd.isoformat(), last_col=last_col)
    hs.set_widths(ws, {1: 10, 2: 18, 3: 10, 4: 12, 5: 12, 6: 12})

    headers = ["계약", "거래처", "계정", "잔여(월)", "만기", "연환산"]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j <= 2 else hs.CENTER)
    r += 1

    # bucket 순서대로 정렬 + bucket 소계
    order = {b: i for i, b in enumerate(_BUCKETS + (_ROLLING,))}
    rows_sorted = sorted(rows, key=lambda x: (order.get(x["bucket"], 9),
                                              x["remaining_m"] if x["remaining_m"] is not None else 999))
    bucket_tot: dict[str, float] = {}
    cur_bucket = None
    for row in rows_sorted:
        if row["bucket"] != cur_bucket:
            cur_bucket = row["bucket"]
            hs.set_cell(ws, r, 1, "[%s]" % cur_bucket, role="soft", align=hs.LEFT)
            r += 1
        hs.set_cell(ws, r, 1, row["contract_id"], role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, row["counterparty"], role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 3, row["account_id"], role="label", align=hs.CENTER)
        hs.set_cell(ws, r, 4, row["remaining_m"] if row["remaining_m"] is not None else "—",
                    role="calc", number_format=hs.FMT_INT, align=hs.CENTER)
        hs.set_cell(ws, r, 5, row["expiry"], role="calc", align=hs.CENTER)
        hs.set_cell(ws, r, 6, row["annualized"], role="calc", number_format=hs.FMT_INT)
        bucket_tot[row["bucket"]] = bucket_tot.get(row["bucket"], 0.0) + row["annualized"]
        r += 1

    # 버킷별 소계 + 총계
    r += 1
    hs.section_header(ws, r, "버킷별 연환산 소계", last_col=last_col)
    r += 1
    grand = 0.0
    for b in _BUCKETS + (_ROLLING,):
        if b in bucket_tot:
            hs.set_cell(ws, r, 1, b, role="label", align=hs.LEFT)
            hs.set_cell(ws, r, 6, bucket_tot[b], role="calc", number_format=hs.FMT_INT)
            grand += bucket_tot[b]
            r += 1
    hs.set_cell(ws, r, 1, "총계(전체 active)", role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 6, grand, role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG

    if data.commentary:
        cr = r + 2
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        r = cr

    hs.report_footer(ws, r + 1, source="계약 마스터(약정 등록부)",
                     prepared_by="FP&A", last_col=last_col)
    wb._fpna_meta = {"bucket_tot": bucket_tot, "grand": grand,
                     "active_total": sum(x["annualized"] for x in rows)}
    return wb


def qc(wb: openpyxl.Workbook, data: MaturityWallInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    # R3 tie: Σ버킷 == Σ전체 active 연환산
    vc.assert_tie_out(rep, meta["active_total"], meta["grand"], tol=1e-6,
                      name="R3 bucket_tie")
    # 연환산 재계산 대조(build 와 동일 _rows 경로)
    _, _, rows = _rows(data)
    recomputed = sum(x["annualized"] for x in rows)
    rep.add("연환산 재계산", abs(recomputed - meta["active_total"]) < 1e-6, "")
    rep.add("단위 표기", bool(data.unit))
    return rep


__all__ = ["TYPE", "MaturityWallInput", "golden_sample", "build", "qc"]
