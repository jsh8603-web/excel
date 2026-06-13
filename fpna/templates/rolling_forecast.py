"""fpna.templates.rolling_forecast — 롤링 포캐스트(실적 + 전망 갱신)."""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401
import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna.templates.base import QCReport, qc_no_formula_errors
from fpna.dims import AccountingCalendar, Fact
from fpna.view_contract import assert_time_ruler

TYPE = "rolling_forecast"


@dataclass
class ForecastInput:
    title: str = "롤링 포캐스트"
    subtitle: str = ""
    unit: str = "₩mn"
    periods: list = field(default_factory=list)
    actual_until: int = 0           # 인덱스 < actual_until = 실적, 이후 = 전망
    series: dict = field(default_factory=dict)   # {metric: [값,...]}
    # --- R1 시간축 전수성(선택) -------------------------------------------
    # cal_coords 를 주면 periods 가 캘린더 연속 ruler 인지 검증한다(R1).
    # 각 원소 = (fy, period). start/end 사이 결측 기간이 있으면 R1 FAIL.
    # 비우면 R1 은 skip(레거시 Q1~Q4 라벨 호환).
    cal_coords: list = field(default_factory=list)   # list[(fy, period)]
    fiscal_year_start_month: int = 1
    # --- 컷오버 grain(선택) ----------------------------------------------
    # 명시 분류 set 를 주면 그대로, 비우면 actual_until 로 [0,actual_until)=실적.
    # 같은 기간 인덱스가 양쪽에 있으면 actual+forecast 이중계상 → 게이트 FAIL.
    actual_idx: list = field(default_factory=list)    # list[int]
    forecast_idx: list = field(default_factory=list)  # list[int]

    @property
    def has_calendar(self) -> bool:
        return bool(self.cal_coords) and len(self.cal_coords) == len(self.periods)

    def _classes(self):
        """(actual set, forecast set) — 명시 없으면 actual_until 로 파생."""
        if self.actual_idx or self.forecast_idx:
            return set(self.actual_idx), set(self.forecast_idx)
        n = len(self.periods)
        return set(range(self.actual_until)), set(range(self.actual_until, n))


def golden_sample() -> ForecastInput:
    # 캘린더 연속 12 period(FY2025-P01..P12) — R1 전수성 검증 가능.
    # actual_until=6 → P01~P06 실적, P07~P12 전망(컷오버 = P06/P07 경계, 중첩 없음).
    coords = [(2025, p) for p in range(1, 13)]
    periods = ["FY2025-P%02d" % p for p in range(1, 13)]
    return ForecastInput(
        title="롤링 포캐스트 (골든샘플)", subtitle="구조 검증용 — 더미", unit="₩mn",
        periods=periods, actual_until=6,
        series={"매출": [100, 105, 110, 108, 112, 118, 120, 125, 130, 128, 132, 140]},
        cal_coords=coords,
    )


def build(data: ForecastInput, *, mode="create", base_path=None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = hs.safe_sheet_title("Forecast")
    n = len(data.periods); last_col = 1 + n
    hs.set_widths(ws, {1: 18, **{j: 11 for j in range(2, last_col + 1)}})
    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")

    hs.set_cell(ws, r, 1, "지표 (단위: %s)" % data.unit, role="header", align=hs.LEFT)
    for j, p in enumerate(data.periods, start=2):
        tag = " (A)" if (j - 2) < data.actual_until else " (F)"
        hs.set_cell(ws, r, j, str(p) + tag, role="header")
    r += 1
    data_start = r
    for metric, vals in data.series.items():
        hs.set_cell(ws, r, 1, metric, role="label", align=hs.LEFT)
        for j, v in enumerate(vals, start=2):
            # 실적=입력(파랑), 전망=입력(파랑)이되 음영으로 구분
            role = "input"
            fill = None if (j - 2) < data.actual_until else hs.FILL_BAND
            hs.set_cell(ws, r, j, v, role=role, number_format=hs.FMT_INT, fill=fill)
        r += 1
    data_end = r - 1

    # FY 합계
    r += 1
    hs.set_cell(ws, r, 1, "FY 합계", role="total", align=hs.LEFT, bold=True)
    fc = get_column_letter(2); lc = get_column_letter(last_col)
    hs.set_cell(ws, r, 2, "=SUM(%s%d:%s%d)" % (fc, data_start, lc, data_start),
                role="calc", number_format=hs.FMT_INT, bold=True)
    r += 2
    hs.add_line_chart(ws, anchor="A%d" % r, data_min_col=2, data_max_col=last_col,
                      data_min_row=data_start, data_max_row=data_end, cat_col=1,
                      title="실적+전망")
    hs.report_footer(ws, r + 16, source="실적 + 전망(롤링)",
                     prepared_by="FP&A", last_col=last_col)
    return wb


def qc(wb, data: ForecastInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    rep.add("actual_until 범위", 0 <= data.actual_until <= len(data.periods),
            "actual_until=%d" % data.actual_until)

    # --- R1 시간축 전수성: periods 가 캘린더 연속 ruler 인지 ----------------
    # cal_coords 가 있으면 첫~끝 사이 결측 기간(silent 갭)을 잡는다. 갭 = 허위
    # 추세·잘못된 컷오버 위치. (레거시 라벨 모드는 skip.)
    if data.has_calendar:
        cal = AccountingCalendar(fiscal_year_start_month=data.fiscal_year_start_month)
        fact = Fact(grain="1행 = 1 period", grain_keys=("period",),
                    rows=[{"period": lbl} for lbl in data.periods])
        start, end = data.cal_coords[0], data.cal_coords[-1]
        assert_time_ruler(rep, fact, cal, start, end, period_key="period")
    else:
        rep.add("R1 time_ruler", True, "cal_coords 미입력 — 레거시 라벨 모드(skip)")

    # --- 컷오버 grain: 한 기간 = actual XOR forecast(이중계상 금지) ---------
    n = len(data.periods)
    a_set, f_set = data._classes()
    overlap = sorted(a_set & f_set)
    union = a_set | f_set
    ok_partition = (not overlap) and (union == set(range(n)))
    detail = []
    if overlap:
        detail.append("중첩 기간(actual+forecast 이중계상): "
                      + ", ".join(data.periods[i] for i in overlap[:6]))
    if union != set(range(n)):
        missing = sorted(set(range(n)) - union)
        detail.append("미분류 기간: " + ", ".join(data.periods[i] for i in missing[:6]))
    rep.add("컷오버 grain(actual XOR forecast)", ok_partition, "; ".join(detail))

    # --- 컷오버 tie: 실적 마지막 기간 == 전망 첫 기간 직전(연속, 갭 없음) ---
    # 실적 인덱스 max + 1 == 전망 인덱스 min 이어야 컷오버가 끊김 없이 이어진다.
    if a_set and f_set:
        cut_ok = (max(a_set) + 1) == min(f_set)
        rep.add("컷오버 연속(실적끝+1 == 전망시작)", cut_ok,
                "" if cut_ok else "실적끝=%d 전망시작=%d (컷오버 갭/중첩)"
                % (max(a_set), min(f_set)))

    rep.add("단위 표기", bool(data.unit))
    return rep


__all__ = ["TYPE", "ForecastInput", "golden_sample", "build", "qc"]
