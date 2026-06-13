"""
fpna.templates.fc_forward_da — 고정비: 미래 감가상각 투영 (Forward D&A, A5).

확정 capex / 아직 미가동(in-service 대기) 자산의 향후 D&A 를 미래 기간에 투영한다.
CFO 질문 "내년 감가상각 얼마 잡혀 있나, 신규 capex 가동되면 P&L 영향은?"에 답한다.

- grain = "1행 = 1 Asset(확정 capex/미가동분) × 1 미래 기간".
- 불변식: Σ미래 D&A == cost − salvage (전수 투영 합 = 감가대상액, R3 tie).
  in_service 이후만 상각(그 전 dep=0). 1차월 proration(일할).
- 엣지: in_service 지연 → flag. C11(fc_depreciation actuals)과 배타(이중계상 금지) —
  여긴 *미래(forecast)*만, 과거 실적은 fc_depreciation_schedule 소관.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Asset, Fact, DOMAIN_LABEL
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_forward_da"


@dataclass
class ForwardCapex:
    """확정 capex 1건(미가동 가능). dims.Asset + 1차월 proration."""
    asset: Asset
    first_period_factor: float = 1.0   # 가동 1차월 일할(0<f≤1)


@dataclass
class ForwardDaInput:
    title: str = "고정비 — 미래 감가상각 투영 (Forward D&A)"
    subtitle: str = "확정 capex·미가동 자산의 향후 D&A (forecast 전용)"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)        # 투영 시작 기간(미래)
    end: tuple = (2026, 12)
    capex: list = field(default_factory=list)   # list[ForwardCapex]
    commentary: list = field(default_factory=list)


def _ordinal(fyp: tuple) -> int:
    return fyp[0] * 12 + (fyp[1] - 1)


def _build_fact(inp: ForwardDaInput):
    """Asset × 미래기간 전수 tidy Fact. in_service 이후만 dep, 1차월 proration."""
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    periods = cal.periods(inp.start, inp.end)
    first_ord, last_ord = periods[0].ordinal, periods[-1].ordinal
    rows: list[dict] = []
    delayed = []           # in_service 가 투영창 밖(지연) flag
    for fc in inp.capex:
        a = fc.asset
        in_ord = _ordinal(a.in_service) if a.in_service else first_ord
        if a.in_service and in_ord > last_ord:
            delayed.append(a.asset_no)      # 투영창 내 가동 없음
        # 가동 시점 0-base 인덱스 기준 스케줄(1차월 proration)
        n = max(last_ord - in_ord + 1, a.life_months, 0)
        sch = finance.depreciation_schedule_ext(
            a.acq_cost, a.salvage, a.life_months, n_periods=n, start_index=0,
            first_period_factor=fc.first_period_factor)
        for p in periods:
            offset = p.ordinal - in_ord
            if offset < 0:
                opening, dep, closing = a.acq_cost, 0.0, a.acq_cost
            elif offset < len(sch):
                opening, dep, closing = sch[offset]
            else:
                opening, dep, closing = a.salvage, 0.0, a.salvage
            rows.append({"asset_no": a.asset_no, "period": p.label, "domain": a.domain,
                         "opening": opening, "dep": dep, "closing": closing})
    fact = Fact("1행 = 1 Asset × 1 미래 기간", ("asset_no", "period"), rows)
    return cal, periods, fact, delayed


def golden_sample() -> ForwardDaInput:
    """구조 골든 — 즉시가동 1건 + 지연가동(투영창 내) 1건 + 1차월 일할 1건."""
    capex = [
        ForwardCapex(Asset("CX-01", "설비공급A", "fixed_parts",
                           acq_cost=24_000, life_months=24, salvage=0.0,
                           in_service=(2024, 1))),
        ForwardCapex(Asset("CX-02", "차량리스B", "vehicle",
                           acq_cost=36_000, life_months=36, salvage=0.0,
                           in_service=(2024, 7)),
                     first_period_factor=0.5),   # 7월 반월 가동
    ]
    return ForwardDaInput(capex=capex, start=(2024, 1), end=(2026, 12),
                          commentary=["CX-02 7월 가동(1차월 반월) — 그 전 D&A 0",
                                      "actuals(과거 상각)는 fc_depreciation_schedule 소관 (이중계상 금지)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: ForwardDaInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, fact, delayed = _build_fact(data)
    nP = len(periods)
    last_col = 1 + nP + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("ForwardDA")

    r = hs.title_block(ws, data.title,
                       (data.subtitle + ("  ·  단위 " + data.unit if data.unit else "")).strip(" ·"),
                       last_col=last_col)
    header_row = r
    hs.style_sheet(ws, freeze="B%d" % (header_row + 1))
    hs.set_widths(ws, {1: 22, last_col: 14})
    for c in range(2, last_col):
        ws.column_dimensions[get_column_letter(c)].width = 10

    hs.set_cell(ws, r, 1, "확정 capex (분야)", role="header", align=hs.LEFT)
    for j, p in enumerate(periods, start=2):
        hs.set_cell(ws, r, j, p.label, role="header")
    hs.set_cell(ws, r, last_col, "합계", role="header")
    r += 1

    by_asset: dict[str, dict[str, float]] = {}
    for row in fact.rows:
        by_asset.setdefault(row["asset_no"], {})[row["period"]] = row["dep"]

    data_start = r
    for fc in data.capex:
        a = fc.asset
        label = "%s (%s)" % (a.asset_no, DOMAIN_LABEL.get(a.domain, a.domain))
        hs.set_cell(ws, r, 1, label, role="label", align=hs.LEFT)
        for j, p in enumerate(periods, start=2):
            hs.set_cell(ws, r, j, by_asset[a.asset_no][p.label], role="calc",
                        number_format=hs.FMT_INT_DASH)
        c0, c1 = get_column_letter(2), get_column_letter(last_col - 1)
        hs.set_cell(ws, r, last_col, "=SUM(%s%d:%s%d)" % (c0, r, c1, r),
                    role="calc", number_format=hs.FMT_INT, bold=True)
        r += 1
    data_end = r - 1

    hs.set_cell(ws, r, 1, "합계 (미래 D&A)", role="total", align=hs.LEFT)
    for j in range(2, last_col + 1):
        cl = get_column_letter(j)
        hs.set_cell(ws, r, j, "=SUM(%s%d:%s%d)" % (cl, data_start, cl, data_end),
                    role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    total_row = r

    # _RECON: Σ미래 D&A(투영창 내) — 투영창이 내용연수를 다 못 담으면 ≤ depreciable.
    projected = sum(row["dep"] for row in fact.rows)
    depreciable = sum(fc.asset.acq_cost - fc.asset.salvage for fc in data.capex)
    recon = vc.recon_block(
        n_input=len(data.capex), n_output=len(fact.rows),
        src_sum=depreciable, out_sum=projected,
        completeness="capex %d × 기간 %d 전수" % (len(data.capex), nP),
        accuracy="정액법 미래 투영 (in_service 이후만, 1차월 proration)",
        cutoff="forecast 전용 — actuals 는 fc_depreciation_schedule(이중계상 금지)",
    )
    rec_top = total_row + 2
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    # in_service 지연 flag
    if delayed:
        nxt = hs.section_header(ws, nxt + 1, "가동 지연 flag (투영창 내 미가동)",
                                last_col=last_col)
        hs.set_cell(ws, nxt, 1, "지연: " + ", ".join(delayed), role="soft", align=hs.LEFT)
        ws.merge_cells(start_row=nxt, start_column=1, end_row=nxt, end_column=last_col)
        nxt += 1

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1

    wb._fpna_meta = {"cal": cal, "periods": periods, "fact": fact,
                     "projected": projected, "depreciable": depreciable,
                     "delayed": delayed}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: ForwardDaInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal, periods, fact = meta["cal"], meta["periods"], meta["fact"]

    # R8 grain + R1 시간축 전수
    vc.assert_grain(rep, fact)
    vc.assert_time_ruler(rep, fact, cal, data.start, data.end, period_key="period")
    vc.assert_no_silent_drop(rep, fact, expected_n=len(data.capex) * len(periods))

    # A5 tie: 투영창이 내용연수 전체를 담으면 Σ미래 D&A == cost−salvage.
    #   담지 못하면(투영창 < 잔여내용연수) projected < depreciable 이 정상 → 자산별 판정.
    first_ord, last_ord = periods[0].ordinal, periods[-1].ordinal
    for fc in data.capex:
        a = fc.asset
        in_ord = _ordinal(a.in_service) if a.in_service else first_ord
        got = sum(r["dep"] for r in fact.rows if r["asset_no"] == a.asset_no)
        covered_months = last_ord - in_ord + 1
        if covered_months >= a.life_months:
            # 투영창이 내용연수 전체 포함 → 전수 합 = depreciable
            expected = a.acq_cost - a.salvage
            ok = finance.approx_equal(got, expected, abs_=1e-6)
            rep.add("A5 미래D&A합:%s" % a.asset_no, ok,
                    "" if ok else "투영=%.6g 기대=%.6g" % (got, expected))
        else:
            # 부분 커버 → projected ≤ depreciable (음수/과상각 아님)
            ok = (got <= a.acq_cost - a.salvage + 1e-6) and (got >= -1e-6)
            rep.add("A5 부분투영 정합:%s" % a.asset_no, ok,
                    "" if ok else "투영=%.6g > 감가대상=%.6g (과상각)"
                    % (got, a.acq_cost - a.salvage))

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "ForwardCapex", "ForwardDaInput", "golden_sample", "build", "qc"]
