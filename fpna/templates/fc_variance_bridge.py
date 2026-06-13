"""
fpna.templates.fc_variance_bridge — 고정비 Variance Bridge(Walk).

전기/예산 대비 고정비 차이를 5요인으로 분해하는 워터폴 — 보고의 본체.
  계약변경 / 신규자산 가동 / 일회성(원상복구·중도해지) / 물가·환율 연동 / 잔차

불변식:
  - bridge 합 == (종료 - 시작) (R3 tie_out, tol=0). 잔차가 차이를 흡수한다.
  - variance 는 시나리오 축의 차이로만(R9): 비교 모집단(cost center)이 정렬돼야 한다.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_variance_bridge"

# 고정비 brige 표준 5요인(보고 일관성).
STANDARD_FACTORS = (
    "계약변경", "신규자산 가동", "일회성(원상복구·중도해지)", "물가·환율 연동", "잔차",
)


@dataclass
class BridgeFactor:
    name: str
    amount: float    # 기여(+증가 / -감소)
    # C12: timing(기간귀속·phasing) vs permanent(구조적·런레이트) 분류.
    #   "residual" = 잔차 버킷(명시). 미지정 = "permanent"(보수적 기본).
    kind: str = "permanent"   # "timing" | "permanent" | "residual"


_FACTOR_KINDS = ("timing", "permanent", "residual")


@dataclass
class FixedCostBridgeInput:
    title: str = "고정비 Variance Bridge"
    subtitle: str = "예산 대비 고정비 변동 요인 분해"
    unit: str = "₩mn"
    base_label: str = "예산"
    end_label: str = "실적"
    base_value: float = 0.0
    end_value: float = 0.0
    factors: list = field(default_factory=list)        # list[BridgeFactor]
    actual_keys: list = field(default_factory=list)    # R9 모집단(cost center 키)
    budget_keys: list = field(default_factory=list)
    commentary: list = field(default_factory=list)


def golden_sample() -> FixedCostBridgeInput:
    """구조 골든 — 잔차가 (end-base) 차이를 정확히 흡수해 tie_out PASS."""
    base = 1_000.0
    factors = [
        BridgeFactor("계약변경", +30.0, kind="permanent"),
        BridgeFactor("신규자산 가동", +45.0, kind="permanent"),
        BridgeFactor("일회성(원상복구·중도해지)", -20.0, kind="timing"),
        BridgeFactor("물가·환율 연동", +12.0, kind="permanent"),
        BridgeFactor("잔차", +3.0, kind="residual"),
    ]
    end = base + sum(f.amount for f in factors)
    keys = [("CC10",), ("CC20",), ("CC30",)]
    return FixedCostBridgeInput(
        base_value=base, end_value=end, factors=factors,
        actual_keys=list(keys), budget_keys=list(keys),
        commentary=[
            "신규자산 가동 +45 = 설비 1건 정상 가동(런레이트 진입)",
            "일회성 -20 = 임차 중도해지 원상복구비 환입",
        ],
    )


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: FixedCostBridgeInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Bridge")
    last_col = 4
    hs.set_widths(ws, {1: 28, 2: 12, 3: 12, 4: 12})

    r = hs.title_block(ws, data.title,
                       (data.subtitle + ("  ·  단위 " + data.unit if data.unit else "")).strip(" ·"),
                       last_col=last_col)
    hs.style_sheet(ws, freeze="A%d" % (r + 1))

    # 워터폴 보조영역(cat / base 투명받침 / value) + 누적 표시
    hs.section_header(ws, r, "변동 브리지 (Bridge)", last_col=last_col)
    r += 1
    for j, h in enumerate(("구간 / 요인", "base", "표시값", "누적"), start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1

    data_start = r
    cum = data.base_value
    # 시작
    hs.set_cell(ws, r, 1, data.base_label, role="label", align=hs.LEFT)
    hs.set_cell(ws, r, 2, 0, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 3, data.base_value, role="input", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 4, cum, role="calc", number_format=hs.FMT_INT)
    r += 1
    # 요인들 (floating bar)
    for f in data.factors:
        signed = f.amount
        base = cum if signed >= 0 else cum + signed
        cum += signed
        hs.set_cell(ws, r, 1, f.name, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, base, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 3, abs(signed), role="input", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 4, cum, role="calc", number_format=hs.FMT_INT)
        r += 1
    # 종료
    hs.set_cell(ws, r, 1, data.end_label, role="total", align=hs.LEFT)
    hs.set_cell(ws, r, 2, 0, role="calc", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 3, cum, role="total", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, 4, cum, role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    data_end = r

    # 워터폴 차트 (base 투명 + value 채움). 보조 base_col=2, value_col=3.
    anchor = "%s%d" % (get_column_letter(last_col + 1), data_start - 1)
    hs.add_waterfall(ws, anchor=anchor, data_min_row=data_start, data_max_row=data_end,
                     base_col=2, value_col=3, cat_col=1, title="고정비 브리지")

    # _RECON
    out_sum = data.base_value + sum(f.amount for f in data.factors)
    recon = vc.recon_block(
        n_input=len(data.factors) + 2, n_output=len(data.factors) + 2,
        src_sum=data.end_value, out_sum=out_sum,
        completeness="5요인 분해 (계약변경·신규자산·일회성·물가환율·잔차)",
        accuracy="bridge 합 = 종료-시작 (잔차 흡수)",
        cutoff="%s → %s" % (data.base_label, data.end_label),
    )
    rec_top = data_end + 2
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)

    # C12: timing vs permanent 분류 소계 + 잔차 버킷 명시.
    timing_sum = sum(f.amount for f in data.factors if f.kind == "timing")
    perm_sum = sum(f.amount for f in data.factors if f.kind == "permanent")
    resid_sum = sum(f.amount for f in data.factors if f.kind == "residual")
    kt = rec_top + len(recon) + 2
    kt = hs.section_header(ws, kt, "변동 성격 분류 (Timing vs Permanent)", last_col=last_col)
    for label, val in (("timing(기간귀속)", timing_sum),
                       ("permanent(구조적·런레이트)", perm_sum),
                       ("residual(잔차)", resid_sum)):
        hs.set_cell(ws, kt, 1, label, role="label", align=hs.LEFT)
        hs.set_cell(ws, kt, 3, val, role="calc", number_format=hs.FMT_INT)
        kt += 1

    if data.commentary:
        cr = kt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1

    wb._fpna_meta = {"out_sum": out_sum, "end_value": data.end_value,
                     "timing_sum": timing_sum, "perm_sum": perm_sum,
                     "resid_sum": resid_sum}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: FixedCostBridgeInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)

    # R3 tie_out: base + Σfactors == end_value (tol=0)
    out_sum = data.base_value + sum(f.amount for f in data.factors)
    vc.assert_tie_out(rep, data.end_value, out_sum, tol=0.0, name="R3 bridge_tie_out")

    # R9 시나리오 모집단 정렬 (cost center 키)
    if data.actual_keys or data.budget_keys:
        vc.assert_scenario_aligned(rep, data.actual_keys, data.budget_keys)

    # C12-1: kind 값 유효성(timing/permanent/residual) — 오타 = 분류 누수.
    bad_kind = [f.name for f in data.factors if f.kind not in _FACTOR_KINDS]
    rep.add("C12 factor kind 유효", not bad_kind,
            "" if not bad_kind else "잘못된 kind: " + ", ".join(bad_kind))

    # C12-2: 잔차 버킷 명시 — 요인이 있으면 residual 버킷이 정확히 정의돼야(은닉 잔차 금지).
    resid_factors = [f for f in data.factors if f.kind == "residual"]
    rep.add("C12 잔차 버킷 명시", (not data.factors) or len(resid_factors) >= 1,
            "" if (not data.factors or resid_factors) else "잔차 버킷(kind=residual) 부재")

    # C12-3: 잔차 유의성 flag(soft) — 잔차/총변동 > 임계면 요인 분해 불충분 경고.
    #   R3 tie_out 이 균형은 보장하므로 passed 를 깎지 않고 노출만(자문 R3: 정직한 emit).
    resid = wb._fpna_meta.get("resid_sum", 0.0)
    total_move = abs(data.end_value - data.base_value)
    if total_move > 0:
        ratio = abs(resid) / total_move
        rep.add("C12 잔차 유의(soft)", True,
                "잔차/총변동=%.1f%% (>30%% 시 요인 분해 보강 권고)" % (ratio * 100.0)
                if ratio > 0.30 else "잔차 비중 %.1f%%" % (ratio * 100.0))

    # 워터폴 차트 1개 존재
    rep.add("워터폴 차트", len(wb.active._charts) >= 1,
            "" if wb.active._charts else "차트 없음")
    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "BridgeFactor", "FixedCostBridgeInput", "STANDARD_FACTORS",
           "golden_sample", "build", "qc"]
