"""
fpna.templates.dupont_roic — ROIC / DuPont 분해 (C3 빠진 템플릿).

수익성을 곱셈 분해로 진단한다. CFO/IR 질문 "ROE 가 마진 때문인지, 자산회전
때문인지, 레버리지 때문인지? ROIC 가 WACC 를 넘나(가치창출)?"에 답한다.

- grain = "1행 = 1 기간(또는 엔티티)".
- DuPont(3-step): ROE = 순이익률 × 자산회전율 × 재무레버리지
    = (NI/Sales) × (Sales/Assets) × (Assets/Equity).
- ROIC = NOPAT / Invested Capital = EBIT×(1−tax) / (Debt+Equity).
- 불변식(곱셈 정합): 세 요소의 곱 == ROE 직접계산값 (R3 tie, 곱셈 닫힘).
  0분모 → R17 NA(0/inf 박제 금지). ROIC−WACC spread 로 가치창출 표시.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "dupont_roic"


@dataclass
class DupontPeriod:
    """1 기간의 DuPont/ROIC 원천 계정."""
    period: str
    net_income: float
    sales: float
    assets: float
    equity: float
    ebit: float
    debt: float
    tax_rate: float = 0.22
    wacc: float | None = None      # 가치창출 spread 비교용(없으면 spread skip)


@dataclass
class DupontInput:
    title: str = "ROIC / DuPont 분해 (DuPont)"
    subtitle: str = "ROE = 순이익률 × 자산회전 × 레버리지 · ROIC vs WACC"
    unit: str = "₩"
    rows_in: list = field(default_factory=list)    # list[DupontPeriod]
    commentary: list = field(default_factory=list)


def _compute(data: DupontInput):
    """기간별 DuPont 3요소 + ROE + ROIC + spread. ledger 동반(R17)."""
    ledger = vc.AnomalyLedger()
    rows = []
    surfaced = 0
    for d in data.rows_in:
        npm, r1 = vc.ratio_or_na(d.net_income, d.sales)            # 순이익률
        ato, r2 = vc.ratio_or_na(d.sales, d.assets)                # 자산회전율
        lev, r3 = vc.ratio_or_na(d.assets, d.equity)               # 재무레버리지
        roe, r4 = vc.ratio_or_na(d.net_income, d.equity)           # ROE 직접
        nopat = d.ebit * (1.0 - d.tax_rate)
        roic, r5 = vc.ratio_or_na(nopat, d.debt + d.equity)        # ROIC
        for reason, name in ((r1, "순이익률"), (r2, "자산회전율"), (r3, "레버리지"),
                             (r4, "ROE"), (r5, "ROIC")):
            if reason is not None:
                ledger.add(grain=(d.period, name), period=d.period,
                           anomaly_type="RATIO_NA", detail=reason)
                surfaced += 1
        product = None
        if None not in (npm, ato, lev):
            product = npm * ato * lev                              # 곱셈 분해
        spread = (roic - d.wacc) if (roic is not None and d.wacc is not None) else None
        rows.append({"period": d.period, "npm": npm, "ato": ato, "lev": lev,
                     "roe": roe, "roe_product": product, "roic": roic,
                     "nopat": nopat, "wacc": d.wacc, "spread": spread})
    return rows, ledger, surfaced


def golden_sample() -> DupontInput:
    """구조 골든 — 정상 3기간 + equity=0(레버리지/ROE ZERO_DENOM) 1기간."""
    rows = [
        DupontPeriod("FY2023", net_income=800.0, sales=10_000.0, assets=8_000.0,
                     equity=4_000.0, ebit=1_200.0, debt=3_000.0, tax_rate=0.22, wacc=0.08),
        DupontPeriod("FY2024", net_income=1_000.0, sales=11_000.0, assets=8_500.0,
                     equity=4_500.0, ebit=1_400.0, debt=3_200.0, tax_rate=0.22, wacc=0.08),
        DupontPeriod("FY2025", net_income=1_150.0, sales=12_000.0, assets=9_000.0,
                     equity=5_000.0, ebit=1_550.0, debt=3_400.0, tax_rate=0.22, wacc=0.085),
        # equity=0 → 레버리지/ROE ZERO_DENOM (R17 NA)
        DupontPeriod("FY2026E", net_income=900.0, sales=12_500.0, assets=9_500.0,
                     equity=0.0, ebit=1_600.0, debt=3_600.0, tax_rate=0.22, wacc=0.085),
    ]
    return DupontInput(rows_in=rows,
                       commentary=["FY2026E equity=0 → 레버리지/ROE R17 ZERO_DENOM",
                                   "ROE = 순이익률 × 자산회전 × 레버리지 (곱셈 닫힘)",
                                   "ROIC − WACC > 0 = 가치창출(EVA 양)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: DupontInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    rows, ledger, surfaced = _compute(data)
    last_col = 8

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("DuPont")
    hs.set_widths(ws, {1: 12, 2: 11, 3: 11, 4: 11, 5: 10, 6: 10, 7: 10, 8: 12})

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col)

    headers = ["기간", "순이익률", "자산회전", "레버리지", "ROE(곱)", "ROE(직접)",
               "ROIC", "ROIC−WACC"]
    for j, h in enumerate(headers, start=1):
        hs.set_cell(ws, r, j, h, role="header", align=hs.LEFT if j == 1 else hs.CENTER)
    r += 1

    def _pct(row, col, key, fmt=hs.FMT_PCT1):
        v = row[key]
        if v is None:
            hs.set_cell(ws, r, col, "NA", role="soft", align=hs.CENTER)
            ws.cell(row=r, column=col).font = hs.font(hs.NEG_FG, bold=True)
        else:
            hs.set_cell(ws, r, col, v, role="calc", number_format=fmt)

    for row in rows:
        hs.set_cell(ws, r, 1, row["period"], role="label", align=hs.LEFT)
        _pct(row, 2, "npm")
        # 자산회전·레버리지는 배수 → FMT_NUM2
        if row["ato"] is None:
            hs.set_cell(ws, r, 3, "NA", role="soft", align=hs.CENTER)
        else:
            hs.set_cell(ws, r, 3, row["ato"], role="calc", number_format=hs.FMT_NUM2)
        if row["lev"] is None:
            hs.set_cell(ws, r, 4, "NA", role="soft", align=hs.CENTER)
            ws.cell(row=r, column=4).font = hs.font(hs.NEG_FG, bold=True)
        else:
            hs.set_cell(ws, r, 4, row["lev"], role="calc", number_format=hs.FMT_NUM2)
        _pct(row, 5, "roe_product")
        _pct(row, 6, "roe")
        _pct(row, 7, "roic")
        _pct(row, 8, "spread")
        r += 1

    # _RECON: ROE(곱) == ROE(직접) 정합 (곱셈 닫힘) — 정상 기간 합산 차이
    prod_ok_sum = sum(row["roe_product"] for row in rows if row["roe_product"] is not None)
    roe_dir_sum = sum(row["roe"] for row in rows
                      if row["roe"] is not None and row["roe_product"] is not None)
    recon = vc.recon_block(
        n_input=len(data.rows_in), n_output=len(rows),
        src_sum=roe_dir_sum, out_sum=prod_ok_sum,
        completeness="기간 %d 전수 (결측 R17 NA)" % len(rows),
        accuracy="ROE(곱) == ROE(직접) — 곱셈 분해 닫힘",
        cutoff="ROIC = NOPAT/(Debt+Equity), spread = ROIC−WACC",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_PCT2)
    nxt = rec_top + len(recon) + 2

    if len(ledger):
        nxt = hs.section_header(ws, nxt + 1, "이상치 대장 (Anomaly Ledger · RATIO_NA)",
                                last_col=last_col)
        for j, h in enumerate(("기간·지표", "유형", "사유"), start=1):
            hs.set_cell(ws, nxt, j, h, role="header", align=hs.LEFT)
        nxt += 1
        for lr in ledger.rows:
            hs.set_cell(ws, nxt, 1, " · ".join(str(g) for g in lr["grain"]),
                        role="label", align=hs.LEFT)
            hs.set_cell(ws, nxt, 2, lr["anomaly_type"], role="soft", align=hs.LEFT)
            hs.set_cell(ws, nxt, 3, lr["detail"], role="soft", align=hs.LEFT)
            nxt += 1

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="재무제표 (손익 · BS)",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 기간", ("period",), [{"period": row["period"]} for row in rows])
    wb._fpna_meta = {"fact": fact, "rows": rows, "anomaly_ledger": ledger,
                     "surfaced_flags": surfaced}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: DupontInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain (period 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # 곱셈 분해 닫힘: ROE(곱) == ROE(직접) (정상 기간만, 곱셈 정합)
    closure_ok = True
    for row in meta["rows"]:
        if row["roe_product"] is not None and row["roe"] is not None:
            if abs(row["roe_product"] - row["roe"]) > 1e-9:
                closure_ok = False
    rep.add("DuPont 곱셈 닫힘(ROE곱=ROE직접)", closure_ok,
            "" if closure_ok else "곱셈 분해 ≠ 직접 ROE")

    # ROIC 재계산 대조 (NOPAT/(Debt+Equity))
    roic_ok = True
    for d, row in zip(data.rows_in, meta["rows"]):
        nopat = d.ebit * (1.0 - d.tax_rate)
        expected, reason = vc.ratio_or_na(nopat, d.debt + d.equity)
        if reason is None and row["roic"] is not None:
            if abs(expected - row["roic"]) > 1e-9:
                roic_ok = False
    rep.add("ROIC = NOPAT/(Debt+Equity)", roic_ok, "" if roic_ok else "ROIC 재계산 불일치")

    # Anomaly 2층 보존: |ledger| == surfaced flag (NA 은폐 금지)
    vc.assert_anomaly_conserved(rep, meta["anomaly_ledger"], meta["surfaced_flags"])

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "DupontPeriod", "DupontInput", "golden_sample", "build", "qc"]
