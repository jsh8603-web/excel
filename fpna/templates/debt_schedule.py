"""
fpna.templates.debt_schedule — 부채 스케줄(이자·리볼버·cash sweep) (C3 빠진 템플릿).

CFO/대출약정 질문 "기간별 이자비용, 의무상환 후 잉여현금으로 얼마 조기상환
(sweep)되나, 부족하면 리볼버로 얼마 끌어쓰나?"에 답한다.

- grain = "1행 = 1 tranche × 1 기간" (term loan / revolver).
- 롤 불변식(R3 chain): closing = opening − mandatory − sweep + draw.
  closing(t) == opening(t+1) (잔액 연속). 이자 = opening × rate.
  sweep = min(잉여현금, opening 잔액)  / revolver draw = max(0, 현금부족).
- 표시 = wide(tranche × 기간), 내부 = tidy Fact. 잔액 roll tie 가 핵심.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "debt_schedule"


@dataclass
class DebtTranche:
    """차입 1건. term=의무상환 있는 정기, revolver=현금에 따라 draw/sweep."""
    tranche_id: str
    label: str
    kind: str                  # "term" | "revolver"
    opening: float             # 첫 기간 기초잔액
    rate: float                # 기간 이자율(월)
    mandatory: list = field(default_factory=list)  # 기간별 의무상환(term)
    sweep_enabled: bool = False                     # cash sweep 대상 여부


@dataclass
class DebtScheduleInput:
    title: str = "부채 스케줄 (Debt Schedule)"
    subtitle: str = "이자 · 의무상환 · cash sweep · 리볼버 (잔액 roll tie)"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)
    end: tuple = (2024, 12)
    tranches: list = field(default_factory=list)         # list[DebtTranche]
    cash_available: list = field(default_factory=list)   # 기간별 sweep 가용 잉여현금
    commentary: list = field(default_factory=list)


def _roll(data: DebtScheduleInput, nP: int):
    """tranche × 기간 롤. (rolled, totals) 반환.

    각 기간:
      1) term tranche: mandatory 상환 → sweep(잉여현금 잔여분, 우선순위 입력순)
      2) revolver: 현금 부족 시 draw(+) / 잉여 시 paydown(−)
    잉여현금(cash_available[t])을 term sweep 에 우선 배분(잔액 한도 clamp).
    """
    rolled = {t.tranche_id: [] for t in data.tranches}
    # 기초 잔액
    book = {t.tranche_id: t.opening for t in data.tranches}
    tot_interest = tot_mand = tot_sweep = tot_draw = 0.0
    for i in range(nP):
        cash = data.cash_available[i] if i < len(data.cash_available) else 0.0
        # term tranche 먼저(의무상환 + sweep)
        for t in data.tranches:
            if t.kind != "term":
                continue
            opening = book[t.tranche_id]
            interest = opening * t.rate
            mand = t.mandatory[i] if i < len(t.mandatory) else 0.0
            mand = min(max(mand, 0.0), max(opening, 0.0))      # 잔액 초과 금지
            after_mand = opening - mand
            sweep = 0.0
            if t.sweep_enabled and cash > 0:
                sweep = min(cash, max(after_mand, 0.0))        # 잉여현금·잔액 한도
                cash -= sweep
            closing = after_mand - sweep
            rolled[t.tranche_id].append(
                {"opening": opening, "interest": interest, "mandatory": mand,
                 "sweep": sweep, "draw": 0.0, "closing": closing})
            book[t.tranche_id] = closing
            tot_interest += interest
            tot_mand += mand
            tot_sweep += sweep
        # revolver(잔여 현금으로 paydown, 부족 시 draw 는 외부 cash 모델이 결정 →
        #   여기선 음수 cash 를 draw 로 해석)
        for t in data.tranches:
            if t.kind != "revolver":
                continue
            opening = book[t.tranche_id]
            interest = opening * t.rate
            draw = paydown = 0.0
            if cash < 0:
                draw = -cash                                    # 현금 부족 → 리볼버 인출
                cash = 0.0
            elif cash > 0:
                paydown = min(cash, max(opening, 0.0))
                cash -= paydown
            closing = opening + draw - paydown
            rolled[t.tranche_id].append(
                {"opening": opening, "interest": interest, "mandatory": paydown,
                 "sweep": 0.0, "draw": draw, "closing": closing})
            book[t.tranche_id] = closing
            tot_interest += interest
            tot_mand += paydown
            tot_draw += draw
    totals = {"interest": tot_interest, "mandatory": tot_mand,
              "sweep": tot_sweep, "draw": tot_draw,
              "opening": sum(t.opening for t in data.tranches),
              "closing": sum(book.values())}
    return rolled, totals


def _build_fact(data: DebtScheduleInput):
    cal = AccountingCalendar(fiscal_year_start_month=data.fy_start_month)
    periods = cal.periods(data.start, data.end)
    rolled, totals = _roll(data, len(periods))
    rows = []
    for t in data.tranches:
        for p, rr in zip(periods, rolled[t.tranche_id]):
            rows.append({"tranche_id": t.tranche_id, "period": p.label, **rr})
    fact = Fact("1행 = 1 tranche × 1 기간", ("tranche_id", "period"), rows)
    return cal, periods, fact, rolled, totals


def golden_sample() -> DebtScheduleInput:
    """구조 골든 — term loan(sweep on) + revolver. 잉여현금이 sweep/paydown 구동."""
    tranches = [
        DebtTranche("TL-A", "Term Loan A", "term", opening=12_000.0, rate=0.005,
                    mandatory=[100.0] * 12, sweep_enabled=True),
        DebtTranche("RCF", "리볼버(RCF)", "revolver", opening=2_000.0, rate=0.006,
                    mandatory=[], sweep_enabled=False),
    ]
    # 잉여현금: 초반 부족(리볼버 draw) → 후반 잉여(sweep/paydown)
    cash = [-300.0, 0.0, 200.0, 500.0, 500.0, 800.0,
            1_000.0, 1_000.0, 1_000.0, 1_500.0, 2_000.0, 2_000.0]
    return DebtScheduleInput(tranches=tranches, cash_available=cash,
                             start=(2024, 1), end=(2024, 12),
                             commentary=["1월 현금부족 → 리볼버 300 draw",
                                         "하반기 잉여현금 → Term Loan A cash sweep 가속"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: DebtScheduleInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, fact, rolled, totals = _build_fact(data)
    nP = len(periods)
    last_col = 1 + nP + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("DebtSchedule")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    hs.set_widths(ws, {1: 26, last_col: 12})
    for c in range(2, last_col):
        ws.column_dimensions[get_column_letter(c)].width = 9

    hs.set_cell(ws, r, 1, "차입 / 흐름", role="header", align=hs.LEFT)
    for j, p in enumerate(periods, start=2):
        hs.set_cell(ws, r, j, p.label, role="header")
    hs.set_cell(ws, r, last_col, "합계", role="header")
    r += 1

    flows = (("기초잔액", "opening", "calc"), ("이자", "interest", "calc"),
             ("의무상환/paydown(−)", "mandatory", "calc"),
             ("Cash Sweep(−)", "sweep", "calc"), ("리볼버 Draw(+)", "draw", "input"),
             ("기말잔액", "closing", "total"))
    for t in data.tranches:
        rr = rolled[t.tranche_id]
        hs.section_header(ws, r, "%s [%s]" % (t.label, t.kind), last_col=last_col)
        r += 1
        for fname, key, role in flows:
            hs.set_cell(ws, r, 1, fname, role="label", align=hs.LEFT)
            for j, cell in enumerate(rr, start=2):
                hs.set_cell(ws, r, j, cell[key], role=role, number_format=hs.FMT_INT_DASH)
            if key in ("interest", "mandatory", "sweep", "draw"):
                ssum = sum(c[key] for c in rr)
                hs.set_cell(ws, r, last_col, ssum, role="calc",
                            number_format=hs.FMT_INT, bold=True)
            r += 1
        r += 1

    # _RECON: Σopening − Σmandatory − Σsweep + Σdraw == Σclosing (roll tie)
    out_sum = (totals["opening"] - totals["mandatory"] - totals["sweep"]
               + totals["draw"])
    recon = vc.recon_block(
        n_input=len(data.tranches), n_output=len(fact.rows),
        src_sum=totals["closing"], out_sum=out_sum,
        completeness="tranche %d × 기간 %d 전수" % (len(data.tranches), nP),
        accuracy="closing = opening − 상환 − sweep + draw (잔액 roll tie)",
        cutoff="이자 = opening × rate (기초잔액 기준)",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="대출 약정서 · 자금수지(현금 모델)",
                     prepared_by="FP&A", last_col=last_col)
    wb._fpna_meta = {"cal": cal, "periods": periods, "fact": fact,
                     "rolled": rolled, "totals": totals, "roll_out": out_sum}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# T2 바인딩 (from_tidy) — module-level                                         #
#   conserves 는 deferred(자문 C6): 보고 총계(closing)는 sweep/draw 현금 캐스   #
#   케이드를 거친 _roll 의 산물이라, INPUT 만으로 독립 재산출하려면 _roll 로직   #
#   전체를 복제해야 한다(복잡로직 중복 = provenance 이득 없음). 잔액 roll tie    #
#   (Σopening−Σ상환−Σsweep+Σdraw==Σclosing)는 이미 qc 가 강제하므로 여기선      #
#   conserves 미구현으로 deferred 명시(silent cap 금지).                        #
# --------------------------------------------------------------------------- #
GRAIN = ("tranche_id", "period")               # 1행 = 1 tranche × 1 기간
REQUIRED = ("tranches",)
UNIT_POLICY = {"tranches.opening": float, "tranches.rate": float}


def from_tidy(rows) -> DebtScheduleInput:
    """tidy rows(tranche × period) → DebtScheduleInput. mandatory·현금 벡터 재조립.

    행 컬럼: tranche_id, period, [label, kind, opening, rate, sweep_enabled,
      mandatory, cash_available].
    tranche 속성(label/kind/opening/rate/sweep)은 같은 tranche 첫 기간 행에서 취득.
    mandatory 는 period 정렬로 벡터화. cash_available 는 period 단위(아무 tranche
    행에서나 동일 값) — period 별 최초 등장값으로 벡터화.
    """
    from fpna.binding import _coerce
    from itertools import groupby
    srt = sorted(rows, key=lambda r: (str(r.get("tranche_id")), str(r.get("period"))))
    # period 순서(첫 등장) + period별 cash
    period_order = []
    cash_by_period = {}
    for r in rows:
        p = str(r.get("period"))
        if p not in cash_by_period:
            period_order.append(p)
            cash_by_period[p] = _coerce(r.get("cash_available"), float) or 0.0
    period_order.sort()
    tranches = []
    for tid, grp in groupby(srt, key=lambda r: _coerce(r.get("tranche_id"), str)):
        grp = list(grp)
        by_p = {str(r.get("period")): r for r in grp}
        first = grp[0]
        mand = [_coerce(by_p.get(p, {}).get("mandatory"), float) or 0.0
                for p in period_order]
        tranches.append(DebtTranche(
            tranche_id=tid,
            label=_coerce(first.get("label"), str) or tid,
            kind=_coerce(first.get("kind"), str) or "term",
            opening=_coerce(first.get("opening"), float) or 0.0,
            rate=_coerce(first.get("rate"), float) or 0.0,
            mandatory=mand,
            sweep_enabled=_coerce(first.get("sweep_enabled"), bool) or False,
        ))
    cash = [cash_by_period[p] for p in period_order]
    return DebtScheduleInput(tranches=tranches, cash_available=cash)


def qc(wb: openpyxl.Workbook, data: DebtScheduleInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal, periods, fact = meta["cal"], meta["periods"], meta["fact"]
    totals = meta["totals"]

    # R8 grain + R1 시간축 전수
    vc.assert_grain(rep, fact)
    vc.assert_time_ruler(rep, fact, cal, data.start, data.end, period_key="period")

    # R3 roll tie: Σopening − Σ상환 − Σsweep + Σdraw == Σclosing
    vc.assert_tie_out(rep, totals["closing"], meta["roll_out"], tol=1e-6,
                      name="R3 balance_roll_tie")

    # chain: 각 tranche 기간 내 closing(t)=opening(t+1) + closing 식 정합 + 잔액 음수 금지
    chain_ok = eq_ok = nonneg_ok = True
    for tid, rr in meta["rolled"].items():
        for i, c in enumerate(rr):
            recomputed = (c["opening"] - c["mandatory"] - c["sweep"] + c["draw"])
            if abs(recomputed - c["closing"]) > 1e-6:
                eq_ok = False
            if c["closing"] < -1e-9:
                nonneg_ok = False
            if i + 1 < len(rr) and abs(c["closing"] - rr[i + 1]["opening"]) > 1e-6:
                chain_ok = False
    rep.add("잔액 chain 연속(closing=opening+1)", chain_ok, "" if chain_ok else "chain 단절")
    rep.add("closing 식 정합", eq_ok, "" if eq_ok else "closing 재계산 불일치")
    rep.add("잔액 음수 금지(clamp)", nonneg_ok, "" if nonneg_ok else "음수 잔액(과상환)")

    # 이자 재계산 대조 (opening × rate)
    int_ok = True
    rate = {t.tranche_id: t.rate for t in data.tranches}
    for tid, rr in meta["rolled"].items():
        for c in rr:
            if abs(c["interest"] - c["opening"] * rate[tid]) > 1e-6:
                int_ok = False
    rep.add("이자 = 기초잔액 × rate", int_ok, "" if int_ok else "이자 재계산 불일치")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "DebtTranche", "DebtScheduleInput", "golden_sample", "build", "qc",
           "GRAIN", "REQUIRED", "UNIT_POLICY", "from_tidy"]
