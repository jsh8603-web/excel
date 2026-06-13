"""
fpna.templates.metric_table — 일반 라벨+단일숫자 표 (General Metric Line-Items).

대부분의 실무 요청은 화려한 분석모형이 아니라 "항목(글자) + 숫자 한두 개"의 평범한
명세/내역/집계표다 — 물량표, 임금 원가 명세, 상각비 명세, 투자비(capex) 내역 등.
이런 요청이 특정 분석 템플릿(variance/investment_appraisal 등)에 안 맞으면 기존엔
pnl_3statement 기본값으로 떨어졌다. 이 템플릿이 그 빈자리를 **상황 프로파일 + 필드
단위 선언적 검증**으로 채운다.

설계(레퍼런스 어휘 차용, 코드 미반입):
  - 필드 단위 계약 = Frictionless Table Schema field constraints + dbt-utils
    (not_null / accepted_range / accepted_values) + Great Expectations
    (expect_column_values_to_be_between/_in_set) 어휘. 전부 stdlib + dataclass 재구현.
  - grain = "1행 = 1 항목(label)" (R8, Kimball: declare grain first).
  - tie-out(R3) = 선언된 합계 vs 컬럼 합 차이 = 0 (tol=0). 미선언이면 검사 안 함.
  - 정직성: 결측(nullable=True) → NO_DATA 셀로 *노출*(저장 막지 않음).
            계약 위반(범위/부호/허용값/필수결측) → VIOLATION → QC FAIL(저장 보류).
            (= garbage-in 은 막는다. fc_driver_unitcost 의 NA emit 철학과 정합:
             '발견'은 노출, '은폐'와 '오염'은 차단.)

프로파일(situation):
  volume(물량) · labor_cost(임금 원가) · depreciation(상각비) · capex(투자비) · mixed(혼합)
각 프로파일 = 컬럼 FieldSpec 묶음(부호·범위·단위·결측허용 기본값 박제).

★범용 — 위 프로파일은 *자주 쓰는 예시 프리셋*일 뿐이다. 이 템플릿은 특정 키워드(상각비·
물량 등) 전용이 아니라, **mixed 프로파일 + 임의 FieldSpec 으로 어떤 라벨+단일숫자든**
받는다(예: 거리 km, 온도 ℃, 건수, 점수 — 재무·비재무 무관). suggest_profile 이 키워드를
못 맞히면 mixed 로 떨어져 부호 무관 범용 금액 1필드를 주고, 사용자는 FieldSpec 을 직접
정의해 어떤 단위·계약이든 선언할 수 있다. "명세/내역/집계/항목표"가 디스패처 진입 키워드.
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors, qc_totals

TYPE = "metric_table"

# 필드 유형 → 기본 (부호, number_format)
_KIND_DEFAULTS = {
    "quantity": ("+", hs.FMT_NUM1),   # 물량/수량
    "currency": ("+", hs.FMT_INT),    # 금액(원가/상각비/투자비) — 기본 ≥0
    "count":    ("+", hs.FMT_INT),    # 인원/건수(정수)
    "ratio":    ("any", hs.FMT_PCT1), # 비율
    "number":   ("any", hs.FMT_NUM1), # 부호 무관 일반 수치
}


@dataclass
class FieldSpec:
    """값 컬럼 1개의 선언적 계약(필드 단위 검증 규칙)."""
    name: str                              # 컬럼 헤더
    kind: str = "number"                   # quantity/currency/count/ratio/number
    unit: str = ""                         # 단위 표기(대/㎡/₩/명/%)
    sign: str | None = None                # '+'/'-'/'any' (None → kind 기본)
    min_value: float | None = None         # accepted_range 하한
    max_value: float | None = None         # accepted_range 상한
    inclusive: bool = True                 # 경계 포함 여부(dbt-utils inclusive)
    nullable: bool = True                  # not_null 위반 여부(False=필수)
    accepted_values: list | None = None    # accepted_values/isin (정수 코드/등급 등)
    number_format: str | None = None       # None → kind 기본
    tie_total: float | None = None         # 이 컬럼 합과 대조할 선언 합계(R3 tie-out)

    def resolved_sign(self) -> str:
        if self.sign is not None:
            return self.sign
        return _KIND_DEFAULTS.get(self.kind, ("any", hs.FMT_NUM1))[0]

    def resolved_fmt(self) -> str:
        if self.number_format is not None:
            return self.number_format
        return _KIND_DEFAULTS.get(self.kind, ("any", hs.FMT_NUM1))[1]


@dataclass
class MetricRow:
    """1 항목(label) + 컬럼별 단일 숫자(None=결측)."""
    label: str
    values: dict = field(default_factory=dict)   # {field_name: number|None}


@dataclass
class MetricTableInput:
    title: str = "일반 항목표 (Metric Line-Items)"
    subtitle: str = "항목별 단일 수치 · 필드 단위 계약 검증 + 합계 tie-out"
    unit: str = "₩"
    profile: str = "mixed"                 # provenance(어떤 상황 프로파일로 만들었는지)
    fields: list = field(default_factory=list)   # list[FieldSpec]
    rows: list = field(default_factory=list)     # list[MetricRow]
    commentary: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# 상황 프로파일 — 상황별 FieldSpec 묶음(부호·단위·범위·결측 기본값 박제)        #
# --------------------------------------------------------------------------- #
def profile_fields(profile: str, unit: str = "₩") -> list[FieldSpec]:
    """상황 프로파일 이름 → 컬럼 FieldSpec 리스트.

    실데이터 바인딩 래퍼/사용자는 이걸 시작점으로 잡고 min/max/tie_total 만 채우면 된다.
    """
    p = (profile or "mixed").lower()
    if p == "volume":            # 물량 — 음수 물량 금지, 결측 허용
        return [FieldSpec("물량", kind="quantity", unit="개", nullable=True),
                FieldSpec("단위", kind="number", unit="", nullable=True,
                          number_format=None)]
    if p == "labor_cost":        # 임금 원가 — 금액·인원 ≥0, per-head 비율은 ratio
        return [FieldSpec("인건비", kind="currency", unit=unit, nullable=False),
                FieldSpec("인원", kind="count", unit="명", nullable=False, min_value=0)]
    if p == "depreciation":      # 상각비 — 금액 ≥0
        return [FieldSpec("상각비", kind="currency", unit=unit, nullable=False)]
    if p == "capex":             # 투자비 — 금액 ≥0
        return [FieldSpec("투자비", kind="currency", unit=unit, nullable=False)]
    # mixed(일반 혼합) — 금액 부호 무관(환입/조정 가능)
    return [FieldSpec("금액", kind="currency", unit=unit, sign="any", nullable=True)]


_PROFILES = ("volume", "labor_cost", "depreciation", "capex", "mixed")

# 디스패처/스킬이 참조할 프로파일 키워드(요청 텍스트 → 프로파일 제안).
PROFILE_KEYWORDS = {
    "volume":       ["물량", "수량", "통수", "건수", "qty", "volume", "수량표", "물량표"],
    "labor_cost":   ["임금", "인건비", "급여", "노무비", "labor", "payroll", "임금원가"],
    "depreciation": ["상각비", "감가상각비", "상각 명세", "depreciation 명세", "감가 명세"],
    "capex":        ["투자비", "capex", "설비투자", "자본적지출", "투자 내역", "투자비 명세"],
}


def suggest_profile(text: str) -> str:
    """요청 텍스트 → 가장 잘 맞는 상황 프로파일(없으면 mixed)."""
    low = (text or "").lower()
    best, best_hits = "mixed", 0
    for prof, kws in PROFILE_KEYWORDS.items():
        hits = sum(1 for k in kws if k.lower() in low)
        if hits > best_hits:
            best, best_hits = prof, hits
    return best


# --------------------------------------------------------------------------- #
# golden — 구조 골든(의미 없는 더미). 정상 + 결측(NO_DATA) 포함, 위반 0.        #
#   tie-out 선언과 컬럼 합이 정확히 일치하도록 구성(R3 tol=0 PASS).             #
# --------------------------------------------------------------------------- #
def golden_sample() -> MetricTableInput:
    fields = [
        FieldSpec("투자비", kind="currency", unit="₩", nullable=False,
                  min_value=0, tie_total=180.0),
        FieldSpec("물량", kind="quantity", unit="대", nullable=True, min_value=0),
        FieldSpec("등급", kind="count", unit="", nullable=True,
                  accepted_values=[1, 2, 3]),
    ]
    rows = [
        MetricRow("항목 A", {"투자비": 100.0, "물량": 10.0, "등급": 1}),
        MetricRow("항목 B", {"투자비": 80.0, "물량": None, "등급": 2}),   # 물량 결측 → NO_DATA
    ]
    return MetricTableInput(
        title="일반 항목표 (구조 골든 — 의미 없는 더미)",
        subtitle="투자비 tie-out=180 · 물량 결측은 NO_DATA 노출 · 등급∈{1,2,3}",
        unit="₩", profile="capex", fields=fields, rows=rows,
        commentary=["골든은 구조 더미(재무 의미 없음).",
                    "물량 B = 결측 → NO_DATA(저장 막지 않음, 정직 노출).",
                    "투자비 합 100+80=180 = 선언 tie_total → R3 tie-out PASS."])


# --------------------------------------------------------------------------- #
# 검증 코어 — 필드 단위 계약 체크(stdlib 재구현)                                #
#   반환: cells = [{label, field, value, status, detail}], status ∈            #
#          {OK, NO_DATA, VIOLATION}. col_sums = {field: Σ(정상+0)}.            #
# --------------------------------------------------------------------------- #
def _check_field(spec: FieldSpec, value):
    """단일 셀 1개에 대한 계약 검증 → (status, detail)."""
    if value is None:
        if spec.nullable:
            return "NO_DATA", "결측(nullable)"
        return "VIOLATION", "필수값 결측(not_null 위반)"
    # 허용값 집합(accepted_values / isin)
    if spec.accepted_values is not None and value not in spec.accepted_values:
        return "VIOLATION", "허용값 아님 %r∉%s" % (value, spec.accepted_values)
    # 부호 규약
    sgn = spec.resolved_sign()
    if sgn == "+" and value < 0:
        return "VIOLATION", "부호 위반(≥0 기대, 값=%g)" % value
    if sgn == "-" and value > 0:
        return "VIOLATION", "부호 위반(≤0 기대, 값=%g)" % value
    # 범위(accepted_range, inclusive)
    if spec.min_value is not None:
        bad = value < spec.min_value if spec.inclusive else value <= spec.min_value
        if bad:
            return "VIOLATION", "하한 위반(min=%g%s, 값=%g)" % (
                spec.min_value, "" if spec.inclusive else " 미포함", value)
    if spec.max_value is not None:
        bad = value > spec.max_value if spec.inclusive else value >= spec.max_value
        if bad:
            return "VIOLATION", "상한 위반(max=%g%s, 값=%g)" % (
                spec.max_value, "" if spec.inclusive else " 미포함", value)
    return "OK", ""


def _validate(inp: MetricTableInput):
    cells = []
    col_sums = {f.name: 0.0 for f in inp.fields}
    for row in inp.rows:
        for spec in inp.fields:
            v = row.values.get(spec.name)
            status, detail = _check_field(spec, v)
            cells.append({"label": row.label, "field": spec.name,
                          "value": v, "status": status, "detail": detail})
            if v is not None and isinstance(v, (int, float)):
                col_sums[spec.name] += float(v)
    return cells, col_sums


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: MetricTableInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    if not data.fields:
        data.fields = profile_fields(data.profile, data.unit)
    cells, col_sums = _validate(data)
    cell_idx = {(c["label"], c["field"]): c for c in cells}
    n_fields = len(data.fields)
    last_col = 1 + n_fields + 1   # 항목 + 값들 + 사유

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("MetricTable")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col)
    widths = {1: 26}
    for j in range(2, 2 + n_fields):
        widths[j] = 13
    widths[last_col] = 22
    hs.set_widths(ws, widths)

    # 헤더 — 단위를 헤더 텍스트에 노출(필드 계약 가독화)
    hs.set_cell(ws, r, 1, "항목", role="header", align=hs.LEFT)
    for j, spec in enumerate(data.fields, start=2):
        head = spec.name + (" (%s)" % spec.unit if spec.unit else "")
        hs.set_cell(ws, r, j, head, role="header", align=hs.CENTER)
    hs.set_cell(ws, r, last_col, "검증(사유)", role="header", align=hs.LEFT)
    r += 1

    surfaced = 0           # NO_DATA + VIOLATION 노출 셀 수(정직성 보존 카운트)
    n_violation = 0
    for row in data.rows:
        hs.set_cell(ws, r, 1, row.label, role="label", align=hs.LEFT)
        row_notes = []
        for j, spec in enumerate(data.fields, start=2):
            c = cell_idx[(row.label, spec.name)]
            if c["status"] == "OK":
                hs.set_cell(ws, r, j, c["value"], role="calc",
                            number_format=spec.resolved_fmt())
            elif c["status"] == "NO_DATA":
                hs.set_cell(ws, r, j, "NO_DATA", role="soft", align=hs.CENTER)
                surfaced += 1
                row_notes.append("%s: 결측" % spec.name)
            else:  # VIOLATION
                hs.set_cell(ws, r, j, c["value"] if c["value"] is not None else "NA",
                            role="soft", align=hs.CENTER)
                ws.cell(row=r, column=j).font = hs.font(hs.NEG_FG, bold=True)
                surfaced += 1
                n_violation += 1
                row_notes.append("%s: %s" % (spec.name, c["detail"]))
        hs.set_cell(ws, r, last_col, "; ".join(row_notes) if row_notes else "—",
                    role="soft", align=hs.LEFT_WRAP)
        r += 1

    # 합계 행 + tie-out(_RECON) — 선언 tie_total 있는 컬럼만 대조
    r += 1
    hs.set_cell(ws, r, 1, "합계 (Σ 정상·0)", role="total", align=hs.LEFT)
    tie_specs = []
    for j, spec in enumerate(data.fields, start=2):
        hs.set_cell(ws, r, j, col_sums[spec.name], role="total",
                    number_format=spec.resolved_fmt(), bold=True)
        if spec.tie_total is not None:
            tie_specs.append(spec)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    r += 1

    if tie_specs:
        r += 1
        r = hs.section_header(ws, r, "대사 (_RECON · tie-out, 차이=0 기대)",
                              last_col=last_col)
        for spec in tie_specs:
            diff = col_sums[spec.name] - spec.tie_total
            hs.set_cell(ws, r, 1, "%s: 컬럼합 vs 선언합" % spec.name,
                        role="label", align=hs.LEFT)
            hs.set_cell(ws, r, 2, col_sums[spec.name], role="calc",
                        number_format=spec.resolved_fmt())
            hs.set_cell(ws, r, 3, spec.tie_total, role="calc",
                        number_format=spec.resolved_fmt())
            hs.set_cell(ws, r, 4, diff, role="calc",
                        number_format=spec.resolved_fmt())
            if abs(diff) > 0:
                ws.cell(row=r, column=4).font = hs.font(hs.NEG_FG, bold=True)
            r += 1

    if data.commentary:
        r += 1
        r = hs.section_header(ws, r, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, r, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
            r += 1

    hs.report_footer(ws, r + 1, source="항목 원장 · 필드 계약(profile=%s)" % data.profile,
                     prepared_by="FP&A", last_col=last_col)

    fact = Fact("1행 = 1 항목(label)", ("label",),
                [{"label": rw.label} for rw in data.rows])
    wb._fpna_meta = {"fact": fact, "cells": cells, "col_sums": col_sums,
                     "surfaced_flags": surfaced, "n_violation": n_violation,
                     "tie_specs": [(s.name, col_sums[s.name], s.tie_total) for s in tie_specs]}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: MetricTableInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain — 항목 라벨 유일성(중복 = silent merge 위험)
    vc.assert_grain(rep, meta["fact"])

    # 필드 단위 계약 재검증(빌드 메타 신뢰 안 하고 입력에서 재계산)
    cells, col_sums = _validate(data)
    n_viol = sum(1 for c in cells if c["status"] == "VIOLATION")
    n_nodata = sum(1 for c in cells if c["status"] == "NO_DATA")
    # 계약 위반(범위/부호/허용값/필수결측) → FAIL(저장 보류). garbage-in 차단.
    rep.add("필드계약 위반 0건", n_viol == 0,
            "" if n_viol == 0 else "위반 %d건: %s" % (
                n_viol, "; ".join("%s/%s %s" % (c["label"], c["field"], c["detail"])
                                  for c in cells if c["status"] == "VIOLATION")[:200]))

    # R3 tie-out — 선언 tie_total 있는 컬럼은 컬럼합과 차이=0 (tol=0)
    for spec in data.fields:
        if spec.tie_total is not None:
            qc_totals("tie:%s" % spec.name, col_sums[spec.name], spec.tie_total,
                      rep, tol=0.0)

    # 정직성 2층 — 결측/위반 셀이 전부 산출물에 노출됐는지(은폐 금지)
    rep.add("anomaly 노출 보존", meta["surfaced_flags"] == (n_viol + n_nodata),
            "" if meta["surfaced_flags"] == (n_viol + n_nodata) else
            "surfaced=%d != viol+nodata=%d (은폐)" % (
                meta["surfaced_flags"], n_viol + n_nodata))

    # 단위 표기 — 테이블 unit 또는 필드별 unit 중 하나면 OK(비재무 혼합단위 범용 허용:
    #   거리 km·온도 ℃ 처럼 테이블 단일 unit 이 무의미하면 FieldSpec.unit 로 충분).
    has_unit = bool(data.unit) or any(f.unit for f in data.fields)
    rep.add("단위 표기", has_unit, "" if has_unit else "테이블·필드 unit 모두 비어있음")
    return rep


__all__ = ["TYPE", "FieldSpec", "MetricRow", "MetricTableInput",
           "profile_fields", "suggest_profile", "PROFILE_KEYWORDS",
           "golden_sample", "build", "qc"]
