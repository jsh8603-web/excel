"""
fpna.templates.fc_prepaid_rollforward — 고정비: 선급비용 롤포워드 (A6).

선급비용(선급임차료·선급보험료 등)의 기초→가산→상각→기말 흐름을 기간별로 펼친다.
CFO 질문 "선급 잔액 추이, 월 상각 인식 맞나?"에 답한다.

- grain = "1행 = 1 (CC × Account × prepaid_item) × 1 기간".
- 불변식(R3 chain): ending = beg + add − amort. ending(t) == beg(t+1)(연속).
  amort ≤ beg + add (음수 금지 — 잔액 초과 상각 차단).
- 엣지: 기중 addition 부분상각, 조기해지 가속상각 + clamp(잔액 0 까지만).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_prepaid_rollforward"


@dataclass
class PrepaidItem:
    """선급비용 1건. 기초잔액 + 기간별 (add, 계획 amort)."""
    cost_center: str
    account: str
    item: str               # "선급임차료" 등
    label: str
    opening: float          # 첫 기간 기초잔액
    adds: list = field(default_factory=list)    # 기간별 가산(period 1:1)
    amorts: list = field(default_factory=list)  # 기간별 상각(계획). clamp 로 잔액 보호


@dataclass
class PrepaidInput:
    title: str = "고정비 — 선급비용 롤포워드 (Prepaid Roll-Forward)"
    subtitle: str = "기초 + 가산 − 상각 = 기말 (chain 연속)"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)
    end: tuple = (2024, 12)
    items: list = field(default_factory=list)   # list[PrepaidItem]
    commentary: list = field(default_factory=list)


def _periods(inp: PrepaidInput):
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    return cal, cal.periods(inp.start, inp.end)


def _roll(item: PrepaidItem, nP: int):
    """기간별 (beg, add, amort, end) 롤포워드. amort 는 잔액(beg+add)으로 clamp."""
    rows = []
    beg = item.opening
    for i in range(nP):
        add = item.adds[i] if i < len(item.adds) else 0.0
        want = item.amorts[i] if i < len(item.amorts) else 0.0
        avail = beg + add
        amort = min(max(want, 0.0), max(avail, 0.0))   # 음수 금지 + 잔액 초과 금지(clamp)
        end = avail - amort
        rows.append((beg, add, amort, end))
        beg = end
    return rows


def _build_fact(inp: PrepaidInput):
    cal, periods = _periods(inp)
    nP = len(periods)
    rows: list[dict] = []
    rolled: dict[str, list] = {}
    for it in inp.items:
        rr = _roll(it, nP)
        rolled[it.item + "|" + it.cost_center] = rr
        for p, (beg, add, amort, end) in zip(periods, rr):
            rows.append({"cost_center": it.cost_center, "account": it.account,
                         "item": it.item, "period": p.label,
                         "beg": beg, "add": add, "amort": amort, "end": end})
    fact = Fact("1행 = 1 (CC × Account × prepaid_item) × 1 기간",
                ("cost_center", "account", "item", "period"), rows)
    return cal, periods, fact, rolled


def golden_sample() -> PrepaidInput:
    """구조 골든 — 정상 상각 1건 + 기중 가산 1건 + 조기해지 가속(clamp) 1건."""
    items = [
        # 선급임차: 기초 1200, 매월 100 상각(12개월) → 기말 0
        PrepaidItem("CC10", "1530", "선급임차료", "선급임차료 (CC10)",
                    opening=1_200.0, adds=[0] * 12, amorts=[100] * 12),
        # 선급보험: 기초 0, 4월 가산 600, 이후 100씩 상각(부분기간)
        PrepaidItem("CC20", "1531", "선급보험료", "선급보험료 (CC20·기중가산)",
                    opening=0.0,
                    adds=[0, 0, 0, 600, 0, 0, 0, 0, 0, 0, 0, 0],
                    amorts=[0, 0, 0, 100, 100, 100, 100, 100, 100, 0, 0, 0]),
        # 선급유지보수: 기초 300, 6월 조기해지로 전액 가속상각(잔액>요청이면 clamp)
        PrepaidItem("CC30", "1532", "선급유지보수", "선급유지보수 (CC30·조기해지)",
                    opening=300.0, adds=[0] * 12,
                    amorts=[25, 25, 25, 25, 25, 999, 0, 0, 0, 0, 0, 0]),  # 6월 999>잔액 → clamp
    ]
    return PrepaidInput(items=items, start=(2024, 1), end=(2024, 12),
                        commentary=["CC30 6월 조기해지 가속상각 — 요청 999 > 잔액 → clamp(잔액까지만)",
                                    "amort ≤ beg+add 불변식: 음수 잔액 차단"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: PrepaidInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, fact, rolled = _build_fact(data)
    nP = len(periods)
    # 각 item 블록: 기초/가산/상각/기말 4행 × 기간 열
    last_col = 1 + nP + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("Prepaid")

    r = hs.title_block(ws, data.title,
                       (data.subtitle + ("  ·  단위 " + data.unit if data.unit else "")).strip(" ·"),
                       last_col=last_col)
    header_row = r
    hs.style_sheet(ws, freeze="B%d" % (header_row + 1))
    hs.set_widths(ws, {1: 24, last_col: 12})
    for c in range(2, last_col):
        ws.column_dimensions[get_column_letter(c)].width = 9

    hs.set_cell(ws, r, 1, "선급 항목 / 흐름", role="header", align=hs.LEFT)
    for j, p in enumerate(periods, start=2):
        hs.set_cell(ws, r, j, p.label, role="header")
    hs.set_cell(ws, r, last_col, "합계", role="header")
    r += 1

    grand_amort = 0.0
    final_endings: dict[str, float] = {}
    for it in data.items:
        rr = rolled[it.item + "|" + it.cost_center]
        hs.section_header(ws, r, it.label, last_col=last_col)
        r += 1
        # 4 행: 기초/가산/상각/기말
        flows = (("기초", 0, "calc"), ("가산(+)", 1, "input"),
                 ("상각(−)", 2, "calc"), ("기말", 3, "total"))
        for fname, idx, role in flows:
            hs.set_cell(ws, r, 1, fname, role="label", align=hs.LEFT)
            for j, (beg, add, amort, end) in enumerate(rr, start=2):
                val = (beg, add, amort, end)[idx]
                hs.set_cell(ws, r, j, val, role=role, number_format=hs.FMT_INT_DASH)
            if idx == 2:    # 상각 합계
                amort_sum = sum(x[2] for x in rr)
                hs.set_cell(ws, r, last_col, amort_sum, role="calc",
                            number_format=hs.FMT_INT, bold=True)
                grand_amort += amort_sum
            r += 1
        final_endings[it.item + "|" + it.cost_center] = rr[-1][3]
        r += 1   # 블록 간 여백

    # _RECON: 전체 기초 + 전체 가산 − 전체 상각 == 전체 기말 (보존)
    sum_open = sum(it.opening for it in data.items)
    sum_add = sum(sum(x[1] for x in rolled[it.item + "|" + it.cost_center]) for it in data.items)
    sum_amort = grand_amort
    sum_end = sum(final_endings.values())
    recon = vc.recon_block(
        n_input=len(data.items), n_output=len(fact.rows),
        src_sum=sum_open + sum_add, out_sum=sum_amort + sum_end,
        completeness="항목 %d × 기간 %d 전수 롤포워드" % (len(data.items), nP),
        accuracy="ending = beg + add − amort (chain 연속)",
        cutoff="amort ≤ beg+add (clamp, 음수 잔액 차단)",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1

    wb._fpna_meta = {"cal": cal, "periods": periods, "fact": fact, "rolled": rolled,
                     "sum_open": sum_open, "sum_add": sum_add,
                     "sum_amort": sum_amort, "sum_end": sum_end}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: PrepaidInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal, periods, fact = meta["cal"], meta["periods"], meta["fact"]

    # R8 grain + R1 시간축 전수
    vc.assert_grain(rep, fact)
    vc.assert_time_ruler(rep, fact, cal, data.start, data.end, period_key="period")

    # A6-1 보존: Σ(기초+가산) == Σ(상각+기말)
    vc.assert_tie_out(rep, meta["sum_open"] + meta["sum_add"],
                      meta["sum_amort"] + meta["sum_end"], tol=1e-6,
                      name="A6 rollforward_tie")

    # A6-2 chain: 각 항목 ending(t) == beg(t+1) + amort 음수 금지 + amort ≤ beg+add
    chain_ok = True
    nonneg_ok = True
    clamp_ok = True
    for key, rr in meta["rolled"].items():
        for i, (beg, add, amort, end) in enumerate(rr):
            if amort < -1e-9:
                nonneg_ok = False
            if amort > beg + add + 1e-9:        # 잔액 초과 상각 → clamp 위반
                clamp_ok = False
            if abs((beg + add - amort) - end) > 1e-9:
                chain_ok = False
            if i + 1 < len(rr) and abs(end - rr[i + 1][0]) > 1e-9:  # ending(t)=beg(t+1)
                chain_ok = False
    rep.add("A6 chain 연속(ending=beg+1)", chain_ok, "" if chain_ok else "chain 단절")
    rep.add("A6 상각 음수 금지", nonneg_ok, "" if nonneg_ok else "음수 상각")
    rep.add("A6 amort≤beg+add(clamp)", clamp_ok, "" if clamp_ok else "잔액 초과 상각")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "PrepaidItem", "PrepaidInput", "golden_sample", "build", "qc"]
