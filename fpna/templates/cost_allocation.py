"""
fpna.templates.cost_allocation — 공통비 배부 (C3 빠진 템플릿).

공통/간접비 풀(pool)을 배부기준(driver: headcount·면적·매출 등)으로 수익부서에
나눠 준다. CFO 질문 "본사 공통비를 어떤 기준으로 어느 부서에 얼마 배부했나,
배부 전후 합이 보존되나(누수 0)?"에 답한다.

- grain = "1행 = 1 (pool × cost_center)" 배부 결과.
- 배부: alloc(pool→cc) = pool_amount × driver(cc) / Σdriver.
- 불변식(R11): Σ배부 == Σ풀 (배부 전후 보존, view_contract.assert_allocation_conserves).
  Σdriver=0 인 풀은 배부 불가 → UNALLOCATED 로 명시(silent drop 금지).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import house_style as hs
from fpna import view_contract as vc
from fpna.dims import Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "cost_allocation"


@dataclass
class CostPool:
    """배부 대상 공통비 풀 1건."""
    pool_id: str
    label: str
    amount: float
    driver: str                # 배부기준 이름(예: "인원수")


@dataclass
class AllocationInput:
    title: str = "공통비 배부 (Cost Allocation)"
    subtitle: str = "풀 × 배부기준 → 수익부서 (배부 전후 보존 · 누수 0)"
    unit: str = "₩"
    cost_centers: list = field(default_factory=list)   # 배부 대상 CC code 리스트
    cc_labels: dict = field(default_factory=dict)      # {cc: 표시명}
    pools: list = field(default_factory=list)          # list[CostPool]
    # {pool_id: {cc: driver_weight}} — 풀별 CC 배부기준 값
    driver_weights: dict = field(default_factory=dict)
    commentary: list = field(default_factory=list)


def _allocate(data: AllocationInput):
    """풀별 CC 배부 + 미배부(driver 합 0). 메타 반환.

    alloc[pool][cc] = amount × w(cc) / Σw. Σw=0 이면 전액 UNALLOCATED.
    잔여 흡수(rounding): 마지막 CC 가 풀 합과의 차이를 흡수해 풀별 Σ=amount 보존.
    """
    alloc: dict[str, dict[str, float]] = {}
    unalloc: dict[str, float] = {}
    cc_tot: dict[str, float] = {cc: 0.0 for cc in data.cost_centers}
    pool_sum = sum(p.amount for p in data.pools)
    for p in data.pools:
        weights = data.driver_weights.get(p.pool_id, {})
        total_w = sum(weights.get(cc, 0.0) for cc in data.cost_centers)
        alloc[p.pool_id] = {}
        if total_w <= 0:
            unalloc[p.pool_id] = p.amount          # 배부 불가 → 명시
            continue
        assigned = 0.0
        ccs = list(data.cost_centers)
        for k, cc in enumerate(ccs):
            if k == len(ccs) - 1:
                a = p.amount - assigned            # 잔여 흡수(보존)
            else:
                a = p.amount * weights.get(cc, 0.0) / total_w
                assigned += a
            alloc[p.pool_id][cc] = a
            cc_tot[cc] += a
    allocated_sum = sum(sum(v.values()) for v in alloc.values())
    unalloc_sum = sum(unalloc.values())
    return {"alloc": alloc, "unalloc": unalloc, "cc_tot": cc_tot,
            "pool_sum": pool_sum, "allocated_sum": allocated_sum,
            "unalloc_sum": unalloc_sum,
            "post_sum": allocated_sum + unalloc_sum}


def golden_sample() -> AllocationInput:
    """구조 골든 — 2 풀 × 3 CC + driver 합 0 풀 1건(UNALLOCATED 명시)."""
    ccs = ["CC10", "CC20", "CC30"]
    pools = [
        CostPool("HR", "인사 공통비", 6_000.0, "인원수"),
        CostPool("IT", "IT 공통비", 4_500.0, "PC수"),
        CostPool("FAC", "시설 공통비(배부기준 부재)", 1_000.0, "면적"),
    ]
    weights = {
        "HR": {"CC10": 10, "CC20": 20, "CC30": 30},
        "IT": {"CC10": 5, "CC20": 10, "CC30": 15},
        "FAC": {"CC10": 0, "CC20": 0, "CC30": 0},   # Σw=0 → UNALLOCATED
    }
    return AllocationInput(
        cost_centers=ccs,
        cc_labels={"CC10": "영업1팀", "CC20": "영업2팀", "CC30": "생산팀"},
        pools=pools, driver_weights=weights,
        commentary=["FAC 풀: 배부기준 합 0 → UNALLOCATED 명시(silent drop 금지)",
                    "배부 전후 보존: Σ배부 + Σ미배부 == Σ풀 (누수 0)"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: AllocationInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    m = _allocate(data)
    nCC = len(data.cost_centers)
    last_col = 1 + nCC + 2                          # 풀 + CC들 + 미배부 + 합계

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("CostAllocation")
    hs.set_widths(ws, {1: 24, last_col: 12, last_col - 1: 11})
    for c in range(2, last_col - 1):
        ws.column_dimensions[get_column_letter(c)].width = 11

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")

    hs.set_cell(ws, r, 1, "풀 \\ 배부기준", role="header", align=hs.LEFT)
    for j, cc in enumerate(data.cost_centers, start=2):
        hs.set_cell(ws, r, j, data.cc_labels.get(cc, cc), role="header")
    hs.set_cell(ws, r, last_col - 1, "미배부", role="header")
    hs.set_cell(ws, r, last_col, "풀 합계", role="header")
    r += 1

    data_start = r
    for p in data.pools:
        hs.set_cell(ws, r, 1, "%s [%s]" % (p.label, p.driver), role="label", align=hs.LEFT)
        for j, cc in enumerate(data.cost_centers, start=2):
            a = m["alloc"].get(p.pool_id, {}).get(cc)
            if a is None:
                hs.set_cell(ws, r, j, "—", role="soft", align=hs.CENTER)
            else:
                hs.set_cell(ws, r, j, a, role="calc", number_format=hs.FMT_INT)
        un = m["unalloc"].get(p.pool_id, 0.0)
        if un:
            hs.set_cell(ws, r, last_col - 1, un, role="soft", number_format=hs.FMT_INT)
            ws.cell(row=r, column=last_col - 1).font = hs.font(hs.NEG_FG, bold=True)
        else:
            hs.set_cell(ws, r, last_col - 1, 0, role="soft", number_format=hs.FMT_INT_DASH)
        hs.set_cell(ws, r, last_col, p.amount, role="calc", number_format=hs.FMT_INT, bold=True)
        r += 1
    data_end = r - 1

    # CC 합계행
    hs.set_cell(ws, r, 1, "부서별 배부 합계", role="total", align=hs.LEFT)
    for j, cc in enumerate(data.cost_centers, start=2):
        hs.set_cell(ws, r, j, m["cc_tot"][cc], role="total", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, last_col - 1, m["unalloc_sum"], role="total", number_format=hs.FMT_INT)
    hs.set_cell(ws, r, last_col, m["post_sum"], role="total", number_format=hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    r += 1

    # _RECON: 배부전(Σ풀) == 배부후(Σ배부 + Σ미배부)
    recon = vc.recon_block(
        n_input=len(data.pools), n_output=len(data.pools) * nCC,
        src_sum=m["pool_sum"], out_sum=m["post_sum"],
        completeness="풀 %d × CC %d 전수 배부" % (len(data.pools), nCC),
        accuracy="Σ배부 + Σ미배부 == Σ풀 (배부 전후 보존)",
        cutoff="driver 합 0 풀 → UNALLOCATED 명시",
    )
    rec_top = r + 1
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    nxt = rec_top + len(recon) + 2

    if data.commentary:
        cr = nxt + 1
        cr = hs.section_header(ws, cr, "코멘터리", last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        nxt = cr

    hs.report_footer(ws, nxt + 1, source="공통비 원장 · 배부기준 마스터",
                     prepared_by="FP&A", last_col=last_col)
    fact = Fact("1행 = 1 (pool × cost_center)", ("pool_id", "cost_center"),
                [{"pool_id": p.pool_id, "cost_center": cc}
                 for p in data.pools for cc in data.cost_centers])
    wb._fpna_meta = {"fact": fact, **m}
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: AllocationInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta

    # R8 grain (pool × cost_center 중복 금지)
    vc.assert_grain(rep, meta["fact"])

    # R11 배부 전후 보존: Σ풀 == Σ배부 + Σ미배부 (누수 0)
    vc.assert_allocation_conserves(rep, meta["pool_sum"], meta["post_sum"], tol=1e-6)

    # 풀별 Σ배부(+미배부) == 풀 amount (개별 보존)
    pool_ok = True
    for p in data.pools:
        got = sum(meta["alloc"].get(p.pool_id, {}).values()) + meta["unalloc"].get(p.pool_id, 0.0)
        if abs(got - p.amount) > 1e-6:
            pool_ok = False
    rep.add("풀별 배부 보존", pool_ok, "" if pool_ok else "풀 합 불일치")

    # 미배부 명시(silent drop 금지): driver 합 0 풀이 unalloc 으로 노출됐는지
    zero_pools = [p.pool_id for p in data.pools
                  if sum(data.driver_weights.get(p.pool_id, {}).get(cc, 0.0)
                         for cc in data.cost_centers) <= 0]
    surfaced = all(pid in meta["unalloc"] for pid in zero_pools)
    rep.add("미배부 명시(UNALLOCATED)", surfaced,
            "" if surfaced else "driver 합 0 풀이 silent drop 됨")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "CostPool", "AllocationInput", "golden_sample", "build", "qc"]
