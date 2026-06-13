"""
fpna.templates.fc_runrate_normalized — 고정비: 정규화 run-rate / 연환산 (A2).

월별 실적에서 one-off(단발 대형계상)를 robust 마스킹하고, 계절성을 조정한 뒤
"정상 월 런레이트"와 그 연환산을 산출한다. CFO 질문 "이 비용의 진짜 베이스라인은
얼마, 1회성 빼면 연환산 얼마?"에 답한다.

- grain = "1행 = 1 CostCenter × 1 Account" (집계 단위. 시계열은 열로 전개).
- 불변식: normalized = actual − Σone_off (tie, R3). annualized = monthly × factor
  (12 하드코딩 금지 — active_months 기준 월런레이트의 연환산, 이중연환산 방지).
- one-off 마스킹은 fpna.finance.normalized_run_rate(robust_mask × seasonal).
- raw vs normalized 병기(자문 ③) + over_masking flag(soft).
"""
from __future__ import annotations

from dataclasses import dataclass, field

import fpna._bootstrap  # noqa: F401

import openpyxl
from openpyxl.utils import get_column_letter

from fpna import finance, house_style as hs
from fpna import view_contract as vc
from fpna.dims import AccountingCalendar, Fact
from fpna.templates.base import QCReport, qc_no_formula_errors

TYPE = "fc_runrate_normalized"


@dataclass
class RunRateLine:
    """1 (CostCenter × Account) 의 월별 실적 시계열."""
    cost_center: str
    account: str
    label: str               # 표시 라벨(예: "본사임차 (CC10·6010)")
    monthly: list = field(default_factory=list)   # 기간순 실적 (period 와 1:1)


@dataclass
class RunRateInput:
    title: str = "고정비 — 정규화 Run-Rate / 연환산"
    subtitle: str = "one-off 마스킹(robust) · 계절 조정 후 월 런레이트와 연환산"
    unit: str = "₩"
    fy_start_month: int = 1
    start: tuple = (2024, 1)
    end: tuple = (2024, 12)
    lines: list = field(default_factory=list)     # list[RunRateLine]
    deseasonalize: bool = True
    annualize_factor: int = 12
    commentary: list = field(default_factory=list)


def _periods(inp: RunRateInput):
    cal = AccountingCalendar(fiscal_year_start_month=inp.fy_start_month)
    return cal, cal.periods(inp.start, inp.end)


def _result_for(inp: RunRateInput, line: RunRateLine) -> finance.RunRateResult:
    return finance.normalized_run_rate(
        list(line.monthly), deseasonalize=inp.deseasonalize,
        annualize_factor=inp.annualize_factor)


def _build_fact(inp: RunRateInput):
    """tidy Fact (CC×Account×period). grain 정합 검증용."""
    cal, periods = _periods(inp)
    rows: list[dict] = []
    for ln in inp.lines:
        for p, v in zip(periods, ln.monthly):
            rows.append({"cost_center": ln.cost_center, "account": ln.account,
                         "period": p.label, "value": v})
    fact = Fact("1행 = 1 CC × 1 Account × 1 기간",
                ("cost_center", "account", "period"), rows)
    return cal, periods, fact


def golden_sample() -> RunRateInput:
    """구조 골든 — 한 라인은 단발 one-off 포함(마스킹), 다른 라인은 평탄."""
    lines = [
        # 본사 임차: 평탄 100, 7월에 1회성 보증금정산 +900 → 마스킹 기대
        RunRateLine("CC10", "6010", "본사임차 (CC10·6010)",
                    [100, 100, 100, 100, 100, 100, 1000, 100, 100, 100, 100, 100]),
        # SaaS: 단가 step(상반기 50 → 하반기 70) — one-off 아님(마스킹 X 기대)
        RunRateLine("CC20", "6210", "SaaS구독 (CC20·6210)",
                    [50, 50, 50, 50, 50, 50, 70, 70, 70, 70, 70, 70]),
    ]
    return RunRateInput(lines=lines, start=(2024, 1), end=(2024, 12),
                        deseasonalize=False,   # 골든은 계절성 미적용(단순 검증)
                        commentary=["본사임차 7월 +900 = 보증금 정산 1회성 → 런레이트 제외"])


# --------------------------------------------------------------------------- #
# build                                                                       #
# --------------------------------------------------------------------------- #
def build(data: RunRateInput, *, mode: str = "create", base_path=None) -> openpyxl.Workbook:
    cal, periods, fact = _build_fact(data)
    nP = len(periods)
    # 열: 라벨 + 기간들 + [실적합 · 1회성 · 정규화합 · 활성월 · 월런레이트 · 연환산]
    summary_cols = ("실적합", "1회성(−)", "정규화합", "활성월", "월런레이트", "연환산")
    last_col = 1 + nP + len(summary_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hs.safe_sheet_title("RunRate")

    r = hs.report_frame(ws, data.title, subtitle=data.subtitle,
                        unit=data.unit, last_col=last_col, freeze_col="B")
    header_row = r
    hs.set_widths(ws, {1: 26})
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11

    # 헤더
    hs.set_cell(ws, r, 1, "비용 라인 (CC·Account)", role="header", align=hs.LEFT)
    for j, p in enumerate(periods, start=2):
        hs.set_cell(ws, r, j, p.label, role="header")
    for k, h in enumerate(summary_cols):
        hs.set_cell(ws, r, 2 + nP + k, h, role="header")
    r += 1

    results: list[tuple[RunRateLine, finance.RunRateResult]] = []
    surfaced = 0
    masked_label = "▣"        # 마스킹된 셀 표식(surfaced)
    for ln in data.lines:
        res = _result_for(data, ln)
        results.append((ln, res))
        masked = set(res.masked_index)
        hs.set_cell(ws, r, 1, ln.label, role="label", align=hs.LEFT)
        for j, v in enumerate(ln.monthly):
            col = 2 + j
            if j in masked:
                # one-off: 값 + 마스킹 표식(은폐 금지 — surfaced flag)
                hs.set_cell(ws, r, col, v, role="soft",
                            number_format=hs.FMT_INT_DASH, align=hs.CENTER)
                ws.cell(row=r, column=col).value = "%s %s" % (masked_label, v)
                surfaced += 1
            else:
                hs.set_cell(ws, r, col, v, role="calc", number_format=hs.FMT_INT_DASH)
        # 마스킹 셀 자체가 surfaced; over-masking 은 qc soft 로 별도 노출.
        base = 2 + nP
        hs.set_cell(ws, r, base + 0, res.actual_total, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, base + 1, res.one_off_total, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, base + 2, res.normalized_total, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, base + 3, res.active_months, role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, base + 4, res.monthly_run_rate, role="calc", number_format=hs.FMT_NUM1)
        hs.set_cell(ws, r, base + 5, res.annualized, role="total",
                    number_format=hs.FMT_INT, bold=True)
        r += 1

    # 합계행(연환산·정규화합 등)
    hs.set_cell(ws, r, 1, "합계", role="total", align=hs.LEFT)
    base = 2 + nP
    for k in range(len(summary_cols)):
        if k == 3:                                # 활성월은 합계 무의미 → 공란
            continue
        tot = sum(getattr(res, ("actual_total", "one_off_total", "normalized_total",
                                "active_months", "monthly_run_rate", "annualized")[k])
                  for _, res in results)
        hs.set_cell(ws, r, base + k, tot, role="total",
                    number_format=hs.FMT_NUM1 if k == 4 else hs.FMT_INT)
    for j in range(1, last_col + 1):
        ws.cell(row=r, column=j).border = hs.BORDER_TOP_STRONG
    total_row = r

    # _RECON: Σnormalized + Σone_off == Σactual (tie)
    sum_actual = sum(res.actual_total for _, res in results)
    sum_oneoff = sum(res.one_off_total for _, res in results)
    sum_norm = sum(res.normalized_total for _, res in results)
    recon = vc.recon_block(
        n_input=len(data.lines), n_output=len(fact.rows),
        src_sum=sum_actual, out_sum=sum_norm + sum_oneoff,
        excluded={"one-off 마스킹": sum(len(res.masked_index) for _, res in results)},
        completeness="라인 %d × 기간 %d 전수" % (len(data.lines), nP),
        accuracy="normalized = actual − Σone_off (robust mask)",
        cutoff="월 런레이트 = normalized / active_months (12 하드코딩 금지)",
    )
    rec_top = total_row + 2
    hs.section_header(ws, rec_top, "대사 (Reconciliation)", last_col=last_col)
    hs.write_matrix(ws, rec_top + 1, 1, ["대사 항목", "값"], recon, value_fmt=hs.FMT_INT)
    end_row = rec_top + len(recon) + 1

    if data.commentary:
        cr = rec_top + len(recon) + 3
        cr = hs.section_header(ws, cr, "코멘터리 (▣ = 정규화 시 제외한 one-off)",
                               last_col=last_col)
        for line in data.commentary:
            hs.set_cell(ws, cr, 1, "• " + line, role="soft", align=hs.LEFT_WRAP)
            ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=last_col)
            cr += 1
        end_row = cr

    hs.report_footer(ws, end_row + 1, source="고정비 원장(월별)",
                     prepared_by="FP&A", last_col=last_col)

    # ⚠ one-off 마스킹은 anomaly 가 아니라 정상 정규화 절차 → anomaly_ledger 미노출.
    #   (spine _base_owned_gate 의 anomaly_conserved 는 ledger 노출 시에만 작동.)
    #   마스킹 셀 표식 보존은 qc 의 "one-off 마스킹 surfaced" 체크가 담당.
    wb._fpna_meta = {
        "cal": cal, "periods": periods, "fact": fact, "results": results,
        "sum_actual": sum_actual, "sum_oneoff": sum_oneoff, "sum_norm": sum_norm,
        "masked_surfaced": surfaced,
    }
    return wb


# --------------------------------------------------------------------------- #
# qc                                                                          #
# --------------------------------------------------------------------------- #
def qc(wb: openpyxl.Workbook, data: RunRateInput) -> QCReport:
    rep = QCReport(TYPE)
    qc_no_formula_errors(wb, rep)
    meta = wb._fpna_meta
    cal, periods, fact = meta["cal"], meta["periods"], meta["fact"]

    # R8 grain + R1 시간축 전수
    vc.assert_grain(rep, fact)
    vc.assert_time_ruler(rep, fact, cal, data.start, data.end, period_key="period")
    vc.assert_no_silent_drop(rep, fact, expected_n=len(data.lines) * len(periods))

    # A2 tie: Σnormalized + Σone_off == Σactual (tol=0)
    vc.assert_tie_out(rep, meta["sum_actual"], meta["sum_norm"] + meta["sum_oneoff"],
                      tol=1e-6, name="A2 normalized_tie")

    # 라인별 normalized 재계산 대조 + annualized = monthly × factor 검증(이중연환산 방지)
    for ln, res in meta["results"]:
        fresh = finance.normalized_run_rate(
            list(ln.monthly), deseasonalize=data.deseasonalize,
            annualize_factor=data.annualize_factor)
        ok_norm = finance.approx_equal(fresh.normalized_total, res.normalized_total)
        rep.add("정규화 재계산:%s" % ln.cost_center, ok_norm, "")
        # annualized = monthly_run_rate × factor (factor=12, active_months 로 나눈 월값의 연환산)
        ok_ann = finance.approx_equal(res.annualized,
                                      res.monthly_run_rate * data.annualize_factor)
        rep.add("연환산=월런레이트×%d:%s" % (data.annualize_factor, ln.cost_center),
                ok_ann, "" if ok_ann else "이중연환산 의심")

    # surfaced flag == Σ마스킹 셀 수(은폐 금지 — 마스킹을 표식 없이 빼면 부정직)
    total_masked = sum(len(res.masked_index) for _, res in meta["results"])
    rep.add("one-off 마스킹 surfaced", meta["masked_surfaced"] == total_masked,
            "" if meta["masked_surfaced"] == total_masked
            else "surfaced=%d 마스킹=%d" % (meta["masked_surfaced"], total_masked))

    # over-masking soft 경고(>50% 마스킹 시 런레이트 신뢰 저하) — passed 안 깎음(emit)
    for ln, res in meta["results"]:
        if res.over_masking:
            rep.add("over-masking(soft):%s" % ln.cost_center, True,
                    "마스킹 비율 과다(>50%) — 런레이트 신뢰 점검")

    rep.add("단위 표기", bool(data.unit), "" if data.unit else "unit 비어있음")
    return rep


__all__ = ["TYPE", "RunRateLine", "RunRateInput", "golden_sample", "build", "qc"]
