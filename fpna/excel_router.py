#!/usr/bin/env python3
"""
fpna/excel_router.py — "엑셀" 트리거 진입점. 주어진 .xlsx 를 *파일 검사*로 경로 자동 분기.

사람이 만든 .xlsx 포함, 스킬로 처리 가능한 상황이면 코드가 스스로 모드를 정하고 검수한다
(에이전트 판단 X). main.py excel <파일> [--fix] 가 호출.

분기 신호(파일 내부 검사):
  · 영역 마커(design_zones 센티넬) 또는 .contract.json(blocks) 존재 → **zone**(정형블록+freehand 혼합):
      design_audit.zone_findings strict(정형블록 hard) + 장식/정렬/폰트 검사. --fix=restyle_zone.
  · 그 외(우리 마커 없음 = 사람/외부 파일) → **external**(비파괴 검수):
      design_findings(장식/숫자정렬/폰트) + --fix=restyle_inplace(서식만, 값·수식 불변).
      ※ 심층 재계산(은폐 에러)·golden 은 freehand 스킬 `xlsx_doctor --external`(COM/formulas) 담당.
"""
from __future__ import annotations

import json
import os
import re

from fpna import _bootstrap  # noqa: F401  vendor 주입
from openpyxl import load_workbook

from fpna import design_zones as dz
from fpna import design_audit as da


def _contract_path(path):
    return path.rsplit(".", 1)[0] + ".contract.json"


# 콤마/통화/괄호음수가 텍스트로 저장된 셀(SUM 깨짐의 상류 원인) — 사람/통계청 파일 빈출.
_NUMTEXT = re.compile(r"^\s*[₩$€£]?\s*\(?\s*[-+]?\d{1,3}(,\d{3})+(\.\d+)?\s*\)?\s*$")


def _numbers_as_text(wb):
    """숫자인데 텍스트로 저장된 셀 수집(콤마·괄호음수·통화기호). SUM/참조 깨짐 원인."""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and _NUMTEXT.match(c.value):
                    out.append("%s!%s=%r" % (ws.title, c.coordinate, c.value))
    return out


def classify_excel_file(path):
    """파일 검사 → (mode, ctx). mode ∈ {zone, external}. ctx 는 zone 시 {'contract':...}."""
    wb = load_workbook(path)
    has_zone = any(dz.read_band_maps(ws)[2] is not None for ws in wb.worksheets)
    cp = _contract_path(path)
    contract = None
    if os.path.isfile(cp):
        try:
            obj = json.load(open(cp, encoding="utf-8"))
            if obj.get("blocks"):
                contract = obj
        except (ValueError, OSError):
            contract = None
    if has_zone or contract:
        return "zone", {"contract": contract or {"blocks": {}, "band_map": {}}}
    return "external", {}


def run_excel(path, *, fix=False):
    """파일을 분기 모드로 검수(+--fix 시 비파괴 정규화). 반환 (mode, ok, report)."""
    path = os.path.abspath(path)
    if not os.path.isfile(path):
        return "none", False, ["파일 없음: %s" % path]
    mode, ctx = classify_excel_file(path)
    wb = load_workbook(path)
    report = []
    legacy = da.design_findings(wb)
    decor = legacy["decoration"]

    if mode == "zone":
        contract = ctx["contract"]
        z = da.zone_findings(wb, contract)
        hard = z["resolved_drift"] + z["unsealed"] + z["unknown_block"]
        ok = not (hard or decor)
        report.append("[zone] 정형블록 strict — drift=%d unsealed=%d unknown=%d, 장식문자=%d" % (
            len(z["resolved_drift"]), len(z["unsealed"]), len(z["unknown_block"]), len(decor)))
        for k in ("resolved_drift", "unsealed", "unknown_block"):
            for it in z[k][:6]:
                report.append("  ✗ [%s] %s" % (k, it))
        if fix and not ok:
            actions = da.restyle_zone(wb, contract)
            wb.save(path)
            report.append("[fix] restyle_zone %d건(값 불변): %s" % (len(actions), actions[:4]))
    else:
        rest = legacy["num_align"] + legacy["font"] + legacy["annotation"]
        numtext = _numbers_as_text(wb)
        # 장식문자 + 숫자-텍스트저장(SUM 깨짐)은 hard. 정렬/폰트/주석은 자문성(보고만).
        ok = not (decor or numtext)
        report.append("[external] 사람/외부 파일 비파괴 검수 — 장식(hard)=%d 숫자텍스트(hard)=%d / 정렬·폰트·주석(자문)=%d"
                      % (len(decor), len(numtext), len(rest)))
        for x in decor[:5]:
            report.append("  ✗ 장식 %s" % x)
        for x in numtext[:6]:
            report.append("  ✗ 숫자-텍스트(SUM깨짐) %s" % x)
        for x in rest[:6]:
            report.append("  ⚠ %s (자문)" % x)
        report.append("  · 숫자-텍스트는 ingest(누더기→tidy) 또는 값 재기입 필요. 심층 재계산·golden = freehand 스킬 `xlsx_doctor --external`.")
        if fix:
            changes = da.restyle_inplace(wb)
            wb.save(path)
            report.append("[fix] restyle_inplace %d건(서식만, 값·수식 불변): %s" % (len(changes), changes[:4]))
    return mode, ok, report
