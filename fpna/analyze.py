"""
fpna.analyze — 다중시트 워크북 구조 스캔 + 템플릿/팩 추천.

"이 파일 어떻게 처리하지?"의 진입점. 워크북의 각 시트를 첫 N행만 읽어 infer 로
컬럼 의미를 추론하고, recommend 로 시트별 착지 템플릿을 정한다. 시트가 여러 개면
연동(pack) 후보를 안내한다. ⛔ 도메인 사전 0 — 시트 역할은 *구조*(시계열/항목·금액/
식별자 목록)로만 태그한다(손익/BS 같은 도메인 라벨 X).

★빌드 경유 주의: analyze 는 *읽기 전용 구조 스캔*이다. 실제 산출은 반드시 run_report
스파인 경유(main.py render/pack/report, 또는 autobind.build_checked)로 만들어야 QC
게이트(계약 위반·결측 은폐·tie)가 걸린다. mod.build(inp) 직접 호출은 스파인을 우회해
검증 없이 wb 를 만든다(저장 금지 — 우회 경로). 또한 analyze 는 data_only 로 값만 보므로
시트 간 수식 연결(SUMIFS 등)은 감지하지 못한다 — 연결 보존이 필요하면 원본 fill 모드
또는 pack(공유 facts 재현)을 쓰되, 원본 라이브 수식 링크 자체는 재현되지 않음(값 정합만).
"""
from __future__ import annotations

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna.dispatcher import recommend_from_roles
from fpna.infer import infer_columns, summarize


def _cross_sheet_refs(path: str) -> dict:
    """워크북 수식에서 시트 간 참조 감지 → {시트: {참조하는 다른 시트들}}.

    analyze 기본 경로는 data_only(값만) 라 시트 간 수식 연결을 못 본다. 여기서 수식 모드로
    한 번 더 열어 '시트명!' / '시트명'! 패턴(SUMIFS·VLOOKUP·=Sheet2!A1 등 크로스시트
    의존 = 연동 신호)을 스캔한다. 연결이 있으면 _recommend 가 pack 을 강제 유도한다.
    """
    refs: dict = {}
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return refs
    try:
        names = list(wb.sheetnames)
        for name in names:
            hit: set = set()
            for row in wb[name].iter_rows():
                for c in row:
                    v = c.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    for other in names:
                        if other != name and ((other + "!") in v or ("'%s'!" % other) in v):
                            hit.add(other)
            if hit:
                refs[name] = hit
    finally:
        wb.close()
    return refs


def _sheet_to_rows(ws, sample: int) -> tuple:
    """워크시트 → (headers, list[dict]). 첫 비어있지 않은 행=헤더, 이후 데이터."""
    headers = None
    out: list = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            if any(c is not None for c in row):
                headers = [str(c) if c is not None else "col%d" % j
                           for j, c in enumerate(row)]
            continue
        if len(out) >= sample:
            break
        if all(c is None for c in row):
            continue
        out.append({headers[j]: row[j] for j in range(min(len(headers), len(row)))})
    return headers, out


def _shape_tag(summ: dict) -> str:
    """구조 태그(도메인 무관)."""
    if summ["time"] and summ["has_measure"]:
        return "시계열형"
    if summ["has_measure"] and (summ["dimension"] or summ["id"]):
        return "항목·금액형"
    if not summ["has_measure"]:
        return "식별자 목록형"
    return "수치형"


def analyze_workbook(path: str, *, sample: int = 50) -> dict:
    """워크북 경로 → {sheets: [...], recommendation: {...}}. 시트별 추론 + 워크북 추천."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheets: list = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            headers, rows = _sheet_to_rows(ws, sample)
            if not rows:
                sheets.append({"sheet": name, "shape": "빈/비정형",
                               "template": None, "summary": None})
                continue
            summ = summarize(infer_columns(rows))
            rec = recommend_from_roles(summ)
            sheets.append({"sheet": name, "shape": _shape_tag(summ),
                           "template": rec.template, "reason": rec.reason,
                           "summary": summ})
    finally:
        wb.close()
    cross = _cross_sheet_refs(path)        # 시트 간 수식 연결(연동 신호) 감지
    return {"sheets": sheets, "cross_refs": cross,
            "recommendation": _recommend(sheets, cross)}


def _recommend(sheets: list, cross: dict | None = None) -> dict:
    """시트 조합 → 워크북 수준 추천(연동/단일/다중/없음).

    ★cross(시트 간 수식 연결)가 있으면 최우선으로 'linked' 판정 — 독립 분해하면 원본
    SUMIFS 등 연결이 끊기므로 pack 을 강제 유도한다(인지 못 해 쪼개는 사고 차단).
    """
    cross = cross or {}
    data_sheets = [s for s in sheets if s["template"]]
    if not data_sheets:
        return {"kind": "none", "detail": "데이터 시트 없음(빈/비정형) — ingest 로 정형화 먼저"}
    # ★연동 감지: 시트 간 수식 참조가 하나라도 있으면 pack 필수(독립 분해 금지)
    if cross:
        n_links = sum(len(v) for v in cross.values())
        pairs = "; ".join("%s→{%s}" % (k, ",".join(sorted(v))) for k, v in sorted(cross.items()))
        return {"kind": "linked",
                "detail": "★시트 간 연결 %d개 발견(%s) — 이 워크북은 연동돼 있다. 각 시트를 "
                "독립 분해하면 SUMIFS 등 연결이 끊긴다. pack(공유 facts 재현, packs.md)으로 "
                "묶어라(단일/시트별 분해 금지). 원본 라이브 수식 자체는 재현 안 됨 — 값 정합만."
                % (n_links, pairs)}
    if len(data_sheets) == 1:
        s = data_sheets[0]
        return {"kind": "single", "template": s["template"],
                "detail": "단일 데이터 시트 '%s' → %s (py main.py render %s --csv ...)"
                % (s["sheet"], s["template"], s["template"])}
    listed = ", ".join("%s=%s" % (s["sheet"], s["template"]) for s in data_sheets)
    return {"kind": "multi",
            "detail": "%d개 데이터 시트 → 시트들이 *연동*(공유 가정·크로스시트 tie)이면 "
            "pack(packs.md 카탈로그), 독립이면 시트별 단일. 시트별 추천: %s"
            % (len(data_sheets), listed)}


__all__ = ["analyze_workbook", "_sheet_to_rows", "_shape_tag", "_recommend",
           "_cross_sheet_refs"]
