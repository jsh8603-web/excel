"""
fpna.report — 다중시트 Report 오케스트레이터 (회계법인 워크페이퍼 제본).

자문 3R §B: 회계법인 형식의 본질 = "제본 그 자체"가 아니라 **워크페이퍼 referencing
구조 = 크로스시트 tie-out**(모든 숫자가 소스 스케줄로 추적). 단일시트 = 분석가 작업파일.

설계(자문 R2/R3, 채택):
- 단일시트 sheet-builder 들을 묶는 상위 오케스트레이터. 각 builder 는 ws 에 직접 쓰고
  facts(dict) 를 반환 → 오케스트레이터가 **Python 메모리값**으로 크로스시트 tie 계산.
- 크로스시트 tie = ConserveSpec 재사용(source = 여러 시트 facts 평탄화). 표시용 수식은
  부가일 뿐, 게이트는 메모리값 bool(openpyxl 은 수식 재계산 못 함 → 수식 read=None).
- 차트는 **LAST + single-pass**(openpyxl read-modify-write 가 차트 드롭).
- DimContext(정규 차원) 하향주입 — 시트들이 grain 독립정의하면 조인키 불일치.
- 폴리시 번들: 표지/목차/탭색/하이퍼링크(법인 룩 지각가치).
- n=1 = 단일시트(기존 28종 경로 특수케이스).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from typing import Callable

import fpna._bootstrap  # noqa: F401

import openpyxl

from fpna import house_style as hs
from fpna.conserve import eval_specs
from fpna.templates.base import QCReport, qc_no_formula_errors

# 섹션 type → 탭색(자문 R2: 분석/상세/검증 구분). '#' 없는 RRGGBB.
TAB_COLORS = {
    "cover": "2E5A87", "toc": "5B7FA6", "summary": "2E5A87",
    "detail": "8A8A8A", "check": "C07A2E", "ref": "4E8A6B",
}


@dataclass
class DimContext:
    """정규 차원(conformed dimension) — 오케스트레이터가 소유해 시트에 하향주입.

    시트들이 grain 을 독립정의하면 조인키가 어긋난다(자문 R2). build_report 가
    sheets 의 grain 합집합으로 이 컨텍스트를 구성해 각 builder 에 ctx 로 주입하면,
    시트들이 "같은 축 정의"를 공유한다. 빈 축 = 차원 미선언(패스스루와 동일 자유도).

    periods  : 기간 축 키(예: ("2024-01", "2024-02")).
    accounts : 계정 축 키.
    entities : 엔티티 축 키.
    """
    periods: tuple = ()
    accounts: tuple = ()
    entities: tuple = ()

    @classmethod
    def from_spec(cls, spec) -> "DimContext":
        """ReportSpec.sheets 의 grain 축 *키*를 합집합으로 모아 DimContext 구성.

        SheetSpec.grain 은 {축: (값,...)} 매핑. 알려진 축(periods/accounts/entities)
        키만 합집합(정렬·중복제거)으로 누적한다. grain 미선언 시트는 기여 0 →
        빈 축이면 차원 미강제(하위호환)."""
        acc: dict[str, set] = {"periods": set(), "accounts": set(), "entities": set()}
        for s in getattr(spec, "sheets", []) or []:
            grain = getattr(s, "grain", None) or {}
            for axis in acc:
                vals = grain.get(axis)
                if vals:
                    acc[axis].update(vals)
        return cls(periods=tuple(sorted(acc["periods"])),
                   accounts=tuple(sorted(acc["accounts"])),
                   entities=tuple(sorted(acc["entities"])))


@dataclass
class SheetSpec:
    """1개 시트 선언.

    name    : 시트 탭명(+ 목차 표시).
    builder : (ws, ctx) -> dict[str,float] facts. ws 에 직접 작성, 크로스tie용 facts 반환.
    section : TAB_COLORS 키(탭색·순서 힌트).
    title   : 목차/표지 표시 제목(없으면 name).
    grain   : {축: (값,...)} 시트 차원 선언(예: {"accounts": ("6010","6020")}).
              build_report 가 DimContext(합집합) 구성에 쓰고, qc_report 가 크로스tie
              시트 간 grain 정합(축 키 공유)을 1차 검증한다. 미선언(빈 dict)이면
              차원 무강제(하위호환).
    """
    name: str
    builder: Callable
    section: str = "detail"
    title: str = ""
    grain: dict = field(default_factory=dict)


@dataclass
class ReportSpec:
    """다중시트 리포트 선언."""
    title: str
    sheets: list                      # list[SheetSpec]
    cross_specs: list = field(default_factory=list)   # list[ConserveSpec] over flat facts
    subtitle: str = ""
    as_of: str = ""
    unit: str = "₩"
    source: str = ""
    prepared_by: str = "FP&A"


# 섹션 순서: 표지·목차 먼저, 검증 마지막.
_ORDER = {"cover": 0, "toc": 1, "summary": 2, "detail": 5, "ref": 7, "check": 9}


def _flatten(facts: dict) -> dict:
    """{sheet: {k:v}} → {'sheet.k': v}. 크로스시트 tie source."""
    return {"%s.%s" % (sn, k): v for sn, d in (facts or {}).items()
            for k, v in (d or {}).items()}


def build_report(spec: ReportSpec, ctx=None) -> openpyxl.Workbook:
    """sheet-builder 들을 single-pass 로 묶어 다중시트 wb 생성.

    순서: 데이터 시트 빌드(facts 수집) → 검증 시트(메모리값 tie) → 표지/목차 → 정렬 →
    (차트는 builder 내부에서 자기 시트 마지막에). fullCalcOnLoad 는 스파인 저장 시 부여.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)                      # 기본 빈 시트 제거

    if ctx is None:                           # R4: 패스스루 stub 탈피 — 정규 차원 하향주입
        ctx = DimContext.from_spec(spec)     # sheets grain 합집합(없으면 빈 축)

    titles = _resolve_titles(spec)           # s.name → 31자·유일 시트 제목(충돌 dedup)
    facts: dict = {}
    for s in spec.sheets:
        ws = wb.create_sheet(titles[s.name])
        color = TAB_COLORS.get(s.section)
        if color:
            ws.sheet_properties.tabColor = color
        facts[s.name] = dict(s.builder(ws, ctx) or {})

    flat = _flatten(facts)
    _build_check_sheet(wb, spec, flat)
    _build_cover(wb, spec)
    _build_toc(wb, spec, titles)
    _reorder(wb, spec, titles)

    wb.calculation.fullCalcOnLoad = True       # artifact-gap 완화(B도 스파인과 동일 규율)
    wb._fpna_meta = {"facts": facts, "flat": flat, "titles": titles, "dim_ctx": ctx}
    return wb


def _resolve_titles(spec) -> dict:
    """s.name → 31자 truncate + 유일성 보장 시트 제목. 긴 한글명 충돌 시 silent
    rename(openpyxl) 으로 목차 하이퍼링크·정렬이 깨지는 것을 방지(엣지 리뷰)."""
    used: set = set()
    out: dict = {}
    for s in spec.sheets:
        base = hs.safe_sheet_title(s.name)
        title, k = base, 2
        while title in used:
            suf = "~%d" % k
            title = base[:31 - len(suf)] + suf
            k += 1
        used.add(title)
        out[s.name] = title
    return out


def _build_check_sheet(wb, spec, flat) -> None:
    """크로스시트 tie 결과를 **Python 메모리값 bool** 로 기록(수식 아님)."""
    ws = wb.create_sheet(hs.safe_sheet_title("검증"))
    ws.sheet_properties.tabColor = TAB_COLORS["check"]
    r = hs.report_frame(ws, "크로스시트 검증 (Tie-out)", subtitle="모든 숫자가 소스로 추적되는가",
                        unit=spec.unit, as_of=spec.as_of, last_col=4)
    hs.set_widths(ws, {1: 28, 2: 16, 3: 16, 4: 10})
    for j, h in enumerate(["검증 항목", "독립 재계산", "보고값", "tie"], start=1):
        hs.set_cell(ws, r, j, h, role="header")
    r += 1
    for name, lhs, rhs, tol in eval_specs(spec.cross_specs, flat, flat):
        ok = (lhs is not None) and (rhs is not None) and (abs(lhs - rhs) <= tol)
        hs.set_cell(ws, r, 1, name, role="label", align=hs.LEFT)
        hs.set_cell(ws, r, 2, lhs if lhs is not None else "—", role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 3, rhs if rhs is not None else "—", role="calc", number_format=hs.FMT_INT)
        hs.set_cell(ws, r, 4, "OK" if ok else "XX", role="total",
                    align=hs.CENTER, bold=True)
        r += 1
    hs.report_footer(ws, r + 1, source="크로스시트 = Python 메모리값 대조(표시수식 무의존)",
                     prepared_by=spec.prepared_by, last_col=4)


def _build_cover(wb, spec) -> None:
    ws = wb.create_sheet(hs.safe_sheet_title("표지"))
    ws.sheet_properties.tabColor = TAB_COLORS["cover"]
    hs.style_sheet(ws, freeze=None)
    nxt = hs.title_block(ws, spec.title, spec.subtitle, row=3, last_col=6)
    hs.meta_header(ws, nxt + 1, unit=spec.unit, as_of=spec.as_of)
    hs.set_cell(ws, nxt + 3, 1, "작성: %s" % spec.prepared_by, role="soft", align=hs.LEFT)
    if spec.source:
        hs.set_cell(ws, nxt + 4, 1, "출처: %s" % spec.source, role="soft", align=hs.LEFT)
    hs.set_widths(ws, {1: 14, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14})


def _build_toc(wb, spec, titles) -> None:
    """목차 — 각 시트로 내부 하이퍼링크(#'시트명'!A1). titles=유일 제목 맵."""
    ws = wb.create_sheet(hs.safe_sheet_title("목차"))
    ws.sheet_properties.tabColor = TAB_COLORS["toc"]
    r = hs.title_block(ws, "목차 (Index)", row=2, last_col=4) + 1
    hs.set_widths(ws, {1: 6, 2: 36})
    n = 1
    ordered = sorted(spec.sheets, key=lambda s: _ORDER.get(s.section, 5))
    for s in ordered:
        label = (s.title or s.name)
        sheet_title = titles[s.name]
        hs.set_cell(ws, r, 1, n, role="soft", align=hs.CENTER)
        cell = hs.set_cell(ws, r, 2, label, role="link", align=hs.LEFT)
        cell.hyperlink = "#'%s'!A1" % sheet_title       # 한글·공백 단일따옴표
        r += 1
        n += 1
    # 검증 시트 링크
    hs.set_cell(ws, r, 1, n, role="soft", align=hs.CENTER)
    cell = hs.set_cell(ws, r, 2, "검증 (Tie-out)", role="link", align=hs.LEFT)
    cell.hyperlink = "#'%s'!A1" % hs.safe_sheet_title("검증")
    hs.style_sheet(ws, freeze=None)


def _reorder(wb, spec, titles) -> None:
    """표지(0)·목차(1)·요약·상세·검증(last) 순으로 시트 정렬. titles=유일 제목 맵."""
    title2sec = {titles[s.name]: s.section for s in spec.sheets}

    def key(ws):
        t = ws.title
        if t.startswith("표지"):
            return 0
        if t.startswith("목차"):
            return 1
        if t.startswith("검증"):
            return 99
        return _ORDER.get(title2sec.get(t, "detail"), 5)
    wb._sheets.sort(key=key)        # openpyxl internal — 버전 민감(품질 리뷰 인지)
    wb.active = 0


def _grain_conflict(spec) -> str:
    """크로스tie 에 쓰이는 시트들이 호환 grain(같은 축 키 공유)인지 1차 검증.

    가벼운 게이트(과설계 금지): 두 시트가 *같은 축*에 grain 을 선언했는데 그 축의
    키 교집합이 비면 조인 불가 → 불일치. cross_specs 가 참조하는 시트만 대상으로 하되,
    raw_sum_fn(lambda)은 introspect 불가하므로 grain 을 선언한 모든 시트쌍을 본다.
    축 미선언/단일 시트면 항상 호환(빈 문자열=정합)."""
    graned = [s for s in (getattr(spec, "sheets", []) or [])
              if getattr(s, "grain", None)]
    for i in range(len(graned)):
        for j in range(i + 1, len(graned)):
            ga, gb = graned[i].grain, graned[j].grain
            for axis in set(ga) & set(gb):           # 공통 축만
                ka, kb = set(ga.get(axis) or ()), set(gb.get(axis) or ())
                if ka and kb and not (ka & kb):      # 둘 다 키 있는데 교집합 0
                    return "시트간 grain 불일치: '%s'↔'%s' 축 %s 키 공유 0" % (
                        graned[i].name, graned[j].name, axis)
    return ""


def qc_report(wb, spec: ReportSpec) -> QCReport:
    """리포트 QC — 수식에러 0 + 크로스시트 tie(메모리값) + grain 정합. 스파인이 호출."""
    rep = QCReport("report:" + spec.title)
    qc_no_formula_errors(wb, rep)
    conflict = _grain_conflict(spec)              # R4: 시트간 grain 1차 정합
    rep.add("grain 정합", not conflict, conflict)
    flat = (getattr(wb, "_fpna_meta", None) or {}).get("flat", {})   # build_report 외 경로 방어
    for name, lhs, rhs, tol in eval_specs(spec.cross_specs, flat, flat):
        if lhs is None or rhs is None:
            rep.add("크로스tie:%s" % name, False,
                    "raw 계산 실패" if lhs is None else "reported 부재")
        else:
            rep.add("크로스tie:%s" % name, abs(lhs - rhs) <= tol,
                    "" if abs(lhs - rhs) <= tol else "독립=%.6g 보고=%.6g" % (lhs, rhs))
    return rep


__all__ = ["SheetSpec", "ReportSpec", "DimContext", "build_report", "qc_report",
           "TAB_COLORS"]
