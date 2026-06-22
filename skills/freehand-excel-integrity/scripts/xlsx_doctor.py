#!/usr/bin/env python3
"""
xlsx_doctor.py — 클로드 코드가 *직접 작성한* .xlsx 를 검사(+서식 자동수리)하는 단일파일.

의존성: openpyxl 만(템플릿/repo 불필요). 클로드 코드가 어떤 작업 폴더에서 엑셀을
프리핸드로 만들든, 끝에 이 스크립트를 돌려 "깨진 산출"을 스스로 잡고 고치게 한다.
per-change subagent 검수를 대체하는 deterministic 게이트.

무엇을 잡나 (W26 스크린샷 버그 클래스 기준)
  [1] 수식 에러 리터럴(#REF!/#VALUE!/#DIV/0! …)        신뢰 · 수리불가→재생성
  [2] 숫자영역 텍스트 침투(헤더 "AUG" 누수 / 주석 누수)  신뢰 · 수리불가→재생성
  [3] 열 내 서식 불균형(엉뚱한 셀만 다른 서식)            휴리스틱 · --fix 수리
  [4] 가드 없는 나눗셈(#VALUE!/#DIV/0! 후보, CPU/CPP)     자문성
  [5] cross-foot 의심(소계 ≠ 자식합)                      자문성(부호규약 모름)

수리 경계(정직)
  ▶ 서식(3)은 결정적 수리 가능.
  ▶ 값/텍스트(1·2·4·5)는 dead 파일만으로 자동수리 불가 — 올바른 값은 원천에서
    다시 와야 한다. 클로드 코드가 *재작성*하는 게 정답(아래 가이드 참조).

사용
  python xlsx_doctor.py report.xlsx            # 검사만(exit 0/1)
  python xlsx_doctor.py report.xlsx --fix      # 서식 수리 → report.fixed.xlsx
  python xlsx_doctor.py report.xlsx --fix --inplace
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter, defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl 필요: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

ERROR_LITERALS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!",
                  "#NUM!", "#SPILL!", "#CALC!", "#GETTING_DATA"}
HEADER_TOKENS = {"APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC",
                 "JAN", "FEB", "MAR", "Q1", "Q2", "Q3", "Q4", "FY",
                 "1Q", "2Q", "3Q", "4Q"}


def _isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _isf(v):
    return isinstance(v, str) and v.startswith("=")


# --------------------------------------------------------------------------- #
def scan_error_cells(wb):
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERROR_LITERALS:
                    bad.append("%s!%s=%s" % (ws.title, c.coordinate, c.value))
    return bad


def text_in_numeric_columns(wb, min_majority=3):
    """숫자-다수 열의 텍스트 침투. 헤더토큰은 데이터 시작행보다 아래면 누수,
    그 외 텍스트는 데이터 구간(시작행 이상)이면 누수."""
    out = []
    for ws in wb.worksheets:
        bycol = defaultdict(lambda: {"num_rows": [], "txt": []})
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None or _isf(v):
                    continue
                if _isnum(v):
                    bycol[c.column]["num_rows"].append(c.row)
                elif isinstance(v, str) and v.strip():
                    bycol[c.column]["txt"].append((c.row, c.coordinate, v))
        for col, d in bycol.items():
            if len(d["num_rows"]) < min_majority:
                continue
            first = min(d["num_rows"])
            for r, coord, v in d["txt"]:
                tok = v.strip().upper() in HEADER_TOKENS
                if tok and r > first:
                    out.append((ws.title, coord, v, "헤더토큰 누수"))
                elif not tok and r >= first:
                    out.append((ws.title, coord, v, "주석/텍스트 누수"))
    return out


def column_format_outliers(wb, min_majority=3):
    out = []
    for ws in wb.worksheets:
        bycol = defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                if _isnum(c.value):
                    bycol[c.column].append((c.coordinate, c.number_format))
        for col, items in bycol.items():
            if len(items) < min_majority + 1:
                continue
            cnt = Counter(nf for _, nf in items)
            (top_nf, top_n), = cnt.most_common(1)
            if top_n < min_majority:
                continue
            for coord, nf in items:
                if nf != top_nf and cnt[nf] <= max(1, top_n // 3):
                    out.append({"sheet": ws.title, "col": col, "coord": coord,
                                "found": nf, "expected": top_nf})
    return out


def unguarded_divisions(wb):
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not (isinstance(v, str) and v.startswith("=") and "/" in v):
                    continue
                up = v.upper()
                # 가드 인식: IFERROR/IFNA/ISNUMBER 또는 IF(...분모=0...) 래핑.
                if "IFERROR" in up or "ISNUMBER" in up or "IFNA" in up:
                    continue
                if up.startswith("=IF(") and ("=0" in up or "<>0" in up or ">0" in up):
                    continue   # IF 로 분모 0 가드한 정상 비율
                out.append("%s!%s=%s" % (ws.title, c.coordinate, v))
    return out


def crossfoot_suspects(wb, abs_tol=0.5):
    susp = []
    for ws in wb.worksheets:
        bycol = defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                if _isnum(c.value):
                    bycol[c.column].append((c.row, c.value, c))
        for col, items in bycol.items():
            items.sort()
            for i, (r, v, cell) in enumerate(items):
                totalish = bool(getattr(cell.font, "bold", False)) and \
                    getattr(cell.border, "top", None) is not None and \
                    getattr(cell.border.top, "style", None) is not None
                if not totalish or i == 0:
                    continue
                run, pr = [], r
                for rr, vv, _ in reversed(items[:i]):
                    if rr == pr - 1:
                        run.append(vv); pr = rr
                    else:
                        break
                if len(run) >= 2 and abs(sum(run) - v) > abs_tol:
                    susp.append("%s!%s(=%g, 위 %d셀 합=%g)" %
                                (ws.title, cell.coordinate, v, len(run), sum(run)))
    return susp


# --------------------------------------------------------------------------- #
# 사이드카 계약(생성시 _fpna_meta 의 프리핸드 버전) — 값 차원 불변식 포팅       #
#   <파일>.contract.json 로 클로드 코드가 "약속"을 선언하면 doctor 가 대조.      #
# --------------------------------------------------------------------------- #
def _split_ref(ref, default_sheet):
    """'SHEET!A1' / 'A1' → (sheet, addr)."""
    if "!" in ref:
        s, a = ref.split("!", 1)
        return s, a
    return default_sheet, ref


def _cells_in(ws, a1):
    """'A1' 또는 'A1:B5' → 셀 객체 리스트(평탄)."""
    from openpyxl.utils import range_boundaries
    c0, r0, c1, r1 = range_boundaries(a1)
    out = []
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            out.append(ws.cell(row=r, column=c))
    return out


def _totalish_cells(wb):
    """굵게+상단테두리 숫자셀(합계 관습) 좌표 집합."""
    s = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if not _isnum(c.value):
                    continue
                if bool(getattr(c.font, "bold", False)) and \
                        getattr(c.border, "top", None) is not None and \
                        getattr(c.border.top, "style", None) is not None:
                    s.add("%s!%s" % (ws.title, c.coordinate))
    return s


def _division_cells(wb):
    s = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and "/" in c.value:
                    s.add("%s!%s" % (ws.title, c.coordinate))
    return s


def contract_coverage(wb, contract):
    """계약 커버리지(약한 고리 보완): 합계형 셀·나눗셈 셀이 있는데 ties/ratios 에
    선언이 없으면 WARN. 계약이 *무엇을 안 보는지*를 드러내 작성자가 빠뜨리지 않게.
    반환 [(WARN, msg), ...]."""
    out = []
    contract = contract or {}
    declared_t = set()
    for t in contract.get("ties", []):
        ds = contract.get("sheet") or (wb.worksheets[0].title if wb.worksheets else "")
        s, a = _split_ref(t["total"], ds)
        declared_t.add("%s!%s" % (s, a))
    declared_r = set()
    for r in contract.get("ratios", []):
        ds = contract.get("sheet") or (wb.worksheets[0].title if wb.worksheets else "")
        ref = r["cell"] if isinstance(r, dict) else r
        s, a = _split_ref(ref, ds)
        declared_r.add("%s!%s" % (s, a))

    totals = _totalish_cells(wb)
    divs = _division_cells(wb)
    uncovered_t = totals - declared_t
    uncovered_r = divs - declared_r
    if uncovered_t:
        out.append(("WARN", "합계형 셀 %d개가 ties 미선언: %s (소계 정합 미검증)"
                    % (len(uncovered_t), ", ".join(sorted(uncovered_t)[:6]))))
    if uncovered_r:
        out.append(("WARN", "나눗셈 셀 %d개가 ratios 미선언: %s (#VALUE! 가드 미검증)"
                    % (len(uncovered_r), ", ".join(sorted(uncovered_r)[:6]))))
    return out


def check_contract(wb, contract):
    """사이드카 계약 검증. 반환 [(레벨, 메시지), ...]. 레벨 ∈ {FAIL, WARN}."""
    out = []
    by_sheet = {ws.title: ws for ws in wb.worksheets}
    default_sheet = contract.get("sheet") or (wb.worksheets[0].title if wb.worksheets else "")

    # 1) tie-out (R3/R10/R11/R14): total == Σparts (값) 또는 == expected(독립 N-version).
    for t in contract.get("ties", []):
        ts, ta = _split_ref(t["total"], default_sheet)
        name = t.get("name", t["total"])
        if ts not in by_sheet:
            out.append(("FAIL", "tie '%s': 시트 없음" % name)); continue
        total_cell = _cells_in(by_sheet[ts], ta)[0]
        tv = total_cell.value
        tol = t.get("tol", 0.5)
        # (a) 작성자가 소스에서 독립 계산한 기대총계 — 정적 값 워크북의 진짜 N-version
        if "expected" in t:
            if _isnum(tv):
                if abs(tv - t["expected"]) > tol:
                    out.append(("FAIL", "tie '%s': 렌더 total=%g ≠ expected(독립)=%g"
                                % (name, tv, t["expected"])))
            else:
                out.append(("FAIL", "tie '%s': total 셀이 숫자 아님(%r) — 정적 값 기대" % (name, tv)))
        # (b) parts 와의 내부 정합(렌더 합계 == 렌더 자식합)
        if t.get("parts"):
            ps, pa = _split_ref(t["parts"], default_sheet)
            if ps in by_sheet:
                parts = [c.value for c in _cells_in(by_sheet[ps], pa) if _isnum(c.value)]
                if _isf(tv):
                    rng = pa.split("!")[-1].upper().replace(" ", "")
                    if not ("SUM(" in tv.upper().replace(" ", "") and rng in tv.upper().replace(" ", "")):
                        out.append(("FAIL", "tie '%s': total 수식이 parts(%s) 를 SUM 으로 안 덮음"
                                    % (name, pa.split("!")[-1])))
                elif _isnum(tv):
                    if abs(sum(parts) - tv) > tol:
                        out.append(("FAIL", "tie '%s': total=%g ≠ Σparts=%g (%d개)"
                                    % (name, tv, sum(parts), len(parts))))
                elif "expected" not in t:
                    out.append(("FAIL", "tie '%s': total 셀이 숫자/수식 아님(%r)" % (name, tv)))

    # 2) grain 유일성(R8): 선언 key 범위에 중복 라벨 금지(침묵 병합 차단)
    g = contract.get("grain")
    if g and g.get("region"):
        gs, ga = _split_ref(g["region"], default_sheet)
        ws = by_sheet.get(gs)
        if ws:
            vals = [c.value for c in _cells_in(ws, ga)
                    if isinstance(c.value, str) and c.value.strip()]
            dup = [k for k, n in Counter(vals).items() if n > 1]
            if dup:
                out.append(("FAIL", "grain 중복(침묵 병합 위험): %s" % ", ".join(dup[:6])))

    # 3) ratio 완전성(R17): 정적 값이면 num/den 재계산 대조, 수식이면 가드 확인.
    for ref in contract.get("ratios", []):
        if isinstance(ref, dict):
            # 값 모드: {cell, num, den, tol} — .py 계산 정적 비율의 독립 재계산
            cs, ca = _split_ref(ref["cell"], default_sheet)
            ws = by_sheet.get(cs)
            if not ws:
                continue
            cv = _cells_in(ws, ca)[0].value
            ns, na = _split_ref(ref["num"], default_sheet)
            ds, da = _split_ref(ref["den"], default_sheet)
            num = _cells_in(by_sheet[ns], na)[0].value if ns in by_sheet else None
            den = _cells_in(by_sheet[ds], da)[0].value if ds in by_sheet else None
            tol = ref.get("tol", 0.01)
            if not (_isnum(den) and den != 0):
                # 분모 무효 → 셀은 NA 센티넬(문자열)이거나 빈칸이어야(숫자/에러 금지)
                if _isnum(cv) or (isinstance(cv, str) and cv in ERROR_LITERALS):
                    out.append(("FAIL", "ratio %s: 분모 무효인데 값/에러(%r) — 'NA' 센티넬 권장"
                                % (ref["cell"], cv)))
            elif _isnum(num):
                if not _isnum(cv):
                    out.append(("FAIL", "ratio %s: 숫자 아님(%r), 기대 %g" % (ref["cell"], cv, num/den)))
                elif abs(cv - num/den) > max(tol, abs(num/den)*tol):
                    out.append(("FAIL", "ratio %s: %g ≠ num/den=%g (독립 재계산)"
                                % (ref["cell"], cv, num/den)))
            continue
        # 문자열 ref: 수식 가드 모드(라이브 수식 워크북)
        rs, ra = _split_ref(ref, default_sheet)
        ws = by_sheet.get(rs)
        if not ws:
            continue
        v = _cells_in(ws, ra)[0].value
        if _isf(v):
            up = v.upper()
            if not ("ISNUMBER" in up or "IFERROR" in up or "IFNA" in up):
                out.append(("FAIL", "ratio %s 가드 없음(#VALUE! 위험): %s" % (ref, v)))
        elif isinstance(v, str) and v in ERROR_LITERALS:
            out.append(("FAIL", "ratio %s 이미 에러: %s" % (ref, v)))

    # 4) FieldSpec(metric_table): 범위·부호·허용값 (binding/metric_table 포팅)
    for fs in contract.get("fields", []):
        s, a = _split_ref(fs["region"], default_sheet)
        ws = by_sheet.get(s)
        if not ws:
            continue
        cells = _cells_in(ws, a)
        nums = [(c.coordinate, c.value) for c in cells if _isnum(c.value)]
        sign = fs.get("sign")
        for coord, v in nums:
            if sign in ("+", "nonneg") and v < 0:
                out.append(("FAIL", "field %s!%s 음수(기대 %s): %g" % (s, coord, sign, v)))
            elif sign in ("-", "nonpos") and v > 0:
                out.append(("FAIL", "field %s!%s 양수(기대 %s): %g" % (s, coord, sign, v)))
            if "min" in fs and v < fs["min"]:
                out.append(("FAIL", "field %s!%s < min %s: %g" % (s, coord, fs["min"], v)))
            if "max" in fs and v > fs["max"]:
                out.append(("FAIL", "field %s!%s > max %s: %g" % (s, coord, fs["max"], v)))
        if "accepted_values" in fs:
            allowed = set(fs["accepted_values"])
            for c in cells:
                if c.value is not None and not _isf(c.value) and c.value not in allowed:
                    out.append(("FAIL", "field %s!%s 허용값 밖: %r" % (s, c.coordinate, c.value)))

    # 5) 시나리오 정합(R9): actual/budget 라벨 집합이 같은 모집단인지
    sc = contract.get("scenario")
    if sc and sc.get("actual") and sc.get("budget"):
        def _labels(ref):
            s, a = _split_ref(ref, default_sheet)
            ws = by_sheet.get(s)
            return {c.value for c in _cells_in(ws, a)
                    if isinstance(c.value, str) and c.value.strip()} if ws else set()
        a_keys, b_keys = _labels(sc["actual"]), _labels(sc["budget"])
        ao, bo = a_keys - b_keys, b_keys - a_keys
        if ao or bo:
            out.append(("FAIL", "scenario 모집단 불일치(0처리 금지, LEFT/RIGHT_ONLY 노출): "
                        "actual-only %s, budget-only %s"
                        % (sorted(ao)[:4], sorted(bo)[:4])))

    # 6) 행수 보존(R7 no_silent_drop): 선언 region 의 비어있지 않은 라벨 수 == n
    en = contract.get("expected_n")
    if en and en.get("region") and "n" in en:
        s, a = _split_ref(en["region"], default_sheet)
        ws = by_sheet.get(s)
        if ws:
            got = sum(1 for c in _cells_in(ws, a)
                      if c.value is not None and str(c.value).strip())
            if got != en["n"]:
                out.append(("FAIL", "행수 불일치(누락/추가): 기대 %d ≠ 실제 %d" % (en["n"], got)))

    # 7) 수식 방향/칼럼 참조(E): region 의 각 행 수식이 =<left><행><op><right><행>
    from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
    for fc in contract.get("formula_refs", []):
        s, a = _split_ref(fc["region"], default_sheet)
        ws = by_sheet.get(s)
        if not ws:
            continue
        op = fc.get("op", "-")
        left = fc["left"] if isinstance(fc["left"], int) else column_index_from_string(fc["left"])
        right = fc["right"] if isinstance(fc["right"], int) else column_index_from_string(fc["right"])
        c0, r0, c1, r1 = range_boundaries(a)
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                v = ws.cell(row=r, column=c).value
                if not _isf(v):
                    continue
                want = "=%s%d%s%s%d" % (get_column_letter(left), r, op,
                                       get_column_letter(right), r)
                if v.replace("$", "").replace(" ", "").upper() != want.upper():
                    out.append(("FAIL", "formula %s!%s 참조 불일치: 기대 %s ≠ %s"
                                % (s, ws.cell(row=r, column=c).coordinate, want, v)))

    # 8) 단위/스케일 정합(unit): 선언 region 의 nonzero 값이 같은 자릿수대인지(스케일 혼용 차단)
    import math
    for u in contract.get("units", []):
        s, a = _split_ref(u["region"], default_sheet)
        ws = by_sheet.get(s)
        if not ws:
            continue
        mags = [math.floor(math.log10(abs(c.value))) for c in _cells_in(ws, a)
                if _isnum(c.value) and c.value != 0]
        if mags and (max(mags) - min(mags)) >= 3:
            out.append(("FAIL", "unit %s: 스케일 혼용 의심(자릿수 %d~%d; %s 일관성 확인)"
                        % (u["region"], min(mags), max(mags), u.get("unit", "?"))))

    # 9) 시간축/기간 연속성(R1 time_ruler): 선언 header 범위가 기대 기간과 정확히 일치
    for p in contract.get("periods", []):
        s, a = _split_ref(p["header"], default_sheet)
        ws = by_sheet.get(s)
        if not ws:
            continue
        got = [c.value for c in _cells_in(ws, a)]
        exp = p["expected"]
        if got != exp:
            # 무엇이 틀렸는지 구체화: 누락/중복/순서
            missing = [x for x in exp if x not in got]
            dup = [x for x in set(got) if got.count(x) > 1]
            detail = []
            if missing:
                detail.append("누락 %s" % missing[:5])
            if dup:
                detail.append("중복 %s" % dup[:5])
            if not detail:
                detail.append("순서/값 불일치: 기대 %s ≠ %s" % (exp, got))
            out.append(("FAIL", "periods %s: %s" % (p["header"], "; ".join(detail))))

    return out


# --------------------------------------------------------------------------- #
_NUMTEXT_RE = re.compile(r"^\(?\s*[₩$€£¥]?\s*-?\d{1,3}(,\d{3})+(\.\d+)?\s*\)?%?$|^\(?\s*[₩$€£¥]\s*-?\d+(\.\d+)?\s*\)?$|^-?\d+(\.\d+)?%$")


def numbers_as_text(wb):
    """콤마/통화/괄호음수/% 형태의 '서식된 숫자가 텍스트로' 저장된 셀(SUM 무시 → 0).
    bare '5000'/'2024' 는 id/연도일 수 있어 제외(고정밀)."""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and _NUMTEXT_RE.match(v.strip()):
                    out.append("%s!%s=%r" % (ws.title, c.coordinate, v))
    return out


def recalc_smell(path):
    """수식 캐시 상태를 분리: none=계산 안 됨(생성직후 정상·재계산이면 채워짐),
    zero=캐시가 0(xlsxwriter 기본/은폐 가능 — 더 의심). 반환 (none_list, zero_list)."""
    try:
        from openpyxl import load_workbook as _lw
        wf = _lw(path); wd = _lw(path, data_only=True)
    except Exception:
        return [], []
    none_cells, zero_cells = [], []
    for ws in wf.worksheets:
        wsd = wd[ws.title] if ws.title in wd.sheetnames else None
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    cached = wsd[c.coordinate].value if wsd is not None else None
                    if cached is None:
                        none_cells.append("%s!%s" % (ws.title, c.coordinate))
                    elif cached == 0:
                        zero_cells.append("%s!%s=%s" % (ws.title, c.coordinate, c.value))
    return none_cells, zero_cells


def source_numbers(src_path):
    """소스(csv/xlsx)의 모든 숫자를 진실집합으로(원배율 + 천/백만 스케일 변형 포함)."""
    vals = set()
    def add(x):
        if _isnum(x):
            for s in (1, 1_000, 1_000_000, 0.001, 1e-6):
                vals.add(round(float(x) * s, 2))
    if src_path.lower().endswith((".csv", ".tsv")):
        import csv
        with open(src_path, encoding="utf-8-sig", newline="") as f:
            for row in csv.reader(f, delimiter="\t" if src_path.endswith(".tsv") else ","):
                for cell in row:
                    try:
                        add(float(str(cell).replace(",", "").replace("(", "-").replace(")", "")))
                    except Exception:
                        pass
    else:
        from openpyxl import load_workbook as _lw
        wb = _lw(src_path, data_only=True)
        for ws in wb.worksheets:
            for r in ws.iter_rows(values_only=True):
                for c in r:
                    add(c)
    return vals


def provenance_untraced(wb, src_vals, declared=frozenset(), tol=0.01):
    """워크북의 정적 숫자 중 소스 진실집합에 추적되지 않는 셀(=날조/오타 의심).
    선언된 합계/비율(declared)·수식·0 은 파생이므로 제외."""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not _isnum(v) or v == 0:
                    continue
                ref = "%s!%s" % (ws.title, c.coordinate)
                if ref in declared:
                    continue
                if round(v, 2) not in src_vals:
                    out.append("%s=%s" % (ref, v))
    return out


def accessibility(wb):
    """접근성/투명성 자문(보수적·저오탐): 데이터 병합·기본 시트명·숨김 시트 데이터.
    치명 아님 — 공유 산출물 품질 신호."""
    out = []
    for ws in wb.worksheets:
        # 기본 시트명
        if re.match(r"^Sheet\d*$", ws.title or ""):
            out.append("기본 시트명 '%s'(의미있는 이름 권장)" % ws.title)
        # 숨김 시트에 데이터
        if getattr(ws, "sheet_state", "visible") != "visible":
            has_data = any(c.value is not None for row in ws.iter_rows() for c in row)
            if has_data:
                out.append("숨김 시트 '%s'에 데이터(투명성/감사 주의)" % ws.title)
        # 숫자를 포함한 병합(스크린리더·합계 방해). 텍스트-only 제목 병합은 제외.
        for mr in ws.merged_cells.ranges:
            try:
                vals = [ws.cell(row=r, column=c).value
                        for r in range(mr.min_row, mr.max_row + 1)
                        for c in range(mr.min_col, mr.max_col + 1)]
            except Exception:
                continue
            if any(_isnum(v) for v in vals):
                out.append("%s!%s 숫자 셀 병합(스크린리더/SUM 방해)" % (ws.title, str(mr)))
    return out


def _rel_form(formula, row):
    """수식의 행참조를 자기행 기준 상대오프셋으로 정규화(=C5-B5 @r5 → =C[0]-B[0])."""
    def repl(m):
        col, r = m.group(1), int(m.group(2))
        return "%s[%+d]" % (col, r - row)
    return re.sub(r"\$?([A-Z]{1,3})\$?(\d+)", repl, formula.replace("$", "").replace(" ", "").upper())


def formula_consistency(wb, min_run=3):
    """한 열의 수식들이 동일한 상대형(fill-down)인지. 한 행만 형태가 다르면 파손 후보."""
    from collections import Counter, defaultdict
    out = []
    for ws in wb.worksheets:
        bycol = defaultdict(list)
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("=") and ":" not in c.value:
                    bycol[c.column].append((c.row, c.coordinate, _rel_form(c.value, c.row)))
        for col, items in bycol.items():
            if len(items) < min_run:
                continue
            forms = Counter(f for _, _, f in items)
            (top, n), = forms.most_common(1)
            if n < min_run - 1:
                continue  # 지배형 없음 → 의도적 혼합
            for r, coord, f in items:
                if f != top and forms[f] <= max(1, n // 3):
                    out.append("%s!%s (열 다수형 ≠ 이 셀; fill-down 파손 의심)" % (ws.title, coord))
    return out


def doctor(path, fix=False, inplace=False, contract_path=None, do_recalc=False, source=None):
    path = os.path.abspath(path)
    if not os.path.isfile(path):
        print("파일 없음:", path); return 2
    wb = load_workbook(path)

    err = scan_error_cells(wb)
    numtext = numbers_as_text(wb)
    fconsist = formula_consistency(wb)
    a11y = accessibility(wb)
    stray = text_in_numeric_columns(wb)
    outliers = column_format_outliers(wb)
    unguarded = unguarded_divisions(wb)
    crossfoot = crossfoot_suspects(wb)

    # 사이드카 계약 자동탐색: <파일>.contract.json
    if contract_path is None:
        cand = path.rsplit(".", 1)[0] + ".contract.json"
        contract_path = cand if os.path.isfile(cand) else None
    contract_findings = []
    coverage = []
    contract_obj = {}
    if contract_path and os.path.isfile(contract_path):
        import json
        with open(contract_path, encoding="utf-8") as f:
            contract_obj = json.load(f)
        contract_findings = check_contract(wb, contract_obj)
    coverage = contract_coverage(wb, contract_obj) if contract_path else []

    print("=" * 60)
    print("XLSX DOCTOR:", os.path.basename(path))
    print("=" * 60)
    fatal = False

    print("\n[1] 수식 에러 리터럴: %s" % ("없음" if not err else "%d건" % len(err)))
    for x in err[:12]:
        print("    ✗", x)
    if err:
        fatal = True
        print("    → 자동수리 불가. 참조/분모 오류 — 재작성 필요.")

    print("\n[2] 숫자영역 텍스트 침투: %s" % ("없음" if not stray else "%d건" % len(stray)))
    for s, coord, v, kind in stray[:12]:
        print("    ✗ %s!%s = %r  [%s]" % (s, coord, v, kind))
    if stray:
        fatal = True
        print("    → #VALUE! 의 상류 원인(분모 오염). 값 자동수리 불가 — 재작성 필요.")

    print("\n[3] 열 내 서식 불균형: %s" % ("없음" if not outliers else "%d건" % len(outliers)))
    for o in outliers[:12]:
        print("    ✗ %s!%s = %r (열 다수서식 %r)" %
              (o["sheet"], o["coord"], o["found"], o["expected"]))

    print("\n[4] 가드 없는 나눗셈: %s  [자문성]" %
          ("없음" if not unguarded else "%d건" % len(unguarded)))
    for x in unguarded[:12]:
        print("    ⚠", x)
    if unguarded:
        print("    → CPU/CPP 류는 =IF(OR(NOT(ISNUMBER(분모)),분모=0),\"NA\",분자/분모) 로.")

    print("\n[5] cross-foot 의심: %s  [자문성·부호규약 모름→오탐 가능]" %
          ("없음" if not crossfoot else "%d건" % len(crossfoot)))
    for x in crossfoot[:12]:
        print("    ⚠", x)

    print("\n[6] 사이드카 계약(tie/grain/ratio/field/scenario/formula): %s" % (
        "선언 없음" if contract_path is None else
        ("통과 ✓" if not contract_findings else "%d건 위반" % len(contract_findings))))
    for lvl, msg in contract_findings[:12]:
        print("    %s %s" % ("✗" if lvl == "FAIL" else "⚠", msg))
        if lvl == "FAIL":
            fatal = True

    print("\n[7] 계약 커버리지(미선언 합계/비율): %s  [자문성]" %
          ("해당없음(계약 미사용)" if not contract_path else
           ("완전" if not coverage else "%d건" % len(coverage))))
    for _lvl, msg in coverage[:12]:
        print("    ⚠", msg)
    if coverage:
        print("    → 합계/비율은 ties/ratios 로 선언해야 값 검증이 걸린다(미선언=무검증).")

    print("\n[8] 숫자-텍스트 저장(SUM 무시→0): %s" % ("없음" if not numtext else "%d건" % len(numtext)))
    for x in numtext[:12]:
        print("    ✗", x)
    if numtext:
        fatal = True
        print("    → 콤마/통화/괄호음수가 텍스트로 저장됨. 숫자로 기입(서식은 number_format 으로).")

    none_cells, zero_cells = recalc_smell(path)
    print("\n[9] 재계산 상태: 미계산 %d · 캐시0 %d  [자문성]" % (len(none_cells), len(zero_cells)))
    if none_cells:
        print("    · 미계산 %d개(캐시 None) — 생성 직후 정상. Excel/--recalc 시 채워짐." % len(none_cells))
    for x in zero_cells[:8]:
        print("    ⚠ %s (캐시=0 — 소비자가 0으로 봄; 재계산/value= 확인)" % x)
    if zero_cells:
        print("    → 캐시 0 은 xlsxwriter value= 누락 등 — write_formula(..., value=) 또는 재계산.")

    print("\n[11] 수식 fill-down 일관성: %s" % ("일관" if not fconsist else "%d건" % len(fconsist)))
    for x in fconsist[:10]:
        print("    ✗", x)
    if fconsist:
        fatal = True
        print("    → 같은 열인데 한 행만 참조형이 다름(엉뚱한 칼럼/방향). 동일 상대수식으로 채울 것.")

    print("\n[13] 접근성/투명성: %s  [자문성]" % ("양호" if not a11y else "%d건" % len(a11y)))
    for x in a11y[:10]:
        print("    ⚠", x)

    if source:
        try:
            declared = set()
            for t in contract_obj.get("ties", []):
                ds = contract_obj.get("sheet") or wb.worksheets[0].title
                s, a = _split_ref(t["total"], ds); declared.add("%s!%s" % (s, a))
            for r0 in contract_obj.get("ratios", []):
                ds = contract_obj.get("sheet") or wb.worksheets[0].title
                ref = r0["cell"] if isinstance(r0, dict) else r0
                s, a = _split_ref(ref, ds); declared.add("%s!%s" % (s, a))
            sv = source_numbers(source)
            untraced = provenance_untraced(wb, sv, declared)
            print("\n[12] 출처추적(소스 대조): %s  [자문성]" %
                  ("전부 추적됨" if not untraced else "%d건 미추적" % len(untraced)))
            for x in untraced[:12]:
                print("    ✗ %s (소스에 없음 — 날조/오타 의심)" % x)
            if untraced:
                fatal = True
                print("    → --source 는 '모든 숫자가 추적돼야'의 선언. 파생이면 ties/ratios 로 선언해 제외.")
        except Exception as e:
            print("\n[12] 출처추적: 건너뜀(%s)" % e)

    if do_recalc:
        try:
            import recalc as _rc
            rr = _rc.recalc(path)
            errs, stale = rr.get("errors", []), rr.get("stale", [])
            print("\n[10] 실재계산(%s): 에러 %d · stale %d  %s" %
                  (rr.get("engine"), len(errs), len(stale), rr.get("note", "")))
            for c, code in errs[:10]:
                print("    ✗ %s = %s (은폐된 진짜 에러)" % (c, code)); fatal = True
            for c, old, new in stale[:10]:
                print("    ✗ %s: 캐시=%r → 계산=%s (캐시값이 틀림 — 소비자가 본 값 ≠ 진짜)" % (c, old, new))
                fatal = True
            if rr.get("engine") == "formulas" and (errs or stale):
                print("    (formulas 엔진은 일부 함수 시맨틱이 Excel 과 다름 — pywin32/LibreOffice 가 정확)")
        except Exception as e:
            print("\n[10] 실재계산: 건너뜀(%s)" % e)

    n_fixed = 0
    if fix and outliers:
        bysheet = {ws.title: ws for ws in wb.worksheets}
        for o in outliers:
            bysheet[o["sheet"]][o["coord"]].number_format = o["expected"]
            n_fixed += 1
        out = path if inplace else path.rsplit(".", 1)[0] + ".fixed.xlsx"
        wb.save(out)
        print("\n[FIX] 서식 %d건 수리 → %s (값/텍스트 이슈는 미수리)" %
              (n_fixed, os.path.basename(out)))

    print("\n" + "-" * 60)
    if not (fatal or outliers):
        print("진단: 치명 이슈 없음 ✓"); return 0
    if fix and not fatal and n_fixed:
        print("진단: 서식 수리 완료. 값/텍스트 치명 이슈 없음 ✓"); return 0
    print("진단: 이슈 발견 — 위 항목 확인.")
    return 1


def main(argv=None):
    ap = argparse.ArgumentParser(description="클로드 코드가 만든 .xlsx 검사/수리(openpyxl만 필요)")
    ap.add_argument("path")
    ap.add_argument("--fix", action="store_true")
    ap.add_argument("--inplace", action="store_true")
    ap.add_argument("--contract", default=None, help="사이드카 계약 JSON 경로(기본: <파일>.contract.json 자동탐색)")
    ap.add_argument("--recalc", action="store_true", help="헤드리스 재계산으로 은폐 에러·stale 캐시 검출(recalc.py 필요)")
    ap.add_argument("--source", default=None, help="출처추적: 모든 숫자가 이 소스(csv/xlsx)로 추적되는지 대조")
    a = ap.parse_args(argv)
    return doctor(a.path, fix=a.fix, inplace=a.inplace, contract_path=a.contract,
                  do_recalc=a.recalc, source=a.source)


if __name__ == "__main__":
    sys.exit(main())
