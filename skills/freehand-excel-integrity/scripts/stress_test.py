#!/usr/bin/env python3
"""
stress_test.py — 4 백엔드 버킷별 모의 워크북으로 게이트를 스트레스(규칙 확정용).

각 버킷의 대표 실패모드를 *고의로* 심은 픽스처를 만들고, xlsx_doctor 가 잡는지 단언한다.
헤드리스에서 돌아가는 openpyxl/xlsxwriter 는 실측, xlwings/COM 은 Excel 없으면 SKIP
(그 자체가 "이 버킷은 Excel 필요" 규칙의 증거).

실행: python3 stress_test.py
"""
import os
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
DOCTOR = os.path.join(HERE, "xlsx_doctor.py")


def run_doctor(path):
    r = subprocess.run([sys.executable, DOCTOR, path], capture_output=True,
                       text=True, encoding="utf-8", errors="replace")
    return r.returncode, (r.stdout or "")


def expect(out, needle, case):
    ok = needle in out
    print("  %s %s : '%s'" % ("✅" if ok else "❌", case, needle))
    return ok


def bucket_openpyxl(d):
    import openpyxl
    p = os.path.join(d, "op.xlsx")
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    # 합계=0 버그 + 텍스트-숫자 + 캐시없는 수식
    for i, v in enumerate([100, 200, 300]):
        ws.cell(row=2 + i, column=2, value=v).number_format = "#,##0"
    ws.cell(row=5, column=2, value=0).number_format = "#,##0"       # 소계 0(버그)
    ws.cell(row=6, column=2, value="1,234")                        # 숫자-텍스트
    ws.cell(row=7, column=2, value="=SUM(B2:B4)")                  # 캐시 None
    wb.save(p)
    import json
    json.dump({"sheet": "S", "ties": [{"name": "subtotal", "total": "S!B5", "parts": "S!B2:B4"}]},
              open(p.rsplit(".", 1)[0] + ".contract.json", "w"))
    _, out = run_doctor(p)
    base = all([
        expect(out, "[8] 숫자-텍스트", "openpyxl: 숫자-텍스트 탐지(fatal)"),
        expect(out, "tie 'subtotal'", "openpyxl: 소계 미합산 탐지(fatal)"),
    ])
    # 캐시 None(생성직후)은 양성 → --recalc 로 진짜 값 surfacing 확인
    r = subprocess.run([sys.executable, DOCTOR, p, "--recalc"], capture_output=True,
                       text=True, encoding="utf-8", errors="replace")
    rc = expect(r.stdout or "", "[10] 실재계산", "openpyxl: --recalc 로 stale/에러 surfacing")
    return base and rc


def bucket_xlsxwriter(d):
    try:
        import xlsxwriter
    except ImportError:
        print("  ⏭  xlsxwriter 미설치 — SKIP"); return None
    p = os.path.join(d, "xw.xlsx")
    wb = xlsxwriter.Workbook(p); ws = wb.add_worksheet("S")
    ws.write_number(0, 1, 100); ws.write_number(1, 1, 200)
    ws.write_formula(2, 1, "=SUM(B1:B2)")          # value= 미동봉 → 캐시 0
    wb.close()
    _, out = run_doctor(p)
    return expect(out, "캐시=0", "xlsxwriter: 캐시0 silent-zero 탐지")


def bucket_xlwings(d):
    try:
        import xlwings  # noqa
    except ImportError:
        print("  ⏭  xlwings/Excel 미가용 — SKIP (이 버킷은 Excel 필요 = 규칙 증거)")
        return None
    print("  ⏭  (Excel 런타임에서만 의미; 여기선 로직 검증 생략)")
    return None


def bucket_com(d):
    try:
        import win32com  # noqa
    except ImportError:
        print("  ⏭  pywin32/COM 미가용 — SKIP (이 버킷은 Windows+Excel 필요 = 규칙 증거)")
        return None
    return None


def main():
    results = {}
    with tempfile.TemporaryDirectory() as d:
        print("[openpyxl 버킷]"); results["openpyxl"] = bucket_openpyxl(d)
        print("[xlsxwriter 버킷]"); results["xlsxwriter"] = bucket_xlsxwriter(d)
        print("[xlwings 버킷]"); results["xlwings"] = bucket_xlwings(d)
        print("[pywin32 COM 버킷]"); results["com"] = bucket_com(d)
    ran = {k: v for k, v in results.items() if v is not None}
    print("\n결과: %d/%d 버킷 통과 (SKIP %d)" %
          (sum(1 for v in ran.values() if v), len(ran), len(results) - len(ran)))
    return 0 if all(ran.values()) else 1


if __name__ == "__main__":
    sys.exit(main())
