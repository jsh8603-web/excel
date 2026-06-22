#!/usr/bin/env python3
"""
recalc_check.py — 수식 재계산 검증 폴백 체인.

수식 구동 워크북은 셀에 적힌 `=SUM/=A-B` 가 실제로 에러 없이 계산되는지 확인해야 한다
(디스크의 수식 텍스트만 보는 doctor 로는 #VALUE!/#DIV/0! 를 못 잡는 경우가 있다). 환경에
따라 가용한 엔진이 다르므로 폴백 체인으로 처리한다:

  ① pywin32 있음            → verify_xlsx.py (실 Excel COM 재계산)
  ② 없고 LibreOffice 있음   → soffice --headless 변환 재계산 후 에러 스캔
  ③ 둘 다 없고 수식 워크북  → `formulas` 패키지로 평가 (커버리지 한계 경고)
  ④ 정적 값 워크북          → 재계산 불요 (doctor 가 expected/num·den 로 오프라인 검증)

이 스크립트는 검증 보조 도구다. `formulas`/pywin32 는 런타임 의존성이 아니라
'있으면 쓰는' 선택 엔진이며, 부재해도 깔끔히 빠진다(무설치 환경 배려).

의존성: stdlib + openpyxl(정적/수식 판별). 스킬 번들.
Exit: 0 = 통과/재계산 불요 / 1 = 재계산 에러 발견 / 2 = 엔진 없음(미검증 경고).
"""
from __future__ import annotations
import sys, os, re, shutil, subprocess, argparse, tempfile, importlib.util

_HERE = os.path.dirname(os.path.abspath(__file__))
for _cand in ("../vendor", "../../vendor", "../../../vendor"):
    _p = os.path.join(_HERE, _cand)
    if os.path.isdir(_p):
        sys.path.insert(0, _p)
        break
from openpyxl import load_workbook  # noqa: E402

_ERR = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def _mod(name: str) -> bool:
    try:
        return importlib.util.find_spec(name) is not None
    except (ImportError, ValueError):
        return False


def _soffice() -> str | None:
    for exe in ("soffice", "libreoffice"):
        p = shutil.which(exe)
        if p:
            return p
    return None


def detect_engines() -> dict:
    """가용한 재계산 엔진 탐지 → preflight 가 게이트 분기에 사용."""
    return {"pywin32": _mod("win32com"), "soffice": _soffice() is not None, "formulas": _mod("formulas")}


def is_formula_workbook(path: str) -> bool:
    wb = load_workbook(path, data_only=False, keep_links=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        return True
    finally:
        wb.close()
    return False


def _scan_errors_openpyxl(path: str) -> list[str]:
    """data_only 로 (캐시된) 계산값을 읽어 에러 리터럴 스캔."""
    out = []
    wb = load_workbook(path, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and any(e in c.value for e in _ERR):
                        out.append(f"{ws.title}!{c.coordinate}={c.value}")
    finally:
        wb.close()
    return out


def via_soffice(path: str) -> tuple[int, list[str]]:
    exe = _soffice()
    tmp = tempfile.mkdtemp()
    try:
        # LibreOffice 는 변환(로드) 시 수식을 재계산하고 결과를 캐시에 굽는다.
        subprocess.run([exe, "--headless", "--calc", "--convert-to",
                        "xlsx:Calc MS Excel 2007 XML", "--outdir", tmp, path],
                       check=False, timeout=120,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        out_path = os.path.join(tmp, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
        if not os.path.isfile(out_path):
            return 2, ["soffice 변환 산출 없음 — 재계산 미검증"]
        errs = _scan_errors_openpyxl(out_path)
        return (1, errs) if errs else (0, [])
    except Exception as e:
        return 2, [f"soffice 실행 실패: {e}"]


def via_formulas(path: str) -> tuple[int, list[str]]:
    try:
        import formulas  # type: ignore
        xl = formulas.ExcelModel().loads(path).finish()
        sol = xl.calculate()
        errs = []
        for k, v in sol.items():
            val = getattr(v, "value", v)
            sval = str(val)
            if any(e in sval for e in _ERR):
                errs.append(f"{k}={sval[:40]}")
        # 커버리지 한계: formulas 는 일부 함수 미지원 — 경고로 동반.
        return (1, errs) if errs else (0, ["formulas 평가 통과 (단, 일부 함수 미지원 가능 — 커버리지 한계)"])
    except Exception as e:
        return 2, [f"formulas 평가 실패: {e} — 재계산 미검증"]


def check(path: str) -> int:
    if not os.path.isfile(path):
        print("파일 없음:", path)
        return 2
    if not is_formula_workbook(path):
        print("recalc: 정적 값 워크북 → 재계산 불요 (doctor 가 expected/num·den 로 오프라인 검증)")
        return 0

    eng = detect_engines()
    if eng["pywin32"]:
        print("recalc: pywin32 → verify_xlsx.py 위임")
        return subprocess.call([sys.executable, os.path.join(_HERE, "verify_xlsx.py"), path])
    if eng["soffice"]:
        print("recalc: LibreOffice(soffice) --headless 재계산")
        rc, msgs = via_soffice(path)
    elif eng["formulas"]:
        print("recalc: `formulas` 패키지 평가 (커버리지 한계)")
        rc, msgs = via_formulas(path)
    else:
        print("recalc: ⚠ 엔진 없음(pywin32/LibreOffice/formulas) — 수식 결과 미검증(커버리지 한계)")
        return 2
    for m in msgs:
        print(("  ✗ " if rc == 1 else "  · ") + m)
    print("recalc PASS" if rc == 0 else ("recalc FAIL" if rc == 1 else "recalc WARN(미검증)"))
    return rc


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--detect", action="store_true", help="가용 엔진만 출력")
    a = ap.parse_args(argv)
    if a.detect:
        for k, v in detect_engines().items():
            print(f"  {k}: {'있음' if v else '없음'}")
        return 0
    return check(a.path)


if __name__ == "__main__":
    sys.exit(main())
