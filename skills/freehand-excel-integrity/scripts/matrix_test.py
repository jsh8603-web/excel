#!/usr/bin/env python3
"""
matrix_test.py — 12+ 시나리오 × 4 백엔드 매트릭스. 오탐(FP)/미탐(FN)을 동시 검출해 함수를 깎는다.

백엔드: openpyxl·xlsxwriter 는 실측. xlwings·COM 은 Excel 부재 → 산출물은 openpyxl 로 동등
생성(프록시; 게이트는 *파일*을 보므로 동등). 백엔드 고유 위험(lifecycle/part)은 별도 smoke.

각 셀: doctor 판정(clean='치명 이슈 없음', dirty='이슈 발견')을 기대와 대조.
  FP = clean 인데 dirty 판정 / FN = dirty 인데 clean 판정.
"""
import json
import os
import subprocess
import sys
import tempfile

HERE = os.path.dirname(os.path.abspath(__file__))
DOCTOR = os.path.join(HERE, "xlsx_doctor.py")
import openpyxl
try:
    import xlsxwriter
    HAVE_XW = True
except ImportError:
    HAVE_XW = False


def verdict(path, flags):
    r = subprocess.run([sys.executable, DOCTOR, path] + flags, capture_output=True, text=True)
    o = r.stdout
    return "dirty" if "이슈 발견" in o else ("clean" if "치명 이슈 없음" in o else "?")


# ---- 빌더: backend(op|xw) 로 파일 작성. 반환 False=해당 백엔드 미적용 ----
def b_clean_table(path, be):
    if be == "xw" and not HAVE_XW: return False
    rows = [("Seg", "Units", "Sales")] + [("A", 100, 1000), ("B", 200, 2000), ("C", 150, 1500)]
    if be == "xw":
        wb = xlsxwriter.Workbook(path); ws = wb.add_worksheet("S")
        for i, r in enumerate(rows):
            for j, v in enumerate(r): ws.write(i, j, v)
        wb.close()
    else:
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
        for r in rows: ws.append(list(r))
        wb.save(path)
    return True

def b_clean_formula(path, be):
    # 합계 =SUM, 가드 비율; 정적 데이터. (xw 는 value 동봉)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S"
    ws.append(["Seg", "Units", "Sales", "Margin"])
    for i, (n, u, s) in enumerate([("A",100,1000),("B",200,2000)], start=2):
        ws.cell(row=i,column=1,value=n); ws.cell(row=i,column=2,value=u); ws.cell(row=i,column=3,value=s)
        ws.cell(row=i,column=4,value='=IF(B%d=0,"NA",C%d/B%d)'%(i,i,i))
    ws.cell(row=4,column=2,value="=SUM(B2:B3)"); ws.cell(row=4,column=3,value="=SUM(C2:C3)")
    wb.save(path); return True

def b_text_in_num(path, be):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="S"
    for i,v in enumerate([100,200,300],2): ws.cell(row=i,column=2,value=v)
    ws.cell(row=5,column=2,value="합계 메모")          # 숫자열에 텍스트
    wb.save(path); return True

def b_numtext(path, be):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="S"
    ws["B2"]="1,234"; ws["B3"]="5,678"                  # 콤마 텍스트
    wb.save(path); return True

def b_merged_title(path, be):
    # 정상: 제목만 병합(텍스트) → FP 나면 안 됨
    wb = openpyxl.Workbook(); ws = wb.active; ws.title="Report"
    ws["A1"]="2026 Report"; ws.merge_cells("A1:D1")
    for i,v in enumerate([100,200,300],3): ws.cell(row=i,column=2,value=v)
    wb.save(path); return True

def b_multisheet(path, be):
    wb = openpyxl.Workbook(); ws=wb.active; ws.title="Summary"
    ws.append(["K","V"]); ws.append(["a",1]); ws.append(["b",2])
    ws2=wb.create_sheet("Detail"); ws2.append(["x",10]); ws2.append(["y",20])
    wb.save(path); return True

def b_xw_nocache(path, be):
    if be != "xw" or not HAVE_XW: return False
    wb=xlsxwriter.Workbook(path); ws=wb.add_worksheet("S")
    ws.write_number(0,1,100); ws.write_number(1,1,200)
    ws.write_formula(2,1,"=SUM(B1:B2)")                 # value 누락 → 캐시0
    wb.close(); return True

def b_subtotal0(path, be):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="S"
    for i,v in enumerate([532,479,527],2): ws.cell(row=i,column=2,value=v)
    ws.cell(row=5,column=2,value=0)                     # 소계 0(버그) → contract tie
    wb.save(path); return True

def b_scale_mix(path, be):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="S"
    ws["B2"]=1200; ws["B3"]=1500; ws["B4"]=1_300_000    # 스케일 혼용 → contract units
    wb.save(path); return True

def b_period_gap(path, be):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="S"
    for j,h in enumerate(["APR","MAY","Q2","Q1","Q2","FY"],2): ws.cell(row=12,column=j,value=h)
    wb.save(path); return True

def b_fill_break(path, be):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="S"
    for r in range(2,8): ws.cell(row=r,column=4,value="=C%d-B%d"%(r,r))
    ws.cell(row=5,column=4,value="=D5-B5")              # fill-down 파손
    wb.save(path); return True

def b_fabricated(path, be):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="S"
    ws["B2"]=100; ws["B3"]=200; ws["B4"]=999999.42      # 소스에 없는 날조 → --source
    wb.save(path); return True


SCEN = [
    # name, builder, contract, flags, expect
    ("clean_table",     b_clean_table,   None, [], "clean"),
    ("clean_formula",   b_clean_formula, None, [], "clean"),
    ("clean_merged_title", b_merged_title, None, [], "clean"),
    ("clean_multisheet", b_multisheet,   None, [], "clean"),
    ("text_in_numeric", b_text_in_num,   None, [], "dirty"),
    ("numbers_as_text", b_numtext,       None, [], "dirty"),
    ("xw_nocache",      b_xw_nocache,    None, ["--recalc"], "dirty"),
    ("subtotal_zero",   b_subtotal0,     {"sheet":"S","ties":[{"name":"sub","total":"S!B5","parts":"S!B2:B4"}]}, [], "dirty"),
    ("scale_mix",       b_scale_mix,     {"sheet":"S","units":[{"region":"S!B2:B4","unit":"천원"}]}, [], "dirty"),
    ("period_gap",      b_period_gap,    {"sheet":"S","periods":[{"header":"S!B12:G12","expected":["APR","MAY","JUN","Q1","Q2","FY"]}]}, [], "dirty"),
    ("fill_break",      b_fill_break,    None, [], "dirty"),
    ("fabricated",      b_fabricated,    None, ["--source","SRC"], "dirty"),
]

BACKENDS = ["openpyxl", "xlsxwriter", "xlwings*", "com*"]  # * = openpyxl 프록시(Excel 부재)


def main():
    fp, fn, total, skip = [], [], 0, 0
    # 출처추적용 소스(날조 시나리오)
    with tempfile.TemporaryDirectory() as d:
        src = os.path.join(d, "src.xlsx")
        sb = openpyxl.Workbook(); ss = sb.active
        for v in [100,200,300,532,479,527,1000,2000,1500,150]: ss.append([v])
        sb.save(src)
        print("%-18s | %s" % ("scenario", " ".join("%-11s" % b for b in BACKENDS)))
        print("-"*70)
        for name, builder, contract, flags, expect in SCEN:
            cells = []
            for be in BACKENDS:
                real_be = "xw" if be == "xlsxwriter" else "op"  # xlwings*/com* → openpyxl 프록시
                path = os.path.join(d, "%s_%s.xlsx" % (name, be.strip("*")))
                if not builder(path, real_be):
                    cells.append("N/A"); continue
                if contract:
                    json.dump(contract, open(path.rsplit(".",1)[0]+".contract.json","w"))
                fl = [f if f != "SRC" else src for f in flags]
                v = verdict(path, fl)
                total += 1
                ok = (v == expect)
                if not ok and expect == "clean": fp.append("%s/%s" % (name, be))
                if not ok and expect == "dirty": fn.append("%s/%s" % (name, be))
                cells.append(("✅" if ok else "❌") + v)
            print("%-18s | %s" % (name, " ".join("%-11s" % c for c in cells)))
    print("-"*70)
    print("실행 %d셀 | FP(정상→dirty) %d | FN(버그→clean) %d" % (total, len(fp), len(fn)))
    if fp: print("  FP:", ", ".join(fp))
    if fn: print("  FN:", ", ".join(fn))
    return 0 if not (fp or fn) else 1


if __name__ == "__main__":
    sys.exit(main())
