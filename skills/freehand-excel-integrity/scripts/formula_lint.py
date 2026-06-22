#!/usr/bin/env python3
"""
formula_lint.py — 수식 일관성 린터: 같은 열의 수식이 행마다 같은 상대참조 형태인지.

fill-down(아래로 채우기) 으로 깐 수식 열은 행이 내려가도 **상대참조 형태가 동일**해야
한다. 예: C2=`=A2-B2`, C3=`=A3-B3`, … 는 모두 "현재행 A − 현재행 B" 형태다. 누가
한 셀만 `=A3-B2`(B 가 한 행 위 고정) 로 잘못 깔면 fill-down 이 깨진 것 — 값은 그럴듯해
보여도 행마다 다른 셀을 가리킨다. 이 린터는 각 열 수식을 **자기 행 기준 상대 오프셋**으로
정규화해, 같은 열에서 소수만 형태가 어긋난 셀을 fill-down 파손으로 적발한다.

의존성: openpyxl + stdlib (이식 가능). 스킬 번들.
쓰임:  python3 scripts/formula_lint.py <file.xlsx>
Exit:  0 = 일관 / 1 = 파손 의심 셀 발견.
"""
from __future__ import annotations
import sys, os, re, argparse
from collections import defaultdict, Counter

for _cand in ("../vendor", "../../vendor", "../../../vendor"):
    _p = os.path.join(os.path.dirname(os.path.abspath(__file__)), _cand)
    if os.path.isdir(_p):
        sys.path.insert(0, _p)
        break
from openpyxl import load_workbook  # noqa: E402
from openpyxl.utils import coordinate_to_tuple  # noqa: E402

_REF = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def normalize(formula: str, cell_row: int) -> str:
    """수식 안 모든 셀참조의 행번호를 '자기 행 기준 상대 오프셋' 으로 치환.

    절대행($있음)은 'A$=절대값' 으로 보존(고정 의도이므로 형태 비교에 그대로 둠).
    상대행은 'A[r±k]' (k=참조행−셀행) 로 — fill-down 이면 열마다 k 가 동일해진다.
    """
    def repl(m):
        col_abs, col, row_abs, row = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if row_abs:  # 절대행 = 고정 의도
            return f"{col}$ABS{row}"
        return f"{col}[r{row - cell_row:+d}]"
    return _REF.sub(repl, formula.replace(" ", ""))


def lint(path: str) -> list[str]:
    wb = load_workbook(path, data_only=False, keep_links=True)
    issues: list[str] = []
    try:
        for ws in wb.worksheets:
            # 열 → [(row, normalized_form, raw)]
            by_col: dict[int, list[tuple[int, str, str]]] = defaultdict(list)
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if isinstance(v, str) and v.startswith("="):
                        r, col = coordinate_to_tuple(c.coordinate)
                        by_col[col].append((r, normalize(v, r), v))
            for col, entries in by_col.items():
                if len(entries) < 3:
                    continue  # 형태 다수결을 논하려면 최소 3개
                forms = Counter(e[1] for e in entries)
                majority, maj_n = forms.most_common(1)[0]
                if maj_n == len(entries):
                    continue  # 전부 동일 = 일관
                # 다수 형태와 어긋난 소수 셀을 파손 의심으로 보고
                for r, form, raw in entries:
                    if form != majority and forms[form] < maj_n:
                        issues.append(
                            f"{ws.title}!{_col_letter(col)}{r}: fill-down 형태 어긋남  "
                            f"'{raw}'  (열 다수 형태와 불일치 — 참조 열/행 고정 오류 의심)"
                        )
    finally:
        wb.close()
    return issues


def _col_letter(idx: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    a = ap.parse_args(argv)
    issues = lint(a.path)
    if not issues:
        print("FORMULA LINT PASS: 열별 수식 상대참조 형태 일관")
        return 0
    for i in issues:
        print(f"  ✗ {i}")
    print(f"FORMULA LINT FAIL: {len(issues)}건 (fill-down 파손 의심)")
    return 1


if __name__ == "__main__":
    sys.exit(main())
