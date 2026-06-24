#!/usr/bin/env python3
"""
house_style_min.py — repo fpna/house_style.py 의 최소 미러(vendor). freehand 가 같은 룩을
적용하도록 set_cell(role)/title_font/brand_header/FMT_*/BORDER_* 를 동일 API 로 제공한다.

목적: 디자인 *준수*를 생성 단계에서 강제. freehand 가 색·폰트·정렬을 직접 박지 말고 이 헬퍼만
쓰면, repo 파이프라인(run_report→render→house_style)과 같은 결과가 나온다. xlsx_doctor [14]
린터는 이 모듈의 토큰(허용 폰트 크기·역할 색)을 읽어 *이 엔진이 내는 산출엔 침묵*한다.

repo SSOT 와 동기화할 것(룩 변경은 repo house_style.py 가 정본). 여기 값은 graphite 테마 기준.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- 타이포 스케일(repo SIZE_* 미러) ----
SIZE_BODY = 10
SIZE_SMALL = 9
SIZE_EYEBROW = 9
SIZE_TITLE = 20          # 영문 헤드라인
SIZE_SUBTITLE = 10.5     # 한글 부제
SIZE_SECTION = 11
ALLOWED_SIZES = {SIZE_BODY, SIZE_SMALL, SIZE_EYEBROW, SIZE_TITLE, SIZE_SUBTITLE, SIZE_SECTION}

# ---- graphite 테마 팔레트(repo DEFAULT_THEME 미러) ----
INK = "111418"; INK_SOFT = "5B616B"
ACCENT = "12404A"; ACCENT_SOFT = "BBD3D6"
EYEBROW_FG = "12404A"
HEADER_BAND = "EEF2F3"; HEADER_FG = "111418"; BAND = "F6F8F8"
POS_FG = "1F7A1F"; NEG_FG = "C0392B"
INPUT_FG = "0B66C2"; LINK_FG = "1F7A1F"; CALC_FG = INK
TITLE_FONT = "Calibri Light"; BODY_FONT = "맑은 고딕"; BODY_FONT_LATIN = "Calibri"
# 역할→색(린터 교차검증용)
ROLE_FG = {"calc": CALC_FG, "input": INPUT_FG, "link": LINK_FG, "label": INK,
           "soft": INK_SOFT, "header": HEADER_FG, "total": INK}

# ---- 숫자 서식(repo FMT_* 미러) ----
FMT_INT = "#,##0;(#,##0)"
FMT_INT_DASH = '#,##0;(#,##0);"-"'
FMT_NUM1 = "#,##0.0;(#,##0.0)"
FMT_PCT1 = "0.0%;(0.0%)"
FMT_MULT = '0.0"x"'
FMT_INT_MN = "#,##0,,;(#,##0,,)"
FMT_KPI_ARROW = '[Green]"▲"0.0%;[Red]"▼"0.0%;0.0%'
FMT_DATE = "yyyy-mm-dd"

# ---- 정렬(라벨 좌·숫자 우) ----
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")

# ---- 보더(합계 상단 단일 = IB 관례) ----
_thin = Side(style="thin", color="D8DCE0"); _med = Side(style="medium", color="AAB0B7")
_dbl = Side(style="double", color="AAB0B7")
BORDER_TOP_STRONG = Border(top=_med)
BORDER_TOTAL = Border(top=_med, bottom=_dbl)

FILL_HEADER = PatternFill("solid", fgColor=HEADER_BAND)


def font(color=None, *, bold=False, size=SIZE_BODY, italic=False, name=None):
    return Font(name=name or BODY_FONT, size=size, bold=bold, italic=italic, color=color or INK)


def title_font(*, size=SIZE_TITLE, bold=False, color=None):
    return Font(name=TITLE_FONT, size=size, bold=bold, color=color or INK)


def set_cell(ws, row, col, value=None, *, role="calc", number_format=None,
             align=None, bold=False, border=None, fill=None, size=SIZE_BODY):
    """값+서식 일괄. role ∈ {calc,input,link,label,soft,header,total}. 숫자 우측·라벨 좌측."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = {
        "calc": font(CALC_FG, bold=bold, size=size),
        "input": font(INPUT_FG, bold=bold, size=size),
        "link": font(LINK_FG, bold=bold, size=size),
        "label": font(INK, bold=bold, size=size),
        "soft": font(INK_SOFT, bold=bold, size=size),
        "header": font(HEADER_FG, bold=True, size=SIZE_SMALL),
        "total": font(INK, bold=True, size=size),
    }.get(role, font(CALC_FG, bold=bold, size=size))
    if role == "header":
        cell.fill = FILL_HEADER
        cell.alignment = align or CENTER
        cell.border = Border(bottom=Side(style="medium", color=ACCENT))
    else:
        cell.alignment = align or (LEFT if role in ("label", "soft") else RIGHT)
    if fill is not None:
        cell.fill = fill
    if number_format:
        cell.number_format = number_format
    if border is not None:
        cell.border = border
    return cell


def brand_header(ws, title_en, subtitle_kr="", *, eyebrow="", row=1, col=1):
    """빅4 표지 관례: ① eyebrow(영문 대문자) ② 영문 헤드라인 ③ 한글 부제 ④ 액센트 룰.
    '메인 제목은 영문' 규칙을 코드로 강제."""
    r = row
    if eyebrow:
        c = ws.cell(row=r, column=col, value=eyebrow.upper())
        c.font = font(EYEBROW_FG, bold=True, size=SIZE_EYEBROW, name=BODY_FONT_LATIN)
        c.alignment = LEFT; r += 1
    c = ws.cell(row=r, column=col, value=title_en)
    c.font = title_font(size=SIZE_TITLE); c.alignment = LEFT; r += 1
    if subtitle_kr:
        c = ws.cell(row=r, column=col, value=subtitle_kr)
        c.font = font(INK_SOFT, size=SIZE_SUBTITLE); c.alignment = LEFT; r += 1
    return r  # 다음 사용 가능 행


def meta_header(ws, row, col, *, unit="₩mn", currency="KRW", basis="US GAAP", asof=""):
    """단위·통화·회계기준·기준일 한 줄(회색 소프트)."""
    txt = "Unit: %s   Currency: %s   Basis: %s%s" % (unit, currency, basis,
                                                     ("   As of: %s" % asof if asof else ""))
    c = ws.cell(row=row, column=col, value=txt)
    c.font = font(INK_SOFT, size=SIZE_SMALL); c.alignment = LEFT
    return row + 1


def _infer_role(cell):
    """기존 셀의 폰트 색으로 role 추정(편집 시 역할 유지)."""
    try:
        rgb = cell.font.color.rgb if cell.font and cell.font.color else None
        code = rgb[-6:].upper() if isinstance(rgb, str) else None
    except Exception:
        code = None
    inv = {v.upper(): k for k, v in ROLE_FG.items()}
    if code in inv:
        return inv[code]
    # 정렬로 라벨/숫자 보조 추정
    if cell.alignment and cell.alignment.horizontal == "left":
        return "label"
    return "calc"


def edit_cell(ws, ref, value, *, role=None, number_format=None):
    """기존 셀을 *역할·서식 유지*하며 값만 교체. 직접 ws[ref]=v 대신 이걸 써서 드리프트 차단.
    role 미지정 시 기존 폰트색에서 추정, number_format 미지정 시 기존 유지."""
    c = ws[ref] if isinstance(ref, str) else ws.cell(row=ref[0], column=ref[1])
    r = role or _infer_role(c)
    nf = number_format or (c.number_format if c.number_format and c.number_format != "General" else None)
    prior_empty = c.value is None
    size = SIZE_BODY if prior_empty else (c.font.size if (c.font and c.font.size in ALLOWED_SIZES) else SIZE_BODY)
    return set_cell(ws, c.row, c.column, value, role=r, number_format=nf, size=size)


def style_sheet(ws, *, freeze="A2", zoom=100):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = zoom
    if freeze:
        ws.freeze_panes = freeze


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
