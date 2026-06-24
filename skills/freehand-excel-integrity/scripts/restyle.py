#!/usr/bin/env python3
"""
restyle.py — 외부 입수 .xlsx 를 house_style 표준으로 *비파괴* 정규화(서식만).

보장: 숫자·수식 셀의 값은 절대 바꾸지 않는다(정규화 후 검증으로 단언). 바꾸는 것은
  · 숫자 셀의 좌/가운데 정렬 → 우측정렬
  · 비표준 폰트 크기 → house_style 스케일로 스냅
  · 라벨 텍스트의 장식문자(별표/마크다운 강조/장식선) 제거(텍스트는 의미 보존)
헤더 밴드·색 역할은 추정 위험이 있어 여기선 건드리지 않고 [14] 가 보고만 한다.

사용: python3 restyle.py <in.xlsx> [-o out.xlsx]   (기본 out=<in>.restyled.xlsx)
"""
from __future__ import annotations

import re
import sys

import house_style_min as hs

_DECOR_STRIP = re.compile(r"^\s*[\*★☆■◆▶●=\-_]{1,}\s*|\s*[\*★☆■◆▶●=\-_]{1,}\s*$")
_MD_EMPH = re.compile(r"\*\*(.+?)\*\*")
_ALLOWED = sorted(hs.ALLOWED_SIZES)


def _snap_size(s):
    if s in hs.ALLOWED_SIZES:
        return s
    if s > hs.SIZE_TITLE:
        return hs.SIZE_TITLE
    return min(_ALLOWED, key=lambda a: abs(a - s))


def _clean_label(text):
    t = _MD_EMPH.sub(r"\1", text)
    prev = None
    while prev != t:
        prev = t
        t = _DECOR_STRIP.sub("", t)
    return t.strip()


def _is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def _isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def restyle(in_path: str, out_path: str | None = None):
    from openpyxl import load_workbook
    out_path = out_path or (in_path.rsplit(".", 1)[0] + ".restyled.xlsx")
    wb = load_workbook(in_path)
    # 정규화 전 숫자/수식 값 스냅샷(불변 검증용)
    before = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if _isnum(c.value) or _is_formula(c.value):
                    before[(ws.title, c.coordinate)] = c.value

    changes = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                # 1) 숫자 셀 우측정렬(값·서식코드 불변)
                if _isnum(v) and c.alignment and c.alignment.horizontal in ("left", "center"):
                    c.alignment = hs.RIGHT
                    changes.append("%s!%s 정렬→우측" % (ws.title, c.coordinate))
                # 2) 비표준 폰트 크기 스냅(값 불변)
                if c.font and c.font.size and c.font.size not in hs.ALLOWED_SIZES and v not in (None, ""):
                    old = c.font.size
                    ns = _snap_size(old)
                    c.font = hs.Font(name=c.font.name, size=ns, bold=c.font.bold,
                                     italic=c.font.italic, color=c.font.color)
                    changes.append("%s!%s 폰트 %g→%g" % (ws.title, c.coordinate, old, ns))
                # 3) 라벨 텍스트 장식 제거(숫자/수식은 제외 → 값 안전)
                if isinstance(v, str) and not _is_formula(v):
                    cleaned = _clean_label(v)
                    if cleaned != v and cleaned:
                        c.value = cleaned
                        changes.append("%s!%s 라벨 정리 %r→%r" % (ws.title, c.coordinate, v[:24], cleaned[:24]))

    # 불변 검증: 숫자/수식 값이 그대로인지 단언
    violated = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                key = (ws.title, c.coordinate)
                if key in before and c.value != before[key]:
                    violated.append("%s!%s 값 변경됨!(%r→%r)" % (ws.title, c.coordinate, before[key], c.value))
    if violated:
        raise AssertionError("restyle 가 값을 바꿈(버그): " + "; ".join(violated[:5]))

    wb.save(out_path)
    return out_path, changes


def main(argv=None):
    import argparse
    ap = argparse.ArgumentParser(description="외부 엑셀 비파괴 디자인 정규화(서식만)")
    ap.add_argument("path")
    ap.add_argument("-o", "--out", default=None)
    a = ap.parse_args(argv)
    out, changes = restyle(a.path, a.out)
    print("정규화 %d건 → %s (값·수식 불변 검증 통과)" % (len(changes), out))
    for x in changes[:20]:
        print("  ·", x)
    if len(changes) > 20:
        print("  · ... 외 %d건" % (len(changes) - 20))
    return 0


if __name__ == "__main__":
    sys.exit(main())
